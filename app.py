"""
Streamlit web app — Global Mining Stock Undervaluation Screener (market set via MARKET env)
Run: streamlit run app.py
"""
import os
import sys
import logging
import threading
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import plotly.io as pio
import streamlit as st

# Global chart style — light "research report" template, mobile-friendly:
# compact margins, horizontal legend on top, smaller base fonts.
_CHART_COLORWAY = [
    "#1a3a5c", "#2c6e9e", "#4a8ab5", "#e67e22",
    "#16a34a", "#dc2626", "#7c3aed", "#0891b2",
]
_tmpl = go.layout.Template(pio.templates["plotly_white"])
_tmpl.layout.update(
    colorway=_CHART_COLORWAY,
    font=dict(size=12),
    # automargin keeps the title clear of multi-row top legends
    title=dict(font=dict(size=15), x=0.0, xanchor="left",
               y=1.0, yanchor="top", automargin=True, pad=dict(t=8, b=8)),
    margin=dict(l=10, r=10, t=56, b=10),
    legend=dict(
        orientation="h",
        yanchor="bottom", y=1.02,
        xanchor="left", x=0.0,
        font=dict(size=11),
        title=dict(font=dict(size=11), side="left"),
    ),
    hoverlabel=dict(font_size=12),
    xaxis=dict(automargin=True, title=dict(font=dict(size=12))),
    yaxis=dict(automargin=True, title=dict(font=dict(size=12))),
)
pio.templates["miner_light"] = _tmpl
pio.templates.default = "miner_light"
px.defaults.color_discrete_sequence = _CHART_COLORWAY

# Ensure project root on path
sys.path.insert(0, os.path.dirname(__file__))


def _report_to_pdf(markdown_text: str, title: str = "Mining Report") -> bytes:
    """Convert a markdown report string to a PDF byte stream using reportlab."""
    import io
    import re
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_name = "Helvetica"
    # Only register font if not already registered (Streamlit reruns the script each interaction)
    if "CJK" not in pdfmetrics.getRegisteredFontNames():
        for fp in [
            r"C:\Windows\Fonts\msyh.ttc",
            r"C:\Windows\Fonts\simhei.ttf",
            r"C:\Windows\Fonts\simsun.ttc",
        ]:
            if os.path.exists(fp):
                try:
                    pdfmetrics.registerFont(TTFont("CJK", fp))
                    break
                except Exception:
                    pass
    if "CJK" in pdfmetrics.getRegisteredFontNames():
        font_name = "CJK"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
        title=title,
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontName=font_name, fontSize=16, spaceAfter=10, textColor=colors.HexColor("#1a3a5c"))
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], fontName=font_name, fontSize=13, spaceAfter=6,  textColor=colors.HexColor("#2c6e9e"))
    h3 = ParagraphStyle("H3", parent=styles["Heading3"], fontName=font_name, fontSize=11, spaceAfter=4,  textColor=colors.HexColor("#4a8ab5"))
    body = ParagraphStyle("Body",   parent=styles["Normal"], fontName=font_name, fontSize=9, spaceAfter=4, leading=14)
    bull = ParagraphStyle("Bullet", parent=body, leftIndent=12, spaceAfter=3)

    story = []
    for line in markdown_text.splitlines():
        line = re.sub(r"<!--.*?-->", "", line).strip()
        if not line:
            story.append(Spacer(1, 4))
            continue
        safe = line.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        safe = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", safe)
        safe = re.sub(r"\*(.+?)\*",     r"<i>\1</i>", safe)
        if line.startswith("# "):
            story.append(Paragraph(safe[2:], h1))
        elif line.startswith("## "):
            story.append(Paragraph(safe[3:], h2))
        elif line.startswith("### "):
            story.append(Paragraph(safe[4:], h3))
        elif line.startswith("---"):
            story.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey, spaceAfter=6))
        elif line.startswith("- ") or line.startswith("• "):
            story.append(Paragraph(f"• {safe[2:]}", bull))
        else:
            story.append(Paragraph(safe, body))

    doc.build(story)
    return buf.getvalue()

from data.universe import ALL_COMMODITIES, ALL_STAGES, get_ticker_meta
import data.snl_client as snl_client
from data.database import (
    init_db, load_latest, load_history, load_prev_scores,
    load_sector_trends, load_commodity_price_history,
    load_return_matrix, load_backtest_data,
    last_refresh,
    get_watchlist, add_to_watchlist, remove_from_watchlist,
    get_watchlist_note, update_watchlist_note,
    get_score_weights, save_score_weights,
    get_positions, upsert_position, delete_position,
    get_price_target, set_price_target, get_all_price_targets,
    save_filter_preset, load_filter_presets, delete_filter_preset,
    add_transaction, get_transactions, delete_transaction,
)
from scheduler.jobs import run_daily_refresh, start_scheduler
import config
from auth import check_auth, current_user, logout

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s %(levelname)s %(message)s")

# ── Authentication gate (must be called before any other st.* calls) ──
check_auth()


@st.cache_data(ttl=1800)   # cache news 30 minutes per ticker
def _fetch_news(ticker: str) -> list[dict]:
    """Fetch latest Yahoo Finance news for a ticker. Returns list of article dicts."""
    try:
        import yfinance as yf
        news = yf.Ticker(ticker).news or []
        return news[:8]   # cap at 8 headlines
    except Exception:
        return []


@st.cache_data(ttl=3600)   # cache benchmark prices 1 hour
def _fetch_benchmark(symbol: str, period: str = "1y") -> pd.DataFrame:
    """Fetch daily close prices for a benchmark ETF (GDX, GDXJ, etc.)."""
    try:
        import yfinance as yf
        hist = yf.Ticker(symbol).history(period=period, auto_adjust=True)
        if hist.empty:
            return pd.DataFrame()
        hist = hist[["Close"]].rename(columns={"Close": "price"})
        hist.index = pd.to_datetime(hist.index).tz_localize(None)
        hist.index.name = "date"
        return hist.reset_index()
    except Exception:
        return pd.DataFrame()


# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title=f"{config.MARKET_FLAG} {config.MARKET_NAME} Mining Screener",
    page_icon="⛏️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── One-time startup (scheduler + DB) ─────────────────────────────────────────
@st.cache_resource
def _startup():
    init_db()
    # On Streamlit Cloud the in-app scheduler is disabled (DISABLE_SCHEDULER=true):
    # the daily refresh runs in GitHub Actions instead, and cloud storage is
    # ephemeral so an in-container refresh would be lost on restart anyway.
    if os.getenv("DISABLE_SCHEDULER", "").lower() == "true":
        return None
    sched = start_scheduler()
    return sched

_startup()

# ── Load persisted score weights into session state (once per session) ─────────
if "weights_loaded" not in st.session_state:
    _sw = get_score_weights()
    st.session_state["w_val"] = _sw["valuation"]
    st.session_state["w_hlt"] = _sw["health"]
    st.session_state["w_mom"] = _sw["momentum"]
    st.session_state["w_min"] = _sw["mining"]
    st.session_state["w_com"] = _sw["commodity"]
    st.session_state["w_stg"] = _sw["stage"]
    st.session_state["weights_loaded"] = True
if "weight_gen" not in st.session_state:
    st.session_state["weight_gen"] = 0

# ── CSS — light "research report" design system ───────────────────────────────
# Palette matches the PDF report (#1a3a5c navy / #2c6e9e mid-blue).
st.markdown("""
<style>
:root {
    --navy:    #1a3a5c;
    --navy-2:  #2c6e9e;
    --ink:     #172033;
    --muted:   #5b6b7f;
    --line:    #e3e9f0;
    --bg-soft: #f4f7fa;
}

/* ── Global chrome ── */
#MainMenu, footer { visibility: hidden; }
header[data-testid="stHeader"] { background: transparent; }
h1 { color: var(--navy) !important; font-weight: 800; letter-spacing: -0.5px; }
h2, h3 { color: var(--navy) !important; }
hr { border-color: var(--line); }

/* ── Sidebar: light panel ── */
div[data-testid="stSidebar"] {
    background: var(--bg-soft);
    border-right: 1px solid var(--line);
}
div[data-testid="stSidebar"] [data-testid="stMetricValue"] {
    font-size: 22px; color: var(--navy); font-weight: 700;
}
div[data-testid="stSidebar"] [data-testid="stMetricLabel"] { color: var(--muted); }

/* ── Summary metric cards ── */
.metric-card {
    background: #ffffff; color: var(--ink);
    border: 1px solid var(--line); border-top: 3px solid var(--navy);
    border-radius: 10px; padding: 14px 10px; text-align: center;
    box-shadow: 0 1px 3px rgba(23,32,51,0.06);
}
.metric-card h3 { margin: 0; font-size: 26px; color: var(--navy); letter-spacing: -0.5px; }
.metric-card p  { margin: 4px 0 0; font-size: 12px; color: var(--muted); opacity: 1; }

/* ── Grade colours ── */
.grade-strong { color: #16a34a; font-weight: 700; }
.grade-buy    { color: var(--navy-2); font-weight: 700; }
.grade-watch  { color: #b45309; font-weight: 700; }
.grade-neutral{ color: #ea580c; }
.grade-avoid  { color: #dc2626; }

/* ── Tabs: research-report underline style, swipeable on small screens ── */
div[data-baseweb="tab-list"] {
    gap: 2px; border-bottom: 1px solid var(--line);
    overflow-x: auto; -webkit-overflow-scrolling: touch;
}
button[data-baseweb="tab"] { font-weight: 600; white-space: nowrap; }

/* ── Buttons & inputs ── */
div.stButton > button { border-radius: 8px; font-weight: 600; }

/* ── Dataframes ── */
div[data-testid="stDataFrame"] { font-size: 13px; }

/* ── Opportunity cards ── */
.opp-card {
    background-color: #ffffff !important;
    border: 1px solid var(--line) !important;
    border-left: 3px solid var(--navy) !important;
    border-radius: 10px !important;
    padding: 14px 16px !important;
    box-shadow: 0 1px 3px rgba(23,32,51,0.06) !important;
    color: var(--ink) !important;
}
.opp-card * { color: inherit; }
.opp-card .opp-name  { font-size: 16px; font-weight: 800; color: var(--navy) !important; margin: 8px 0 3px; }
.opp-card .opp-sub   { font-size: 12px; color: var(--muted) !important; font-weight: 500; }
.opp-card .opp-kpis  { font-size: 13px; color: var(--ink) !important; font-weight: 600; }
.opp-card .opp-foot  { font-size: 12px; color: var(--muted) !important; margin-top: 4px; }
.opp-card .opp-div   { height: 1px; background-color: var(--line) !important; margin: 10px 0; }

/* ── Mobile (≤640px): keep columns side-by-side as a 2-up grid ── */
@media (max-width: 640px) {
    .block-container { padding: 0.8rem 0.8rem 3rem !important; }
    h1 { font-size: 1.45rem !important; }
    div[data-testid="stHorizontalBlock"] {
        flex-direction: row !important;
        flex-wrap: wrap !important;
        gap: 8px !important;
    }
    div[data-testid="stHorizontalBlock"] > div[data-testid="stColumn"],
    div[data-testid="stHorizontalBlock"] > div[data-testid="column"] {
        flex: 1 1 calc(50% - 8px) !important;
        min-width: calc(50% - 8px) !important;
        width: auto !important;
    }
    .metric-card { padding: 10px 6px; }
    .metric-card h3 { font-size: 19px; }
    .metric-card p  { font-size: 11px; }
    div[data-testid="stDataFrame"] { font-size: 12px; }
    /* Summary cards: compact 3-up grid on phones */
    .st-key-summary_cards div[data-testid="stColumn"],
    .st-key-summary_cards div[data-testid="column"] {
        flex: 1 1 calc(33.3% - 8px) !important;
        min-width: calc(33.3% - 8px) !important;
    }
    /* Lists inside expanders read better stacked full-width */
    div[data-testid="stExpander"] div[data-testid="stColumn"],
    div[data-testid="stExpander"] div[data-testid="column"] {
        flex: 1 1 100% !important;
        min-width: 100% !important;
    }
}
</style>
""", unsafe_allow_html=True)

# ── Header ─────────────────────────────────────────────────────────────────────
col_title, col_refresh = st.columns([5, 1])
with col_title:
    st.markdown(f"# ⛏️ {config.MARKET_NAME} Mining Stock Screener")
    st.caption(f"{config.MARKET_EXCH} | All commodities | All stages | Last refresh: **{last_refresh()}**")
# On Streamlit Cloud (DISABLE_SCHEDULER=true) in-app refresh is disabled:
# Yahoo blocks the cloud IP, so a refresh there writes garbage into the
# ephemeral DB. Data is refreshed by the GitHub Actions workflow instead.
_IS_CLOUD = os.getenv("DISABLE_SCHEDULER", "").lower() == "true"

with col_refresh:
    st.markdown("<br>", unsafe_allow_html=True)
    if _IS_CLOUD:
        st.caption("🔄 Auto-refresh daily ~08:15 Sydney (GitHub Actions)")
    elif st.button("🔄 Refresh Now", width="stretch"):
        with st.spinner("Fetching live data… this takes ~2 min"):
            run_daily_refresh()
        st.cache_data.clear()   # invalidate _load() cache so new DB rows appear
        st.success("Refresh complete!")
        st.rerun()

st.divider()

# ── Load data ─────────────────────────────────────────────────────────────────
@st.cache_data(ttl=300)
def _load():
    return load_latest()


# ── SNL runtime overlay ────────────────────────────────────────────────────────
# Fills missing spg_aisc_per_oz / spg_cash_cost_oz / spg_production_oz
# from live Snowflake for tickers in our SNL mapping — never overwrites
# existing values, never writes to SQLite (trial §3.2.2 compliant).

@st.cache_data(ttl=3600, show_spinner=False)
def _snl_batch() -> dict:
    """Fetch SNL batch metrics; cached 1 h to avoid hammering Snowflake."""
    if not snl_client.is_configured():
        return {}
    try:
        return snl_client.get_batch_metrics("2024Y")
    except Exception:
        return {}


def _apply_snl_overlay(df: pd.DataFrame) -> pd.DataFrame:
    """
    For each row where a spg_ column is NaN but SNL has a value,
    fill it in from the live Snowflake batch.
    Priority commodity per ticker: Gold > Silver > Copper > other.
    """
    batch = _snl_batch()
    if not batch:
        return df

    _PRIORITY = ["Gold", "Silver", "Copper", "Zinc", "Nickel",
                 "Uranium", "PGM", "Iron Ore"]

    for i, row in df.iterrows():
        ticker = row.get("ticker", i)
        snl_key = snl_client.get_snl_key(str(ticker))
        if not snl_key or snl_key not in batch:
            continue

        comm_data = batch[snl_key]  # {commodity: row_dict}

        def _has_aisc(r: dict) -> bool:
            return any(
                r.get(k) is not None and float(r.get(k) or 0) > 0
                for k in ("AISC_OZ", "AISC_T", "AISC_LB")
            )

        # Prefer the DB commodity; fall back to best-covered commodity
        db_comm = str(row.get("commodity", "")).split("/")[0]
        db_row  = comm_data.get(db_comm)
        chosen  = (
            (db_row if db_row and _has_aisc(db_row) else None)
            or next((comm_data[c] for c in _PRIORITY if c in comm_data and _has_aisc(comm_data[c])), None)
            or comm_data.get(db_comm)
            or next((comm_data[c] for c in _PRIORITY if c in comm_data), None)
        )
        if not chosen:
            continue

        def _fill(col: str, snl_field: str) -> None:
            if col in df.columns and (pd.isna(df.at[i, col])):
                v = chosen.get(snl_field)
                if v is not None:
                    try:
                        df.at[i, col] = float(v)
                    except (TypeError, ValueError):
                        pass

        # oz-based (gold/silver/PGM)
        _fill("spg_aisc_per_oz",       "AISC_OZ")
        _fill("spg_cash_cost_oz",      "CASH_COST_OZ")
        _fill("spg_production_oz",     "PROD_OZ")
        _fill("spg_realized_price_oz", "REALIZED_PRICE_OZ")
        # tonne-based (copper/nickel/zinc)
        _fill("spg_aisc_per_t",        "AISC_T")
        _fill("spg_cash_cost_t",       "CASH_COST_T")
        _fill("spg_production_t",      "PROD_T")
        _fill("spg_realized_price_t",  "REALIZED_PRICE_T")
        # lb-based (uranium)
        _fill("spg_aisc_per_lb",       "AISC_LB")
        _fill("spg_cash_cost_lb",      "CASH_COST_LB")
        _fill("spg_production_lb",     "PROD_LB")
        _fill("spg_realized_price_lb", "REALIZED_PRICE_LB")

    # ── Local mine-life and global-rank from SQLite ────────────────────────────
    _RANK_PRIORITY = ["Gold", "Silver", "Copper", "Nickel", "Zinc",
                      "Uranium", "PGM", "Iron Ore"]
    try:
        mine_life_data   = snl_client.get_batch_mine_life_local()
        global_rank_data = snl_client.get_batch_global_rank_local()
    except Exception as _e:
        import logging as _log
        _log.getLogger(__name__).warning("Local SNL mine-life/rank query failed: %s", _e)
        mine_life_data   = {}
        global_rank_data = {}

    if "spg_mine_life" not in df.columns:
        df["spg_mine_life"] = float("nan")
    if "spg_global_rank" not in df.columns:
        df["spg_global_rank"] = float("nan")

    for i, row in df.iterrows():
        ticker  = row.get("ticker", i)
        snl_key = snl_client.get_snl_key(str(ticker))
        if not snl_key:
            continue

        if pd.isna(df.at[i, "spg_mine_life"]):
            ml = mine_life_data.get(snl_key)
            if ml and 0 < ml < 100:
                df.at[i, "spg_mine_life"] = ml

        if pd.isna(df.at[i, "spg_global_rank"]):
            ranks = global_rank_data.get(snl_key, {})
            if ranks:
                chosen_rank = next(
                    (ranks[c] for c in _RANK_PRIORITY if c in ranks), None
                ) or min(ranks.values())
                df.at[i, "spg_global_rank"] = chosen_rank

    return df


# ── SNL SQLite enrichment ──────────────────────────────────────────────────────
# Reads our local mining_screener.db (synced by snl_sync.py) and builds a
# per-ticker enrichment DataFrame: AISC, R&R oz/lb, in-situ value, global rank.

@st.cache_data(ttl=3600, show_spinner=False)
def _snl_sqlite_enrichment() -> pd.DataFrame:
    """Return a DataFrame (one row per ticker) of SNL metrics from local SQLite."""
    import sqlite3 as _sl3
    import json as _js2

    _db  = str(config.DB_PATH)
    _map = os.path.join(os.path.dirname(__file__), "_asx_snl_ticker_mapping.json")
    if not os.path.exists(_db):
        return pd.DataFrame()
    try:
        with open(_map) as _f:
            _mapping = _js2.load(_f)
    except Exception:
        return pd.DataFrame()

    _key2ticker = {str(v["snl_key"]): k for k, v in _mapping.items()}
    _PRIO = ["Gold", "Silver", "Platinum", "Palladium", "PGM",
             "Copper", "Nickel", "Zinc", "Uranium", "Iron Ore"]

    def _best(group):
        if group.empty:
            return None
        sub = group[group["period"] == group["period"].max()]
        for _c in _PRIO:
            _r = sub[sub["commodity"] == _c]
            if not _r.empty:
                return _r.iloc[0]
        return sub.iloc[0]

    try:
        _conn = _sl3.connect(_db)

        _prod = pd.read_sql_query(
            "SELECT snl_key,period,commodity,prod_oz,prod_t,prod_lb,"
            "aisc_oz,aisc_t,aisc_lb,cash_cost_oz,cash_cost_t,cash_cost_lb,"
            "realized_price_oz,realized_price_t,realized_price_lb,revenue_m "
            "FROM snl_company_production WHERE period>='2022Y'", _conn)
        _prod["snl_key"] = _prod["snl_key"].astype(str)

        _rr_prec = pd.read_sql_query(
            "SELECT snl_key,period,commodity,grade_gpt,"
            "contained_reserves_oz,contained_rr_oz "
            "FROM snl_company_rr WHERE period>='2022Y' "
            "AND commodity IN ('Gold','Silver','Platinum','Palladium','PGM')", _conn)
        _rr_prec["snl_key"] = _rr_prec["snl_key"].astype(str)

        _rr_base = pd.read_sql_query(
            "SELECT snl_key,period,commodity,grade_pct,"
            "contained_reserves_lb,contained_rr_lb "
            "FROM snl_company_rr WHERE period>='2022Y' "
            "AND commodity NOT IN ('Gold','Silver','Platinum','Palladium','PGM')", _conn)
        _rr_base["snl_key"] = _rr_base["snl_key"].astype(str)

        _ins = pd.read_sql_query(
            "SELECT i.snl_key,i.insitu_reserves_m,i.insitu_rr_m "
            "FROM snl_company_insitu i "
            "INNER JOIN (SELECT snl_key,MAX(period) mp FROM snl_company_insitu "
            "GROUP BY snl_key) m ON i.snl_key=m.snl_key AND i.period=m.mp", _conn)
        _ins["snl_key"] = _ins["snl_key"].astype(str)

        _rnk = pd.read_sql_query(
            "SELECT r.snl_key,r.global_rank "
            "FROM snl_company_ranking r "
            "INNER JOIN (SELECT snl_key,MAX(period) mp FROM snl_company_ranking "
            "WHERE commodity='Gold' AND ownership_method='Controlled' "
            "GROUP BY snl_key) m ON r.snl_key=m.snl_key AND r.period=m.mp "
            "WHERE r.commodity='Gold' AND r.ownership_method='Controlled'", _conn)
        _rnk["snl_key"] = _rnk["snl_key"].astype(str)

        # ── Projections: forward production / AISC guidance ───────────────────
        # Use the two most recent estimate periods ≥ current year as "forward".
        _cur_yr = pd.Timestamp.utcnow().year
        _proj = pd.read_sql_query(
            "SELECT snl_key, estimate_period, description, "
            "prod_high_oz, prod_low_oz, prod_high_t, prod_low_t, "
            "prod_high_lb, prod_low_lb, "
            "aisc_high_oz, aisc_low_oz, aisc_high_t, aisc_low_t "
            "FROM snl_company_projections "
            f"WHERE estimate_period >= '{_cur_yr - 1}Y' "
            "ORDER BY snl_key, estimate_period", _conn)
        _proj["snl_key"] = _proj["snl_key"].astype(str)

        # ── Property studies: best IRR / NPV per company ──────────────────────
        # Rank: FS > PFS > PEA > Mine Plan. Pick highest-ranked most-recent study.
        _study_rank = {"Full Feasibility": 4, "Feasibility": 4,
                       "Prefeasibility": 3, "Preliminary Economic Assessment": 2,
                       "Mine Plan": 1}
        _studies_raw = pd.read_sql_query(
            "SELECT o.snl_key, s.property_name, s.study_type, s.study_year, "
            "s.posttax_irr_pct, s.posttax_npv_m, s.initial_capex_m, s.mine_life_yrs "
            "FROM snl_property_studies s "
            "JOIN snl_property_owner o ON s.property_id = o.property_id "
            "WHERE s.posttax_irr_pct IS NOT NULL OR s.posttax_npv_m IS NOT NULL "
            "ORDER BY o.snl_key, s.study_year DESC", _conn)
        _studies_raw["snl_key"] = _studies_raw["snl_key"].astype(str)
        _studies_raw["_rank"] = _studies_raw["study_type"].map(
            lambda t: _study_rank.get(t, 0))

        _conn.close()

    except Exception as _ex:
        logging.warning(f"SNL SQLite read: {_ex}")
        return pd.DataFrame()

    _rows = []
    for _sk, _tk in _key2ticker.items():
        _r = {"ticker": _tk}

        _bp = _best(_prod[_prod["snl_key"] == _sk])
        if _bp is not None:
            _r.update({
                "snl_aisc_oz": _bp.get("aisc_oz"),   "snl_aisc_t": _bp.get("aisc_t"),
                "snl_aisc_lb": _bp.get("aisc_lb"),   "snl_cc_oz":  _bp.get("cash_cost_oz"),
                "snl_cc_t":    _bp.get("cash_cost_t"),"snl_cc_lb":  _bp.get("cash_cost_lb"),
                "snl_prod_oz": _bp.get("prod_oz"),    "snl_prod_t": _bp.get("prod_t"),
                "snl_prod_lb": _bp.get("prod_lb"),    "snl_rp_oz":  _bp.get("realized_price_oz"),
                "snl_revenue_m": _bp.get("revenue_m"),
            })

        _bprec = _best(_rr_prec[_rr_prec["snl_key"] == _sk])
        if _bprec is not None:
            _r.update({
                "snl_grade_gpt": _bprec.get("grade_gpt"),
                "snl_rsv_oz":    _bprec.get("contained_reserves_oz"),
                "snl_rr_oz":     _bprec.get("contained_rr_oz"),
            })

        _bbase = _best(_rr_base[_rr_base["snl_key"] == _sk])
        if _bbase is not None:
            _r.update({
                "snl_grade_pct": _bbase.get("grade_pct"),
                "snl_rsv_lb":    _bbase.get("contained_reserves_lb"),
                "snl_rr_lb":     _bbase.get("contained_rr_lb"),
            })

        _ig = _ins[_ins["snl_key"] == _sk]
        if not _ig.empty:
            _r["snl_insitu_rr_m"]  = _ig.iloc[0]["insitu_rr_m"]
            _r["snl_insitu_rsv_m"] = _ig.iloc[0]["insitu_reserves_m"]

        _rkg = _rnk[_rnk["snl_key"] == _sk]
        if not _rkg.empty:
            _r["snl_global_rank"] = _rkg.iloc[0]["global_rank"]

        # ── Projections: forward vs actual production growth ───────────────────
        _pk = _proj[_proj["snl_key"] == _sk].copy()
        if not _pk.empty:
            # Most recent forward estimate period
            _latest_fwd = _pk["estimate_period"].max()
            _fwd_row = _pk[_pk["estimate_period"] == _latest_fwd]
            # Prefer prod_oz description; fall back to any row
            _fwd_prod = _fwd_row[
                _fwd_row["description"].str.contains("Production|production", na=False)
            ]
            if _fwd_prod.empty:
                _fwd_prod = _fwd_row
            _fr = _fwd_prod.iloc[0]
            # Forward production midpoint (oz)
            _fhi = _fr["prod_high_oz"]
            _flo = _fr["prod_low_oz"]
            if _fhi and _flo and float(_fhi) > 1000 and float(_flo) > 1000:
                _r["snl_fwd_prod_oz"] = (float(_fhi) + float(_flo)) / 2
            # Forward AISC midpoint ($/oz)
            _ahi = _fr["aisc_high_oz"]
            _alo = _fr["aisc_low_oz"]
            if _ahi and _alo and float(_ahi) > 0:
                _r["snl_fwd_aisc_oz"] = (float(_ahi) + float(_alo)) / 2

        # Production growth %: forward vs latest actual (snl_prod_oz already set)
        # Sanity guard: only compute when both values are in a plausible oz range
        # (1 koz – 15 Moz) and the ratio is reasonable (0.1× – 4×).
        # This filters unit-mislabelled rows (copper/nickel reported in lb as prod_oz).
        _actual = _r.get("snl_prod_oz")
        _fwd    = _r.get("snl_fwd_prod_oz")
        if (_actual and _fwd
                and float(_actual) > 1_000
                and float(_actual) < 15_000_000
                and float(_fwd) > 1_000
                and float(_fwd) < 15_000_000):
            _ratio = float(_fwd) / float(_actual)
            if 0.1 <= _ratio <= 4.0:
                _r["snl_prod_growth_pct"] = round(
                    (_ratio - 1.0) * 100, 1)

        # ── Best study IRR per company ─────────────────────────────────────────
        _stk = _studies_raw[_studies_raw["snl_key"] == _sk]
        if not _stk.empty:
            # Pick highest study_rank, then most recent year
            _stk_sorted = _stk.sort_values(["_rank", "study_year"], ascending=[False, False])
            _best_study = _stk_sorted.iloc[0]
            _irr = _best_study.get("posttax_irr_pct")
            _npv = _best_study.get("posttax_npv_m")
            if _irr is not None:
                _r["snl_best_irr"]        = float(_irr)
                _r["snl_best_irr_type"]   = _best_study.get("study_type", "")
                _r["snl_best_irr_year"]   = int(_best_study.get("study_year") or 0)
                _r["snl_best_irr_prop"]   = _best_study.get("property_name", "")
            if _npv is not None:
                _r["snl_best_npv_m"]      = float(_npv)

        _rows.append(_r)

    return pd.DataFrame(_rows)


def _apply_sqlite_snl(df: pd.DataFrame) -> pd.DataFrame:
    """Merge SQLite SNL enrichment into df, fill spg_ gaps, add derived metrics."""
    _enrich = _snl_sqlite_enrichment()
    if _enrich.empty:
        return df

    df = df.merge(_enrich, on="ticker", how="left")

    # Fill existing spg_ gaps where Snowflake overlay left NaN
    for _spg, _snl in [
        ("spg_aisc_per_oz",          "snl_aisc_oz"),
        ("spg_aisc_per_t",           "snl_aisc_t"),
        ("spg_aisc_per_lb",          "snl_aisc_lb"),
        ("spg_cash_cost_oz",         "snl_cc_oz"),
        ("spg_cash_cost_t",          "snl_cc_t"),
        ("spg_cash_cost_lb",         "snl_cc_lb"),
        ("spg_production_oz",        "snl_prod_oz"),
        ("spg_production_t",         "snl_prod_t"),
        ("spg_production_lb",        "snl_prod_lb"),
        ("spg_realized_price_oz",    "snl_rp_oz"),
        ("spg_grade_gpt",            "snl_grade_gpt"),
        ("spg_grade_pct",            "snl_grade_pct"),
        ("spg_contained_reserves_oz","snl_rsv_oz"),
        ("spg_contained_reserves_lb","snl_rsv_lb"),
    ]:
        if _spg in df.columns and _snl in df.columns:
            _mask = df[_spg].isna() & df[_snl].notna()
            df.loc[_mask, _spg] = pd.to_numeric(df.loc[_mask, _snl], errors="coerce")

    # ── Derived metrics ────────────────────────────────────────────────────────
    _ev = pd.to_numeric(df.get("enterprise_value", pd.Series(dtype=float)), errors="coerce")
    _mc = pd.to_numeric(df.get("market_cap",       pd.Series(dtype=float)), errors="coerce")

    if "snl_rr_oz" in df.columns:
        _rr = pd.to_numeric(df["snl_rr_oz"], errors="coerce")
        df["snl_ev_per_oz_rr"]  = np.where(
            _rr.notna() & (_rr > 0) & _ev.notna() & (_ev > 0),
            (_ev / _rr).round(0), np.nan)
        df["snl_rr_koz"]  = _rr.div(1000).round(0)

    if "snl_rsv_oz" in df.columns:
        _rsv = pd.to_numeric(df["snl_rsv_oz"], errors="coerce")
        df["snl_ev_per_oz_rsv"] = np.where(
            _rsv.notna() & (_rsv > 0) & _ev.notna() & (_ev > 0),
            (_ev / _rsv).round(0), np.nan)
        df["snl_rsv_koz"] = _rsv.div(1000).round(0)

    if "snl_insitu_rr_m" in df.columns:
        _iv = pd.to_numeric(df["snl_insitu_rr_m"], errors="coerce")
        df["snl_p_insitu"] = np.where(
            _iv.notna() & (_iv > 0) & _mc.notna() & (_mc > 0),
            (_mc / (_iv * 1e6)).round(3), np.nan)

    if "snl_rr_lb" in df.columns:
        df["snl_rr_mlb"] = pd.to_numeric(df["snl_rr_lb"], errors="coerce").div(1e6).round(1)

    # ── Forward production guidance: fill spg_production_oz gap with fwd estimate
    # Only fill if actual is missing AND forward looks reasonable (>0, <10M oz to filter noise)
    if "snl_fwd_prod_oz" in df.columns and "spg_production_oz" in df.columns:
        _fwd = pd.to_numeric(df["snl_fwd_prod_oz"], errors="coerce")
        _fwd_valid = _fwd.notna() & (_fwd > 5000) & (_fwd < 10_000_000)
        _miss = df["spg_production_oz"].isna()
        df.loc[_miss & _fwd_valid, "spg_production_oz"] = _fwd[_miss & _fwd_valid]

    # ── snl_insitu_rsv_m → spg_reserves_m gap fill (in-situ reserves value in $M)
    if "snl_insitu_rsv_m" in df.columns and "spg_reserves_m" in df.columns:
        _irsv = pd.to_numeric(df["snl_insitu_rsv_m"], errors="coerce")
        _miss2 = df["spg_reserves_m"].isna()
        df.loc[_miss2 & _irsv.notna(), "spg_reserves_m"] = _irsv[_miss2 & _irsv.notna()]

    # Ensure scorer-facing columns exist
    for _col in ("snl_prod_growth_pct", "snl_best_irr", "snl_p_insitu",
                 "snl_ev_per_oz_rr", "snl_fwd_aisc_oz"):
        if _col not in df.columns:
            df[_col] = np.nan
        df[_col] = pd.to_numeric(df[_col], errors="coerce")

    return df


df = _load()

if df.empty:
    st.info(f"⏳ First data build for {config.MARKET_NAME} is running in GitHub Actions "
            "(up to ~2 h for large markets). The app updates automatically when it lands — "
            "check back soon.")
    st.stop()

# ── SNL overlay: fill missing spg_ fields from live Snowflake
df = _apply_snl_overlay(df)

# ── SNL SQLite overlay: enrich with locally-cached SNL data + derived metrics
df = _apply_sqlite_snl(df)

# Merge previous-snapshot scores → score_delta + grade_prev
_prev = load_prev_scores()
if not _prev.empty:
    df = df.merge(_prev, on="ticker", how="left")
    df["score_delta"] = (df["score_composite"] - df["score_prev"]).round(1)
else:
    df["score_delta"] = float("nan")
    df["grade_prev"]  = None

# Ensure numeric columns
_num_cols = [
    "score_composite", "score_valuation", "score_health",
    "score_momentum", "score_mining", "score_commodity", "score_stage",
    "market_cap", "price", "price_to_book", "ev_ebitda",
    "ev_revenue", "p_cf", "debt_to_equity", "current_ratio",
    "cash_pct_mcap", "net_debt_m", "rsi",
    "wk52_position", "pct_from_52hi",
    "spg_p_nav", "spg_reserves_m", "spg_resources_m",
    "spg_aisc_per_oz", "spg_aisc_per_t", "spg_aisc_per_lb", "spg_aisc_margin",
    "spg_grade_gpt", "spg_grade_pct",
    "spg_cash_cost_oz", "spg_cash_cost_t", "spg_cash_cost_lb",
    "spg_production_oz", "spg_production_t", "spg_production_lb",
    "spg_realized_price_oz", "spg_realized_price_t", "spg_realized_price_lb",
    "spg_contained_reserves_oz", "spg_contained_reserves_lb",
    "spg_reserve_life",
    "dividend_yield", "return_1m", "return_3m",
    "analyst_target_mean", "analyst_count", "analyst_rec_mean", "analyst_upside",
    "univ_rank",
    "return_on_equity", "operating_margins", "gross_margins",
    "spg_mine_life", "spg_global_rank",
    # SNL SQLite-derived metrics
    "snl_ev_per_oz_rr", "snl_ev_per_oz_rsv", "snl_p_insitu",
    "snl_rr_koz", "snl_rsv_koz", "snl_rr_mlb",
    "snl_global_rank", "snl_revenue_m", "snl_insitu_rr_m",
    "snl_aisc_oz", "snl_aisc_t", "snl_aisc_lb",
]
for c in _num_cols:
    if c in df.columns:
        df[c] = pd.to_numeric(df[c], errors="coerce")

# S&P data coverage flag: True when at least one SPG field is present
_spg_flag_cols = ["spg_p_nav", "spg_aisc_per_oz", "spg_aisc_per_t",
                  "spg_aisc_per_lb", "spg_reserves_m"]
df["has_spg"] = df[[c for c in _spg_flag_cols if c in df.columns]].notna().any(axis=1)

# FCF Yield (%) = Free Cash Flow / Market Cap × 100  — only for positive FCF
if "free_cf" in df.columns and "market_cap" in df.columns:
    _fcf_mask = df["free_cf"].notna() & df["market_cap"].notna() & (df["market_cap"] > 0) & (df["free_cf"] > 0)
    df["fcf_yield"] = np.where(_fcf_mask, (df["free_cf"] / df["market_cap"] * 100).round(1), np.nan)
else:
    df["fcf_yield"] = np.nan

# Primary Grade display column — picks g/t for precious metals, % for base metals
# Shows whichever value is available (gpt takes precedence for gold/silver companies)
if "spg_grade_gpt" in df.columns or "spg_grade_pct" in df.columns:
    _gpt = df.get("spg_grade_gpt", pd.Series(np.nan, index=df.index))
    _pct = df.get("spg_grade_pct", pd.Series(np.nan, index=df.index))
    df["grade_primary"] = np.where(_gpt.notna(), _gpt, _pct)
    # unit column: "g/t" when gpt is used, "%" when pct is used, None when both missing
    df["grade_unit"] = np.where(_gpt.notna(), "g/t", np.where(_pct.notna(), "%", None))
    # Pre-formatted string for screener table display (unit appended)
    df["grade_display"] = df.apply(
        lambda r: f"{r['grade_primary']:.3f} g/t" if r.get("grade_unit") == "g/t" and pd.notna(r.get("grade_primary"))
                  else (f"{r['grade_primary']:.3f}%" if r.get("grade_unit") == "%" and pd.notna(r.get("grade_primary"))
                  else "—"),
        axis=1,
    )
else:
    df["grade_primary"] = np.nan
    df["grade_unit"] = None
    df["grade_display"] = "—"

# EV/Reserves — Enterprise Value as a multiple of in-situ reserves value
# Ratio < 1 means you're buying the enterprise at a discount to in-ground assets
if "enterprise_value" in df.columns and "spg_reserves_m" in df.columns:
    _er_mask = (df["enterprise_value"].notna() & df["spg_reserves_m"].notna()
                & (df["spg_reserves_m"] > 0) & (df["enterprise_value"] > 0))
    df["ev_reserves"] = np.where(
        _er_mask,
        (df["enterprise_value"] / (df["spg_reserves_m"] * 1e6)).round(2),
        np.nan,
    )
else:
    df["ev_reserves"] = np.nan

# ── EV / oz Production  ($/oz) ─────────────────────────────────────────────
# EV (raw $) / production (oz) — enterprise_value is stored in raw dollars by Yahoo.
# Gold benchmark: ~$2,000–15,000/oz for mid-tier; majors ~$30,000–50,000/oz.
if "enterprise_value" in df.columns and "spg_production_oz" in df.columns:
    _ep_mask = (df["enterprise_value"].notna() & df["spg_production_oz"].notna()
                & (df["spg_production_oz"] > 0) & (df["enterprise_value"] > 0))
    df["ev_per_oz_prod"] = np.where(
        _ep_mask,
        (df["enterprise_value"] / df["spg_production_oz"]).round(0),
        np.nan,
    )
else:
    df["ev_per_oz_prod"] = np.nan

# ── EV / oz Reserve  ($/oz) ────────────────────────────────────────────────
# EV per oz in the ground. Gold benchmark: ~$100–500/oz reserve.
if "enterprise_value" in df.columns and "spg_contained_reserves_oz" in df.columns:
    _eor_mask = (df["enterprise_value"].notna() & df["spg_contained_reserves_oz"].notna()
                 & (df["spg_contained_reserves_oz"] > 0) & (df["enterprise_value"] > 0))
    df["ev_per_oz_reserve"] = np.where(
        _eor_mask,
        (df["enterprise_value"] / df["spg_contained_reserves_oz"]).round(0),
        np.nan,
    )
else:
    df["ev_per_oz_reserve"] = np.nan

# ── EV / lb Production  ($/lb) — copper & uranium ─────────────────────────
if "enterprise_value" in df.columns and "spg_production_lb" in df.columns:
    _eplb_mask = (df["enterprise_value"].notna() & df["spg_production_lb"].notna()
                  & (df["spg_production_lb"] > 0) & (df["enterprise_value"] > 0))
    df["ev_per_lb_prod"] = np.where(
        _eplb_mask,
        (df["enterprise_value"] / df["spg_production_lb"]).round(4),
        np.nan,
    )
else:
    df["ev_per_lb_prod"] = np.nan

# ── EV / t Production  ($/t) — iron ore & base metals ─────────────────────
if "enterprise_value" in df.columns and "spg_production_t" in df.columns:
    _ept_mask = (df["enterprise_value"].notna() & df["spg_production_t"].notna()
                 & (df["spg_production_t"] > 0) & (df["enterprise_value"] > 0))
    df["ev_per_t_prod"] = np.where(
        _ept_mask,
        (df["enterprise_value"] / df["spg_production_t"]).round(2),
        np.nan,
    )
else:
    df["ev_per_t_prod"] = np.nan

# ── EV / lb Reserve  ($/lb) — uranium & copper ─────────────────────────────
if "enterprise_value" in df.columns and "spg_contained_reserves_lb" in df.columns:
    _elb_mask = (df["enterprise_value"].notna() & df["spg_contained_reserves_lb"].notna()
                 & (df["spg_contained_reserves_lb"] > 0) & (df["enterprise_value"] > 0))
    df["ev_per_lb_reserve"] = np.where(
        _elb_mask,
        (df["enterprise_value"] / df["spg_contained_reserves_lb"]).round(4),
        np.nan,
    )
else:
    df["ev_per_lb_reserve"] = np.nan

# ── Realized price premium/discount vs spot (%) ────────────────────────────
# Positive = realized above spot (premium); negative = hedged below spot
def _realized_vs_spot(row):
    comm = str(row.get("commodity", "")).lower()
    spots = row.get("_spots", {}) if hasattr(row, "get") else {}
    if comm in ("gold", "silver") and pd.notna(row.get("spg_realized_price_oz")):
        spot = spots.get("Gold") if comm == "gold" else spots.get("Silver")
        if spot and spot > 0:
            return round((row["spg_realized_price_oz"] / spot - 1) * 100, 1)
    elif comm == "uranium" and pd.notna(row.get("spg_realized_price_lb")):
        spot = spots.get("Uranium")
        if spot and spot > 0:
            return round((row["spg_realized_price_lb"] / spot - 1) * 100, 1)
    return np.nan

# Production display — koz for precious metals, kt for base metals, Mlb for uranium
def _production_display(row):
    if pd.notna(row.get("spg_production_oz")):
        return f"{row['spg_production_oz'] / 1000:.0f} koz"
    if pd.notna(row.get("spg_production_t")):
        return f"{row['spg_production_t'] / 1000:.0f} kt"
    if pd.notna(row.get("spg_production_lb")):
        return f"{row['spg_production_lb'] / 1e6:.2f} Mlb"
    return "—"

df["production_display"] = df.apply(_production_display, axis=1)

# Reserve life display
df["reserve_life_display"] = df["spg_reserve_life"].apply(
    lambda x: f"{x:.1f} yr" if pd.notna(x) and x > 0 else "—"
) if "spg_reserve_life" in df.columns else "—"

# ── Analyst consensus upside (%) ───────────────────────────────────────────
# Upside to consensus mean price target: (target / price − 1) × 100
# Requires both analyst_target_mean and current price to be available.
if "analyst_target_mean" in df.columns and "price" in df.columns:
    _at_mask = (df["analyst_target_mean"].notna() & df["price"].notna()
                & (df["price"] > 0) & (df["analyst_count"].fillna(0) >= 1))
    df["analyst_upside"] = np.where(
        _at_mask,
        np.clip((df["analyst_target_mean"] / df["price"] - 1) * 100, -200, 500).round(1),
        np.nan,
    )
else:
    df["analyst_upside"] = np.nan

# Upside to NAV (%)
# Primary: S&P P/NAV ratio  → upside = (1/P_NAV - 1) × 100
# Fallback: Brokerage EV/NAV → same formula (approximate when EV ≈ mkt cap)
# Only show when trading at a discount (positive upside); clamp to ±300%
if "spg_p_nav" in df.columns:
    _pnav_mask = df["spg_p_nav"].notna() & (df["spg_p_nav"] > 0.05)
    _upside_spg = np.where(
        _pnav_mask,
        ((1 / df["spg_p_nav"]) - 1) * 100,
        np.nan,
    )
else:
    _upside_spg = np.full(len(df), np.nan)

df["upside_to_nav"] = np.where(
    np.isfinite(_upside_spg),
    np.clip(_upside_spg, -300, 300).round(1),
    np.nan,
)
df["nav_source"] = np.where(np.isfinite(_upside_spg), "S&P", None)

# Peer P/B relative upside — works without S&P data
# "If this company re-rated to its commodity peers' median P/B, what's the gain?"
# Uses primary commodity as peer group. Positive = trading at a discount to peers.
if "price_to_book" in df.columns:
    df["_primary_comm"] = df["commodity"].str.split("/").str[0].str.strip()
    _peer_pb_median = (
        df[df["price_to_book"] > 0]
        .groupby("_primary_comm")["price_to_book"]
        .transform("median")
    )
    _pb_upside_mask = (
        df["price_to_book"].notna() &
        _peer_pb_median.notna() &
        (df["price_to_book"] > 0) &
        (_peer_pb_median > 0)
    )
    df["pb_peer_upside"] = np.where(
        _pb_upside_mask,
        np.clip(((_peer_pb_median / df["price_to_book"]) - 1) * 100, -300, 300).round(1),
        np.nan,
    )
    df.drop(columns=["_primary_comm"], inplace=True)
else:
    df["pb_peer_upside"] = np.nan

# ── Universe rank & peer rank ───────────────────────────────────────────────────
# Rank within the full universe (1 = highest score). Used as a quick context
# metric in the screener table — lower rank = better opportunity in the universe.
df["univ_rank"] = df["score_composite"].rank(ascending=False, method="min").astype(int)
df["univ_n"]    = len(df)

# Peer-group ranking — recompute if not persisted in DB (older snapshots)
# peer_group / peer_rank / peer_n / peer_pct are written by scorer.compute_scores()
# and saved via upsert_snapshot(); fall back to live computation when absent.
if "peer_rank" not in df.columns or df["peer_rank"].isna().all():
    def _sb(stage: str) -> str:
        s = str(stage)
        if "Royalty"  in s: return "Royalty"
        if "Producer" in s: return "Producer"
        if "Developer"in s: return "Developer"
        if "Explorer" in s: return "Explorer"
        return "Other"
    _comm_p = df["commodity"].str.split("/").str[0].str.strip()
    _stage_b = df["stage"].apply(_sb)
    _peer_fine = _comm_p + " · " + _stage_b
    _fine_sizes = _peer_fine.map(_peer_fine.value_counts())
    df["peer_group"] = np.where(_fine_sizes >= 3, _peer_fine, _stage_b)
    df.loc[_stage_b == "Royalty", "peer_group"] = "Royalty"
    df["peer_n"]    = df.groupby("peer_group")["score_composite"].transform("count").astype(int)
    df["peer_rank"] = df.groupby("peer_group")["score_composite"].rank(ascending=False, method="min").astype(int)
    df["peer_pct"]  = df.groupby("peer_group")["score_composite"].rank(ascending=True, pct=True).mul(100).round(0).astype(int)

# Formatted display column: "3 / 9  Gold·Producer"
df["peer_rank_display"] = df.apply(
    lambda r: f"{int(r['peer_rank'])} / {int(r['peer_n'])}" if pd.notna(r.get("peer_rank")) else "—",
    axis=1,
)

# ── Sidebar filters ────────────────────────────────────────────────────────────
with st.sidebar:
    # ── Global Quick Search ────────────────────────────────────────────────────
    st.markdown("## 🔍 Quick Lookup")
    _qs_options = [""] + [
        f"{r['name']}  ({r['ticker']})"
        for _, r in df.sort_values("name").iterrows()
    ]
    _qs_sel = st.selectbox(
        "Search company", _qs_options,
        format_func=lambda x: "Type to search…" if x == "" else x,
        key="qs_company",
        label_visibility="collapsed",
    )
    if _qs_sel:
        _qs_tk = _qs_sel.split("(")[-1].rstrip(")")
        _qs_row = df[df["ticker"] == _qs_tk]
        if not _qs_row.empty:
            _qs = _qs_row.iloc[0]
            _qs_grade  = _qs.get("grade", "—")
            _qs_score  = _qs.get("score_composite")
            _qs_price  = _qs.get("price")
            _qs_delta  = _qs.get("score_delta")
            _qs_upside = _qs.get("upside_to_nav") or _qs.get("pb_peer_upside")
            _qs_an_up  = _qs.get("analyst_upside")
            _qs_an_ct  = _qs.get("analyst_count")
            _qs_comm   = _qs.get("commodity", "—")
            _qs_stage  = _qs.get("stage", "—")
            _qs_mcap   = _qs.get("market_cap")

            _qs_score_str = f"{_qs_score:.1f}/100" if pd.notna(_qs_score) else "—"
            _qs_price_str = f"${_qs_price:.3f}" if pd.notna(_qs_price) else "—"
            _qs_delta_str = (f"  {_qs_delta:+.1f}" if pd.notna(_qs_delta) else "")
            _qs_mcap_str  = (f"${_qs_mcap/1e9:.2f}B" if pd.notna(_qs_mcap) and _qs_mcap >= 1e9
                             else f"${_qs_mcap/1e6:.0f}M" if pd.notna(_qs_mcap) else "—")
            _qs_peer_rank = _qs.get("peer_rank")
            _qs_peer_n    = _qs.get("peer_n")
            _qs_peer_grp  = _qs.get("peer_group", "")
            _qs_peer_str  = (f"#{int(_qs_peer_rank)} of {int(_qs_peer_n)} in {_qs_peer_grp}"
                             if pd.notna(_qs_peer_rank) and pd.notna(_qs_peer_n) else "")

            # Grade colour
            _qs_grade_color = {
                "🟢 Strong Buy": "#16a34a", "🔵 Buy": "#2c6e9e",
                "🟡 Watch": "#b45309", "🟠 Neutral": "#ea580c", "🔴 Avoid": "#dc2626",
            }.get(_qs_grade, "#5b6b7f")

            st.markdown(
                f"<div style='background:#ffffff;border:1px solid #e3e9f0;"
                f"border-left:3px solid #1a3a5c;border-radius:10px;padding:12px 14px;"
                f"margin-bottom:8px'>"
                f"<div style='font-size:13px;font-weight:700;color:{_qs_grade_color}'>"
                f"{_qs_grade}</div>"
                f"<div style='font-size:15px;font-weight:800;color:#1a3a5c;margin:2px 0'>"
                f"{_qs.get('name', _qs_tk)}</div>"
                f"<div style='font-size:11px;color:#5b6b7f'>{_qs_comm} · {_qs_stage}</div>"
                f"<hr style='border-color:#e3e9f0;margin:8px 0'>"
                f"<div style='display:flex;justify-content:space-between;"
                f"font-size:13px;color:#172033'>"
                f"<span>Score <b>{_qs_score_str}</b>{_qs_delta_str}</span>"
                f"<span>Price <b>{_qs_price_str}</b></span>"
                f"</div>"
                f"<div style='display:flex;justify-content:space-between;"
                f"font-size:12px;color:#5b6b7f;margin-top:4px'>"
                f"<span>Mkt Cap {_qs_mcap_str}</span>"
                + (f"<span style='color:#16a34a'>↑ {_qs_upside:+.0f}% upside</span>"
                   if pd.notna(_qs_upside) and _qs_upside > 0
                   else f"<span style='color:#dc2626'>{_qs_upside:+.0f}% vs NAV</span>"
                   if pd.notna(_qs_upside) else "<span></span>") +
                f"</div>"
                + (f"<div style='font-size:11px;color:#5b6b7f;margin-top:3px'>"
                   f"Analyst target: <span style='color:#2c6e9e'>{_qs_an_up:+.0f}%</span>"
                   f" ({int(_qs_an_ct)} analysts)</div>"
                   if pd.notna(_qs_an_up) and pd.notna(_qs_an_ct) and _qs_an_ct >= 1 else "")
                + (f"<div style='font-size:11px;color:#b45309;margin-top:3px'>"
                   f"⚡ {_qs_peer_str}</div>"
                   if _qs_peer_str else "")
                + f"</div>",
                unsafe_allow_html=True,
            )
            # Pre-select in Company Detail tab via session state
            st.session_state["qs_jump_name"] = _qs_sel.strip()
    st.divider()

    # ── Live spot prices ───────────────────────────────────────────────────────
    st.markdown("## 📈 Spot Prices")
    _spot_display = [
        ("Gold",    "USD/oz",    "🥇"),
        ("Silver",  "USD/oz",    "🥈"),
        ("Copper",  "USD/t",     "🔶"),
        ("Zinc",    "USD/t",     "⬜"),
        ("Nickel",  "USD/t",     "🔩"),
        ("Uranium", "USD/lb",    "⚛️"),
    ]
    # Latest stored prices (written by the daily refresh); config values are
    # only the last-resort fallback and can be months stale.
    _spot_db: dict = {}
    _spot_date = None
    try:
        _sph = load_commodity_price_history(days=14)
        if not _sph.empty:
            _sp_latest = _sph.sort_values("price_date").groupby("commodity").last()
            _spot_db = _sp_latest["price"].to_dict()
            _spot_date = _sph["price_date"].max()
    except Exception:
        pass
    _spot_cols = st.columns(2)
    for i, (comm, unit, icon) in enumerate(_spot_display):
        price = _spot_db.get(comm) or config.COMMODITY_SPOT.get(comm)
        label = f"{icon} {comm}"
        value = f"${price:,.0f}" if price else "—"
        _spot_cols[i % 2].metric(label, value, help=unit)
    if _spot_date is not None:
        st.caption(f"As of {_spot_date}")
    st.divider()

    # ── Filter Presets ─────────────────────────────────────────────────────────
    st.markdown("## 💾 Filter Presets")
    _all_presets = load_filter_presets()
    _preset_names = list(_all_presets.keys())

    _prs_col1, _prs_col2 = st.columns([3, 1])
    _sel_preset = _prs_col1.selectbox(
        "Load preset", ["— none —"] + _preset_names,
        key="preset_selector", label_visibility="collapsed",
    )
    if _prs_col2.button("📂 Load", use_container_width=True, key="preset_load_btn"):
        if _sel_preset != "— none —" and _sel_preset in _all_presets:
            _p = _all_presets[_sel_preset]
            st.session_state["flt_comm"]    = _p.get("comm", [])
            st.session_state["flt_stage"]   = _p.get("stage", [])
            st.session_state["flt_mcap"]    = tuple(_p.get("mcap", [0, 200_000]))
            st.session_state["flt_score"]   = _p.get("score", 0)
            st.session_state["flt_grades"]  = _p.get("grades", [])
            st.session_state["flt_sort"]    = _p.get("sort_col", "score_composite")
            st.session_state["flt_asc"]     = _p.get("sort_asc", False)
            st.session_state["flt_pnav"]    = _p.get("pnav_max", 5.0)
            st.session_state["flt_aisc"]    = _p.get("min_aisc", 0)
            st.session_state["flt_prod"]    = _p.get("min_prod_koz", 0)
            st.session_state["flt_rli"]     = _p.get("min_reserve_life", 0)
            st.session_state["flt_evoz"]    = _p.get("max_ev_oz_prod", 9_999_999)
            st.session_state["flt_analyst_upside"] = _p.get("min_analyst_upside", 0)
            st.rerun()

    _save_col1, _save_col2, _save_col3 = st.columns([3, 1, 1])
    _new_preset_name = _save_col1.text_input(
        "Preset name", placeholder="e.g. Gold Value", key="preset_name_input",
        label_visibility="collapsed",
    )

    def _current_filters_dict():
        return {
            "comm":       st.session_state.get("flt_comm",   []),
            "stage":      st.session_state.get("flt_stage",  []),
            "mcap":       list(st.session_state.get("flt_mcap", [0, 200_000])),
            "score":      st.session_state.get("flt_score",  0),
            "grades":     st.session_state.get("flt_grades", []),
            "sort_col":   st.session_state.get("flt_sort",   "score_composite"),
            "sort_asc":   st.session_state.get("flt_asc",    False),
            "pnav_max":        st.session_state.get("flt_pnav",   5.0),
            "min_aisc":        st.session_state.get("flt_aisc",   0),
            "min_prod_koz":    st.session_state.get("flt_prod",   0),
            "min_reserve_life": st.session_state.get("flt_rli",   0),
            "max_ev_oz_prod":  st.session_state.get("flt_evoz",   9_999_999),
            "min_analyst_upside": st.session_state.get("flt_analyst_upside", 0),
        }

    if _save_col2.button("💾 Save", use_container_width=True, key="preset_save_btn"):
        if _new_preset_name.strip():
            save_filter_preset(_new_preset_name.strip(), _current_filters_dict())
            st.toast(f"Preset '{_new_preset_name}' saved ✓", icon="💾")
            st.rerun()
        else:
            st.warning("Enter a preset name first.")
    if _save_col3.button("🗑️", use_container_width=True, key="preset_del_btn",
                         help="Delete selected preset"):
        if _sel_preset != "— none —":
            delete_filter_preset(_sel_preset)
            st.toast(f"Preset '{_sel_preset}' deleted", icon="🗑️")
            st.rerun()

    st.divider()

    # ── Watchlist toggle ────────────────────────────────────────────────────────
    _watchlist_tickers = get_watchlist()
    _wl_only = st.toggle(
        f"⭐ Watchlist only ({len(_watchlist_tickers)})",
        value=False,
        help="Show only your starred companies",
    )
    _spg_only = st.toggle(
        "📡 S&P data only",
        value=False,
        help="Show only companies that have S&P Capital IQ / SNL Mining data",
    )

    st.markdown("## 🎛️ Filters")

    # Commodity
    commodities = sorted(df["commodity"].dropna().unique())
    _def_comm = st.session_state.pop("flt_comm", list(commodities))
    # Ensure only valid options remain after a data refresh
    _def_comm = [c for c in _def_comm if c in commodities] or list(commodities)
    sel_comm = st.multiselect(
        "Commodity", commodities,
        default=_def_comm, key="flt_comm",
        help="Primary commodity of the company",
    )

    # Stage
    stages = [s for s in ALL_STAGES if s in df["stage"].values]
    _def_stage = st.session_state.pop("flt_stage", list(stages))
    _def_stage = [s for s in _def_stage if s in stages] or list(stages)
    sel_stage = st.multiselect(
        "Stage", stages, default=_def_stage, key="flt_stage",
    )

    # Market cap
    st.markdown(f"**Market Cap ({config.CURRENCY})**")
    _mcap_max = max(200_000, int(df["market_cap"].max() / 1e6) + 10_000) if "market_cap" in df.columns and df["market_cap"].notna().any() else 200_000
    _def_mcap = st.session_state.pop("flt_mcap", (0, _mcap_max))
    # Clamp stored value to current max (universe can grow)
    _def_mcap = (int(_def_mcap[0]), min(int(_def_mcap[1]), _mcap_max))
    mcap_lo, mcap_hi = st.slider(
        "Market Cap Range (M AUD)",
        min_value=0, max_value=_mcap_max,
        value=_def_mcap, step=500, key="flt_mcap",
    )

    # Composite score minimum
    _def_score = st.session_state.pop("flt_score", 0)
    min_score = st.slider("Min Composite Score", 0, 100, _def_score, 5, key="flt_score")

    # Grade filter
    all_grades = ["🟢 Strong Buy", "🔵 Buy", "🟡 Watch", "🟠 Neutral", "🔴 Avoid"]
    _def_grades = st.session_state.pop("flt_grades", all_grades)
    _def_grades = [g for g in _def_grades if g in all_grades] or all_grades
    sel_grades = st.multiselect("Grade", all_grades, default=_def_grades, key="flt_grades")

    st.divider()
    st.markdown("**Sort by**")
    _sort_options = [
        "score_composite", "score_valuation", "score_health",
        "score_momentum", "score_mining",
        "analyst_upside", "return_3m", "return_1m",
        "spg_p_nav", "spg_aisc_margin", "spg_reserves_m",
        "upside_to_nav",
        "price_to_book", "ev_ebitda", "rsi",
        "wk52_position", "market_cap",
    ]
    _def_sort = st.session_state.pop("flt_sort", "score_composite")
    _sort_idx = _sort_options.index(_def_sort) if _def_sort in _sort_options else 0
    sort_col = st.selectbox(
        "Column", _sort_options, index=_sort_idx, key="flt_sort",
    )
    _def_asc = st.session_state.pop("flt_asc", False)
    sort_asc = st.checkbox("Ascending", value=_def_asc, key="flt_asc")

    # P/NAV filter (S&P data — only relevant for companies with SPG coverage)
    if "spg_p_nav" in df.columns and df["spg_p_nav"].notna().any():
        st.markdown("**P/NAV Filter** *(S&P data)*")
        _def_pnav = st.session_state.pop("flt_pnav", 5.0)
        pnav_max = st.slider(
            "Max P/NAV (< = cheaper)",
            min_value=0.5, max_value=5.0,
            value=min(float(_def_pnav), 5.0), step=0.1, key="flt_pnav",
            help="P/NAV < 1.0 means trading below net asset value",
        )
    else:
        pnav_max = 99.0

    # AISC margin filter (gold companies only)
    if "spg_aisc_margin" in df.columns and df["spg_aisc_margin"].notna().any():
        st.markdown("**Min AISC Margin %** *(gold, S&P data)*")
        _def_aisc = st.session_state.pop("flt_aisc", 0)
        min_aisc_margin = st.slider(
            "Min AISC Margin %",
            min_value=0, max_value=70,
            value=int(_def_aisc), step=5, key="flt_aisc",
            help="AISC margin = (Gold spot − AISC) / Gold spot × 100",
        )
    else:
        min_aisc_margin = 0

    # Min Production filter
    if "spg_production_oz" in df.columns and df["spg_production_oz"].notna().any():
        _prod_max_koz = max(int(df["spg_production_oz"].dropna().max() / 1000), 100)
        st.markdown("**Min Production (koz / yr)**")
        _def_prod = st.session_state.pop("flt_prod", 0)
        min_prod_koz = st.slider(
            "Min Production (koz/yr)",
            min_value=0, max_value=min(_prod_max_koz, 2000),
            value=int(_def_prod), step=10, key="flt_prod",
            help="Filter out companies below this annual production threshold",
        )
    else:
        min_prod_koz = 0

    # Min Reserve Life filter
    if "spg_reserve_life" in df.columns and df["spg_reserve_life"].notna().any():
        st.markdown("**Min Reserve Life (yr)**")
        _def_rli = st.session_state.pop("flt_rli", 0)
        min_reserve_life = st.slider(
            "Min Reserve Life (yr)",
            min_value=0, max_value=30,
            value=int(_def_rli), step=1, key="flt_rli",
            help="Only show companies with mine life ≥ this many years",
        )
    else:
        min_reserve_life = 0

    # Max EV/oz Production filter
    if "ev_per_oz_prod" in df.columns and df["ev_per_oz_prod"].notna().any():
        _ev_oz_cap = 100_000   # fixed cap: covers majors like AEM (~$40k/oz)
        st.markdown("**Max EV/oz Produced ($)**")
        _def_evoz = st.session_state.pop("flt_evoz", _ev_oz_cap)
        max_ev_oz_prod = st.slider(
            "Max EV/oz Produced ($)",
            min_value=500, max_value=_ev_oz_cap,
            value=min(int(_def_evoz), _ev_oz_cap), step=500, key="flt_evoz",
            help="Lower EV/oz = cheaper on production basis; filter out expensive names",
        )
    else:
        max_ev_oz_prod = 9_999_999

    # Min Analyst Upside filter
    if "analyst_upside" in df.columns and df["analyst_upside"].notna().any():
        st.markdown("**Min Analyst Upside %** *(consensus target)*")
        min_analyst_upside = st.slider(
            "Min analyst upside %",
            min_value=0, max_value=200,
            value=0, step=10, key="flt_analyst_upside",
            help="Filter to companies where analysts see ≥ X% upside to consensus target",
        )
    else:
        min_analyst_upside = 0

    st.divider()
    st.markdown("### 🔬 Score Weights")
    st.caption("Adjust how the composite score is calculated")
    _wg = st.session_state["weight_gen"]
    w_val = st.slider("Valuation",      0, 100, st.session_state["w_val"], key=f"sl_val_{_wg}")
    w_hlt = st.slider("Health",         0, 100, st.session_state["w_hlt"], key=f"sl_hlt_{_wg}")
    w_mom = st.slider("Momentum",       0, 100, st.session_state["w_mom"], key=f"sl_mom_{_wg}")
    w_min = st.slider("⛏️ Mining (S&P)", 0, 100, st.session_state["w_min"], key=f"sl_min_{_wg}")
    w_com = st.slider("Commodity",      0, 100, st.session_state["w_com"], key=f"sl_com_{_wg}")
    w_stg = st.slider("Stage",          0, 100, st.session_state["w_stg"], key=f"sl_stg_{_wg}")

    _sw_cols = st.columns(2)
    if _sw_cols[0].button("💾 Save weights", use_container_width=True):
        save_score_weights({"valuation": w_val, "health": w_hlt, "momentum": w_mom,
                            "mining": w_min, "commodity": w_com, "stage": w_stg})
        st.session_state.update(w_val=w_val, w_hlt=w_hlt, w_mom=w_mom,
                                w_min=w_min, w_com=w_com, w_stg=w_stg)
        st.toast("Weights saved ✓", icon="💾")
    if _sw_cols[1].button("↩️ Reset", use_container_width=True):
        _defaults = {"valuation": 30, "health": 20, "momentum": 15,
                     "mining": 25, "commodity": 5, "stage": 5}
        st.session_state.update(w_val=30, w_hlt=20, w_mom=15,
                                w_min=25, w_com=5,  w_stg=5)
        st.session_state["weight_gen"] += 1   # new key suffix → sliders reinitialise
        save_score_weights(_defaults)
        st.rerun()

    st.divider()
    st.markdown("### 🔄 Data Refresh")
    _lr_ts = last_refresh()
    _lr_display = str(_lr_ts)[:16] if _lr_ts and _lr_ts != "Never" else "Never"
    st.caption(f"Last refresh: **{_lr_display}**")
    if _IS_CLOUD:
        st.caption("Refresh runs daily via GitHub Actions (~08:15 Sydney).")
    elif st.button("🔄 Refresh Data Now", use_container_width=True,
                 help="Re-fetch all Yahoo Finance prices, recompute scores (~60 s)"):
        _refresh_placeholder = st.empty()
        _refresh_placeholder.info("⏳ Fetching data… this takes about 60 seconds.")
        try:
            run_daily_refresh()
            st.cache_data.clear()
            _refresh_placeholder.empty()
            st.toast("Data refreshed ✓", icon="🔄")
            st.rerun()
        except Exception as _re:
            _refresh_placeholder.error(f"Refresh failed: {_re}")

    total_w = w_val + w_hlt + w_mom + w_min + w_com + w_stg
    if total_w > 0:
        mining_s = df["score_mining"].fillna(50) if "score_mining" in df.columns else pd.Series(50, index=df.index)
        df["score_composite"] = (
            df["score_valuation"] * w_val / total_w +
            df["score_health"]    * w_hlt / total_w +
            df["score_momentum"]  * w_mom / total_w +
            mining_s              * w_min / total_w +
            df["score_commodity"] * w_com / total_w +
            df["score_stage"]     * w_stg / total_w
        ).round(1)
        # Recompute grade to match the updated score_composite
        def _regrade(s):
            if s >= 75: return "🟢 Strong Buy"
            if s >= 60: return "🔵 Buy"
            if s >= 45: return "🟡 Watch"
            if s >= 30: return "🟠 Neutral"
            return "🔴 Avoid"
        df["grade"] = df["score_composite"].apply(_regrade)

# ── Apply filters ──────────────────────────────────────────────────────────────
_debug_steps: list[tuple[str, int]] = [("DB total", len(df))]

mask = (
    df["commodity"].isin(sel_comm) &
    df["stage"].isin(sel_stage) &
    (df["score_composite"] >= min_score) &
    df["grade"].isin(sel_grades)
)
_debug_steps.append(("commodity+stage+score+grade", int(mask.sum())))

if "market_cap" in df.columns:
    mcap_m = df["market_cap"].fillna(0) / 1e6
    mask &= (mcap_m >= mcap_lo) & (mcap_m <= mcap_hi)
_debug_steps.append((f"market_cap [{mcap_lo},{mcap_hi}]", int(mask.sum())))

if pnav_max < 5.0 and "spg_p_nav" in df.columns:
    # Include rows where P/NAV ≤ max OR P/NAV is missing (non-SPG companies still visible)
    mask &= (df["spg_p_nav"].isna() | (df["spg_p_nav"] <= pnav_max))
_debug_steps.append((f"pnav_max={pnav_max:.1f}", int(mask.sum())))

if min_aisc_margin > 0 and "spg_aisc_margin" in df.columns:
    mask &= (df["spg_aisc_margin"].isna() | (df["spg_aisc_margin"] >= min_aisc_margin))
_debug_steps.append((f"aisc_margin>={min_aisc_margin}", int(mask.sum())))

if min_prod_koz > 0 and "spg_production_oz" in df.columns:
    mask &= (df["spg_production_oz"].isna() | (df["spg_production_oz"] >= min_prod_koz * 1000))
_debug_steps.append((f"prod>={min_prod_koz}koz", int(mask.sum())))

if min_reserve_life > 0 and "spg_reserve_life" in df.columns:
    mask &= (df["spg_reserve_life"].isna() | (df["spg_reserve_life"] >= min_reserve_life))
_debug_steps.append((f"reserve_life>={min_reserve_life}yr", int(mask.sum())))

if max_ev_oz_prod < 9_999_999 and "ev_per_oz_prod" in df.columns:
    mask &= (df["ev_per_oz_prod"].isna() | (df["ev_per_oz_prod"] <= max_ev_oz_prod))
_debug_steps.append((f"ev_oz<={max_ev_oz_prod}", int(mask.sum())))

if min_analyst_upside > 0 and "analyst_upside" in df.columns:
    # Require analyst_upside >= threshold; exclude companies with no analyst coverage
    mask &= (df["analyst_upside"].notna() & (df["analyst_upside"] >= min_analyst_upside))
_debug_steps.append((f"analyst_upside>={min_analyst_upside}", int(mask.sum())))

if _wl_only and _watchlist_tickers:
    mask &= df["ticker"].isin(_watchlist_tickers)
_debug_steps.append(("watchlist", int(mask.sum())))

if _spg_only and "has_spg" in df.columns:
    mask &= df["has_spg"]
_debug_steps.append(("spg_only", int(mask.sum())))

filtered = df[mask].sort_values(sort_col, ascending=sort_asc)

# ── Summary cards (first thing after the header — key overview on mobile) ─────
with st.container(key="summary_cards"):
    c1, c2, c3, c4, c5 = st.columns(5)
    cards = [
        (c1, len(filtered),                                         "Companies shown"),
        (c2, (filtered["grade"] == "🟢 Strong Buy").sum(),         "🟢 Strong Buy"),
        (c3, (filtered["grade"] == "🔵 Buy").sum(),                "🔵 Buy"),
        (c4, round(filtered["score_composite"].mean(), 1),          "Avg Score"),
        (c5, round(filtered["score_composite"].max(), 1),           "Top Score"),
    ]
    for col, val, label in cards:
        with col:
            st.markdown(
                f'<div class="metric-card"><h3>{val}</h3><p>{label}</p></div>',
                unsafe_allow_html=True,
            )
st.markdown("<br>", unsafe_allow_html=True)

# ── Tabs ────────────────────────────────────────────────────────────────────────
_wl_count = len(get_watchlist())
tab_today, tab_table, tab_detail, tab_charts, tab_report, tab_watchlist, tab_snl = st.tabs([
    "📌 Today", "📋 Screener", "🔍 Company Detail", "📊 Analytics",
    "📄 Report", f"⭐ Watchlist ({_wl_count})", "🗄️ SNL Data",
])

# ── TAB 0: Today — the daily answer: what moved, what to look at ────────────
with tab_today:
    # ── Price target alerts ────────────────────────────────────────────────────────
    _all_targets = get_all_price_targets()
    if _all_targets:
        _target_hits = []
        for _tk, _tgt in _all_targets.items():
            _trow = df[df["ticker"] == _tk]
            if _trow.empty:
                continue
            _tprice = _trow.iloc[0].get("price")
            if pd.notna(_tprice) and _tprice > 0:
                _tdist = (_tgt / _tprice - 1) * 100
                _target_hits.append({"ticker": _tk,
                                      "name": _trow.iloc[0].get("name", _tk),
                                      "price": _tprice,
                                      "target": _tgt,
                                      "dist_pct": _tdist})
        _near_target  = [h for h in _target_hits if abs(h["dist_pct"]) <= 10]
        _hit_target   = [h for h in _target_hits if h["dist_pct"] <= 0]
        if _hit_target or _near_target:
            _tgt_parts = []
            for h in sorted(_hit_target, key=lambda x: x["dist_pct"]):
                _tgt_parts.append(
                    f"🎯 **{h['name']}** ({h['ticker']}) hit target "
                    f"${h['target']:.3f} — now {h['dist_pct']:.1f}% past"
                )
            for h in sorted(_near_target, key=lambda x: abs(x["dist_pct"])):
                if h not in _hit_target:
                    _tgt_parts.append(
                        f"⚡ **{h['name']}** ({h['ticker']}) within "
                        f"{abs(h['dist_pct']):.1f}% of target ${h['target']:.3f}"
                    )
            if _tgt_parts:
                st.info("  ·  ".join(_tgt_parts))


    # ── Score Movers Banner ──────────────────────────────────────────────────
    _movers_df = df[df["score_delta"].abs() > 2.5].copy() if "score_delta" in df.columns else pd.DataFrame()
    if not _movers_df.empty:
        _upgrades = _movers_df[_movers_df["score_delta"] > 0].nlargest(5, "score_delta")
        _downgrades = _movers_df[_movers_df["score_delta"] < 0].nsmallest(5, "score_delta")

        _mover_parts_up = []
        for _, _mr in _upgrades.iterrows():
            _col = "#15803d" if _mr.get("score_delta", 0) >= 5 else "#16a34a"
            _mover_parts_up.append(
                f"<span style='color:{_col};font-weight:700'>{_mr.get('name','?')} "
                f"({_mr['ticker']})</span> "
                f"<span style='color:#5b6b7f'>{_mr.get('score_composite',0):.0f} "
                f"(<span style='color:{_col}'>{_mr.get('score_delta',0):+.1f}</span>)</span>"
            )
        _mover_parts_dn = []
        for _, _mr in _downgrades.iterrows():
            _mover_parts_dn.append(
                f"<span style='color:#b91c1c;font-weight:700'>{_mr.get('name','?')} "
                f"({_mr['ticker']})</span> "
                f"<span style='color:#5b6b7f'>{_mr.get('score_composite',0):.0f} "
                f"(<span style='color:#dc2626'>{_mr.get('score_delta',0):+.1f}</span>)</span>"
            )

        if _mover_parts_up or _mover_parts_dn:
            _mover_html = (
                "<div style='background:#f4f7fa;border:1px solid #e3e9f0;"
                "border-left:3px solid #1a3a5c;"
                "padding:8px 12px;border-radius:8px;margin-bottom:8px;"
                "font-size:12px;line-height:1.7'>"
                "<span style='color:#1a3a5c;font-weight:700;margin-right:8px'>📊 SCORE MOVERS</span>"
            )
            if _mover_parts_up:
                _mover_html += "<span style='color:#16a34a;margin-right:4px'>⬆</span>" + "  ·  ".join(_mover_parts_up)
            if _mover_parts_up and _mover_parts_dn:
                _mover_html += "  <span style='color:#cbd5e1;margin:0 6px'>|</span>  "
            if _mover_parts_dn:
                _mover_html += "<span style='color:#dc2626;margin-right:4px'>⬇</span>" + "  ·  ".join(_mover_parts_dn)
            _mover_html += "</div>"
            st.markdown(_mover_html, unsafe_allow_html=True)


    # ── Top Opportunities callout ──────────────────────────────────────────────
    _top_ops = filtered[filtered["grade"] == "🟢 Strong Buy"].nlargest(6, "score_composite")
    if _top_ops.empty:
        _top_ops = filtered[filtered["grade"] == "🔵 Buy"].nlargest(6, "score_composite")

    if not _top_ops.empty:
        st.markdown("#### 🏆 Top Opportunities")
        # Lay cards out 3 per row (2-up grid on mobile via CSS)
        _op_cols = []
        for _ri in range(0, len(_top_ops), 3):
            _op_cols.extend(st.columns(3)[: min(3, len(_top_ops) - _ri)])
        for _ci, (_, _op) in enumerate(zip(_op_cols, _top_ops.itertuples())):
            _op_name   = _op.name if hasattr(_op, "name") else str(_op.ticker)
            _op_ticker = _op.ticker
            _op_grade  = _op.grade if hasattr(_op, "grade") else "—"
            _op_score  = _op.score_composite if hasattr(_op, "score_composite") else 0
            _op_comm   = _op.commodity if hasattr(_op, "commodity") else "—"
            _op_stage  = _op.stage if hasattr(_op, "stage") else "—"
            _op_price  = _op.price if hasattr(_op, "price") and pd.notna(_op.price) else None
            _op_mcap   = _op.market_cap if hasattr(_op, "market_cap") and pd.notna(_op.market_cap) else None
            _op_pnav   = _op.spg_p_nav if hasattr(_op, "spg_p_nav") and pd.notna(_op.spg_p_nav) else None
            _op_delta  = _op.score_delta if hasattr(_op, "score_delta") and pd.notna(_op.score_delta) else None
            _op_upside   = _op.upside_to_nav if hasattr(_op, "upside_to_nav") and pd.notna(_op.upside_to_nav) else None
            _op_an_up    = getattr(_op, "analyst_upside", None)
            _op_an_up    = _op_an_up if pd.notna(_op_an_up) else None
            _op_an_ct    = getattr(_op, "analyst_count", None)
            _op_an_ct    = int(_op_an_ct) if pd.notna(_op_an_ct) else None
            _mcap_fmt  = (f"${_op_mcap/1e9:.1f}B" if _op_mcap and _op_mcap >= 1e9
                          else f"${_op_mcap/1e6:.0f}M" if _op_mcap else "—")
            _kpis = [f"Score <b>{_op_score:.0f}/100</b>"]
            if _op_upside and _op_upside > 0:
                _kpis.append(f"<b>↑{_op_upside:+.0f}% to NAV</b>")
            elif _op_pnav:
                _kpis.append(f"P/NAV {_op_pnav:.2f}x")
            if _op_an_up and _op_an_up >= 20 and _op_an_ct and _op_an_ct >= 2:
                _kpis.append(f"Analyst <b>↑{_op_an_up:.0f}%</b> ({_op_an_ct})")
            if _op_delta: _kpis.append(f"Δ {_op_delta:+.1f}")
            with _op_cols[_ci]:
                _grade_color = (
                    "#16a34a" if "Strong Buy" in _op_grade else
                    "#2563eb" if "Buy"        in _op_grade else
                    "#d97706"
                )
                _grade_bg = (
                    "#dcfce7" if "Strong Buy" in _op_grade else
                    "#dbeafe" if "Buy"        in _op_grade else
                    "#fef3c7"
                )
                st.markdown(
                    f"<div class='opp-card' style='border-top:4px solid {_grade_color};'>"
                    f"<span style='display:inline-block;font-size:11px;font-weight:700;"
                    f"text-transform:uppercase;letter-spacing:0.06em;"
                    f"background-color:{_grade_bg};color:{_grade_color};"
                    f"padding:2px 8px;border-radius:4px;'>{_op_grade}</span>"
                    f"<div class='opp-name'>{_op_name}</div>"
                    f"<div class='opp-sub'>{_op_ticker} &nbsp;·&nbsp; {_op_comm} &nbsp;·&nbsp; {_op_stage}</div>"
                    f"<div class='opp-div'></div>"
                    f"<div class='opp-kpis'>{'&nbsp;&nbsp;·&nbsp;&nbsp;'.join(_kpis)}</div>"
                    f"<div class='opp-foot'>Mkt Cap {_mcap_fmt}"
                    f"{f'&nbsp;&nbsp;·&nbsp;&nbsp;${_op_price:.3f}' if _op_price else ''}</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
        st.markdown("<br>", unsafe_allow_html=True)


    # ── Grade-transition alert bar ────────────────────────────────────────────────
    _GRADE_ORDER = {"🟢 Strong Buy": 5, "🔵 Buy": 4, "🟡 Watch": 3, "🟠 Neutral": 2, "🔴 Avoid": 1}

    if "grade_prev" in df.columns and df["grade_prev"].notna().any():
        _upgrades   = df[
            (df["grade_prev"].notna()) &
            (df["grade"].map(_GRADE_ORDER).fillna(0) > df["grade_prev"].map(_GRADE_ORDER).fillna(0))
        ].sort_values("score_delta", ascending=False)
        _downgrades = df[
            (df["grade_prev"].notna()) &
            (df["grade"].map(_GRADE_ORDER).fillna(0) < df["grade_prev"].map(_GRADE_ORDER).fillna(0))
        ].sort_values("score_delta")

        _top_movers = df[df["score_delta"].notna()].reindex(
            df["score_delta"].abs().sort_values(ascending=False).index
        ).head(5)

        if not _upgrades.empty or not _downgrades.empty:
            _al_label = (
                f"📊 Grade changes since last refresh — "
                f"⬆️ {len(_upgrades)} upgrades · ⬇️ {len(_downgrades)} downgrades"
            )
            with st.expander(_al_label, expanded=False):
                _al_cols = st.columns([2, 2, 3])

                with _al_cols[0]:
                    if not _upgrades.empty:
                        st.success(f"**⬆️ {len(_upgrades)} grade upgrade{'s' if len(_upgrades)>1 else ''}**")
                        for _, _r in _upgrades.head(4).iterrows():
                            st.caption(
                                f"**{_r['name']}** ({_r['ticker']})  "
                                f"{_r['grade_prev']} → {_r['grade']}  "
                                f"[{_r['score_delta']:+.1f}]"
                            )

                with _al_cols[1]:
                    if not _downgrades.empty:
                        st.error(f"**⬇️ {len(_downgrades)} grade downgrade{'s' if len(_downgrades)>1 else ''}**")
                        for _, _r in _downgrades.head(4).iterrows():
                            st.caption(
                                f"**{_r['name']}** ({_r['ticker']})  "
                                f"{_r['grade_prev']} → {_r['grade']}  "
                                f"[{_r['score_delta']:+.1f}]"
                            )

                with _al_cols[2]:
                    if not _top_movers.empty:
                        st.info("**📈 Top movers since last refresh**")
                        for _, _r in _top_movers.iterrows():
                            _d = _r.get("score_delta", 0)
                            if pd.notna(_d) and _d != 0:
                                _arrow = "⬆️" if _d > 0 else "⬇️"
                                st.caption(
                                    f"{_arrow} **{_r['name']}** ({_r['ticker']})  "
                                    f"{_r['grade']}  [{_d:+.1f}]"
                                )
            st.markdown("<br>", unsafe_allow_html=True)


    st.caption("📋 Full universe with filters → **Screener** tab · deep-dive a name → **Company Detail** tab")


# ── TAB 1: Table ───────────────────────────────────────────────────────────────
with tab_table:

    # ── Strategy presets ──────────────────────────────────────────────────────
    st.markdown("**⚡ Quick Filters**")
    _preset_cols_r1 = st.columns(4)
    _preset_cols_r2 = st.columns(4)

    # Define presets: (label, tooltip, filter_func(df) → mask)
    _PRESETS = [
        ("💎 Quality\nProducers",
         "Producers & mid-tiers with AISC margin > 30% or P/NAV < 1.2",
         lambda d: (
             d["stage"].isin(["Major Producer", "Mid-tier Producer", "Producer"]) &
             (
                 (d.get("spg_aisc_margin", pd.Series(dtype=float)).fillna(0) > 30) |
                 (d.get("spg_p_nav",       pd.Series(dtype=float)).fillna(99) < 1.2)
             )
         )),
        ("🔍 Oversold\nGems",
         "Score ≥ 55, RSI < 40, price near 52wk low",
         lambda d: (
             (d["score_composite"] >= 55) &
             (d.get("rsi", pd.Series(dtype=float)).fillna(50) < 40) &
             (d.get("wk52_position", pd.Series(dtype=float)).fillna(50) < 35)
         )),
        ("⚛️ Uranium\nPlay",
         "All uranium companies sorted by score",
         lambda d: d["commodity"].str.contains("Uranium", na=False)),
        ("💰 High\nFCF Yield",
         "Positive FCF and FCF yield ≥ 5%",
         lambda d: (
             d.get("fcf_yield", pd.Series(dtype=float)).fillna(0) >= 5
         )),
        ("🎯 NAV\nDiscount >40%",
         "Companies trading at >40% discount to NAV (upside_to_nav > 40%)",
         lambda d: (
             d.get("upside_to_nav", pd.Series(dtype=float)).fillna(0) > 40
         )),
        ("🧲 Deep\nValue",
         "P/B < 1 or P/NAV < 0.8 — trading below asset value",
         lambda d: (
             (d.get("price_to_book", pd.Series(dtype=float)).fillna(99) < 1.0) |
             (d.get("spg_p_nav",     pd.Series(dtype=float)).fillna(99) < 0.8)
         )),
        ("🎙️ Analyst\nConviction",
         "Score ≥ 50, analyst upside ≥ 50%, ≥ 3 opinions — consensus high-conviction ideas",
         lambda d: (
             (d["score_composite"] >= 50) &
             (d.get("analyst_upside", pd.Series(dtype=float)).fillna(0) >= 50) &
             (d.get("analyst_count",  pd.Series(dtype=float)).fillna(0) >= 3)
         )),
        ("📈 Score\nMomentum",
         "Score improved since last refresh (Δ Score > 3)",
         lambda d: (
             d.get("score_delta", pd.Series(dtype=float)).fillna(0) > 3
         )),
    ]

    _all_preset_cols = _preset_cols_r1 + _preset_cols_r2
    for _pi, (_plabel, _ptip, _pfunc) in enumerate(_PRESETS):
        if _all_preset_cols[_pi].button(_plabel, help=_ptip, use_container_width=True):
            try:
                _pmask = _pfunc(filtered)
                _pcount = int(_pmask.sum())
                st.session_state["_preset_msg"] = (
                    f"Preset applied — {_pcount} companies match. "
                    f"Use sidebar filters to refine further."
                )
                st.session_state["_preset_df"] = filtered[_pmask].copy()
            except Exception:
                pass

    if "_preset_msg" in st.session_state:
        st.info(st.session_state.pop("_preset_msg"))

    if "_preset_df" in st.session_state:
        _pdf = st.session_state.pop("_preset_df")
        if not _pdf.empty:
            filtered = _pdf

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Universe Summary Stats Banner ─────────────────────────────────────────
    _n_total     = len(filtered)
    _n_spg       = int(filtered.get("has_spg", pd.Series(dtype=bool)).fillna(False).sum()) if "has_spg" in filtered.columns else 0
    _n_strong    = int((filtered["grade"] == "🟢 Strong Buy").sum()) if "grade" in filtered.columns else 0
    _n_buy       = int((filtered["grade"] == "🔵 Buy").sum()) if "grade" in filtered.columns else 0
    _avg_score   = filtered["score_composite"].mean() if "score_composite" in filtered.columns else None
    _med_pnav    = filtered["spg_p_nav"].median() if "spg_p_nav" in filtered.columns and filtered["spg_p_nav"].notna().any() else None
    _med_aisc_m  = filtered["spg_aisc_margin"].median() if "spg_aisc_margin" in filtered.columns and filtered["spg_aisc_margin"].notna().any() else None
    _med_rli     = filtered["spg_reserve_life"].median() if "spg_reserve_life" in filtered.columns and filtered["spg_reserve_life"].notna().any() else None
    _n_deep_val  = int((filtered["upside_to_nav"].fillna(0) >= 40).sum()) if "upside_to_nav" in filtered.columns else 0
    _med_analyst_up = (
        filtered.loc[filtered["analyst_count"].fillna(0) >= 2, "analyst_upside"].median()
        if "analyst_upside" in filtered.columns and "analyst_count" in filtered.columns else None
    )

    _stat_items = [
        ("Companies", f"{_n_total}"),
        ("S&P Data", f"{_n_spg}"),
        ("⭐ Buys", f"{_n_strong + _n_buy}  ({_n_strong} Strong)"),
        ("Avg Score", f"{_avg_score:.1f}" if _avg_score is not None and not pd.isna(_avg_score) else "—"),
        ("Median P/NAV", f"{_med_pnav:.2f}x" if _med_pnav is not None and not pd.isna(_med_pnav) else "—"),
        ("Med Analyst ↑", f"{_med_analyst_up:+.0f}%" if _med_analyst_up is not None and not pd.isna(_med_analyst_up) else "—"),
        (">40% NAV Discount", f"{_n_deep_val}"),
        ("Median AISC Mgn", f"{_med_aisc_m:.1f}%" if _med_aisc_m is not None and not pd.isna(_med_aisc_m) else "—"),
    ]
    _stat_cols = st.columns(len(_stat_items))
    for _si, (_slbl, _sval) in enumerate(_stat_items):
        _stat_cols[_si].metric(_slbl, _sval)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── All available display columns (ordered; label → column key mapping inverted below)
    display_cols = {
        "grade":            "Grade",
        "name":             "Company",
        "commodity":        "Commodity",
        "stage":            "Stage",
        "price":            "Price",
        "market_cap":       "Mkt Cap (M)",
        "score_composite":    "Score",
        "univ_rank":          "# Rank",
        "peer_rank_display":  "Peer Rank",
        "score_delta":        "Δ Score",
        "return_1m":        "1M Ret%",
        "return_3m":        "3M Ret%",
        "has_spg":          "S&P",
        "score_valuation":  "Valuation",
        "score_health":     "Health",
        "score_momentum":   "Momentum",
        "score_mining":     "⛏️ Mining",
        "upside_to_nav":    "↑ Upside to NAV%",
        "pb_peer_upside":   "↑ P/B Re-rating%",
        "spg_p_nav":        "P/NAV",
        "spg_aisc_per_oz":  "AISC($/oz)",
        "spg_aisc_margin":  "AISC Margin%",
        "spg_reserves_m":   "Reserves($M)",
        "ev_reserves":      "EV/Reserves",
        "grade_display":    "Ore Grade",
        "fcf_yield":        "FCF Yield%",
        "analyst_upside":   "↑ Analyst Target%",
        "analyst_target_mean": "Analyst Target",
        "analyst_count":    "# Analysts",
        "dividend_yield":   "Div Yield%",
        "price_to_book":    "P/B",
        "ev_ebitda":        "EV/EBITDA",
        "p_cf":             "P/CF",
        "debt_to_equity":   "D/E",
        "current_ratio":    "Curr. Ratio",
        "ev_revenue":       "EV/Rev",
        "cash_pct_mcap":    "Cash %Mkt",
        "net_debt_m":       "Net Debt (M)",
        "rsi":              "RSI",
        "wk52_position":    "52wk Pos%",
        "pct_from_52hi":    "vs 52Hi%",
        "pct_from_52lo":    "vs 52Lo%",
        # New: production, EV multiples, reserve life
        "spg_cash_cost_oz":     "Cash Cost ($/oz)",
        "spg_cash_cost_t":      "Cash Cost ($/t)",
        "spg_cash_cost_lb":     "Cash Cost ($/lb)",
        "production_display":   "Production",
        "ev_per_oz_prod":       "EV/oz Prod",
        "ev_per_oz_reserve":    "EV/oz Rsv",
        "ev_per_lb_reserve":    "EV/lb Rsv",
        "spg_reserve_life":         "Reserve Life (yr)",
        "reserve_life_display":     "Reserve Life",
        "spg_realized_price_oz":    "Realized Price ($/oz)",
        "spg_realized_price_t":     "Realized Price ($/t)",
        "spg_realized_price_lb":    "Realized Price ($/lb)",
        "spg_contained_reserves_oz": "Contained Rsv (oz)",
        "spg_contained_reserves_lb": "Contained Rsv (lb)",
        # SNL SQLite-sourced metrics
        "snl_rr_koz":        "R&R (koz)",
        "snl_rsv_koz":       "Reserves (koz)",
        "snl_rr_mlb":        "R&R (Mlb)",
        "snl_ev_per_oz_rr":  "EV/oz R&R",
        "snl_ev_per_oz_rsv": "EV/oz Reserve",
        "snl_p_insitu":      "P/In-situ NAV",
        "snl_insitu_rr_m":   "In-situ R&R ($M)",
        "snl_global_rank":   "Gold Rank",
        "snl_revenue_m":     "Revenue ($M)",
        # SNL mine_econ and global rank (new)
        "spg_mine_life":     "Mine Life (yr)",
        "spg_global_rank":   "Global Rank",
        "analyst_upside":    "Analyst Upside%",
        # Profitability
        "return_on_equity":  "ROE%",
        "operating_margins": "Op Margin%",
        "gross_margins":     "Gross Margin%",
    }
    # ── Column picker ─────────────────────────────────────────────────────────
    _all_labels   = [v for k, v in display_cols.items() if k in filtered.columns]
    _always_show  = {"Grade", "Company", "Commodity", "Stage", "Score"}
    _default_cols = [
        "Grade", "Company", "Commodity", "Stage", "Price", "Mkt Cap (M)",
        "Score", "# Rank", "Peer Rank", "Δ Score", "1M Ret%",
        "⛏️ Mining",          # Mining sub-score (AISC + reserves + P/NAV)
        "AISC Margin%",       # (Spot − AISC) / Spot — true profitability
        "Ore Grade",          # Primary grade g/t (Au/Ag) or % (Cu/Ni/U3O8/…)
        "Production",         # Attributable annual production (koz / kt / Mlb)
        "EV/oz Prod",         # EV per oz produced — key trading multiple
        "EV/oz R&R",          # EV per oz of total R&R — key SNL metric
        "R&R (koz)",          # Total reserves + resources in koz (SNL)
        "P/In-situ NAV",      # Market cap / SNL in-situ value
        "Reserve Life (yr)",  # Contained reserves ÷ annual production
        "↑ Upside to NAV%", "↑ Analyst Target%", "P/NAV", "EV/EBITDA", "RSI",
    ]
    _default_sel = [c for c in _default_cols if c in _all_labels]

    with st.expander("🗂️ Column selector", expanded=False):
        _col_pick = st.multiselect(
            "Choose columns to display",
            options=_all_labels,
            default=st.session_state.get("_tbl_cols", _default_sel),
            help="Locked columns: Grade, Company, Commodity, Stage, Score",
            key="_col_picker_widget",
        )
        # Always include locked columns
        for _lc in _always_show:
            if _lc in _all_labels and _lc not in _col_pick:
                _col_pick.insert(0, _lc)
        if st.button("↩️ Reset columns", key="_col_reset"):
            st.session_state["_tbl_cols"] = _default_sel
            st.rerun()
    st.session_state["_tbl_cols"] = _col_pick

    # Build label→key reverse map, then filter to chosen columns
    _label_to_key = {v: k for k, v in display_cols.items()}
    _chosen_keys  = [_label_to_key[lbl] for lbl in _col_pick if lbl in _label_to_key]
    available = {k: v for k, v in display_cols.items()
                 if k in filtered.columns and k in _chosen_keys}

    tbl = filtered[list(available.keys())].copy()
    tbl["market_cap"] = (tbl["market_cap"] / 1e6).round(1)
    tbl.columns = list(available.values())
    # Prepend watchlist star column
    tbl.insert(0, "⭐", filtered["ticker"].map(
        lambda t: "⭐" if t in _watchlist_tickers else ""
    ).values)
    tbl.index = range(1, len(tbl) + 1)

    def color_score(val):
        if pd.isna(val): return ""
        if val >= 75:    return "background-color: #dcfce7; color: #14532d; font-weight:600"
        if val >= 60:    return "background-color: #dbeafe; color: #1e3a5f"
        if val >= 45:    return "background-color: #fef9c3; color: #713f12"
        return                  "background-color: #fee2e2; color: #7f1d1d"

    def color_delta(val):
        if pd.isna(val) or val == 0: return ""
        return "color: #22c55e; font-weight:600" if val > 0 else "color: #ef4444; font-weight:600"

    def color_rsi(val):
        """Green = oversold (<30), amber = leaning oversold (30–45), red = overbought (>70)."""
        if pd.isna(val): return ""
        if val < 30: return "background-color: #dcfce7; color: #14532d; font-weight:600"
        if val > 70: return "background-color: #fee2e2; color: #7f1d1d"
        if val < 45: return "background-color: #fef9c3; color: #713f12"
        return ""

    def color_return(val):
        """Soft green for positive returns, soft red for negative."""
        if pd.isna(val) or val == 0: return ""
        return "color: #22c55e; font-weight:600" if val > 0 else "color: #ef4444; font-weight:600"

    def color_upside(val):
        """Green = big discount to NAV; amber = modest discount; grey = at/above NAV."""
        if pd.isna(val): return ""
        if val >= 50:  return "background-color: #dcfce7; color: #14532d; font-weight:700"
        if val >= 25:  return "background-color: #dbeafe; color: #1e3a5f; font-weight:600"
        if val >= 10:  return "color: #22c55e; font-weight:600"
        if val < 0:    return "color: #ef4444"
        return ""

    def color_peer_rank(val):
        """Gold for #1, silver-ish for top-3, muted for rest."""
        if not isinstance(val, str) or "/" not in val:
            return ""
        try:
            rank = int(val.split("/")[0].strip())
        except ValueError:
            return ""
        if rank == 1: return "background-color: #fef3c7; color: #92400e; font-weight:700"
        if rank <= 3: return "color: #94a3b8; font-weight:600"
        return ""

    _score_cols  = [c for c in ["Score", "Valuation", "Health", "Momentum", "⛏️ Mining"] if c in tbl.columns]
    _delta_cols  = ["Δ Score"]           if "Δ Score"           in tbl.columns else []
    _rsi_cols    = ["RSI"]               if "RSI"               in tbl.columns else []
    _ret_cols    = [c for c in ["1M Ret%", "3M Ret%"]         if c in tbl.columns]
    _upside_cols  = [c for c in ["↑ Upside to NAV%", "↑ P/B Re-rating%"] if c in tbl.columns]
    _analyst_cols = ["↑ Analyst Target%"] if "↑ Analyst Target%" in tbl.columns else []
    _peer_cols   = ["Peer Rank"]          if "Peer Rank"         in tbl.columns else []
    styled = (
        tbl.style
        .map(color_score,      subset=_score_cols   if _score_cols   else [])
        .map(color_delta,      subset=_delta_cols   if _delta_cols   else [])
        .map(color_rsi,        subset=_rsi_cols     if _rsi_cols     else [])
        .map(color_return,     subset=_ret_cols     if _ret_cols     else [])
        .map(color_upside,     subset=_upside_cols  if _upside_cols  else [])
        .map(color_upside,     subset=_analyst_cols if _analyst_cols else [])
        .map(color_peer_rank,  subset=_peer_cols    if _peer_cols    else [])
        .format({
        "Price":             "{:.3f}",
        "Mkt Cap (M)":       "{:,.0f}",
        "Score":             "{:.1f}",
        "Δ Score":           lambda x: f"{x:+.1f}" if pd.notna(x) else "—",
        "1M Ret%":           lambda x: f"{x:+.1f}%" if pd.notna(x) else "—",
        "3M Ret%":           lambda x: f"{x:+.1f}%" if pd.notna(x) else "—",
        "S&P":               lambda x: "✅" if x else "—",
        "Valuation":         "{:.1f}",
        "Health":            "{:.1f}",
        "Momentum":          "{:.1f}",
        "⛏️ Mining":         lambda x: f"{x:.1f}" if pd.notna(x) else "—",
        "# Rank":            lambda x: f"#{int(x)}" if pd.notna(x) else "—",
        "↑ Upside to NAV%":  lambda x: f"{x:+.0f}%" if pd.notna(x) else "—",
        "↑ P/B Re-rating%":  lambda x: f"{x:+.0f}%" if pd.notna(x) else "—",
        "P/NAV":             lambda x: f"{x:.2f}x" if pd.notna(x) else "—",
        "AISC($/oz)":        lambda x: f"${x:,.0f}" if pd.notna(x) else "—",
        "AISC Margin%":      lambda x: f"{x:.1f}%" if pd.notna(x) else "—",
        "Reserves($M)":      lambda x: f"${x:,.0f}M" if pd.notna(x) else "—",
        "EV/Reserves":       lambda x: f"{x:.2f}x" if pd.notna(x) else "—",
        "Ore Grade":         lambda x: str(x) if x and x != "—" else "—",
        "Production":        lambda x: str(x) if x and x != "—" else "—",
        "EV/oz Prod":        lambda x: f"${x:,.0f}" if pd.notna(x) else "—",
        "EV/oz Rsv":         lambda x: f"${x:,.0f}" if pd.notna(x) else "—",
        "EV/lb Rsv":         lambda x: f"${x:.3f}" if pd.notna(x) else "—",
        "Cash Cost ($/oz)":  lambda x: f"${x:,.0f}" if pd.notna(x) else "—",
        "Cash Cost ($/t)":   lambda x: f"${x:,.0f}" if pd.notna(x) else "—",
        "Cash Cost ($/lb)":  lambda x: f"${x:.3f}" if pd.notna(x) else "—",
        "Realized Price ($/oz)": lambda x: f"${x:,.0f}" if pd.notna(x) else "—",
        "Realized Price ($/t)":  lambda x: f"${x:,.0f}" if pd.notna(x) else "—",
        "Realized Price ($/lb)":  lambda x: f"${x:.3f}" if pd.notna(x) else "—",
        "Contained Rsv (oz)":     lambda x: f"{x/1e6:,.2f}Moz" if pd.notna(x) and x >= 1e6
                                            else (f"{x/1e3:,.0f}koz" if pd.notna(x) and x >= 1000
                                            else (f"{x:,.0f}oz" if pd.notna(x) else "—")),
        "Contained Rsv (lb)":     lambda x: f"{x/1e9:,.2f}Blb" if pd.notna(x) and x >= 1e9
                                            else (f"{x/1e6:,.0f}Mlb" if pd.notna(x) and x >= 1e6
                                            else (f"{x:,.0f}lb" if pd.notna(x) else "—")),
        "Reserve Life":      lambda x: str(x) if x and x != "—" else "—",
        "Reserve Life (yr)": lambda x: f"{x:.1f} yr" if pd.notna(x) else "—",
        "↑ Analyst Target%": lambda x: f"{x:+.0f}%" if pd.notna(x) else "—",
        "Analyst Target":    lambda x: f"${x:.2f}" if pd.notna(x) else "—",
        "# Analysts":        lambda x: f"{int(x)}" if pd.notna(x) else "—",
        "FCF Yield%":        lambda x: f"{x:.1f}%" if pd.notna(x) else "—",
        "Div Yield%":        lambda x: f"{x:.2f}%" if pd.notna(x) else "—",
        "P/B":               lambda x: f"{x:.2f}" if pd.notna(x) else "—",
        "EV/EBITDA":         lambda x: f"{x:.1f}" if pd.notna(x) else "—",
        "EV/Rev":        lambda x: f"{x:.1f}" if pd.notna(x) else "—",
        "P/CF":          lambda x: f"{x:.1f}" if pd.notna(x) else "—",
        "D/E":           lambda x: f"{x:.0f}" if pd.notna(x) else "—",
        "Cash %Mkt":     lambda x: f"{x:.1f}%" if pd.notna(x) else "—",
        "RSI":           lambda x: f"{x:.0f}" if pd.notna(x) else "—",
        "52wk Pos%":     lambda x: f"{x:.0f}%" if pd.notna(x) else "—",
        "vs 52Lo%":      lambda x: f"{x:+.1f}%" if pd.notna(x) else "—",
        "ROE%":          lambda x: f"{x*100:.1f}%" if pd.notna(x) else "—",
        "Op Margin%":    lambda x: f"{x*100:.1f}%" if pd.notna(x) else "—",
        "Gross Margin%": lambda x: f"{x*100:.1f}%" if pd.notna(x) else "—",
    }, na_rep="—")
    )

    st.dataframe(styled, width="stretch", height=600)

    _dl_col1, _dl_col2 = st.columns([1, 1])
    with _dl_col1:
        csv = tbl.to_csv(index=False).encode("utf-8-sig")
        st.download_button("⬇️ Download CSV", csv, "mining_screener.csv", "text/csv")

    with _dl_col2:
        try:
            import io
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            _GRADE_FILLS = {
                "🟢 Strong Buy": PatternFill("solid", fgColor="14532D"),
                "🔵 Buy":        PatternFill("solid", fgColor="1E3A5F"),
                "🟡 Watch":      PatternFill("solid", fgColor="713F12"),
                "🟠 Neutral":    PatternFill("solid", fgColor="3F0A0A"),
                "🔴 Avoid":      PatternFill("solid", fgColor="3F0A0A"),
            }
            _WHITE_FONT = Font(color="FFFFFF", bold=True)
            _HEADER_FILL = PatternFill("solid", fgColor="0F172A")
            _HEADER_FONT = Font(color="FFFFFF", bold=True)
            _THIN = Side(style="thin", color="374151")
            _BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Mining Screener"

            # Write header
            for col_idx, col_name in enumerate(tbl.columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill   = _HEADER_FILL
                cell.font   = _HEADER_FONT
                cell.border = _BORDER
                cell.alignment = Alignment(horizontal="center")

            # Write data rows
            for row_idx, row_data in enumerate(tbl.itertuples(index=False), start=2):
                grade_val = None
                for col_idx, value in enumerate(row_data, start=1):
                    col_name = tbl.columns[col_idx - 1]
                    # Store raw value (strip emoji formatting)
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = _BORDER
                    cell.alignment = Alignment(horizontal="center" if col_idx > 2 else "left")
                    if col_name == "Grade":
                        grade_val = str(value)
                # Apply grade row colour
                if grade_val and grade_val in _GRADE_FILLS:
                    for col_idx in range(1, len(tbl.columns) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = _GRADE_FILLS[grade_val]
                        ws.cell(row=row_idx, column=col_idx).font = _WHITE_FONT

            # Auto-width columns
            for col_idx, col_name in enumerate(tbl.columns, start=1):
                max_len = max(len(str(col_name)), 8)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 20)

            ws.freeze_panes = "A2"

            # ── Sheet 2: Full Mining Data ──────────────────────────────────
            _spg_export_cols = [c for c in [
                "ticker", "name", "commodity", "stage", "grade",
                "score_composite", "score_mining",
                "spg_p_nav", "upside_to_nav",
                "spg_aisc_per_oz", "spg_aisc_per_t", "spg_aisc_per_lb", "spg_aisc_margin",
                "spg_cash_cost_oz", "spg_cash_cost_t", "spg_cash_cost_lb",
                "spg_production_oz", "spg_production_t", "spg_production_lb",
                "spg_realized_price_oz", "spg_realized_price_t", "spg_realized_price_lb",
                "spg_contained_reserves_oz", "spg_contained_reserves_lb",
                "spg_reserve_life",
                "spg_reserves_m", "spg_resources_m",
                "spg_grade_gpt", "spg_grade_pct",
                "ev_per_oz_prod", "ev_per_oz_reserve", "ev_per_lb_reserve",
                "market_cap", "enterprise_value", "price",
            ] if c in filtered.columns]

            if _spg_export_cols:
                ws2 = wb.create_sheet("Mining Data")
                _spg_tbl = filtered[_spg_export_cols].copy()
                # Header row
                for ci, cname in enumerate(_spg_export_cols, start=1):
                    cell = ws2.cell(row=1, column=ci, value=cname)
                    cell.fill = _HEADER_FILL
                    cell.font = _HEADER_FONT
                    cell.border = _BORDER
                    cell.alignment = Alignment(horizontal="center")
                # Data rows
                for ri, (_, srow) in enumerate(_spg_tbl.iterrows(), start=2):
                    for ci, cname in enumerate(_spg_export_cols, start=1):
                        val = srow[cname]
                        # Coerce numpy types to native Python
                        if pd.isna(val):
                            val = None
                        elif hasattr(val, "item"):
                            val = val.item()
                        cell = ws2.cell(row=ri, column=ci, value=val)
                        cell.border = _BORDER
                        cell.alignment = Alignment(horizontal="center" if ci > 2 else "left")
                # Auto-width
                for ci, cname in enumerate(_spg_export_cols, start=1):
                    ws2.column_dimensions[get_column_letter(ci)].width = min(len(cname) + 2, 22)
                ws2.freeze_panes = "A2"

            _buf = io.BytesIO()
            wb.save(_buf)
            _buf.seek(0)
            st.download_button(
                "⬇️ Download Excel (.xlsx)",
                data=_buf.getvalue(),
                file_name="mining_screener.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except ImportError:
            st.caption("Install `openpyxl` for Excel export")

    # ── AISC Peer Ranking (SNL Live) ──────────────────────────────────────────
    if snl_client.is_configured():
        _aisc_rows = df[df["spg_aisc_per_oz"].notna()].copy()
        if not _aisc_rows.empty:
            st.markdown("---")
            st.markdown("### AISC Cost Ranking — Gold Producers (SNL Live Data)")
            _aisc_rows = _aisc_rows.sort_values("spg_aisc_per_oz")
            # Spot gold for margin line
            from config import COMMODITY_SPOT as _CS
            _gold_spot = _CS.get("Gold", 3300)

            _fig_aisc = px.bar(
                _aisc_rows,
                x="ticker",
                y="spg_aisc_per_oz",
                color="spg_aisc_per_oz",
                color_continuous_scale=["#2ecc71", "#f0a500", "#e74c3c"],
                labels={"spg_aisc_per_oz": "AISC ($/oz)", "ticker": ""},
                hover_data={
                    "name": True,
                    "spg_aisc_per_oz": ":.0f",
                    "spg_cash_cost_oz": ":.0f",
                    "spg_production_oz": ":,.0f",
                } if "name" in _aisc_rows.columns else {},
                title=f"AISC vs Gold Spot ~${_gold_spot:,.0f}/oz  |  {len(_aisc_rows)} producers",
            )
            # Gold spot reference line
            _fig_aisc.add_hline(
                y=_gold_spot,
                line_dash="dot",
                line_color="#FFD700",
                annotation_text=f"Spot ${_gold_spot:,.0f}",
                annotation_position="top right",
            )
            _fig_aisc.update_layout(
                height=340,
                margin=dict(t=50, b=30, l=10, r=10),
                coloraxis_showscale=False,
                xaxis_tickangle=-45,
            )
            st.plotly_chart(_fig_aisc, use_container_width=True)
            st.caption("Source: S&P Global SNL Metals & Mining via Snowflake (live, not stored). "
                       "Sorted by AISC ascending — leftmost = lowest cost = widest margin.")

# ── TAB 2: Charts ──────────────────────────────────────────────────────────────
with tab_charts:
    col_l, col_r = st.columns(2)

    with col_l:
        st.markdown("#### Score Distribution")
        fig = px.histogram(
            filtered, x="score_composite", nbins=20,
            color="stage",
            title="Composite Score Distribution",
            labels={"score_composite": "Score", "count": "# Companies"},
            color_discrete_sequence=px.colors.qualitative.Set2,
        )
        fig.update_layout(showlegend=True, height=350)
        st.plotly_chart(fig, width="stretch")

    with col_r:
        st.markdown("#### Valuation vs Health")
        fig2 = px.scatter(
            filtered.dropna(subset=["score_valuation", "score_health"]),
            x="score_valuation", y="score_health",
            color="commodity", size="score_composite",
            hover_name="name",
            hover_data={"ticker": True, "grade": True, "stage": True},
            title="Valuation Score vs Health Score (bubble = composite score)",
            labels={"score_valuation": "Valuation Score",
                    "score_health":    "Health Score"},
        )
        fig2.add_hline(y=50, line_dash="dot", line_color="gray")
        fig2.add_vline(x=50, line_dash="dot", line_color="gray")
        fig2.update_layout(height=350)
        st.plotly_chart(fig2, width="stretch")

    col_l2, col_r2 = st.columns(2)

    with col_l2:
        st.markdown("#### P/B vs EV/EBITDA")
        plot_df = filtered.dropna(subset=["price_to_book", "ev_ebitda"])
        # Keep only sensible positive values; drop negative EBITDA (developers/explorers)
        # and extreme outliers that wreck the log scale
        plot_df = plot_df[
            (plot_df["price_to_book"] > 0) & (plot_df["price_to_book"] < 20) &
            (plot_df["ev_ebitda"]     > 0.5) & (plot_df["ev_ebitda"]    < 80)
        ]
        if not plot_df.empty:
            fig3 = px.scatter(
                plot_df,
                x="price_to_book", y="ev_ebitda",
                color="stage",
                size="score_composite",
                size_max=18,
                hover_name="name",
                hover_data={"score_composite": True, "commodity": True,
                            "price_to_book": ":.2f", "ev_ebitda": ":.1f"},
                title="P/B vs EV/EBITDA (bottom-left = cheap)",
                labels={"price_to_book": "Price / Book",
                        "ev_ebitda":     "EV / EBITDA"},
            )
            # "Cheap" zone: P/B < 2, EV/EBITDA < 8 — drawn in data coords (linear scale)
            fig3.add_shape(
                type="rect", x0=0, x1=2, y0=0, y1=8,
                fillcolor="rgba(34,197,94,0.10)",
                line=dict(color="rgba(34,197,94,0.6)", dash="dot", width=1.5),
            )
            fig3.add_annotation(
                x=1, y=7.5, text="Value Zone", showarrow=False,
                font=dict(color="#22c55e", size=11),
            )
            fig3.update_layout(
                height=380,
                xaxis=dict(range=[0, min(plot_df["price_to_book"].max() * 1.1, 15)],
                           title="Price / Book"),
                yaxis=dict(range=[0, min(plot_df["ev_ebitda"].max() * 1.1, 60)],
                           title="EV / EBITDA"),
                legend=dict(orientation="v", x=1.01),
            )
            st.plotly_chart(fig3, width="stretch")
            _excluded = filtered[(filtered["ev_ebitda"].notna()) & (filtered["ev_ebitda"] <= 0)]
            if not _excluded.empty:
                st.caption(
                    f"⚠️ {len(_excluded)} companies excluded (negative EBITDA): "
                    + ", ".join(_excluded["name"].head(5).tolist())
                    + ("…" if len(_excluded) > 5 else "")
                )

    with col_r2:
        st.markdown("#### Top 20 by Composite Score")
        top20 = filtered.nlargest(20, "score_composite")
        fig4 = px.bar(
            top20[::-1],
            x="score_composite", y="name",
            color="grade",
            orientation="h",
            color_discrete_map={
                "🟢 Strong Buy": "#22c55e",
                "🔵 Buy":        "#3b82f6",
                "🟡 Watch":      "#eab308",
                "🟠 Neutral":    "#f97316",
                "🔴 Avoid":      "#ef4444",
            },
            labels={"score_composite": "Score", "name": ""},
            title="Top 20 Companies",
        )
        fig4.update_layout(showlegend=False, height=550)
        st.plotly_chart(fig4, width="stretch")

    # ── Score Movers ──────────────────────────────────────────────────────────
    _has_delta = "score_delta" in filtered.columns and filtered["score_delta"].notna().any()
    if _has_delta:
        _delta_nonzero = filtered[filtered["score_delta"].abs() > 0.4].copy()
        if not _delta_nonzero.empty:
            col_mup, col_mdn = st.columns(2)
            _top_ups = _delta_nonzero.nlargest(12, "score_delta")
            _top_dns = _delta_nonzero.nsmallest(12, "score_delta")

            with col_mup:
                st.markdown("#### ⬆️ Score Upgrades (vs prev)")
                if not _top_ups.empty:
                    _fig_ups = px.bar(
                        _top_ups[::-1],
                        x="score_delta", y="name",
                        orientation="h",
                        color="score_composite",
                        color_continuous_scale="RdYlGn",
                        range_color=[0, 100],
                        hover_name="name",
                        hover_data={"ticker": True, "score_composite": ":.1f",
                                    "score_delta": ":+.1f", "grade": True},
                        labels={"score_delta": "Δ Score", "name": "",
                                "score_composite": "Score"},
                        title="",
                    )
                    _fig_ups.update_layout(
                        showlegend=False, height=380,
                        margin=dict(l=10, r=10, t=10, b=10),
                        coloraxis_showscale=False,
                    )
                    st.plotly_chart(_fig_ups, width="stretch")

            with col_mdn:
                st.markdown("#### ⬇️ Score Downgrades (vs prev)")
                if not _top_dns.empty:
                    _fig_dns = px.bar(
                        _top_dns,
                        x="score_delta", y="name",
                        orientation="h",
                        color="score_composite",
                        color_continuous_scale="RdYlGn",
                        range_color=[0, 100],
                        hover_name="name",
                        hover_data={"ticker": True, "score_composite": ":.1f",
                                    "score_delta": ":+.1f", "grade": True},
                        labels={"score_delta": "Δ Score", "name": "",
                                "score_composite": "Score"},
                        title="",
                    )
                    _fig_dns.update_layout(
                        showlegend=False, height=380,
                        margin=dict(l=10, r=10, t=10, b=10),
                        coloraxis_showscale=False,
                    )
                    st.plotly_chart(_fig_dns, width="stretch")

    # ── Analyst Conviction charts ──────────────────────────────────────────────
    _has_analyst = (
        "analyst_upside" in filtered.columns
        and "analyst_count" in filtered.columns
        and filtered["analyst_upside"].notna().any()
    )
    if _has_analyst:
        col_la, col_ra = st.columns(2)

        with col_la:
            st.markdown("#### Analyst Score vs Composite Score")
            _an_df = filtered.dropna(subset=["analyst_upside", "score_composite"])
            _an_df = _an_df[
                (_an_df["analyst_count"].fillna(0) >= 1) &
                (_an_df["analyst_upside"] >= -100) &
                (_an_df["analyst_upside"] <= 300)
            ].copy()
            if not _an_df.empty:
                fig_an = px.scatter(
                    _an_df,
                    x="score_composite",
                    y="analyst_upside",
                    color="commodity",
                    size="analyst_count",
                    size_max=18,
                    hover_name="name",
                    hover_data={
                        "ticker": True,
                        "stage": True,
                        "analyst_count": True,
                        "analyst_upside": ":.1f",
                        "score_composite": ":.1f",
                    },
                    title="Quant Score vs Analyst Upside",
                    labels={
                        "score_composite": "Composite Score",
                        "analyst_upside":  "Analyst Upside %",
                    },
                )
                # Quadrant lines
                fig_an.add_hline(y=50, line_dash="dot", line_color="gray", opacity=0.5)
                fig_an.add_vline(x=50, line_dash="dot", line_color="gray", opacity=0.5)
                # Highlight conviction zone (top-right)
                fig_an.add_shape(
                    type="rect", x0=50, x1=100, y0=50, y1=300,
                    fillcolor="rgba(34,197,94,0.08)",
                    line=dict(color="rgba(34,197,94,0.5)", dash="dot", width=1),
                )
                fig_an.add_annotation(
                    x=75, y=280, text="Conviction Zone",
                    showarrow=False, font=dict(color="#22c55e", size=10),
                )
                fig_an.update_layout(height=400)
                st.plotly_chart(fig_an, width="stretch")
                st.caption(
                    f"Bubble size = # analysts. "
                    f"Top-right quadrant = high quant score AND high analyst upside. "
                    f"Showing {len(_an_df)} companies with analyst coverage."
                )
            else:
                st.info("Not enough analyst data to display this chart.")

        with col_ra:
            st.markdown("#### Top 20 by Analyst Upside")
            _top_an = (
                filtered[
                    (filtered["analyst_upside"].notna()) &
                    (filtered["analyst_count"].fillna(0) >= 2)
                ]
                .nlargest(20, "analyst_upside")
            )
            if not _top_an.empty:
                _top_an_rev = _top_an[::-1].copy()
                _top_an_rev["_label"] = _top_an_rev["analyst_upside"].map(
                    lambda v: f"+{v:.0f}%" if v >= 0 else f"{v:.0f}%"
                )
                fig_an2 = px.bar(
                    _top_an_rev,
                    x="analyst_upside",
                    y="name",
                    color="score_composite",
                    color_continuous_scale="RdYlGn",
                    range_color=[20, 90],
                    orientation="h",
                    text="_label",
                    hover_data={"analyst_count": True, "score_composite": ":.1f",
                                "stage": True, "commodity": True},
                    title="Top 20: Analyst Consensus Upside (≥2 analysts)",
                    labels={
                        "analyst_upside":   "Analyst Upside %",
                        "name":             "",
                        "score_composite":  "Score",
                    },
                )
                fig_an2.update_traces(textposition="outside")
                fig_an2.update_layout(
                    height=550,
                    coloraxis_colorbar=dict(title="Score", thickness=12, len=0.7),
                    xaxis=dict(title="Analyst Upside %"),
                )
                st.plotly_chart(fig_an2, width="stretch")
                st.caption("Bar colour = composite score (green = high). Requires ≥ 2 analyst opinions.")
            else:
                st.info("No analyst upside data available (need ≥ 2 analyst opinions).")

    # ── P/NAV Distribution ────────────────────────────────────────────────────
    _has_pnav = "spg_p_nav" in filtered.columns and filtered["spg_p_nav"].notna().any()
    if _has_pnav:
        col_pnav_l, col_pnav_r = st.columns(2)

        with col_pnav_l:
            st.markdown("#### P/NAV Distribution")
            _pnav_df = filtered[
                filtered["spg_p_nav"].notna() &
                (filtered["spg_p_nav"] > 0) &
                (filtered["spg_p_nav"] < 5)
            ].copy()
            if not _pnav_df.empty:
                fig_pnav = px.histogram(
                    _pnav_df,
                    x="spg_p_nav",
                    nbins=25,
                    color="stage",
                    title="P/NAV Distribution (S&P Capital IQ)",
                    labels={"spg_p_nav": "P/NAV", "count": "# Companies"},
                    color_discrete_sequence=px.colors.qualitative.Set2,
                )
                # Fair value line at 1.0
                fig_pnav.add_vline(
                    x=1.0, line_dash="solid", line_color="rgba(239,68,68,0.7)",
                    line_width=2,
                    annotation_text="NAV = 1.0×",
                    annotation_position="top right",
                    annotation_font_color="#ef4444",
                )
                # Median line
                _pnav_med = _pnav_df["spg_p_nav"].median()
                fig_pnav.add_vline(
                    x=_pnav_med, line_dash="dot", line_color="rgba(59,130,246,0.7)",
                    line_width=1.5,
                    annotation_text=f"Median {_pnav_med:.2f}×",
                    annotation_position="top left",
                    annotation_font_color="#3b82f6",
                )
                fig_pnav.update_layout(height=360, bargap=0.05)
                st.plotly_chart(fig_pnav, width="stretch")
                _below_nav = ((_pnav_df["spg_p_nav"] < 1.0).sum())
                st.caption(
                    f"{len(_pnav_df)} companies with S&P NAV coverage · "
                    f"{_below_nav} ({_below_nav/len(_pnav_df)*100:.0f}%) trading below NAV (P/NAV < 1.0×)"
                )

        with col_pnav_r:
            st.markdown("#### P/NAV vs Composite Score")
            _pnav_sc = _pnav_df.dropna(subset=["score_composite"])
            if not _pnav_sc.empty:
                fig_pnav2 = px.scatter(
                    _pnav_sc,
                    x="spg_p_nav",
                    y="score_composite",
                    color="commodity",
                    size="market_cap",
                    size_max=18,
                    hover_name="name",
                    hover_data={
                        "ticker": True,
                        "stage": True,
                        "spg_p_nav": ":.3f",
                        "score_composite": ":.1f",
                    },
                    title="P/NAV vs Composite Score",
                    labels={
                        "spg_p_nav":       "P/NAV",
                        "score_composite": "Composite Score",
                    },
                )
                # Fair value and score threshold lines
                fig_pnav2.add_vline(x=1.0, line_dash="dot", line_color="rgba(239,68,68,0.5)")
                fig_pnav2.add_hline(y=60, line_dash="dot", line_color="rgba(34,197,94,0.5)",
                                    annotation_text="Buy", annotation_position="right")
                # Deep value zone: P/NAV < 0.7, Score > 60
                fig_pnav2.add_shape(
                    type="rect", x0=0, x1=0.7, y0=60, y1=105,
                    fillcolor="rgba(34,197,94,0.08)",
                    line=dict(color="rgba(34,197,94,0.4)", dash="dot", width=1),
                )
                fig_pnav2.add_annotation(
                    x=0.35, y=100, text="Deep Value",
                    showarrow=False, font=dict(color="#22c55e", size=10),
                )
                fig_pnav2.update_layout(height=360)
                st.plotly_chart(fig_pnav2, width="stretch")

    st.markdown("#### Score Heatmap by Commodity")
    pivot = (
        filtered.groupby(["commodity", "stage"])["score_composite"]
        .mean().unstack(fill_value=np.nan)
    )
    if not pivot.empty:
        n_rows = len(pivot.index)
        n_cols = len(pivot.columns)
        cell_h = max(48, min(72, 600 // n_rows))   # px per row
        cell_w = max(120, min(200, 900 // n_cols))  # px per col
        fig5 = px.imshow(
            pivot.round(1),
            text_auto=True,
            color_continuous_scale="RdYlGn",
            title="Avg Composite Score: Commodity × Stage",
            zmin=20, zmax=80,
            aspect="auto",
        )
        fig5.update_traces(
            xgap=3, ygap=3,
            textfont=dict(size=14, color="white"),
        )
        fig5.update_layout(
            height=n_rows * cell_h + 160,
            margin=dict(l=120, r=80, t=60, b=60),
            xaxis=dict(side="bottom", tickfont=dict(size=13)),
            yaxis=dict(tickfont=dict(size=13)),
            coloraxis_colorbar=dict(thickness=14, len=0.8),
        )
        st.plotly_chart(fig5, width="stretch")

    # ── Sector trend over time ─────────────────────────────────────────────────
    # ── Return Performance by Commodity ──────────────────────────────────────
    _has_ret3m = "return_3m" in filtered.columns and filtered["return_3m"].notna().any()
    if _has_ret3m:
        col_ret_l, col_ret_r = st.columns(2)

        with col_ret_l:
            st.markdown("#### 3M Return: Commodity Median")
            _ret_comm = (
                filtered.dropna(subset=["return_3m"])
                .groupby("commodity", as_index=False)
                .agg(
                    median_ret3m=("return_3m", "median"),
                    n=("return_3m", "count"),
                )
                .sort_values("median_ret3m", ascending=True)
            )
            if not _ret_comm.empty:
                _ret_comm["color"] = _ret_comm["median_ret3m"].apply(
                    lambda v: "#22c55e" if v >= 0 else "#ef4444"
                )
                _ret_comm["label"] = _ret_comm["median_ret3m"].map(
                    lambda v: f"{v:+.1f}%"
                )
                fig_ret = px.bar(
                    _ret_comm,
                    x="median_ret3m",
                    y="commodity",
                    orientation="h",
                    text="label",
                    color="median_ret3m",
                    color_continuous_scale=[
                        [0.0, "#ef4444"], [0.5, "#f3f4f6"], [1.0, "#22c55e"]
                    ],
                    range_color=[-30, 30],
                    hover_data={"n": True},
                    title="Median 3M Price Return by Commodity",
                    labels={"median_ret3m": "3M Return %", "commodity": ""},
                )
                fig_ret.update_traces(textposition="outside")
                fig_ret.update_layout(
                    height=400,
                    coloraxis_showscale=False,
                    xaxis=dict(title="3M Return %", zeroline=True,
                               zerolinecolor="#6b7280", zerolinewidth=1.5),
                )
                fig_ret.add_vline(x=0, line_width=1.5, line_color="#6b7280")
                st.plotly_chart(fig_ret, width="stretch")
                st.caption(f"Median 3-month price return across {len(filtered.dropna(subset=['return_3m']))} companies with data.")

        with col_ret_r:
            st.markdown("#### 1M vs 3M Returns: Momentum Map")
            _has_ret1m = "return_1m" in filtered.columns and filtered["return_1m"].notna().any()
            if _has_ret1m:
                _mom_df = filtered.dropna(subset=["return_1m", "return_3m"]).copy()
                if not _mom_df.empty:
                    fig_mom = px.scatter(
                        _mom_df,
                        x="return_3m",
                        y="return_1m",
                        color="commodity",
                        size="score_composite",
                        size_max=15,
                        hover_name="name",
                        hover_data={
                            "ticker": True,
                            "stage": True,
                            "return_1m": ":.1f",
                            "return_3m": ":.1f",
                        },
                        title="Momentum: 1M vs 3M Returns",
                        labels={"return_3m": "3M Return %", "return_1m": "1M Return %"},
                    )
                    fig_mom.add_hline(y=0, line_dash="dot", line_color="gray", opacity=0.5)
                    fig_mom.add_vline(x=0, line_dash="dot", line_color="gray", opacity=0.5)
                    # Quadrant labels
                    for (tx, ty, label, color) in [
                        (20, 10, "Acceleration", "#22c55e"),
                        (-20, 10, "Recovery", "#3b82f6"),
                        (20, -10, "Fading", "#f97316"),
                        (-20, -10, "Weakness", "#ef4444"),
                    ]:
                        fig_mom.add_annotation(
                            x=tx, y=ty, text=label, showarrow=False,
                            font=dict(color=color, size=9), opacity=0.6,
                        )
                    fig_mom.update_layout(height=400)
                    st.plotly_chart(fig_mom, width="stretch")
                    st.caption("Bubble size = composite score. Top-right = accelerating momentum.")
            else:
                st.info("1M return data not available.")

    st.markdown("#### Sector Score Trends Over Time")
    _trends = load_sector_trends()
    if _trends.empty or _trends["snap_date"].nunique() < 2:
        st.info("Sector trends will appear here after multiple daily refreshes. "
                "Each 🔄 Refresh builds one snapshot; come back tomorrow for the first trend line.")
    else:
        _trends["snap_date"] = pd.to_datetime(_trends["snap_date"])

        # Filter to main commodity groups with enough data
        _main_comms = (
            _trends.groupby("commodity_group")["n_companies"]
            .mean()
            .where(lambda s: s >= 2)
            .dropna()
            .index.tolist()
        )
        _trends_main = _trends[_trends["commodity_group"].isin(_main_comms)]

        _COMM_COLORS = {
            "Gold":     "#FFD700",
            "Silver":   "#C0C0C0",
            "Copper":   "#B87333",
            "Uranium":  "#7CFC00",
            "Nickel":   "#A9A9A9",
            "Lithium":  "#00BFFF",
            "Iron Ore": "#CD5C5C",
            "Zinc":     "#708090",
            "Other":    "#778899",
        }

        fig_trend = px.line(
            _trends_main,
            x="snap_date", y="avg_score",
            color="commodity_group",
            color_discrete_map=_COMM_COLORS,
            markers=True,
            labels={
                "snap_date":      "Date",
                "avg_score":      "Avg Composite Score",
                "commodity_group":"Commodity",
                "n_companies":    "# Companies",
            },
            hover_data={"n_companies": True},
            title="Average Composite Score by Commodity — Historical",
        )
        fig_trend.add_hline(y=60, line_dash="dot", line_color="rgba(59,130,246,0.4)",
                            annotation_text="Buy threshold", annotation_position="right")
        fig_trend.add_hline(y=75, line_dash="dot", line_color="rgba(34,197,94,0.4)",
                            annotation_text="Strong Buy", annotation_position="right")
        fig_trend.update_layout(height=420, legend_title_text="Commodity",
                                hovermode="x unified")
        st.plotly_chart(fig_trend, width="stretch")

        # Mini table: latest avg score per sector — light-theme polished
        _latest_date = _trends["snap_date"].max()
        _latest_raw = (
            _trends[_trends["snap_date"] == _latest_date]
            .sort_values("avg_score", ascending=False)
            .rename(columns={
                "commodity_group": "Commodity",
                "avg_score":       "Avg Score",
                "n_companies":     "Companies",
            })
            [["Commodity", "Avg Score", "Companies"]]
            .reset_index(drop=True)
        )

        def _score_bar_light(val: float, max_val: float = 100) -> str:
            """Return an HTML progress-bar cell for Avg Score (light theme)."""
            if pd.isna(val):
                return "—"
            pct = max(0, min(100, val / max_val * 100))
            if val >= 75:
                bar_color, badge_bg, badge_fg = "#16a34a", "#dcfce7", "#15803d"
            elif val >= 60:
                bar_color, badge_bg, badge_fg = "#2563eb", "#dbeafe", "#1d4ed8"
            elif val >= 45:
                bar_color, badge_bg, badge_fg = "#d97706", "#fef3c7", "#b45309"
            else:
                bar_color, badge_bg, badge_fg = "#dc2626", "#fee2e2", "#b91c1c"
            return (
                f'<div style="display:flex;align-items:center;gap:8px;">'
                f'<div style="flex:1;background:#e5e7eb;border-radius:6px;height:9px;overflow:hidden;">'
                f'<div style="width:{pct:.0f}%;background:{bar_color};height:100%;border-radius:6px;'
                f'transition:width 0.3s;"></div>'
                f'</div>'
                f'<span style="min-width:40px;text-align:right;font-weight:700;font-size:0.83em;'
                f'background:{badge_bg};color:{badge_fg};padding:2px 7px;border-radius:5px;'
                f'border:1px solid {bar_color}22;letter-spacing:0.01em;">'
                f'{val:.1f}</span>'
                f'</div>'
            )

        _tbl_rows_light = []
        for i, row in _latest_raw.iterrows():
            score = row["Avg Score"]
            n     = int(row["Companies"]) if not pd.isna(row["Companies"]) else 0
            bar   = _score_bar_light(score)
            row_bg = "#ffffff" if i % 2 == 0 else "#f9fafb"
            _tbl_rows_light.append(
                f'<tr style="background:{row_bg};border-bottom:1px solid #f3f4f6;">'
                f'<td style="padding:8px 14px;font-weight:600;color:#111827;width:140px;'
                f'white-space:nowrap;">{row["Commodity"]}</td>'
                f'<td style="padding:8px 14px;">{bar}</td>'
                f'<td style="padding:8px 14px;text-align:center;color:#6b7280;'
                f'font-variant-numeric:tabular-nums;font-weight:500;">{n}</td>'
                f'</tr>'
            )

        _tbl_html = (
            '<div style="border:1px solid #e5e7eb;border-radius:10px;overflow:hidden;'
            'box-shadow:0 1px 4px rgba(0,0,0,0.06);background:#fff;">'
            '<table style="width:100%;border-collapse:collapse;font-size:0.875em;">'
            '<thead><tr style="background:#f8fafc;border-bottom:2px solid #e2e8f0;">'
            '<th style="padding:9px 14px;text-align:left;font-weight:700;color:#374151;'
            'font-size:0.75em;text-transform:uppercase;letter-spacing:0.07em;width:140px;">'
            'Commodity</th>'
            '<th style="padding:9px 14px;text-align:left;font-weight:700;color:#374151;'
            'font-size:0.75em;text-transform:uppercase;letter-spacing:0.07em;">'
            'Avg Score</th>'
            '<th style="padding:9px 14px;text-align:center;font-weight:700;color:#374151;'
            'font-size:0.75em;text-transform:uppercase;letter-spacing:0.07em;width:70px;">'
            'Companies</th>'
            '</tr></thead>'
            '<tbody>'
            + "".join(_tbl_rows_light)
            + '</tbody></table></div>'
        )

        st.caption(f"Sector snapshot — {_latest_date.strftime('%Y-%m-%d')}")
        st.markdown(_tbl_html, unsafe_allow_html=True)

        # Macro overlay — gold spot price on sector trend chart
        _gold_hist = load_commodity_price_history(days=180)
        if not _gold_hist.empty:
            _gold_ts = _gold_hist[_gold_hist["commodity"] == "Gold"].copy()
            if not _gold_ts.empty and _gold_ts["price_date"].nunique() >= 2:
                _gold_ts["price_date"] = pd.to_datetime(_gold_ts["price_date"])
                _gold_ts = _gold_ts.sort_values("price_date")

                st.markdown("**📊 Macro Overlay — Gold Spot vs Sector Scores**")
                st.caption("Secondary axis (right): Gold spot price USD/oz. Shows how sector scores correlate with commodity price moves.")

                _macro_comms = st.multiselect(
                    "Commodity groups to overlay",
                    options=_main_comms,
                    default=_main_comms[:3] if len(_main_comms) >= 3 else _main_comms,
                    key="macro_overlay_comms",
                )
                if _macro_comms:
                    _mo_trend = _trends_main[_trends_main["commodity_group"].isin(_macro_comms)]

                    fig_macro = go.Figure()

                    # Score lines (primary y)
                    for _mc in _macro_comms:
                        _mdf = _mo_trend[_mo_trend["commodity_group"] == _mc]
                        fig_macro.add_trace(go.Scatter(
                            x=_mdf["snap_date"], y=_mdf["avg_score"],
                            name=_mc,
                            line=dict(color=_COMM_COLORS.get(_mc, "#94a3b8"), width=2),
                            mode="lines+markers", marker=dict(size=5),
                            yaxis="y1",
                        ))

                    # Gold spot line (secondary y)
                    fig_macro.add_trace(go.Scatter(
                        x=_gold_ts["price_date"], y=_gold_ts["price"],
                        name="Gold Spot (USD/oz)",
                        line=dict(color="rgba(255,215,0,0.7)", width=2, dash="dot"),
                        mode="lines",
                        yaxis="y2",
                    ))

                    fig_macro.update_layout(
                        yaxis=dict(
                            title="Avg Composite Score",
                            range=[0, 100],
                            side="left",
                        ),
                        yaxis2=dict(
                            title="Gold Spot USD/oz",
                            overlaying="y",
                            side="right",
                            showgrid=False,
                        ),
                        xaxis=dict(title="Date"),
                        height=400,
                        hovermode="x unified",
                        legend=dict(orientation="h", y=-0.2),
                        margin=dict(t=10, b=60),
                    )
                    st.plotly_chart(fig_macro, width="stretch")

    # ── Commodity Spot Price Trends ────────────────────────────────────────
    st.markdown("#### Commodity Spot Price Trends")
    _cph = load_commodity_price_history(days=180)
    if _cph.empty or _cph["price_date"].nunique() < 2:
        st.info(
            "Commodity price history will appear here after multiple daily refreshes. "
            "Each 🔄 Refresh stores today's spot prices."
        )
    else:
        _cph["price_date"] = pd.to_datetime(_cph["price_date"])
        # Normalise to 100 at first date so all commodities are on the same axis
        _first_date = _cph["price_date"].min()
        _base = (
            _cph[_cph["price_date"] == _first_date]
            .set_index("commodity")["price"]
        )
        def _normalise(row_c):
            base = _base.get(row_c["commodity"], row_c["price"])
            return round(row_c["price"] / base * 100, 2) if base else row_c["price"]

        _cph["indexed"] = _cph.apply(_normalise, axis=1)

        _COMM_COLORS_SPOT = {
            "Gold":    "#FFD700",
            "Silver":  "#C0C0C0",
            "Copper":  "#B87333",
            "Uranium": "#7CFC00",
            "Nickel":  "#A9A9A9",
            "Zinc":    "#708090",
        }
        _cph_main = _cph[_cph["commodity"].isin(_COMM_COLORS_SPOT)]
        if not _cph_main.empty:
            fig_spot = px.line(
                _cph_main,
                x="price_date", y="indexed",
                color="commodity",
                color_discrete_map=_COMM_COLORS_SPOT,
                markers=False,
                labels={
                    "price_date": "Date",
                    "indexed":    "Price (indexed to 100 at start)",
                    "commodity":  "Commodity",
                },
                title=f"Commodity Spot Prices — Indexed to 100 (base: {_first_date.strftime('%Y-%m-%d')})",
                hover_data={"price": True},
            )
            fig_spot.add_hline(y=100, line_dash="dot", line_color="gray",
                               annotation_text="Base", annotation_position="right")
            fig_spot.update_layout(height=380, hovermode="x unified",
                                   legend_title_text="Commodity")
            st.plotly_chart(fig_spot, width="stretch")

            # Latest prices table
            _spot_latest = (
                _cph.sort_values("price_date")
                .groupby("commodity")
                .last()
                .reset_index()[["commodity", "price"]]
                .rename(columns={"commodity": "Commodity", "price": "Latest Price"})
                .set_index("Commodity")
                .sort_values("Latest Price", ascending=False)
            )
            st.caption("Latest stored spot prices")
            st.dataframe(
                _spot_latest.style.format({"Latest Price": "{:,.2f}"}),
                width="stretch",
                height=min(60 + len(_spot_latest) * 35, 280),
            )

    # ── AISC Cost Curve ────────────────────────────────────────────────────────
    st.markdown("#### ⛏️ AISC Cost Curve")
    _cc_comm = st.selectbox(
        "Commodity",
        ["Gold / Silver", "Copper", "Iron Ore"],
        key="cc_commodity",
        label_visibility="collapsed",
    )
    if _cc_comm == "Gold / Silver":
        _cc_df = df[
            df["spg_aisc_per_oz"].notna() &
            (df["spg_aisc_per_oz"] >= 200) &          # copper/base-metal $/oz artefacts are near $0
            df["commodity"].str.contains("Gold|Silver", case=False, na=False)
        ].copy()
        _cc_y  = "spg_aisc_per_oz"
        _cc_ylabel = "AISC ($/oz)"
        _cc_spot   = config.COMMODITY_SPOT.get("Gold")
        _cc_spot_label = f"Gold spot ~${_cc_spot:,.0f}/oz" if _cc_spot else None
    elif _cc_comm == "Copper":
        _cc_df = df[df["spg_aisc_per_t"].notna() & (df["spg_aisc_per_t"] > 0)
                    & (df["commodity"] == "Copper")].copy()
        _cc_y  = "spg_aisc_per_t"
        _cc_ylabel = "AISC ($/t)"
        _cc_spot   = config.COMMODITY_SPOT.get("Copper")
        _cc_spot_label = f"Copper spot ~${_cc_spot:,.0f}/t" if _cc_spot else None
    else:  # Iron Ore
        _cc_df = df[df["spg_aisc_per_t"].notna() & (df["spg_aisc_per_t"] > 0)
                    & (df["commodity"] == "Iron Ore")].copy()
        _cc_y  = "spg_aisc_per_t"
        _cc_ylabel = "AISC ($/t)"
        _cc_spot   = None
        _cc_spot_label = None

    if _cc_df.empty:
        st.info("No AISC data yet. Run _ciq_spg_mining.py then click 🔄 Refresh Now.")
    else:
        _cc_df = _cc_df.sort_values(_cc_y)
        _cc_df["_aisc_margin"] = (
            ((_cc_spot - _cc_df[_cc_y]) / _cc_spot * 100).round(1)
            if _cc_spot else float("nan")
        )
        _cc_colors = _cc_df["_aisc_margin"].apply(
            lambda m: "#22c55e" if pd.notna(m) and m > 40
            else "#3b82f6" if pd.notna(m) and m > 20
            else "#eab308" if pd.notna(m) and m > 0
            else "#ef4444"
        )

        # Determine cash cost column for this commodity group
        if _cc_comm == "Gold / Silver":
            _cc_cash_col = "spg_cash_cost_oz"
        elif _cc_comm == "Copper":
            _cc_cash_col = "spg_cash_cost_lb"
        else:
            _cc_cash_col = "spg_cash_cost_t"

        _has_cash_cost = (
            _cc_cash_col in _cc_df.columns
            and _cc_df[_cc_cash_col].notna().any()
            and (_cc_df[_cc_cash_col] > 0).any()
        )

        fig_cc = go.Figure()

        if _has_cash_cost:
            # Stacked layout: cash cost (base) + sustaining capex wedge
            _cc_df["_cash_cost"] = pd.to_numeric(_cc_df[_cc_cash_col], errors="coerce")
            _cc_df["_sustaining"] = (_cc_df[_cc_y] - _cc_df["_cash_cost"]).clip(lower=0)
            _cc_df["_cash_cost_clipped"] = _cc_df["_cash_cost"].clip(lower=0)

            # Base: cash cost
            fig_cc.add_trace(go.Bar(
                x=_cc_df["name"],
                y=_cc_df["_cash_cost_clipped"],
                name="Cash Cost",
                marker_color="rgba(59,130,246,0.85)",
                customdata=np.column_stack([
                    _cc_df["name"],
                    _cc_df["_cash_cost_clipped"].round(0),
                    _cc_df[_cc_y].round(0),
                    _cc_df["_aisc_margin"].fillna(float("nan")),
                ]),
                hovertemplate=(
                    "<b>%{customdata[0]}</b><br>"
                    f"Cash Cost: $%{{customdata[1]:,.0f}}<br>"
                    f"AISC: $%{{customdata[2]:,.0f}}<br>"
                    "Margin: %{customdata[3]:.1f}%<extra></extra>"
                ),
            ))
            # Sustaining capex wedge
            fig_cc.add_trace(go.Bar(
                x=_cc_df["name"],
                y=_cc_df["_sustaining"],
                name="Sustaining Capex",
                marker_color=_cc_colors.tolist(),
                text=_cc_df[_cc_y].apply(lambda v: f"${v:,.0f}"),
                textposition="outside",
                customdata=np.column_stack([
                    _cc_df["name"],
                    _cc_df["_sustaining"].round(0),
                    _cc_df[_cc_y].round(0),
                ]),
                hovertemplate=(
                    "<b>%{customdata[0]}</b><br>"
                    f"Sustaining Capex: $%{{customdata[1]:,.0f}}<br>"
                    f"AISC Total: $%{{customdata[2]:,.0f}}<extra></extra>"
                ),
            ))
            fig_cc.update_layout(barmode="stack", showlegend=True,
                                 legend=dict(orientation="h", y=1.05, x=0))
        else:
            # Fallback: simple bar coloured by margin
            fig_cc.add_trace(go.Bar(
                x=_cc_df["name"],
                y=_cc_df[_cc_y],
                marker_color=_cc_colors.tolist(),
                text=_cc_df[_cc_y].apply(lambda v: f"${v:,.0f}"),
                textposition="outside",
                customdata=np.column_stack([
                    _cc_df["name"],
                    _cc_df[_cc_y].round(0),
                    _cc_df["_aisc_margin"].fillna(float("nan")),
                    _cc_df.get("market_cap", pd.Series([float("nan")] * len(_cc_df))).fillna(float("nan")) / 1e6,
                ]),
                hovertemplate=(
                    "<b>%{customdata[0]}</b><br>"
                    f"{_cc_ylabel}: $%{{customdata[1]:,.0f}}<br>"
                    "AISC Margin: %{customdata[2]:.1f}%<br>"
                    "Mkt Cap: $%{customdata[3]:,.0f}M<extra></extra>"
                ),
            ))

        if _cc_spot and _cc_spot_label:
            fig_cc.add_hline(
                y=_cc_spot,
                line_dash="dash",
                line_color="#FFD700",
                line_width=2,
                annotation_text=_cc_spot_label,
                annotation_position="top right",
                annotation_font_color="#FFD700",
            )
        fig_cc.update_layout(
            height=420,
            yaxis_title=_cc_ylabel,
            xaxis_tickangle=-35,
            margin=dict(t=30, b=120),
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)",
        )
        st.plotly_chart(fig_cc, width="stretch")
        _cc_caption = (
            "🔵 Cash Cost  + coloured wedge = Sustaining Capex  →  stacked bar = AISC  |  "
            "Gold line = current spot price  |  Bar colour: 🟢 >40% margin  🔵 20–40%  🟡 0–20%  🔴 above spot"
            if _has_cash_cost else
            "🟢 >40% margin  🔵 20–40%  🟡 0–20%  🔴 above spot  |  Gold line = current spot price"
        )
        st.caption(_cc_caption)

    # ── P/NAV Ladder ───────────────────────────────────────────────────────────
    _pnav_df = df[df["spg_p_nav"].notna() & (df["spg_p_nav"] > 0)
                  & (df["spg_p_nav"] < 10)].copy()
    if not _pnav_df.empty:
        st.markdown("#### 📐 P/NAV Ladder")
        _pnav_df = _pnav_df.sort_values("spg_p_nav")
        _pnav_df["_pnav_color"] = _pnav_df["spg_p_nav"].apply(
            lambda v: "#22c55e" if v < 0.75
            else "#3b82f6" if v < 1.0
            else "#eab308" if v < 1.5
            else "#ef4444"
        )
        fig_pnav = go.Figure()
        fig_pnav.add_vline(x=1.0, line_dash="dash", line_color="#FFD700",
                           line_width=2, annotation_text="NAV (1.0x)",
                           annotation_font_color="#FFD700")
        fig_pnav.add_trace(go.Scatter(
            x=_pnav_df["spg_p_nav"],
            y=_pnav_df["name"],
            mode="markers+text",
            marker=dict(
                size=14,
                color=_pnav_df["_pnav_color"].tolist(),
                symbol="circle",
                line=dict(width=1, color="white"),
            ),
            text=_pnav_df["spg_p_nav"].apply(lambda v: f"{v:.2f}x"),
            textposition="middle right",
            customdata=np.column_stack([
                _pnav_df["name"],
                _pnav_df["spg_p_nav"],
                _pnav_df.get("nav_per_shr", pd.Series([float("nan")] * len(_pnav_df))).fillna(float("nan")),
                _pnav_df["grade"],
            ]),
            hovertemplate=(
                "<b>%{customdata[0]}</b><br>"
                "P/NAV: %{customdata[1]:.2f}x<br>"
                "NAV/Share: $%{customdata[2]:.2f}<br>"
                "Grade: %{customdata[3]}<extra></extra>"
            ),
        ))
        fig_pnav.update_layout(
            height=max(300, len(_pnav_df) * 28),
            xaxis_title="Price / NAV",
            yaxis_title="",
            showlegend=False,
            margin=dict(l=160, r=80, t=30, b=30),
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)",
        )
        fig_pnav.update_xaxes(range=[0, min(_pnav_df["spg_p_nav"].max() * 1.2, 4.5)])
        st.plotly_chart(fig_pnav, width="stretch")
        st.caption("🟢 <0.75x (deep discount)  🔵 0.75–1.0x (discount)  🟡 1.0–1.5x  🔴 >1.5x (premium)")

    # ── EV/oz Production ──────────────────────────────────────────────────────
    if "ev_per_oz_prod" in df.columns and df["ev_per_oz_prod"].notna().any():
        st.markdown("#### 💰 EV/oz Annual Production — Valuation Efficiency")
        st.caption(
            "Enterprise value per ounce of annual attributable production. "
            "Lower = cheaper on a production basis. Sorted ascending — left = better value."
        )
        _ev_oz_comms = sorted(
            df.dropna(subset=["ev_per_oz_prod"])["commodity"]
            .str.split("/").str[0].str.strip().unique()
        )
        _ev_oz_sel = st.selectbox(
            "Commodity group",
            ["All"] + _ev_oz_comms,
            key="evoz_comm_sel",
        )
        _ev_oz_df = df[df["ev_per_oz_prod"].notna() & (df["ev_per_oz_prod"] > 0)].copy()
        if _ev_oz_sel != "All":
            _ev_oz_df = _ev_oz_df[_ev_oz_df["commodity"].str.startswith(_ev_oz_sel)]
        _ev_oz_df = _ev_oz_df.sort_values("ev_per_oz_prod").head(40)

        if not _ev_oz_df.empty:
            # Peer-median reference line
            _ev_oz_med = _ev_oz_df["ev_per_oz_prod"].median()
            # Colour by quartile: green (Q1 cheapest) → red (Q4 most expensive)
            _q25 = _ev_oz_df["ev_per_oz_prod"].quantile(0.25)
            _q75 = _ev_oz_df["ev_per_oz_prod"].quantile(0.75)
            _ev_oz_df["_evoz_color"] = _ev_oz_df["ev_per_oz_prod"].apply(
                lambda v: "#22c55e" if v <= _q25
                else "#3b82f6" if v <= _ev_oz_med
                else "#eab308" if v <= _q75
                else "#ef4444"
            )
            _ev_oz_df["_label"] = _ev_oz_df["name"] + " (" + _ev_oz_df["ticker"] + ")"

            fig_evoz = go.Figure()
            fig_evoz.add_trace(go.Bar(
                x=_ev_oz_df["_label"],
                y=_ev_oz_df["ev_per_oz_prod"],
                marker_color=_ev_oz_df["_evoz_color"].tolist(),
                text=_ev_oz_df["ev_per_oz_prod"].apply(lambda v: f"${v:,.0f}"),
                textposition="outside",
                customdata=np.column_stack([
                    _ev_oz_df["name"],
                    _ev_oz_df["ev_per_oz_prod"].round(0),
                    (_ev_oz_df.get("spg_production_oz", pd.Series([float("nan")] * len(_ev_oz_df))).fillna(0) / 1000).round(1),
                    _ev_oz_df["commodity"],
                ]),
                hovertemplate=(
                    "<b>%{customdata[0]}</b><br>"
                    "EV/oz Produced: $%{customdata[1]:,.0f}<br>"
                    "Production: %{customdata[2]:,.1f} koz/yr<br>"
                    "Commodity: %{customdata[3]}<extra></extra>"
                ),
            ))
            fig_evoz.add_hline(
                y=_ev_oz_med,
                line_dash="dash", line_color="#94a3b8", line_width=1.5,
                annotation_text=f"Median ${_ev_oz_med:,.0f}",
                annotation_position="top right",
                annotation_font_color="#94a3b8",
            )
            fig_evoz.update_layout(
                height=420,
                yaxis_title="EV / oz Annual Production ($)",
                xaxis_tickangle=-35,
                showlegend=False,
                margin=dict(t=30, b=140),
                plot_bgcolor="rgba(0,0,0,0)",
                paper_bgcolor="rgba(0,0,0,0)",
            )
            st.plotly_chart(fig_evoz, width="stretch")
            st.caption(
                "🟢 Q1 (cheapest)  🔵 Q2  🟡 Q3  🔴 Q4 (most expensive)  "
                "· Dashed line = group median"
            )

    # ── Reserve Life Bar Chart ─────────────────────────────────────────────────
    if "spg_reserve_life" in df.columns and df["spg_reserve_life"].notna().any():
        st.markdown("#### ⏳ Reserve Life Index")
        st.caption(
            "Contained reserves ÷ annual production — estimated years until reserves are exhausted. "
            "≥10 years considered healthy; <5 years indicates near-term depletion risk."
        )
        _rli_df = df[df["spg_reserve_life"].notna() & (df["spg_reserve_life"] > 0)].copy()
        _rli_comms = sorted(
            _rli_df["commodity"].str.split("/").str[0].str.strip().unique()
        )
        _rli_sel = st.selectbox(
            "Commodity group",
            ["All"] + _rli_comms,
            key="rli_comm_sel",
        )
        if _rli_sel != "All":
            _rli_df = _rli_df[_rli_df["commodity"].str.startswith(_rli_sel)]
        _rli_df = _rli_df.sort_values("spg_reserve_life", ascending=False).head(40)

        if not _rli_df.empty:
            _rli_df["_rli_color"] = _rli_df["spg_reserve_life"].apply(
                lambda v: "#22c55e" if v >= 15
                else "#3b82f6" if v >= 10
                else "#eab308" if v >= 5
                else "#ef4444"
            )
            _rli_df["_label"] = _rli_df["name"] + " (" + _rli_df["ticker"] + ")"

            fig_rli = go.Figure()
            fig_rli.add_trace(go.Bar(
                x=_rli_df["_label"],
                y=_rli_df["spg_reserve_life"],
                marker_color=_rli_df["_rli_color"].tolist(),
                text=_rli_df["spg_reserve_life"].apply(lambda v: f"{v:.1f} yr"),
                textposition="outside",
                customdata=np.column_stack([
                    _rli_df["name"],
                    _rli_df["spg_reserve_life"].round(1),
                    (_rli_df.get("spg_production_oz", pd.Series([float("nan")] * len(_rli_df))).fillna(0) / 1000).round(1),
                    _rli_df["commodity"],
                ]),
                hovertemplate=(
                    "<b>%{customdata[0]}</b><br>"
                    "Reserve Life: %{customdata[1]:.1f} yr<br>"
                    "Production: %{customdata[2]:,.1f} koz/yr<br>"
                    "Commodity: %{customdata[3]}<extra></extra>"
                ),
            ))
            # Threshold lines
            fig_rli.add_hline(
                y=10, line_dash="dash", line_color="#3b82f6", line_width=1.5,
                annotation_text="10 yr (healthy)", annotation_position="top right",
                annotation_font_color="#3b82f6",
            )
            fig_rli.add_hline(
                y=5, line_dash="dot", line_color="#ef4444", line_width=1,
                annotation_text="5 yr (caution)", annotation_position="top right",
                annotation_font_color="#ef4444",
            )
            fig_rli.update_layout(
                height=420,
                yaxis_title="Reserve Life (years)",
                xaxis_tickangle=-35,
                showlegend=False,
                margin=dict(t=30, b=140),
                plot_bgcolor="rgba(0,0,0,0)",
                paper_bgcolor="rgba(0,0,0,0)",
            )
            st.plotly_chart(fig_rli, width="stretch")
            st.caption(
                "🟢 ≥15 yr  🔵 10–15 yr  🟡 5–10 yr  🔴 <5 yr (depletion risk)  "
                "· Blue dashed = 10-yr threshold"
            )

    # ── Realized Price vs Spot ────────────────────────────────────────────────
    _rp_cols_avail = {
        "Gold":    ("spg_realized_price_oz", "$/oz"),
        "Silver":  ("spg_realized_price_oz", "$/oz"),
        "Copper":  ("spg_realized_price_lb", "$/lb"),
        "Uranium": ("spg_realized_price_lb", "$/lb"),
        "Iron Ore":("spg_realized_price_t",  "$/t"),
    }
    _rp_comm = st.selectbox(
        "Commodity — Realized Price vs Spot",
        list(_rp_cols_avail.keys()),
        key="rp_comm_sel",
        label_visibility="collapsed",
    ) if True else None

    if _rp_comm:
        _rp_col, _rp_unit = _rp_cols_avail[_rp_comm]
        _spot_key = _rp_comm
        _rp_spot = config.COMMODITY_SPOT.get(_spot_key)

        if _rp_col in df.columns and df[_rp_col].notna().any() and _rp_spot:
            _rp_df = df[
                df[_rp_col].notna() &
                (df[_rp_col] > 0) &
                df["commodity"].str.contains(_rp_comm, case=False, na=False)
            ].copy()

            if not _rp_df.empty:
                st.markdown("#### 💹 Realized Price vs Spot")
                st.caption(
                    f"How much each company received vs the {_rp_comm} spot price "
                    f"({_rp_unit}). Positive = premium to spot; negative = discount. "
                    "Sorted by premium descending."
                )
                _rp_df["_premium_pct"] = ((_rp_df[_rp_col] - _rp_spot) / _rp_spot * 100).round(1)
                _rp_df = _rp_df.sort_values("_premium_pct", ascending=False).head(30)
                _rp_df["_label"] = _rp_df["name"] + " (" + _rp_df["ticker"] + ")"
                _rp_df["_color"] = _rp_df["_premium_pct"].apply(
                    lambda v: "#22c55e" if v >= 5
                    else "#3b82f6" if v >= 0
                    else "#eab308" if v >= -5
                    else "#ef4444"
                )
                fig_rp = go.Figure()
                fig_rp.add_trace(go.Bar(
                    x=_rp_df["_label"],
                    y=_rp_df["_premium_pct"],
                    marker_color=_rp_df["_color"].tolist(),
                    text=[f"{v:+.1f}%" for v in _rp_df["_premium_pct"]],
                    textposition="outside",
                    customdata=np.column_stack([
                        _rp_df["name"],
                        _rp_df[_rp_col].round(2),
                        _rp_df["_premium_pct"].round(1),
                    ]),
                    hovertemplate=(
                        "<b>%{customdata[0]}</b><br>"
                        "Realized: $%{customdata[1]:,.2f}<br>"
                        f"Spot: ${_rp_spot:,.2f}<br>"
                        "Premium: %{customdata[2]:+.1f}%<extra></extra>"
                    ),
                ))
                fig_rp.add_hline(y=0, line_dash="solid", line_color="rgba(23,32,51,0.3)",
                                 line_width=1.5)
                fig_rp.update_layout(
                    height=420,
                    yaxis_title=f"Premium / Discount to {_rp_comm} Spot (%)",
                    xaxis_tickangle=-35,
                    showlegend=False,
                    margin=dict(t=30, b=140),
                    plot_bgcolor="rgba(0,0,0,0)",
                    paper_bgcolor="rgba(0,0,0,0)",
                    xaxis=dict(zeroline=False),
                )
                st.plotly_chart(fig_rp, width="stretch")
                st.caption(
                    f"🟢 ≥+5% premium  🔵 0–5%  🟡 −5–0%  🔴 <−5% discount  "
                    f"· {_rp_comm} spot: ${_rp_spot:,.2f} {_rp_unit}"
                )

    # ── Upside to NAV ranked bar chart ────────────────────────────────────────
    st.markdown("#### 🎯 Upside Potential — Ranked")
    st.caption(
        "**↑ Upside to NAV%** uses S&P Capital IQ P/NAV data (most accurate).  "
        "When NAV data is absent, **↑ P/B Re-rating%** shows the implied gain "
        "if the company re-rated to its commodity peers' median P/B multiple."
    )

    # Prefer NAV upside; fall back to P/B peer upside
    _udf = df.copy()
    _udf["_upside_val"]    = _udf["upside_to_nav"].where(_udf["upside_to_nav"].notna(), _udf.get("pb_peer_upside"))
    _udf["_upside_source"] = np.where(
        _udf["upside_to_nav"].notna(), "NAV (S&P)",
        np.where(_udf.get("pb_peer_upside", pd.Series(dtype=float)).notna(), "P/B peers", None)
    )
    _upside_df = (
        _udf[_udf["_upside_val"].notna()]
        .sort_values("_upside_val", ascending=False)
        .head(30)
        [["name", "ticker", "commodity", "grade", "_upside_val", "spg_p_nav", "_upside_source"]]
        .copy()
        .rename(columns={"_upside_val": "upside_to_nav", "_upside_source": "nav_source"})
    )

    if _upside_df.empty:
        st.info("No upside data available — refresh data to populate.")
    else:
        _upside_df["label"] = _upside_df.apply(
            lambda r: f"{r['name']} ({r['ticker']})", axis=1
        )
        _upside_df["color"] = _upside_df["upside_to_nav"].apply(
            lambda v: ("#22c55e" if v >= 50 else
                       "#3b82f6" if v >= 25 else
                       "#eab308" if v >= 10 else
                       "#f97316" if v >= 0  else
                       "#ef4444")
        )
        _upside_df["upside_fmt"] = _upside_df["upside_to_nav"].apply(
            lambda v: f"{v:+.0f}%"
        )

        fig_upside = go.Figure()
        fig_upside.add_bar(
            x=_upside_df["upside_to_nav"],
            y=_upside_df["label"],
            orientation="h",
            marker_color=_upside_df["color"].tolist(),
            text=_upside_df["upside_fmt"].tolist(),
            textposition="outside",
            customdata=np.column_stack([
                _upside_df["spg_p_nav"].round(2).fillna(0),
                _upside_df["grade"],
                _upside_df["commodity"],
                _upside_df["nav_source"].fillna("—"),
            ]),
            hovertemplate=(
                "<b>%{y}</b><br>"
                "Upside to NAV: %{x:+.0f}%<br>"
                "P/NAV: %{customdata[0]:.2f}x<br>"
                "Grade: %{customdata[1]}<br>"
                "Commodity: %{customdata[2]}<br>"
                "NAV source: %{customdata[3]}<extra></extra>"
            ),
        )
        # Reference lines
        fig_upside.add_vline(x=0,  line_color="white",      line_dash="solid", line_width=1.5)
        fig_upside.add_vline(x=50, line_color="#22c55e",    line_dash="dot",   line_width=1,
                             annotation_text="50%", annotation_position="top")
        fig_upside.add_vline(x=25, line_color="#3b82f6",    line_dash="dot",   line_width=1,
                             annotation_text="25%", annotation_position="top")

        fig_upside.update_layout(
            height=max(300, len(_upside_df) * 26),
            xaxis_title="Upside to 1× NAV (%)",
            yaxis_title="",
            showlegend=False,
            margin=dict(l=20, r=80, t=30, b=30),
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)",
            xaxis=dict(zeroline=True, zerolinecolor="white", zerolinewidth=1.5),
        )
        fig_upside.update_yaxes(autorange="reversed")
        st.plotly_chart(fig_upside, width="stretch")
        st.caption(
            "🟢 >50% upside  🔵 25–50%  🟡 10–25%  🟠 0–10%  🔴 at premium to NAV  "
            "· Source: S&P Capital IQ P/NAV data (brokerage fallback where available)"
        )

    # ── Intra-sector Relative Strength ────────────────────────────────────────
    st.markdown("#### 🏆 Intra-Sector Relative Strength")
    st.caption(
        "Each company's 3-month return vs its commodity-group median — "
        "positive = outperforming peers, negative = lagging. "
        "Only companies with return data shown."
    )
    if "return_3m" in df.columns and df["return_3m"].notna().any():
        _rs_df = df[df["return_3m"].notna()].copy()
        _rs_df["_comm_group"] = _rs_df["commodity"].str.split("/").str[0].str.strip()
        _rs_df["_peer_med_3m"] = _rs_df.groupby("_comm_group")["return_3m"].transform("median")
        _rs_df["rel_strength"] = (_rs_df["return_3m"] - _rs_df["_peer_med_3m"]).round(1)
        _rs_df = _rs_df.drop(columns=["_comm_group", "_peer_med_3m"])

        # Commodity picker for this chart
        _rs_comms = sorted(_rs_df["commodity"].str.split("/").str[0].str.strip().unique())
        _rs_sel = st.selectbox(
            "Commodity group", ["All"] + _rs_comms,
            key="rs_comm_sel",
        )
        if _rs_sel != "All":
            _rs_view = _rs_df[_rs_df["commodity"].str.startswith(_rs_sel)].copy()
        else:
            _rs_view = _rs_df.copy()

        _rs_view = _rs_view.sort_values("rel_strength")
        # Readability cap: with hundreds of names the bars become a smear.
        # Keep the informative tails — 20 strongest + 20 weakest.
        _rs_total = len(_rs_view)
        if _rs_total > 40:
            _rs_view = pd.concat([_rs_view.head(20), _rs_view.tail(20)])
        _rs_view["label"] = _rs_view["name"] + " (" + _rs_view["ticker"] + ")"
        _rs_view["color"] = _rs_view["rel_strength"].apply(
            lambda v: "#22c55e" if v > 5 else "#ef4444" if v < -5 else "#94a3b8"
        )
        _rs_view["comm_grp"] = _rs_view["commodity"].str.split("/").str[0].str.strip()

        _n_rs = len(_rs_view)
        _rs_height = max(350, min(_n_rs * 22 + 60, 900))
        fig_rs = go.Figure()
        fig_rs.add_bar(
            x=_rs_view["rel_strength"],
            y=_rs_view["label"],
            orientation="h",
            marker_color=_rs_view["color"].tolist(),
            text=[f"{v:+.1f}%" for v in _rs_view["rel_strength"]],
            textposition="outside",
            customdata=np.column_stack([
                _rs_view["return_3m"].round(1),
                _rs_view["score_composite"].round(1),
                _rs_view["comm_grp"],
            ]),
            hovertemplate=(
                "<b>%{y}</b><br>"
                "3M Return: %{customdata[0]}%<br>"
                "vs Peers: %{x:+.1f}%<br>"
                "Score: %{customdata[1]}<br>"
                "Group: %{customdata[2]}<extra></extra>"
            ),
        )
        fig_rs.add_vline(x=0, line_dash="solid", line_color="gray", line_width=1)
        fig_rs.update_layout(
            xaxis_title="3M Return vs Commodity-Group Median (%)",
            yaxis=dict(tickfont=dict(size=11)),
            height=_rs_height,
            margin=dict(l=0, r=60, t=10, b=40),
            showlegend=False,
        )
        st.plotly_chart(fig_rs, width="stretch")
        if _rs_total > 40:
            st.caption(f"Showing the 20 strongest and 20 weakest of {_rs_total} companies. "
                       "Pick a commodity group above to see the full list.")

        # Summary table: leaders + laggards per group
        with st.expander("📋 Leaders & Laggards by Group", expanded=False):
            _grp_summary = []
            for _grp, _grp_df in _rs_df.groupby(
                _rs_df["commodity"].str.split("/").str[0].str.strip()
            ):
                _sorted = _grp_df.sort_values("rel_strength", ascending=False)
                for _rank_label, _rank_row in [("🥇 Leader", _sorted.iloc[0]),
                                               ("🔴 Laggard", _sorted.iloc[-1])]:
                    if len(_sorted) >= 2:
                        _grp_summary.append({
                            "Group":      _grp,
                            "Rank":       _rank_label,
                            "Company":    f"{_rank_row['name']} ({_rank_row['ticker']})",
                            "3M Ret%":    round(_rank_row["return_3m"], 1),
                            "vs Peers%":  round(_rank_row["rel_strength"], 1),
                            "Score":      round(_rank_row["score_composite"], 1),
                        })
            if _grp_summary:
                _gs_df = pd.DataFrame(_grp_summary)
                st.dataframe(
                    _gs_df.style
                    .map(lambda v: "color:#22c55e;font-weight:700" if isinstance(v, float) and v > 0
                         else "color:#ef4444" if isinstance(v, float) and v < 0 else "",
                         subset=["3M Ret%", "vs Peers%"])
                    .format({"3M Ret%": "{:+.1f}%", "vs Peers%": "{:+.1f}%", "Score": "{:.1f}"},
                            na_rep="—"),
                    width="stretch", height=300,
                )
    else:
        st.info("Return data not yet available — run a data refresh to populate 1M/3M returns.")

    # ── Value Map ─────────────────────────────────────────────────────────────
    st.markdown("#### 🗺️ Value Map — Quality vs Upside")
    st.caption(
        "**X-axis:** upside potential (NAV discount or P/B re-rating vs peers).  "
        "**Y-axis:** composite quality score.  "
        "**Top-right quadrant** = high quality + high upside = the buy zone."
    )

    # Prefer NAV upside; fall back to peer P/B upside
    _vm_df = df.copy()
    _vm_df["_upside"] = _vm_df["upside_to_nav"].where(
        _vm_df["upside_to_nav"].notna(), _vm_df.get("pb_peer_upside")
    )
    _vm_df["_upside_label"] = np.where(
        _vm_df["upside_to_nav"].notna(), "NAV upside",
        np.where(_vm_df.get("pb_peer_upside", pd.Series(dtype=float)).notna(), "P/B re-rating", None)
    )
    _vm_df = _vm_df[_vm_df["_upside"].notna() & _vm_df["score_composite"].notna()].copy()

    if _vm_df.empty:
        st.info("Not enough data for Value Map — refresh to populate upside metrics.")
    else:
        # Bubble size = market cap (log-scaled, capped)
        _mc_max = _vm_df["market_cap"].max() if "market_cap" in _vm_df.columns else 1
        _vm_df["_bubble"] = (
            (_vm_df["market_cap"].fillna(_mc_max * 0.01) / _mc_max * 35 + 5).clip(5, 40)
            if "market_cap" in _vm_df.columns else 15
        )

        _GRADE_COLORS_VM = {
            "🟢 Strong Buy": "#22c55e",
            "🔵 Buy":        "#3b82f6",
            "🟡 Watch":      "#eab308",
            "🟠 Neutral":    "#f97316",
            "🔴 Avoid":      "#ef4444",
        }
        _vm_df["_color"] = _vm_df["grade"].map(_GRADE_COLORS_VM).fillna("#94a3b8")

        fig_vm = go.Figure()

        # Background quadrant shading
        _x_mid = 25   # upside threshold for "cheap"
        _y_mid = 55   # score threshold for "quality"

        _quad_annotations = [
            dict(x=_x_mid / 2,        y=(_y_mid + 100) / 2,
                 text="🏆 Quality<br>at Premium", showarrow=False,
                 font=dict(size=11, color="rgba(23,32,51,0.35)"), align="center"),
            dict(x=(_x_mid + 150) / 2, y=(_y_mid + 100) / 2,
                 text="🟢 BUY ZONE<br>Quality + Upside", showarrow=False,
                 font=dict(size=13, color="rgba(22,163,74,0.55)"), align="center"),
            dict(x=_x_mid / 2,        y=_y_mid / 2,
                 text="🔴 Avoid", showarrow=False,
                 font=dict(size=11, color="rgba(23,32,51,0.35)"), align="center"),
            dict(x=(_x_mid + 150) / 2, y=_y_mid / 2,
                 text="⚠️ Value Trap?<br>Cheap but Low Quality", showarrow=False,
                 font=dict(size=11, color="rgba(255,180,0,0.4)"), align="center"),
        ]

        for grade in _vm_df["grade"].unique():
            _g_df = _vm_df[_vm_df["grade"] == grade]
            # Label only actionable grades — labelling all 400 tickers is unreadable
            _vm_mode = "markers+text" if grade in ("🟢 Strong Buy", "🔵 Buy") else "markers"
            fig_vm.add_trace(go.Scatter(
                x=_g_df["_upside"],
                y=_g_df["score_composite"],
                mode=_vm_mode,
                name=grade,
                marker=dict(
                    size=_g_df["_bubble"].tolist(),
                    color=_GRADE_COLORS_VM.get(grade, "#94a3b8"),
                    opacity=0.8,
                    line=dict(width=1, color="rgba(23,32,51,0.25)"),
                ),
                text=_g_df["ticker"],
                textposition="top center",
                textfont=dict(size=9, color="rgba(23,32,51,0.65)"),
                customdata=np.column_stack([
                    _g_df["name"],
                    _g_df["commodity"],
                    _g_df["stage"],
                    _g_df["_upside"].round(1),
                    _g_df["_upside_label"].fillna("—"),
                    _g_df["score_composite"].round(1),
                    (_g_df["market_cap"] / 1e6).round(0).fillna(0) if "market_cap" in _g_df.columns else np.zeros(len(_g_df)),
                ]),
                hovertemplate=(
                    "<b>%{customdata[0]}</b> (%{text})<br>"
                    "%{customdata[1]} · %{customdata[2]}<br>"
                    "Upside: <b>%{customdata[3]:+.0f}%</b> (%{customdata[4]})<br>"
                    "Score: <b>%{customdata[5]:.1f}</b><br>"
                    "Mkt Cap: $%{customdata[6]:,.0f}M<extra></extra>"
                ),
            ))

        # Quadrant lines
        _x_range = [min(_vm_df["_upside"].min() - 10, -20), max(_vm_df["_upside"].max() + 10, 100)]
        fig_vm.add_vline(x=_x_mid, line_color="rgba(23,32,51,0.25)", line_dash="dash", line_width=1)
        fig_vm.add_hline(y=_y_mid, line_color="rgba(23,32,51,0.25)", line_dash="dash", line_width=1)

        # Buy-zone shading
        fig_vm.add_shape(
            type="rect",
            x0=_x_mid, x1=_x_range[1],
            y0=_y_mid, y1=100,
            fillcolor="rgba(34,197,94,0.06)",
            line=dict(width=0),
        )

        fig_vm.update_layout(
            height=560,
            xaxis_title="Upside Potential (%)",
            yaxis_title="Composite Score",
            xaxis=dict(range=_x_range, zeroline=True, zerolinecolor="rgba(23,32,51,0.2)"),
            yaxis=dict(range=[0, 105]),
            plot_bgcolor="#ffffff",
            paper_bgcolor="rgba(0,0,0,0)",
            legend=dict(orientation="h", y=-0.12, x=0),
            margin=dict(l=60, r=40, t=30, b=60),
            annotations=_quad_annotations,
        )
        st.plotly_chart(fig_vm, width="stretch")

    # ── Score vs Returns cross-section ────────────────────────────────────────
    st.markdown("#### 📈 Does Score Predict Returns?")
    st.caption(
        "Companies binned by composite score quintile.  "
        "Average 1-month and 3-month price returns per bucket — "
        "a cross-sectional check on whether the screener leads price."
    )

    _ret_cols_avail = [c for c in ["return_1m", "return_3m"] if c in df.columns]
    if not _ret_cols_avail or df[_ret_cols_avail].notna().sum().sum() == 0:
        st.info("No return data yet — refresh data to populate 1M/3M returns.")
    else:
        # Bin by score quintile (label with score range)
        _score_bins  = [0, 20, 40, 55, 70, 100]
        _score_labels = ["0–20", "20–40", "40–55", "55–70", "70–100"]
        _sv_df = df[df["score_composite"].notna()].copy()
        _sv_df["score_bin"] = pd.cut(
            _sv_df["score_composite"],
            bins=_score_bins, labels=_score_labels, right=True,
        )

        _agg = (
            _sv_df.groupby("score_bin", observed=True)
            [[c for c in ["return_1m", "return_3m", "score_composite"]
              if c in _sv_df.columns]]
            .agg({"return_1m": ["mean", "count"], "return_3m": "mean",
                  "score_composite": "mean"})
        )
        _agg.columns = ["Avg 1M Ret%", "N", "Avg 3M Ret%", "Avg Score"]
        _agg = _agg.reset_index().rename(columns={"score_bin": "Score Range"})
        _agg["N"] = _agg["N"].astype(int)

        _ret_left, _ret_right = st.columns(2)

        with _ret_left:
            # Bar chart: avg returns by score bin
            _bar_data = []
            for _, _rr in _agg.iterrows():
                for _metric, _col in [("1M Return%", "Avg 1M Ret%"),
                                       ("3M Return%", "Avg 3M Ret%")]:
                    if pd.notna(_rr.get(_col)):
                        _bar_data.append({
                            "Score Range": str(_rr["Score Range"]),
                            "Metric": _metric,
                            "Return%": round(float(_rr[_col]), 2),
                        })

            if _bar_data:
                _bar_df = pd.DataFrame(_bar_data)
                _fig_ret = px.bar(
                    _bar_df,
                    x="Score Range", y="Return%", color="Metric",
                    barmode="group",
                    color_discrete_map={
                        "1M Return%": "#3b82f6",
                        "3M Return%": "#22c55e",
                    },
                    title="Avg Return by Score Quintile",
                    labels={"Return%": "Avg Return (%)", "Score Range": "Score Bucket"},
                )
                _fig_ret.add_hline(y=0, line_color="rgba(23,32,51,0.3)",
                                   line_dash="solid", line_width=1)
                _fig_ret.update_layout(
                    height=340,
                    plot_bgcolor="rgba(0,0,0,0)",
                    paper_bgcolor="rgba(0,0,0,0)",
                    legend=dict(orientation="h", y=1.1),
                    margin=dict(l=40, r=20, t=60, b=40),
                )
                st.plotly_chart(_fig_ret, width="stretch")

        with _ret_right:
            # Summary table
            _disp_agg = _agg[
                ["Score Range", "N", "Avg Score", "Avg 1M Ret%", "Avg 3M Ret%"]
            ].copy()

            def _color_ret_cell(val):
                if pd.isna(val): return ""
                return "color: #22c55e; font-weight:600" if val > 0 else "color: #ef4444; font-weight:600"

            _agg_styled = (
                _disp_agg.style
                .map(_color_ret_cell, subset=["Avg 1M Ret%", "Avg 3M Ret%"])
                .format({
                    "Avg Score":   "{:.1f}",
                    "Avg 1M Ret%": lambda x: f"{x:+.1f}%" if pd.notna(x) else "—",
                    "Avg 3M Ret%": lambda x: f"{x:+.1f}%" if pd.notna(x) else "—",
                }, na_rep="—")
                .set_properties(**{"text-align": "center"})
            )
            st.markdown("**Returns by Score Bucket**")
            st.dataframe(_agg_styled, width="stretch", height=230)

            # Scatter: score vs 3M return (per company)
            _scat_df = _sv_df[_sv_df["return_3m"].notna()][
                ["name", "ticker", "score_composite", "return_3m",
                 "grade", "commodity"]
            ].copy()
            if not _scat_df.empty:
                _fig_scat = px.scatter(
                    _scat_df,
                    x="score_composite", y="return_3m",
                    color="grade",
                    hover_name="name",
                    hover_data={"ticker": True, "commodity": True,
                                "score_composite": ":.1f", "return_3m": ":.1f"},
                    color_discrete_map={
                        "🟢 Strong Buy": "#22c55e",
                        "🔵 Buy":        "#3b82f6",
                        "🟡 Watch":      "#eab308",
                        "🟠 Neutral":    "#f97316",
                        "🔴 Avoid":      "#ef4444",
                    },
                    title="Score vs 3M Return (per company)",
                    labels={"score_composite": "Score", "return_3m": "3M Return%"},
                )
                _fig_scat.add_hline(y=0, line_color="rgba(23,32,51,0.3)",
                                    line_dash="dash", line_width=1)
                _fig_scat.update_layout(
                    height=300,
                    showlegend=False,
                    plot_bgcolor="rgba(0,0,0,0)",
                    paper_bgcolor="rgba(0,0,0,0)",
                    margin=dict(l=40, r=20, t=60, b=40),
                )
                st.plotly_chart(_fig_scat, width="stretch")

    # ── Commodity Heatmap ─────────────────────────────────────────────────────
    st.markdown("#### 🗃️ Commodity Heatmap — Sector Snapshot")
    st.caption(
        "Each cell = universe median for that commodity group. "
        "Useful for identifying which commodity sectors look cheap or expensive."
    )

    _hm_metrics = [
        ("score_composite",   "Score",         False),   # (col, label, invert)
        ("score_mining",      "Mining Score",  False),
        ("spg_p_nav",         "P/NAV",         True),    # lower = better → invert for color
        ("spg_aisc_margin",   "AISC Mgn%",     False),
        ("upside_to_nav",     "↑NAV Upside%",  False),
        ("ev_per_oz_prod",    "EV/oz Prod",    True),
        ("spg_reserve_life",  "Rsv Life (yr)", False),
        ("return_3m",         "3M Ret%",       False),
        ("price_to_book",     "P/B",           True),
    ]
    _hm_avail = [(c, l, inv) for c, l, inv in _hm_metrics if c in df.columns and df[c].notna().any()]

    if len(_hm_avail) >= 3:
        _hm_df = df.copy()
        _hm_df["_comm_grp"] = _hm_df["commodity"].str.split("/").str[0].str.strip()
        _hm_groups = [g for g in _hm_df["_comm_grp"].dropna().unique()
                      if _hm_df[_hm_df["_comm_grp"] == g].shape[0] >= 2]
        _hm_groups.sort()

        if len(_hm_groups) >= 2:
            _hm_cols   = [l for _, l, _ in _hm_avail]
            _hm_matrix = []   # rows = commodities, cols = metrics
            _hm_raw    = []   # raw values for display

            for _grp in _hm_groups:
                _grp_df = _hm_df[_hm_df["_comm_grp"] == _grp]
                _row_vals = []
                _row_raw  = []
                for _col, _lbl, _inv in _hm_avail:
                    _med = _grp_df[_col].median()
                    _row_raw.append(_med)
                    # Normalise 0–1 within column for colour scale
                    _all_vals = _hm_df[_hm_df["_comm_grp"].isin(_hm_groups)][_col].dropna()
                    if len(_all_vals) < 2 or pd.isna(_med):
                        _row_vals.append(float("nan"))
                    else:
                        _pct = (_all_vals < _med).mean()   # percentile among groups
                        _row_vals.append((1 - _pct) if _inv else _pct)
                _hm_matrix.append(_row_vals)
                _hm_raw.append(_row_raw)

            _hm_z    = np.array(_hm_matrix, dtype=float)
            _hm_raw_arr = np.array(_hm_raw, dtype=float)

            # Format raw values for annotation
            _fmt_funcs = [
                (lambda v: f"{v:.1f}"),   # Score
                (lambda v: f"{v:.1f}"),   # Mining Score
                (lambda v: f"{v:.2f}x"),  # P/NAV
                (lambda v: f"{v:.1f}%"),  # AISC Mgn%
                (lambda v: f"{v:+.0f}%"), # ↑NAV Upside%
                (lambda v: f"${v:,.0f}"), # EV/oz Prod
                (lambda v: f"{v:.1f}yr"), # Rsv Life
                (lambda v: f"{v:+.1f}%"), # 3M Ret%
                (lambda v: f"{v:.2f}"),   # P/B
            ]
            _fmt_funcs = _fmt_funcs[:len(_hm_avail)]

            _text_grid = []
            for ri in range(len(_hm_groups)):
                _text_row = []
                for ci, _ff in enumerate(_fmt_funcs):
                    _rv = _hm_raw_arr[ri, ci]
                    try:
                        _text_row.append(_ff(_rv) if not np.isnan(_rv) else "—")
                    except Exception:
                        _text_row.append("—")
                _text_grid.append(_text_row)

            fig_hm = go.Figure(go.Heatmap(
                z=_hm_z,
                x=_hm_cols,
                y=_hm_groups,
                text=_text_grid,
                texttemplate="%{text}",
                textfont=dict(size=11, color="white"),
                colorscale=[
                    [0.0,  "#7f1d1d"],
                    [0.25, "#b45309"],
                    [0.5,  "#1e3a5f"],
                    [0.75, "#166534"],
                    [1.0,  "#14532d"],
                ],
                zmin=0, zmax=1,
                showscale=False,
                hovertemplate="%{y} · %{x}: %{text}<extra></extra>",
            ))
            fig_hm.update_layout(
                height=max(220, len(_hm_groups) * 42 + 80),
                xaxis=dict(tickfont=dict(size=11), tickangle=-30),
                yaxis=dict(tickfont=dict(size=11), autorange="reversed"),
                margin=dict(l=100, r=20, t=20, b=60),
            )
            st.plotly_chart(fig_hm, width="stretch")
            st.caption(
                "Green = better vs universe peers · Red = weaker · "
                "P/NAV, EV/oz Prod, P/B are inverted (lower = greener). "
                "Min 2 companies per group."
            )
        else:
            st.info("Need ≥ 2 commodity groups with data to build heatmap.")
    else:
        st.info("Not enough metric data for heatmap — refresh data first.")

    # ── Correlation Matrix ─────────────────────────────────────────────────────
    st.markdown("#### 🔗 Metric Correlation Matrix")
    st.caption(
        "Cross-sectional Pearson correlation of key metrics across all companies. "
        "Dark green = strongly positive, dark red = strongly negative. "
        "Helps identify which metrics cluster together."
    )

    _corr_cols = [
        c for c in [
            "score_composite", "score_valuation", "score_health",
            "score_momentum", "score_mining",
            "price_to_book", "ev_ebitda", "spg_p_nav",
            "upside_to_nav", "spg_aisc_margin",
            "debt_to_equity", "current_ratio", "rsi",
            "return_1m", "return_3m", "wk52_position",
            "market_cap",
        ]
        if c in df.columns and df[c].notna().sum() >= 5
    ]
    _corr_labels = {
        "score_composite": "Score",   "score_valuation": "Valuation",
        "score_health":    "Health",   "score_momentum":  "Momentum",
        "score_mining":    "Mining",   "price_to_book":   "P/B",
        "ev_ebitda":       "EV/EBITDA","spg_p_nav":       "P/NAV",
        "upside_to_nav":   "↑Upside",  "spg_aisc_margin": "AISC Mgn",
        "debt_to_equity":  "D/E",      "current_ratio":   "CurrRatio",
        "rsi":             "RSI",      "return_1m":       "1M Ret",
        "return_3m":       "3M Ret",   "wk52_position":   "52wk Pos",
        "market_cap":      "Mkt Cap",
    }
    if len(_corr_cols) >= 3:
        _corr_mat = df[_corr_cols].corr(method="pearson")
        _corr_mat.index   = [_corr_labels.get(c, c) for c in _corr_mat.index]
        _corr_mat.columns = [_corr_labels.get(c, c) for c in _corr_mat.columns]

        fig_corr = go.Figure(go.Heatmap(
            z=_corr_mat.values.round(2),
            x=_corr_mat.columns.tolist(),
            y=_corr_mat.index.tolist(),
            colorscale="RdBu",
            zmid=0, zmin=-1, zmax=1,
            text=_corr_mat.values.round(2),
            texttemplate="%{text:.2f}",
            textfont=dict(size=10, color="white"),
            hovertemplate="%{y} × %{x}: %{z:.2f}<extra></extra>",
            colorbar=dict(thickness=14, len=0.8, title="r"),
        ))
        fig_corr.update_layout(
            xaxis=dict(tickangle=-40, tickfont=dict(size=10)),
            yaxis=dict(tickfont=dict(size=10), autorange="reversed"),
            height=520,
            margin=dict(l=0, r=20, t=10, b=60),
        )
        st.plotly_chart(fig_corr, width="stretch")

        # Highlight strongest non-self correlations for the user
        _corr_pairs = []
        _n = len(_corr_mat)
        for i in range(_n):
            for j in range(i + 1, _n):
                _r = _corr_mat.iloc[i, j]
                if abs(_r) >= 0.5:
                    _corr_pairs.append({
                        "Metric A":  _corr_mat.index[i],
                        "Metric B":  _corr_mat.columns[j],
                        "r":         round(_r, 2),
                        "Strength":  ("🟢 Strong" if abs(_r) >= 0.7
                                      else "🟡 Moderate"),
                        "Direction": "↑↑ Positive" if _r > 0 else "↑↓ Negative",
                    })
        if _corr_pairs:
            _cpdf = pd.DataFrame(_corr_pairs).sort_values("r", key=abs, ascending=False)
            with st.expander("📋 Notable correlations (|r| ≥ 0.5)", expanded=False):
                st.dataframe(_cpdf, width="stretch", height=300)
    else:
        st.info("Not enough metric data to compute correlations — refresh data first.")

    # ── Data Quality Dashboard ─────────────────────────────────────────────────
    st.markdown("#### 🔬 Data Quality Dashboard")
    st.caption(
        "Coverage of key metrics across the universe. "
        "Green = well-covered, amber = partial, red = mostly missing."
    )

    # Define metric groups and their column names
    _dq_groups = {
        "Yahoo Finance": [
            "price", "market_cap", "price_to_book", "ev_ebitda", "ev_revenue",
            "p_cf", "debt_to_equity", "current_ratio", "rsi",
            "wk52_position", "return_1m", "return_3m", "dividend_yield",
        ],
        "S&P / SNL Mining": [
            "spg_p_nav", "spg_reserves_m", "spg_resources_m",
            "spg_aisc_per_oz", "spg_aisc_per_t", "spg_aisc_per_lb",
            "spg_aisc_margin", "spg_grade_gpt", "spg_grade_pct",
            "spg_cash_cost_oz", "spg_cash_cost_t", "spg_cash_cost_lb",
            "spg_production_oz", "spg_production_t", "spg_production_lb",
            "spg_realized_price_oz", "spg_contained_reserves_oz",
            "spg_contained_reserves_lb", "spg_reserve_life", "score_mining",
        ],
        "Computed": [
            "score_composite", "score_valuation", "score_health",
            "score_momentum", "upside_to_nav", "fcf_yield",
            "ev_per_oz_prod", "ev_per_oz_reserve", "ev_per_lb_reserve",
        ],
    }

    _dq_col_grps = st.columns(len(_dq_groups))
    for _gi, (_grp_name, _grp_cols) in enumerate(_dq_groups.items()):
        with _dq_col_grps[_gi]:
            st.markdown(f"**{_grp_name}**")
            for _c in _grp_cols:
                if _c not in df.columns:
                    continue
                _pct = df[_c].notna().mean() * 100
                _bar_color = ("#22c55e" if _pct >= 80 else
                              "#eab308" if _pct >= 40 else "#ef4444")
                _label = _c.replace("spg_", "").replace("score_", "").replace("_", " ")
                st.markdown(
                    f"<div style='margin-bottom:4px'>"
                    f"<span style='font-size:12px'>{_label}</span> "
                    f"<span style='float:right;font-size:12px;color:{_bar_color}'>"
                    f"{_pct:.0f}%</span><br>"
                    f"<div style='background:#e3e9f0;border-radius:3px;height:5px;width:100%'>"
                    f"<div style='background:{_bar_color};height:5px;"
                    f"width:{_pct:.0f}%;border-radius:3px'></div></div></div>",
                    unsafe_allow_html=True,
                )

    st.markdown("<br>", unsafe_allow_html=True)

    # Per-company data completeness heatmap (top N companies by score)
    with st.expander("🗂️ Per-company coverage heatmap", expanded=False):
        _all_dq_cols = [c for grp in _dq_groups.values() for c in grp if c in df.columns]
        _dq_heat = df.nlargest(40, "score_composite")[["name"] + _all_dq_cols].copy()
        _dq_heat = _dq_heat.set_index("name")
        _dq_heat_bin = _dq_heat.notna().astype(int)   # 1 = has data, 0 = missing

        _col_labels = [c.replace("spg_", "").replace("score_", "").replace("_", " ")
                       for c in _dq_heat_bin.columns]

        fig_hm_dq = go.Figure(go.Heatmap(
            z=_dq_heat_bin.values,
            x=_col_labels,
            y=_dq_heat_bin.index.tolist(),
            colorscale=[[0, "#3f0a0a"], [1, "#14532d"]],
            showscale=False,
            hovertemplate="<b>%{y}</b><br>%{x}: %{z}<extra></extra>",
            zmin=0, zmax=1,
        ))
        fig_hm_dq.update_layout(
            xaxis=dict(tickangle=-45, tickfont=dict(size=10)),
            yaxis=dict(tickfont=dict(size=10), autorange="reversed"),
            height=max(500, len(_dq_heat_bin) * 18 + 100),
            margin=dict(l=0, r=0, t=10, b=80),
        )
        st.plotly_chart(fig_hm_dq, width="stretch")
        st.caption("Green = data present, Red = missing. Top 40 companies by composite score shown.")

    # ── Score Backtest ────────────────────────────────────────────────────────
    st.markdown("#### 🔬 Score Backtest — Did High Scores Predict Returns?")
    st.caption(
        "For each consecutive pair of snapshot dates, companies are binned by score quintile "
        "at time T and their actual price return to time T+1 is measured. "
        "Requires ≥ 3 snapshot dates to show meaningful results."
    )

    _bt_raw = load_backtest_data()
    _bt_dates = sorted(_bt_raw["snap_date"].unique()) if not _bt_raw.empty else []

    if len(_bt_dates) < 3:
        _n_needed = 3 - len(_bt_dates)
        st.info(
            f"Backtest needs ≥ 3 snapshot dates. Currently have **{len(_bt_dates)}**. "
            f"Run **🔄 Refresh Data Now** on {_n_needed} more day(s) to unlock this chart."
        )
    else:
        # For each ticker-date pair, compute return to the next snapshot date
        _bt_records = []
        _bt_pivot = _bt_raw.pivot(index="snap_date", columns="ticker", values="price")
        _bt_scores = _bt_raw.pivot(index="snap_date", columns="ticker", values="score_composite")
        _bt_dates_list = sorted(_bt_pivot.index.tolist())

        for _i in range(len(_bt_dates_list) - 1):
            _d0 = _bt_dates_list[_i]
            _d1 = _bt_dates_list[_i + 1]
            for _tk in _bt_pivot.columns:
                _p0 = _bt_pivot.loc[_d0, _tk] if _tk in _bt_pivot.columns else None
                _p1 = _bt_pivot.loc[_d1, _tk] if _tk in _bt_pivot.columns else None
                _sc = _bt_scores.loc[_d0, _tk] if _tk in _bt_scores.columns else None
                if pd.notna(_p0) and pd.notna(_p1) and pd.notna(_sc) and _p0 > 0:
                    _ret = (_p1 / _p0 - 1) * 100
                    _bt_records.append({
                        "date_from": _d0, "date_to": _d1,
                        "ticker": _tk, "score": _sc, "return_pct": _ret,
                    })

        if _bt_records:
            _bt_df = pd.DataFrame(_bt_records)
            # Bin scores into quintiles using labels
            _bt_df["score_bin"] = pd.cut(
                _bt_df["score"], bins=[0, 20, 40, 55, 70, 100],
                labels=["0–20\n(Weakest)", "20–40", "40–55", "55–70", "70–100\n(Strongest)"],
            )

            _bt_agg = (
                _bt_df.groupby("score_bin", observed=True)["return_pct"]
                .agg(["mean", "median", "std", "count"])
                .reset_index()
                .rename(columns={
                    "score_bin": "Score Quintile",
                    "mean":      "Avg Return%",
                    "median":    "Median Return%",
                    "std":       "Std Dev%",
                    "count":     "Observations",
                })
            )

            _bt_col1, _bt_col2 = st.columns(2)

            with _bt_col1:
                _bt_colors = [
                    "#ef4444", "#f97316", "#eab308", "#3b82f6", "#22c55e"
                ]
                fig_bt = go.Figure()
                fig_bt.add_bar(
                    x=_bt_agg["Score Quintile"].astype(str),
                    y=_bt_agg["Avg Return%"].round(2),
                    marker_color=_bt_colors[:len(_bt_agg)],
                    error_y=dict(
                        type="data", array=(_bt_agg["Std Dev%"] / np.sqrt(_bt_agg["Observations"])).round(2),
                        visible=True, color="rgba(23,32,51,0.45)",
                    ),
                    text=[f"{v:+.2f}%" for v in _bt_agg["Avg Return%"]],
                    textposition="outside",
                    name="Avg Return",
                )
                fig_bt.add_hline(y=0, line_dash="dash", line_color="gray", line_width=1)
                fig_bt.update_layout(
                    title="Avg Return by Score Quintile (period-to-period)",
                    yaxis_title="Avg Return % (next snapshot period)",
                    height=340,
                    margin=dict(t=40, b=40, l=0, r=0),
                    showlegend=False,
                )
                st.plotly_chart(fig_bt, width="stretch")

            with _bt_col2:
                # Summary stats table
                _bt_display = _bt_agg.copy()
                _bt_display["Avg Return%"]    = _bt_display["Avg Return%"].round(2)
                _bt_display["Median Return%"] = _bt_display["Median Return%"].round(2)
                _bt_display["Std Dev%"]       = _bt_display["Std Dev%"].round(2)

                def _color_bt_return(val):
                    if pd.isna(val) or val == 0: return ""
                    return "color:#22c55e;font-weight:700" if val > 0 else "color:#ef4444;font-weight:700"

                _bt_tbl_styled = (
                    _bt_display.style
                    .map(_color_bt_return, subset=["Avg Return%", "Median Return%"])
                    .format({
                        "Avg Return%":    "{:+.2f}%",
                        "Median Return%": "{:+.2f}%",
                        "Std Dev%":       "{:.2f}%",
                        "Observations":   "{:.0f}",
                    })
                )
                st.dataframe(_bt_tbl_styled, width="stretch", height=260)

                # Pearson r: score vs return
                _r_val = _bt_df[["score", "return_pct"]].corr().iloc[0, 1]
                _r_color = "#16a34a" if _r_val > 0.15 else "#dc2626" if _r_val < -0.05 else "#b45309"
                st.markdown(
                    f"<div style='background:#ffffff;border:1px solid #e3e9f0;"
                    f"border-top:3px solid #1a3a5c;border-radius:8px;padding:12px;text-align:center'>"
                    f"<div style='font-size:12px;color:#5b6b7f'>Score → Return Correlation (Pearson r)</div>"
                    f"<div style='font-size:28px;font-weight:800;color:{_r_color}'>{_r_val:+.3f}</div>"
                    f"<div style='font-size:11px;color:#5b6b7f'>"
                    f"Based on {len(_bt_df)} observations across {len(_bt_dates_list)-1} period(s)</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

            # Scatter: score at T vs return T→T+1
            with st.expander("🔍 Score vs Forward Return — all observations", expanded=False):
                fig_bt_scat = px.scatter(
                    _bt_df,
                    x="score", y="return_pct",
                    color="date_from",
                    hover_data=["ticker", "score", "return_pct"],
                    labels={"score": "Score at T", "return_pct": "Return to T+1 (%)"},
                    title="Score at Time T vs Price Return to Next Snapshot",
                    trendline="ols",
                )
                fig_bt_scat.add_hline(y=0, line_dash="dash", line_color="gray")
                fig_bt_scat.update_layout(height=400, margin=dict(t=40, b=40))
                st.plotly_chart(fig_bt_scat, width="stretch")
                st.caption(
                    "OLS trendline shows the linear relationship between score and subsequent return. "
                    "As snapshots accumulate over weeks/months, this chart becomes statistically meaningful."
                )
        else:
            st.info("No consecutive snapshot pairs found with valid price data yet.")

    # ── SNL Analytics Section ──────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### SNL Industry Analytics")

    # ── 1. AISC Cost Curve ────────────────────────────────────────────────────
    st.markdown("#### AISC Cost Curve — Producers")
    _cc_df = filtered.copy()
    # Pick the best AISC column per company (oz > t > lb)
    _cc_df["_aisc"] = _cc_df.get("spg_aisc_per_oz", pd.Series(dtype=float))
    if "spg_aisc_per_t" in _cc_df.columns:
        _cc_df["_aisc"] = _cc_df["_aisc"].fillna(_cc_df["spg_aisc_per_t"])
    if "spg_aisc_per_lb" in _cc_df.columns:
        _cc_df["_aisc"] = _cc_df["_aisc"].fillna(_cc_df["spg_aisc_per_lb"])
    _cc_df["_prod"] = _cc_df.get("spg_production_oz", pd.Series(dtype=float))
    if "spg_production_t" in _cc_df.columns:
        _cc_df["_prod"] = _cc_df["_prod"].fillna(_cc_df["spg_production_t"])
    if "spg_production_lb" in _cc_df.columns:
        _cc_df["_prod"] = _cc_df["_prod"].fillna(_cc_df["spg_production_lb"])

    _cc_plot = (
        _cc_df[_cc_df["_aisc"].notna() & (_cc_df["_aisc"] > 0) & (_cc_df["_aisc"] < 5000)]
        .sort_values("_aisc")
        .copy()
    )
    if not _cc_plot.empty:
        _cc_plot["_label"] = _cc_plot["ticker"]
        _gold_spot = None
        try:
            _spot_row = filtered[filtered["commodity"].str.contains("Gold", na=False)].head(1)
            if not _spot_row.empty and "spg_realized_price_oz" in _spot_row.columns:
                _gold_spot = float(_spot_row["spg_realized_price_oz"].iloc[0])
        except Exception:
            pass

        _fig_cc = px.bar(
            _cc_plot,
            x="_label", y="_aisc",
            color="grade",
            color_discrete_map={
                "🟢 Strong Buy": "#22c55e",
                "🔵 Buy":        "#3b82f6",
                "🟡 Watch":      "#eab308",
                "🟠 Neutral":    "#f97316",
                "🔴 Avoid":      "#ef4444",
            },
            hover_name="name",
            hover_data={
                "ticker": True, "stage": True, "commodity": True,
                "_aisc": ":.0f", "_prod": ":,.0f",
                "score_composite": ":.0f",
            },
            labels={"_aisc": "AISC ($/unit)", "_label": ""},
            title="All-In Sustaining Cost by Company (sorted low→high)",
        )
        if _gold_spot and _gold_spot > 0:
            _fig_cc.add_hline(
                y=_gold_spot,
                line_dash="dash", line_color="#FFD700", line_width=2,
                annotation_text=f"Gold spot ~${_gold_spot:,.0f}",
                annotation_position="top right",
                annotation_font_color="#FFD700",
            )
        _fig_cc.update_layout(
            height=420,
            xaxis=dict(tickangle=-45, tickfont=dict(size=10)),
            yaxis_title="AISC ($/unit)",
            showlegend=True,
            legend=dict(orientation="h", y=-0.22),
            bargap=0.25,
        )
        st.plotly_chart(_fig_cc, use_container_width=True)
        st.caption(
            f"AISC shown in native unit ($/oz for precious metals, $/t for base metals, $/lb for uranium). "
            f"{len(_cc_plot)} companies with AISC data. "
            f"Gold spot line shown when realized price data available."
        )
    else:
        st.info("No AISC data available — run snl_sync.py to populate.")

    # ── 2. Valuation Scatter: EV/oz R&R vs P/In-situ NAV ─────────────────────
    st.markdown("#### Valuation Map — EV/oz R&R vs P/In-situ NAV")
    _vs_needed = ["snl_ev_per_oz_rr", "snl_p_insitu"]
    _vs_df = filtered.dropna(subset=[c for c in _vs_needed if c in filtered.columns])
    if len(_vs_needed) == 2 and all(c in filtered.columns for c in _vs_needed):
        _vs_df = _vs_df[
            (_vs_df["snl_ev_per_oz_rr"] > 0) & (_vs_df["snl_ev_per_oz_rr"] < 10000) &
            (_vs_df["snl_p_insitu"] > 0) & (_vs_df["snl_p_insitu"] < 5)
        ].copy()
        if not _vs_df.empty:
            _vs_df["_mcap_m"] = (_vs_df["market_cap"] / 1e6).clip(lower=10)
            _fig_vs = px.scatter(
                _vs_df,
                x="snl_ev_per_oz_rr", y="snl_p_insitu",
                size="_mcap_m", size_max=40,
                color="grade",
                color_discrete_map={
                    "🟢 Strong Buy": "#22c55e",
                    "🔵 Buy":        "#3b82f6",
                    "🟡 Watch":      "#eab308",
                    "🟠 Neutral":    "#f97316",
                    "🔴 Avoid":      "#ef4444",
                },
                hover_name="name",
                hover_data={
                    "ticker": True, "commodity": True, "stage": True,
                    "snl_ev_per_oz_rr": ":,.0f",
                    "snl_p_insitu":     ":.3f",
                    "_mcap_m":          ":,.0f",
                    "score_composite":  ":.0f",
                },
                labels={
                    "snl_ev_per_oz_rr": "EV / oz R&R ($)",
                    "snl_p_insitu":     "P / In-situ NAV",
                    "_mcap_m":          "Mkt Cap ($M)",
                },
                title="Valuation Map (bottom-left = cheapest; bubble = market cap)",
                text="ticker",
            )
            # "Deep value" zone: EV/oz R&R < 150, P/In-situ NAV < 0.3
            _x_med = float(_vs_df["snl_ev_per_oz_rr"].median())
            _y_med = float(_vs_df["snl_p_insitu"].median())
            _fig_vs.add_shape(
                type="rect", x0=0, x1=_x_med, y0=0, y1=_y_med,
                fillcolor="rgba(34,197,94,0.08)",
                line=dict(color="rgba(34,197,94,0.5)", dash="dot", width=1.5),
            )
            _fig_vs.add_annotation(
                x=_x_med * 0.05, y=_y_med * 0.95,
                text="Deep Value Zone",
                showarrow=False,
                font=dict(color="#22c55e", size=11),
                xanchor="left", yanchor="top",
            )
            _fig_vs.add_vline(x=_x_med, line_dash="dot",
                              line_color="rgba(148,163,184,0.5)")
            _fig_vs.add_hline(y=_y_med, line_dash="dot",
                              line_color="rgba(148,163,184,0.5)")
            _fig_vs.update_traces(
                textposition="top center",
                textfont=dict(size=9, color="#374151"),
                marker=dict(opacity=0.85, line=dict(width=1, color="#ffffff")),
            )
            _fig_vs.update_layout(
                height=520,
                xaxis_title="EV / oz R&R ($) — lower is cheaper",
                yaxis_title="P / In-situ NAV — lower is cheaper",
                showlegend=True,
                legend=dict(orientation="h", y=-0.15),
            )
            st.plotly_chart(_fig_vs, use_container_width=True)
            st.caption(
                "EV/oz R&R = Enterprise Value ÷ total Reserves+Resources (oz). "
                "P/In-situ NAV = Market Cap ÷ SNL in-situ value of all deposits. "
                "Dashed lines = universe medians. Green zone = below median on both axes."
            )
        else:
            st.info("Not enough SNL valuation data to plot (need EV/oz R&R and P/In-situ NAV).")
    else:
        st.info("SNL valuation columns not yet available — run snl_sync.py.")

    # ── 3. R&R Rankings ───────────────────────────────────────────────────────
    _rr_col_a, _rr_col_b = st.columns(2)

    with _rr_col_a:
        st.markdown("#### R&R Ranking — Gold (koz)")
        if "snl_rr_koz" in filtered.columns:
            _rr_g = (
                filtered[filtered["snl_rr_koz"].notna() & (filtered["snl_rr_koz"] > 0)]
                .nlargest(20, "snl_rr_koz")
                .sort_values("snl_rr_koz")
                .copy()
            )
            if not _rr_g.empty:
                _fig_rr = px.bar(
                    _rr_g, x="snl_rr_koz", y="ticker",
                    orientation="h",
                    color="grade",
                    color_discrete_map={
                        "🟢 Strong Buy": "#22c55e", "🔵 Buy": "#3b82f6",
                        "🟡 Watch": "#eab308", "🟠 Neutral": "#f97316",
                        "🔴 Avoid": "#ef4444",
                    },
                    hover_name="name",
                    hover_data={"snl_rr_koz": ":,.0f", "score_composite": ":.0f"},
                    labels={"snl_rr_koz": "R&R (koz)", "ticker": ""},
                    title="Total R&R (koz Au-eq) Top 20",
                )
                _fig_rr.update_layout(height=480, showlegend=False, bargap=0.25)
                st.plotly_chart(_fig_rr, use_container_width=True)
            else:
                st.info("No gold R&R data.")
        else:
            st.info("R&R data not available.")

    with _rr_col_b:
        st.markdown("#### R&R Ranking — Base Metals (Mlb)")
        if "snl_rr_mlb" in filtered.columns:
            _rr_b = (
                filtered[filtered["snl_rr_mlb"].notna() & (filtered["snl_rr_mlb"] > 0)]
                .nlargest(20, "snl_rr_mlb")
                .sort_values("snl_rr_mlb")
                .copy()
            )
            if not _rr_b.empty:
                _fig_rr2 = px.bar(
                    _rr_b, x="snl_rr_mlb", y="ticker",
                    orientation="h",
                    color="commodity",
                    hover_name="name",
                    hover_data={"snl_rr_mlb": ":,.0f", "score_composite": ":.0f"},
                    labels={"snl_rr_mlb": "R&R (Mlb)", "ticker": ""},
                    title="Total R&R (Mlb Cu-eq) Top 20",
                )
                _fig_rr2.update_layout(height=480, showlegend=True,
                                       legend=dict(orientation="v", x=1.01), bargap=0.25)
                st.plotly_chart(_fig_rr2, use_container_width=True)
            else:
                st.info("No base metal R&R data.")
        else:
            st.info("R&R (Mlb) data not available.")

# ── TAB 3: Company Detail ──────────────────────────────────────────────────────
with tab_detail:
    all_names = filtered[["name", "ticker"]].apply(
        lambda r: f"{r['name']} ({r['ticker']})", axis=1
    ).tolist()
    if not all_names:
        st.info("No companies match current filters.")
    else:
        # Pre-select from Quick Search if set
        _qs_jump = st.session_state.pop("qs_jump_name", None)
        _detail_default_idx = 0
        if _qs_jump:
            _qs_short = _qs_jump.replace("  ", " ")
            for _i, _n in enumerate(all_names):
                if _qs_short in _n or _n in _qs_short:
                    _detail_default_idx = _i
                    break
        selected = st.selectbox("Select company", all_names, index=_detail_default_idx)
        sel_ticker = selected.split("(")[-1].rstrip(")")
        row = df[df["ticker"] == sel_ticker].iloc[0] if not df[df["ticker"] == sel_ticker].empty else None

        if row is not None:
            c1, c2, c3 = st.columns([1, 1, 2])
            with c1:
                # Watchlist toggle + note
                _in_wl = sel_ticker in get_watchlist()
                _wl_label = "✖️ Remove from Watchlist" if _in_wl else "⭐ Add to Watchlist"
                if st.button(_wl_label, key=f"wl_{sel_ticker}"):
                    if _in_wl:
                        remove_from_watchlist(sel_ticker)
                    else:
                        add_to_watchlist(sel_ticker)
                    st.rerun()
                if _in_wl:
                    _current_note = get_watchlist_note(sel_ticker)
                    _new_note = st.text_area(
                        "📝 Note", value=_current_note,
                        height=68, key=f"note_{sel_ticker}",
                        placeholder="Your thesis / reminder…",
                    )
                    if _new_note != _current_note:
                        update_watchlist_note(sel_ticker, _new_note)

                    # Price target
                    _stored_pt = get_price_target(sel_ticker)
                    _cur_price = row.get("price")
                    _new_pt = st.number_input(
                        "🎯 Price Target",
                        min_value=0.0, max_value=9999.0,
                        value=float(_stored_pt) if _stored_pt else 0.0,
                        step=0.01, format="%.3f",
                        key=f"pt_{sel_ticker}",
                        help="Set 0 to clear the target",
                    )
                    if _new_pt != (_stored_pt or 0.0):
                        set_price_target(sel_ticker, _new_pt if _new_pt > 0 else None)
                    if _new_pt > 0 and pd.notna(_cur_price) and _cur_price > 0:
                        _upside_to_pt = (_new_pt / _cur_price - 1) * 100
                        _pt_color = ("#22c55e" if _upside_to_pt > 5
                                     else "#ef4444" if _upside_to_pt < -5
                                     else "#eab308")
                        _pt_arrow = "↑" if _upside_to_pt > 0 else "↓"
                        st.markdown(
                            f"<span style='color:{_pt_color};font-weight:700'>"
                            f"{_pt_arrow} {abs(_upside_to_pt):.1f}% to target</span>",
                            unsafe_allow_html=True,
                        )

                st.markdown(f"### {row.get('name', sel_ticker)}")
                st.markdown(f"**{row.get('commodity', '')}** · {row.get('stage', '')}")
                st.markdown(f"**Grade:** {row.get('grade', '—')}")
                _delta = row.get("score_delta")
                _delta_str = f"{_delta:+.1f}" if pd.notna(_delta) else None
                st.metric("Composite Score", f"{row.get('score_composite', '—')}/100", delta=_delta_str)
                st.metric("Price", f"${row.get('price', 0):.3f}" if pd.notna(row.get('price')) else "—")
                mcap = row.get('market_cap')
                st.metric("Market Cap",
                    f"${mcap/1e9:.2f}B" if pd.notna(mcap) and mcap >= 1e9
                    else f"${mcap/1e6:.0f}M" if pd.notna(mcap) else "—")
                # Data freshness indicator
                _fetched = row.get("fetched_at")
                if _fetched is not None:
                    try:
                        from datetime import datetime, timezone
                        _ft = pd.to_datetime(_fetched)
                        _now = datetime.now(timezone.utc).replace(tzinfo=None)
                        _age = _now - _ft
                        _h = int(_age.total_seconds() // 3600)
                        _age_str = (f"{_h}h ago" if _h < 48
                                    else f"{_age.days}d ago")
                        st.caption(f"🕐 Data: {_age_str}")
                    except Exception:
                        pass

                # Analyst consensus
                _at_mean  = row.get("analyst_target_mean")
                _at_high  = row.get("analyst_target_high")
                _at_low   = row.get("analyst_target_low")
                _at_count = row.get("analyst_count")
                _at_rec   = row.get("analyst_rec_key", "")
                _cur_px   = row.get("price")
                if pd.notna(_at_mean) and _at_mean > 0:
                    _at_upside = (_at_mean / _cur_px - 1) * 100 if pd.notna(_cur_px) and _cur_px > 0 else None
                    _at_color  = "#16a34a" if (_at_upside or 0) > 5 else "#dc2626" if (_at_upside or 0) < -5 else "#b45309"
                    _at_rec_str = _at_rec.replace("-", " ").title() if _at_rec else ""
                    _at_n_str   = f"  ({int(_at_count)} analysts)" if pd.notna(_at_count) else ""
                    st.markdown(
                        f"<div style='background:#ffffff;border:1px solid #e3e9f0;"
                        f"border-left:3px solid #1a3a5c;border-radius:8px;padding:10px 12px;margin-top:8px'>"
                        f"<div style='font-size:11px;color:#5b6b7f;margin-bottom:4px'>Analyst Consensus{_at_n_str}</div>"
                        f"<div style='font-size:16px;font-weight:700;color:{_at_color}'>${_at_mean:.3f}"
                        + (f"  <span style='font-size:13px'>({_at_upside:+.1f}%)</span>" if _at_upside is not None else "")
                        + f"</div>"
                        f"<div style='font-size:11px;color:#5b6b7f;margin-top:3px'>"
                        f"Range ${_at_low:.3f}–${_at_high:.3f}" if pd.notna(_at_low) and pd.notna(_at_high) else ""
                        + (f" · {_at_rec_str}" if _at_rec_str else "")
                        + f"</div></div>",
                        unsafe_allow_html=True,
                    )

            with c2:
                st.markdown("#### Sub-scores vs Peers")
                _score_keys = ["score_valuation", "score_health", "score_momentum",
                               "score_mining", "score_commodity", "score_stage"]
                _score_labels = ["Valuation", "Health", "Momentum", "⛏️ Mining", "Commodity", "Stage"]
                score_data = {
                    lbl: (row.get(k, 0) if pd.notna(row.get(k)) else 0)
                    for k, lbl in zip(_score_keys, _score_labels)
                }

                # Compute peer median scores for overlay
                _radar_primary_comm = str(row.get("commodity", "")).split("/")[0].strip()
                _radar_peers = df[
                    df["commodity"].str.startswith(_radar_primary_comm) &
                    (df["ticker"] != sel_ticker)
                ]
                _peer_score_data = {}
                for _sk, _sl in zip(_score_keys, _score_labels):
                    if _sk in _radar_peers.columns:
                        _pm = _radar_peers[_sk].median()
                        _peer_score_data[_sl] = float(_pm) if pd.notna(_pm) else 50.0
                    else:
                        _peer_score_data[_sl] = 50.0

                fig_radar = go.Figure()
                fig_radar.add_trace(go.Scatterpolar(
                    r=list(score_data.values()),
                    theta=list(score_data.keys()),
                    fill="toself",
                    fillcolor="rgba(59,130,246,0.2)",
                    line=dict(color="#3b82f6", width=2),
                    name=row.get("name", sel_ticker),
                ))
                if _peer_score_data and not _radar_peers.empty:
                    fig_radar.add_trace(go.Scatterpolar(
                        r=list(_peer_score_data.values()),
                        theta=list(_peer_score_data.keys()),
                        fill="toself",
                        fillcolor="rgba(239,68,68,0.08)",
                        line=dict(color="#ef4444", width=1.5, dash="dot"),
                        name=f"{_radar_primary_comm} median",
                    ))
                fig_radar.update_layout(
                    polar=dict(radialaxis=dict(range=[0, 100])),
                    showlegend=True,
                    legend=dict(orientation="h", y=-0.15, font=dict(size=10)),
                    height=300,
                    margin=dict(l=40, r=40, t=20, b=40),
                )
                st.plotly_chart(fig_radar, width="stretch")

            with c3:
                st.markdown("#### Key Metrics")
                metrics = {
                    "P/Book":          f"{row.get('price_to_book', np.nan):.2f}" if pd.notna(row.get('price_to_book')) else "—",
                    "EV/EBITDA":       f"{row.get('ev_ebitda', np.nan):.1f}x"   if pd.notna(row.get('ev_ebitda'))    else "—",
                    "EV/Revenue":      f"{row.get('ev_revenue', np.nan):.1f}x"  if pd.notna(row.get('ev_revenue'))   else "—",
                    "P/CF":            f"{row.get('p_cf', np.nan):.1f}x"        if pd.notna(row.get('p_cf'))         else "—",
                    "Debt/Equity":     f"{row.get('debt_to_equity', np.nan):.0f}%" if pd.notna(row.get('debt_to_equity')) else "—",
                    "Current Ratio":   f"{row.get('current_ratio', np.nan):.2f}" if pd.notna(row.get('current_ratio')) else "—",
                    "Cash % Mkt Cap":  f"{row.get('cash_pct_mcap', np.nan):.1f}%" if pd.notna(row.get('cash_pct_mcap')) else "—",
                    "Net Debt":        f"${row.get('net_debt_m', np.nan):.0f}M"  if pd.notna(row.get('net_debt_m'))  else "—",
                    "RSI (14d)":       f"{row.get('rsi', np.nan):.0f}"           if pd.notna(row.get('rsi'))         else "—",
                    "52wk Position":   f"{row.get('wk52_position', np.nan):.0f}%" if pd.notna(row.get('wk52_position')) else "—",
                    "vs 52wk High":    f"{row.get('pct_from_52hi', np.nan):.1f}%" if pd.notna(row.get('pct_from_52hi')) else "—",
                }
                # Derived yield / efficiency metrics
                _fcf = row.get("free_cf")
                _mcap_r = row.get("market_cap")
                if pd.notna(_fcf) and pd.notna(_mcap_r) and _mcap_r > 0 and _fcf > 0:
                    metrics["FCF Yield"] = f"{_fcf / _mcap_r * 100:.1f}%"
                _div_yld = row.get("dividend_yield")
                if pd.notna(_div_yld) and _div_yld > 0:
                    metrics["Dividend Yield"] = f"{_div_yld:.2f}%"
                _ev_r = row.get("ev_reserves")
                if pd.notna(_ev_r):
                    metrics["EV/Reserves"] = f"{_ev_r:.2f}x"
                # Profitability (from Yahoo financialData)
                _roe = row.get("return_on_equity")
                if pd.notna(_roe):
                    metrics["Return on Equity"] = f"{_roe * 100:.1f}%"
                _op_mgn = row.get("operating_margins")
                if pd.notna(_op_mgn):
                    metrics["Operating Margin"] = f"{_op_mgn * 100:.1f}%"
                _gr_mgn = row.get("gross_margins")
                if pd.notna(_gr_mgn):
                    metrics["Gross Margin"] = f"{_gr_mgn * 100:.1f}%"
                # S&P Mining data (SPG) — shown when available
                spg_metrics = {}
                p_nav  = row.get("spg_p_nav")
                aisc   = row.get("spg_aisc_per_oz")
                aisc_t = row.get("spg_aisc_per_t")
                aisc_lb = row.get("spg_aisc_per_lb")
                aisc_m = row.get("spg_aisc_margin")
                resv   = row.get("spg_reserves_m")
                rsrc   = row.get("spg_resources_m")
                if pd.notna(p_nav):
                    spg_metrics["P/NAV (S&P)"] = f"{p_nav:.2f}x"
                    # Upside to NAV — key investment thesis metric
                    _nav_upside = row.get("upside_to_nav")
                    if pd.notna(_nav_upside):
                        _nav_src = row.get("nav_source", "S&P")
                        _upside_label = f"↑ Upside to NAV ({_nav_src})"
                        spg_metrics[_upside_label] = (
                            f"+{_nav_upside:.0f}%  🟢 discount"
                            if _nav_upside > 0
                            else f"{_nav_upside:.0f}%  🔴 premium"
                        )
                # Show whichever AISC unit is available
                if pd.notna(aisc):
                    spg_metrics["AISC $/oz (S&P)"] = f"${aisc:,.0f}"
                elif pd.notna(aisc_t):
                    spg_metrics["AISC $/t (S&P)"] = f"${aisc_t:,.0f}"
                elif pd.notna(aisc_lb):
                    spg_metrics["AISC $/lb (S&P)"] = f"${aisc_lb:.2f}"
                if pd.notna(aisc_m):
                    spg_metrics["AISC Margin (S&P)"] = f"{aisc_m:.1f}%"
                # Cash cost (stress-test floor — below AISC)
                _cc_oz = row.get("spg_cash_cost_oz")
                _cc_t  = row.get("spg_cash_cost_t")
                _cc_lb = row.get("spg_cash_cost_lb")
                if pd.notna(_cc_oz):
                    spg_metrics["Cash Cost (S&P)"] = f"${_cc_oz:,.0f}/oz"
                elif pd.notna(_cc_t):
                    spg_metrics["Cash Cost (S&P)"] = f"${_cc_t:,.0f}/t"
                elif pd.notna(_cc_lb):
                    spg_metrics["Cash Cost (S&P)"] = f"${_cc_lb:.2f}/lb"
                if pd.notna(resv):
                    spg_metrics["Reserves $M (S&P)"] = f"${resv:,.0f}M"
                if pd.notna(rsrc):
                    spg_metrics["Resources $M (S&P)"] = f"${rsrc:,.0f}M"
                # Primary grade — show with correct unit
                _grd_gpt = row.get("spg_grade_gpt")
                _grd_pct = row.get("spg_grade_pct")
                if pd.notna(_grd_gpt):
                    spg_metrics["Ore Grade (S&P)"] = f"{_grd_gpt:.3f} g/t"
                elif pd.notna(_grd_pct):
                    spg_metrics["Ore Grade (S&P)"] = f"{_grd_pct:.3f}%"

                # Production (attributable annual)
                _prod_disp = row.get("production_display", "—")
                if _prod_disp and _prod_disp != "—":
                    spg_metrics["Production (S&P)"] = _prod_disp

                # EV multiples
                _ev_oz_p = row.get("ev_per_oz_prod")
                _ev_oz_r = row.get("ev_per_oz_reserve")
                _ev_lb_r = row.get("ev_per_lb_reserve")
                if pd.notna(_ev_oz_p):
                    spg_metrics["EV/oz Produced"] = f"${_ev_oz_p:,.0f}"
                if pd.notna(_ev_oz_r):
                    spg_metrics["EV/oz Reserve"] = f"${_ev_oz_r:,.0f}"
                if pd.notna(_ev_lb_r):
                    spg_metrics["EV/lb Reserve"] = f"${_ev_lb_r:.2f}"

                # Reserve life
                _rli = row.get("spg_reserve_life")
                if pd.notna(_rli) and _rli > 0:
                    spg_metrics["Reserve Life"] = f"{_rli:.1f} yr"

                # Realized price vs AISC (margin context)
                _real_oz = row.get("spg_realized_price_oz")
                _real_lb = row.get("spg_realized_price_lb")
                _real_t  = row.get("spg_realized_price_t")
                if pd.notna(_real_oz):
                    spg_metrics["Realized Price (S&P)"] = f"${_real_oz:,.0f}/oz"
                elif pd.notna(_real_lb):
                    spg_metrics["Realized Price (S&P)"] = f"${_real_lb:.2f}/lb"
                elif pd.notna(_real_t):
                    spg_metrics["Realized Price (S&P)"] = f"${_real_t:,.0f}/t"

                if spg_metrics:
                    st.markdown("**S&P/SNL Mining Data**")
                    spg_cols = st.columns(3)
                    for i, (k, v) in enumerate(spg_metrics.items()):
                        spg_cols[i % 3].metric(k, v)
                    st.markdown("**Standard Metrics**")

                cols_m = st.columns(3)
                for i, (k, v) in enumerate(metrics.items()):
                    cols_m[i % 3].metric(k, v)

            # ── SNL Live Data (Snowflake) ─────────────────────────────────────
            if snl_client.is_configured() and snl_client.get_snl_key(sel_ticker):
                with st.expander("📡 SNL Live Data (Snowflake)", expanded=True):
                    with st.spinner("Querying SNL Metals & Mining..."):
                        _snl = snl_client.get_company_data(sel_ticker)

                    if not _snl.get("found"):
                        st.info("No SNL production/cost data for this company.")
                    else:
                        _pc  = _snl.get("production_costs", [])
                        _gr  = _snl.get("global_rank", [])

                        # ── Latest year per commodity ────────────────────────
                        # Build a pivot: commodity → latest row
                        _latest: dict[str, dict] = {}
                        for r in _pc:
                            comm = r.get("COMMODITY") or r.get("commodity", "")
                            period = r.get("PERIOD") or r.get("period", "")
                            if comm not in _latest or period > _latest[comm].get("PERIOD",""):
                                _latest[comm] = r

                        _rank_map: dict[str, dict] = {}
                        for r in _gr:
                            comm = r.get("COMMODITY") or r.get("commodity", "")
                            period = r.get("PERIOD") or r.get("period", "")
                            if comm not in _rank_map or period > _rank_map[comm].get("PERIOD",""):
                                _rank_map[comm] = r

                        # Display per commodity
                        for comm, r in sorted(_latest.items()):
                            period = r.get("PERIOD") or r.get("period", "")
                            st.markdown(f"**{comm}** — {period}")
                            _cols = st.columns(5)

                            def _fmt_num(v, fmt=",.0f", prefix="$", suffix=""):
                                try:
                                    if v is None or (isinstance(v, float) and np.isnan(v)):
                                        return "—"
                                    return f"{prefix}{float(v):{fmt}}{suffix}"
                                except Exception:
                                    return "—"

                            aisc   = r.get("AISC_OZ") or r.get("aisc_oz")
                            cc     = r.get("CASH_COST_OZ") or r.get("cash_cost_oz")
                            rprice = r.get("REALIZED_PRICE_OZ") or r.get("realized_price_oz")
                            prod   = r.get("PROD_OZ") or r.get("prod_oz")

                            # If no $/oz data, try $/t
                            if aisc is None or (isinstance(aisc, float) and np.isnan(aisc)):
                                aisc   = r.get("AISC_T") or r.get("aisc_t")
                                cc     = r.get("CASH_COST_T") or r.get("cash_cost_t")
                                rprice = r.get("REALIZED_PRICE_T") or r.get("realized_price_t")
                                prod   = r.get("PROD_T") or r.get("prod_t")
                                unit   = "$/t"
                            else:
                                unit = "$/oz"

                            _cols[0].metric("AISC", _fmt_num(aisc) + f" {unit}" if aisc else "—")
                            _cols[1].metric("Cash Cost", _fmt_num(cc) + f" {unit}" if cc else "—")
                            _cols[2].metric("Realized Price", _fmt_num(rprice) + f" {unit}" if rprice else "—")
                            _cols[3].metric("Attributable Prod.", _fmt_num(prod, ",.0f", "", f" {'oz' if unit=='$/oz' else 't'}") if prod else "—")

                            # Global rank
                            rk = _rank_map.get(comm, {})
                            _rank_val  = rk.get("GLOBAL_RANK") or rk.get("global_rank")
                            _share_val = rk.get("WORLD_SHARE_PCT") or rk.get("world_share_pct")
                            if _rank_val:
                                try:
                                    _cols[4].metric("Global Rank", f"#{int(_rank_val)}")
                                except Exception:
                                    _cols[4].metric("Global Rank", str(_rank_val))
                            elif _share_val:
                                try:
                                    _cols[4].metric("World Share", f"{float(_share_val)*100:.2f}%")
                                except Exception:
                                    pass

                        # ── 3-year AISC trend chart ──────────────────────────
                        st.markdown("---")
                        with st.spinner("Loading cost history..."):
                            _hist = snl_client.get_company_aisc_history(sel_ticker)

                        if _hist:
                            _hdf = pd.DataFrame(_hist)
                            # Normalise column names
                            _hdf.columns = [c.upper() for c in _hdf.columns]
                            _gold_hist = _hdf[_hdf["COMMODITY"] == "Gold"] if "COMMODITY" in _hdf.columns else pd.DataFrame()
                            if not _gold_hist.empty and "AISC_OZ" in _gold_hist.columns:
                                _gold_hist = _gold_hist.dropna(subset=["AISC_OZ"])
                                if not _gold_hist.empty:
                                    _fig_h = px.bar(
                                        _gold_hist,
                                        x="PERIOD", y="AISC_OZ",
                                        color_discrete_sequence=["#f0a500"],
                                        labels={"AISC_OZ": "AISC ($/oz)", "PERIOD": "Year"},
                                        title="3-Year Gold AISC History (SNL)",
                                    )
                                    if "CASH_COST_OZ" in _gold_hist.columns:
                                        _fig_h.add_scatter(
                                            x=_gold_hist["PERIOD"],
                                            y=_gold_hist["CASH_COST_OZ"],
                                            name="Cash Cost",
                                            mode="lines+markers",
                                            line=dict(color="#6c9bd2", dash="dot"),
                                        )
                                    if "REALIZED_PRICE_OZ" in _gold_hist.columns:
                                        _fig_h.add_scatter(
                                            x=_gold_hist["PERIOD"],
                                            y=_gold_hist["REALIZED_PRICE_OZ"],
                                            name="Realized Price",
                                            mode="lines+markers",
                                            line=dict(color="#2ecc71"),
                                        )
                                    _fig_h.update_layout(height=300, margin=dict(t=40, b=20))
                                    st.plotly_chart(_fig_h, use_container_width=True)

                        # ── Properties ──────────────────────────────────────
                        st.markdown("---")
                        with st.spinner("Loading properties..."):
                            _props = snl_client.get_property_reserves(sel_ticker)

                        if _props:
                            _pdf = pd.DataFrame(_props)
                            _pdf.columns = [c.upper() for c in _pdf.columns]
                            _show_cols = [c for c in ["PROPERTY_NAME", "STAGE", "STATUS",
                                                       "PRIMARY_COMMODITY", "COUNTRY",
                                                       "STATE_PROVINCE", "PCT_OWN"]
                                          if c in _pdf.columns]
                            if _show_cols:
                                st.markdown(f"**Properties / Assets ({len(_pdf)} records)**")
                                st.dataframe(
                                    _pdf[_show_cols].rename(columns={
                                        "PROPERTY_NAME":     "Property",
                                        "STAGE":             "Stage",
                                        "STATUS":            "Status",
                                        "PRIMARY_COMMODITY": "Commodity",
                                        "COUNTRY":           "Country",
                                        "STATE_PROVINCE":    "Province/State",
                                        "PCT_OWN":           "Ownership %",
                                    }),
                                    use_container_width=True,
                                    hide_index=True,
                                )

                        st.caption("Data: S&P Global Market Intelligence SNL Metals & Mining via Snowflake (live, not stored)")

            # ── SNL Project Portfolio (local SQLite cache) ────────────────────
            _snl_detail_key = None
            try:
                import json as _jd
                with open(os.path.join(os.path.dirname(__file__), "_asx_snl_ticker_mapping.json")) as _jf:
                    _jmap = _jd.load(_jf)
                if sel_ticker in _jmap:
                    _snl_detail_key = str(_jmap[sel_ticker]["snl_key"])
            except Exception:
                pass

            if _snl_detail_key:
                _snl_db = str(config.DB_PATH)
                if os.path.exists(_snl_db):
                    with st.expander("📊 SNL Project Portfolio (local cache)", expanded=True):
                        import sqlite3 as _sd3
                        _sdconn = _sd3.connect(_snl_db)

                        # ── AISC History chart ───────────────────────────────
                        _ph = pd.read_sql_query(
                            "SELECT period, commodity, prod_oz, prod_t, prod_lb, "
                            "aisc_oz, aisc_t, aisc_lb, cash_cost_oz, cash_cost_t, "
                            "realized_price_oz, realized_price_t, revenue_m "
                            "FROM snl_company_production WHERE snl_key=? ORDER BY period",
                            _sdconn, params=(_snl_detail_key,)
                        )
                        if not _ph.empty:
                            st.markdown("**Cost History (SNL)**")
                            # Try gold first; fallback to base metal
                            _ph_gold = _ph[_ph["commodity"] == "Gold"]
                            _ph_use  = _ph_gold if not _ph_gold.empty else _ph
                            _ph_use  = _ph_use.copy()
                            # Pick the right unit columns
                            if _ph_use["aisc_oz"].notna().any():
                                _ha, _hc, _hr = "aisc_oz", "cash_cost_oz", "realized_price_oz"
                                _hu = "$/oz"
                            elif _ph_use["aisc_t"].notna().any():
                                _ha, _hc, _hr = "aisc_t", "cash_cost_t", "realized_price_t"
                                _hu = "$/t"
                            else:
                                _ha, _hc, _hr = None, None, None
                                _hu = ""

                            if _ha and _ph_use[_ha].notna().any():
                                _fig_ph = px.bar(
                                    _ph_use.dropna(subset=[_ha]),
                                    x="period", y=_ha,
                                    color_discrete_sequence=["#f0a500"],
                                    labels={_ha: f"AISC ({_hu})", "period": ""},
                                    title=f"AISC History ({_hu})",
                                )
                                if _hc in _ph_use.columns and _ph_use[_hc].notna().any():
                                    _fig_ph.add_scatter(
                                        x=_ph_use.dropna(subset=[_hc])["period"],
                                        y=_ph_use.dropna(subset=[_hc])[_hc],
                                        name=f"C1 Cash Cost",
                                        mode="lines+markers",
                                        line=dict(color="#6c9bd2", dash="dot"),
                                    )
                                if _hr in _ph_use.columns and _ph_use[_hr].notna().any():
                                    _fig_ph.add_scatter(
                                        x=_ph_use.dropna(subset=[_hr])["period"],
                                        y=_ph_use.dropna(subset=[_hr])[_hr],
                                        name="Realized Price",
                                        mode="lines+markers",
                                        line=dict(color="#22c55e"),
                                    )
                                _fig_ph.update_layout(height=280, margin=dict(t=40, b=20),
                                                      legend=dict(orientation="h", y=-0.25))
                                st.plotly_chart(_fig_ph, use_container_width=True)

                        # ── Projects & Properties ────────────────────────────
                        _pp = pd.read_sql_query(
                            "SELECT g.property_name, g.stage, g.status, g.primary_commodity, "
                            "g.country, g.state_province, "
                            "round(o.pct_own,1) pct_own, "
                            "c.mill_capacity_tpd, c.actual_startup_year "
                            "FROM snl_property_general g "
                            "JOIN snl_property_owner o ON o.property_id=g.property_id "
                            "LEFT JOIN snl_property_capacity c ON c.property_id=g.property_id "
                            "WHERE o.snl_key=? "
                            "ORDER BY g.stage, g.primary_commodity, g.property_name",
                            _sdconn, params=(_snl_detail_key,)
                        )
                        if not _pp.empty:
                            st.markdown(f"**Project Pipeline — {len(_pp)} properties**")
                            _pp.columns = ["Property", "Stage", "Status", "Commodity",
                                           "Country", "Province", "Own%", "Capacity(t/d)", "Startup Yr"]
                            # Stage colour coding
                            _stage_order = [
                                "Mine", "Mine/Mill", "Advanced Development",
                                "Preliminary Feasibility", "Scoping Study",
                                "Resource Definition", "Exploration",
                            ]
                            st.dataframe(
                                _pp.sort_values("Stage",
                                    key=lambda s: s.map({v: i for i, v in enumerate(_stage_order)}).fillna(99)),
                                use_container_width=True, hide_index=True,
                                column_config={
                                    "Own%": st.column_config.NumberColumn(format="%.1f%%"),
                                    "Capacity(t/d)": st.column_config.NumberColumn(format="%,.0f"),
                                },
                            )

                        # ── Feasibility Studies ──────────────────────────────
                        _fs = pd.read_sql_query(
                            "SELECT s.property_name, s.study_type, s.study_year, "
                            "round(s.posttax_npv_m,0) npv_m, "
                            "round(s.posttax_irr_pct,1) irr, "
                            "round(s.mine_life_yrs,0) life_yr, "
                            "round(s.initial_capex_m,0) capex_m, "
                            "round(s.lom_sustaining_m,0) sustaining_m, "
                            "s.npv_discount_pct, s.currency "
                            "FROM snl_property_studies s "
                            "JOIN snl_property_owner o ON o.property_id=s.property_id "
                            "WHERE o.snl_key=? "
                            "ORDER BY s.study_year DESC, s.posttax_npv_m DESC",
                            _sdconn, params=(_snl_detail_key,)
                        )
                        if not _fs.empty:
                            st.markdown("**Feasibility Studies (SNL)**")
                            _fs.columns = ["Property", "Type", "Year", "Post-tax NPV($M)",
                                           "IRR%", "Life(yr)", "Initial CAPEX($M)",
                                           "LOM Sustaining($M)", "Discount%", "Currency"]
                            st.dataframe(
                                _fs,
                                use_container_width=True, hide_index=True,
                                column_config={
                                    "Post-tax NPV($M)": st.column_config.NumberColumn(format="$%,.0f"),
                                    "Initial CAPEX($M)": st.column_config.NumberColumn(format="$%,.0f"),
                                    "LOM Sustaining($M)": st.column_config.NumberColumn(format="$%,.0f"),
                                },
                            )

                        # ── Production Forecast ──────────────────────────────
                        _pf = pd.read_sql_query(
                            "SELECT estimate_period, description, "
                            "round(prod_high_oz/1000,0) hi_koz, round(prod_low_oz/1000,0) lo_koz, "
                            "round(prod_high_t/1000,0)  hi_kt,  round(prod_low_t/1000,0)  lo_kt, "
                            "aisc_high_oz, aisc_low_oz, cash_cost_high_oz, cash_cost_low_oz "
                            "FROM snl_company_projections "
                            "WHERE snl_key=? AND estimate_period>='2024' "
                            "ORDER BY estimate_period",
                            _sdconn, params=(_snl_detail_key,)
                        )
                        if not _pf.empty:
                            st.markdown("**Production & Cost Forecast (SNL)**")
                            _pf.columns = ["Period", "Description",
                                           "Prod High(koz)", "Prod Low(koz)",
                                           "Prod High(kt)",  "Prod Low(kt)",
                                           "AISC High($/oz)", "AISC Low($/oz)",
                                           "C1 High($/oz)",   "C1 Low($/oz)"]
                            st.dataframe(_pf, use_container_width=True, hide_index=True)

                            # Plot forecast range if oz data available
                            _pf_oz = _pf[_pf["Prod High(koz)"].notna() | _pf["AISC High($/oz)"].notna()].copy()
                            if not _pf_oz.empty and _pf_oz["Prod High(koz)"].notna().any():
                                _fc1, _fc2 = st.columns(2)
                                with _fc1:
                                    _fig_pf = go.Figure()
                                    _fig_pf.add_trace(go.Scatter(
                                        x=_pf_oz["Period"].tolist() + _pf_oz["Period"].tolist()[::-1],
                                        y=_pf_oz["Prod High(koz)"].tolist() + _pf_oz["Prod Low(koz)"].tolist()[::-1],
                                        fill="toself", fillcolor="rgba(59,130,246,0.15)",
                                        line=dict(color="rgba(0,0,0,0)"), name="Prod Range",
                                    ))
                                    _fig_pf.add_trace(go.Scatter(
                                        x=_pf_oz["Period"],
                                        y=(_pf_oz["Prod High(koz)"].fillna(0) + _pf_oz["Prod Low(koz)"].fillna(0)) / 2,
                                        mode="lines+markers", name="Midpoint",
                                        line=dict(color="#3b82f6", width=2),
                                    ))
                                    _fig_pf.update_layout(
                                        title="Production Forecast (koz)", height=240,
                                        margin=dict(t=40, b=20), showlegend=False,
                                        xaxis_title="", yaxis_title="koz",
                                    )
                                    st.plotly_chart(_fig_pf, use_container_width=True)

                                with _fc2:
                                    _pf_aisc = _pf_oz[_pf_oz["AISC High($/oz)"].notna()]
                                    if not _pf_aisc.empty:
                                        _fig_pfa = go.Figure()
                                        _fig_pfa.add_trace(go.Scatter(
                                            x=_pf_aisc["Period"].tolist() + _pf_aisc["Period"].tolist()[::-1],
                                            y=_pf_aisc["AISC High($/oz)"].tolist() + _pf_aisc["AISC Low($/oz)"].fillna(0).tolist()[::-1],
                                            fill="toself", fillcolor="rgba(234,179,8,0.15)",
                                            line=dict(color="rgba(0,0,0,0)"), name="AISC Range",
                                        ))
                                        _fig_pfa.add_trace(go.Scatter(
                                            x=_pf_aisc["Period"],
                                            y=(_pf_aisc["AISC High($/oz)"].fillna(0) + _pf_aisc["AISC Low($/oz)"].fillna(0)) / 2,
                                            mode="lines+markers", name="Midpoint",
                                            line=dict(color="#eab308", width=2),
                                        ))
                                        _fig_pfa.update_layout(
                                            title="AISC Forecast ($/oz)", height=240,
                                            margin=dict(t=40, b=20), showlegend=False,
                                            xaxis_title="", yaxis_title="$/oz",
                                        )
                                        st.plotly_chart(_fig_pfa, use_container_width=True)

                        _sdconn.close()
                        st.caption("Source: SNL Metals & Mining via local cache (snl_sync.py)")

            # ── Cost Structure Waterfall ──────────────────────────────────────
            _wf_comm = str(row.get("commodity", "")).lower()
            # Pick the right cost columns based on commodity
            if "gold" in _wf_comm or "silver" in _wf_comm:
                _wf_cc   = row.get("spg_cash_cost_oz")
                _wf_aisc = row.get("spg_aisc_per_oz")
                _wf_spot = config.COMMODITY_SPOT.get("Gold") if "gold" in _wf_comm else config.COMMODITY_SPOT.get("Silver")
                _wf_unit = "$/oz"
            elif "copper" in _wf_comm:
                _wf_cc   = row.get("spg_cash_cost_lb")
                _wf_aisc = row.get("spg_aisc_per_lb")
                _wf_spot = config.COMMODITY_SPOT.get("Copper")
                _wf_unit = "$/lb"
            elif "uranium" in _wf_comm:
                _wf_cc   = row.get("spg_cash_cost_lb")
                _wf_aisc = row.get("spg_aisc_per_lb")
                _wf_spot = config.COMMODITY_SPOT.get("Uranium")
                _wf_unit = "$/lb"
            else:
                _wf_cc   = row.get("spg_cash_cost_t")
                _wf_aisc = row.get("spg_aisc_per_t")
                _wf_spot = None
                _wf_unit = "$/t"

            if pd.notna(_wf_cc) and _wf_cc > 0 and pd.notna(_wf_aisc) and _wf_aisc > 0:
                with st.expander("🏗️ Cost Structure Waterfall", expanded=False):
                    _wf_sustaining = max(_wf_aisc - _wf_cc, 0)
                    _wf_margin     = (_wf_spot - _wf_aisc) if _wf_spot else None

                    # Build waterfall segments
                    _wf_x      = ["Cash Cost", "Sustaining\nCapex", "AISC Total"]
                    _wf_y      = [_wf_cc, _wf_sustaining, 0]          # bar heights
                    _wf_base   = [0, _wf_cc, 0]                        # bar bases
                    _wf_colors = ["#3b82f6", "#eab308", "#94a3b8"]
                    _wf_text   = [f"${_wf_cc:,.0f}", f"+${_wf_sustaining:,.0f}", f"${_wf_aisc:,.0f}"]

                    if _wf_spot and _wf_margin is not None:
                        _margin_color = "#22c55e" if _wf_margin > 0 else "#ef4444"
                        _wf_x.append("Margin\nvs Spot")
                        _wf_y.append(abs(_wf_margin))
                        _wf_base.append(_wf_aisc if _wf_margin > 0 else _wf_aisc + _wf_margin)
                        _wf_colors.append(_margin_color)
                        _wf_text.append(f"{'+'  if _wf_margin > 0 else ''}{_wf_margin:,.0f}")

                    fig_wf = go.Figure()
                    # Invisible base bars
                    fig_wf.add_trace(go.Bar(
                        x=_wf_x, y=_wf_base,
                        marker_color="rgba(0,0,0,0)",
                        showlegend=False,
                        hoverinfo="skip",
                    ))
                    # Visible bars
                    fig_wf.add_trace(go.Bar(
                        x=_wf_x, y=_wf_y,
                        marker_color=_wf_colors,
                        text=_wf_text, textposition="inside",
                        textfont=dict(color="white", size=12),
                        showlegend=False,
                        hovertemplate="%{x}: %{text}<extra></extra>",
                    ))
                    # Connector lines between bars
                    _connector_vals = [_wf_cc, _wf_aisc]
                    for _ci, _cv in enumerate(_connector_vals):
                        fig_wf.add_shape(
                            type="line",
                            x0=_ci + 0.4, x1=_ci + 0.6,
                            y0=_cv, y1=_cv,
                            line=dict(color="rgba(23,32,51,0.4)", width=1, dash="dot"),
                        )
                    if _wf_spot:
                        fig_wf.add_hline(
                            y=_wf_spot,
                            line_dash="dash", line_color="#FFD700", line_width=1.5,
                            annotation_text=f"Spot ${_wf_spot:,.0f}",
                            annotation_position="top right",
                            annotation_font_color="#FFD700",
                        )
                    _aisc_margin_pct = ((_wf_spot - _wf_aisc) / _wf_spot * 100) if _wf_spot else None
                    _cc_margin_pct   = ((_wf_spot - _wf_cc)   / _wf_spot * 100) if _wf_spot else None
                    fig_wf.update_layout(
                        barmode="stack",
                        height=340,
                        yaxis_title=_wf_unit,
                        showlegend=False,
                        margin=dict(t=30, b=20),
                        plot_bgcolor="rgba(0,0,0,0)",
                        paper_bgcolor="rgba(0,0,0,0)",
                    )
                    st.plotly_chart(fig_wf, width="stretch")
                    _wf_caption_parts = [
                        f"🔵 Cash Cost: ${_wf_cc:,.0f}",
                        f"🟡 Sustaining Capex: ${_wf_sustaining:,.0f}",
                        f"Total AISC: ${_wf_aisc:,.0f}",
                    ]
                    if _aisc_margin_pct is not None:
                        _margin_icon = "🟢" if _aisc_margin_pct > 20 else "🟡" if _aisc_margin_pct > 0 else "🔴"
                        _wf_caption_parts.append(f"{_margin_icon} AISC Margin: {_aisc_margin_pct:.1f}%")
                    if _cc_margin_pct is not None:
                        _wf_caption_parts.append(f"Cash Cost Margin: {_cc_margin_pct:.1f}%")
                    st.caption("  ·  ".join(_wf_caption_parts))

            # ── NAV Sensitivity Analysis ──────────────────────────────────────
            _nav_comm  = row.get("commodity", "")
            _nav_pnav  = row.get("spg_p_nav")
            _nav_price = row.get("price")
            _nav_aisc  = row.get("spg_aisc_per_oz")
            _is_gold_silver = any(c in str(_nav_comm) for c in ["Gold", "Silver"])

            if _is_gold_silver and pd.notna(_nav_pnav) and _nav_pnav > 0 and pd.notna(_nav_price):
                with st.expander("🧮 NAV Sensitivity — Gold Price Scenarios", expanded=False):
                    _base_gold = config.COMMODITY_SPOT.get("Gold", 2300)
                    _nav_per_share = _nav_price / _nav_pnav   # current NAV/share

                    st.caption(
                        f"Base gold spot: **${_base_gold:,.0f}/oz** · "
                        f"Current NAV/share: **${_nav_per_share:.3f}** · "
                        f"Current P/NAV: **{_nav_pnav:.2f}x**  \n"
                        "NAV scales linearly with gold price (approximation — actual depends on cost structure)."
                    )

                    _gold_scenarios = [1800, 2000, 2200, 2400, 2600, 2800, 3000, 3200, 3500]
                    _scen_rows = []
                    for _gp in _gold_scenarios:
                        _scale      = _gp / _base_gold
                        _new_nav    = _nav_per_share * _scale
                        _new_pnav   = _nav_price / _new_nav if _new_nav > 0 else None
                        _new_upside = (_new_nav / _nav_price - 1) * 100 if _nav_price > 0 else None
                        _new_aisc_m = ((_gp - _nav_aisc) / _gp * 100
                                       if pd.notna(_nav_aisc) and _nav_aisc > 0 else None)
                        _scen_rows.append({
                            "Gold $/oz":    _gp,
                            "NAV/share":    round(_new_nav, 3),
                            "P/NAV":        round(_new_pnav, 2) if _new_pnav else None,
                            "Upside%":      round(_new_upside, 1) if _new_upside else None,
                            "AISC Margin%": round(_new_aisc_m, 1) if _new_aisc_m else None,
                            "_is_base":     abs(_gp - _base_gold) < 50,
                        })

                    _scen_df = pd.DataFrame(_scen_rows)
                    _display_cols = ["Gold $/oz", "NAV/share", "P/NAV", "Upside%"]
                    if _scen_df["AISC Margin%"].notna().any():
                        _display_cols.append("AISC Margin%")

                    def _color_upside_scen(val):
                        if pd.isna(val): return ""
                        if val >= 50:  return "background-color:#dcfce7;color:#14532d;font-weight:700"
                        if val >= 20:  return "background-color:#dbeafe;color:#1e3a5f;font-weight:600"
                        if val >= 0:   return "color:#16a34a"
                        return "color:#dc2626"

                    def _color_aisc(val):
                        if pd.isna(val): return ""
                        if val >= 50: return "color:#22c55e;font-weight:700"
                        if val >= 30: return "color:#eab308"
                        return "color:#ef4444"

                    _scen_styled = (
                        _scen_df[_display_cols].style
                        .map(_color_upside_scen, subset=["Upside%"])
                        .map(_color_aisc, subset=["AISC Margin%"] if "AISC Margin%" in _display_cols else [])
                        .format({
                            "Gold $/oz":    "${:,.0f}",
                            "NAV/share":    "${:.3f}",
                            "P/NAV":        "{:.2f}x",
                            "Upside%":      lambda x: f"{x:+.1f}%" if pd.notna(x) else "—",
                            "AISC Margin%": lambda x: f"{x:.1f}%" if pd.notna(x) else "—",
                        }, na_rep="—")
                        .apply(lambda _: [
                            "background-color:#dbeafe;color:#1e3a5f" if _scen_df.loc[i, "_is_base"] else ""
                            for i in range(len(_scen_df))
                        ], axis=0)
                    )
                    st.dataframe(_scen_styled, width="stretch", height=380)

                    # Chart: upside% vs gold price
                    _chart_scen = _scen_df[_scen_df["Upside%"].notna()]
                    if not _chart_scen.empty:
                        fig_nav_sens = go.Figure()
                        fig_nav_sens.add_trace(go.Scatter(
                            x=_chart_scen["Gold $/oz"],
                            y=_chart_scen["Upside%"],
                            mode="lines+markers",
                            line=dict(color="#3b82f6", width=2.5),
                            marker=dict(size=7),
                            name="Upside to NAV%",
                            fill="tozeroy",
                            fillcolor="rgba(59,130,246,0.08)",
                        ))
                        fig_nav_sens.add_hline(y=0, line_dash="dash",
                                               line_color="gray", line_width=1)
                        fig_nav_sens.add_vline(
                            x=_base_gold, line_dash="dot",
                            line_color="#eab308",
                            annotation_text=f"Current ${_base_gold:,.0f}",
                            annotation_position="top right",
                        )
                        fig_nav_sens.update_layout(
                            xaxis_title="Gold Price (USD/oz)",
                            yaxis_title="Upside to NAV %",
                            height=260,
                            margin=dict(l=0, r=0, t=10, b=40),
                            showlegend=False,
                        )
                        st.plotly_chart(fig_nav_sens, width="stretch")

                    st.caption(
                        "⚠️ Linear approximation only. Actual NAV depends on "
                        "reserve grades, strip ratios, capex, and discount rate. "
                        "Use S&P Capital IQ for precise scenario modelling."
                    )

            # ── Universe Percentile Ranks ─────────────────────────────────────
            with st.expander("📊 Universe Percentile Ranks", expanded=False):
                st.caption(
                    f"Where **{row.get('name', sel_ticker)}** stands among all "
                    f"{len(df)} screened companies — higher percentile = better rank."
                )
                _prank_metrics = [
                    # (column, label, invert)  invert=True → lower value is better
                    ("score_composite",  "Composite Score",  False),
                    ("score_valuation",  "Valuation Score",  False),
                    ("score_health",     "Health Score",     False),
                    ("score_momentum",   "Momentum Score",   False),
                    ("score_mining",     "Mining Score",     False),
                    ("price_to_book",    "P/B Ratio",        True),
                    ("ev_ebitda",        "EV/EBITDA",        True),
                    ("spg_p_nav",        "P/NAV",            True),
                    ("upside_to_nav",    "↑ Upside to NAV%", False),
                    ("spg_aisc_margin",  "AISC Margin%",     False),
                    ("fcf_yield",        "FCF Yield%",       False),
                    ("dividend_yield",   "Dividend Yield%",  False),
                    ("rsi",              "RSI (oversold)",   True),
                    ("wk52_position",    "52wk Pos (low)",   True),
                    ("return_1m",        "1M Return%",       False),
                    ("return_3m",        "3M Return%",       False),
                ]
                _rank_rows = []
                for _col, _lbl, _inv in _prank_metrics:
                    if _col not in df.columns:
                        continue
                    _val = row.get(_col)
                    _ser = df[_col].dropna()
                    if pd.isna(_val) or len(_ser) < 3:
                        continue
                    _raw_pct = (_ser < float(_val)).mean() * 100
                    _pct = round((100 - _raw_pct) if _inv else _raw_pct, 0)
                    _tier = ("🟢 Top 10%"     if _pct >= 90 else
                             "🔵 Top 25%"     if _pct >= 75 else
                             "🟡 Mid"         if _pct >= 40 else
                             "🟠 Bottom 25%"  if _pct >= 15 else
                             "🔴 Bottom 10%")
                    _rank_rows.append({
                        "Metric": _lbl,
                        "Value":  f"{float(_val):.2f}",
                        "Pct":    int(_pct),
                        "Tier":   _tier,
                    })
                if _rank_rows:
                    _pr_cols = st.columns(2)
                    _half = len(_rank_rows) // 2 + len(_rank_rows) % 2
                    for _ci, _chunk in enumerate([_rank_rows[:_half], _rank_rows[_half:]]):
                        with _pr_cols[_ci]:
                            for _r in _chunk:
                                _bw = _r["Pct"]
                                _bc = ("#22c55e" if _bw >= 75 else
                                       "#3b82f6" if _bw >= 50 else
                                       "#eab308" if _bw >= 25 else "#ef4444")
                                st.markdown(
                                    f"**{_r['Metric']}** ({_r['Value']}) — {_r['Tier']}<br>"
                                    f"<div style='background:#e3e9f0;border-radius:3px;"
                                    f"height:6px;width:100%'><div style='background:{_bc};"
                                    f"height:6px;width:{_bw}%;border-radius:3px'></div></div>",
                                    unsafe_allow_html=True,
                                )
                else:
                    st.info("Not enough data to compute percentile ranks.")

            # ── Score Explainer ───────────────────────────────────────────────
            with st.expander("🔍 Score Explainer — why this score?", expanded=False):
                _expl_cols = st.columns(3)

                def _score_bar(label: str, score, col, hints: list[str]):
                    """Render a mini progress bar + bullet hints."""
                    if pd.isna(score):
                        col.markdown(f"**{label}**: —")
                        return
                    s = float(score)
                    color = "#22c55e" if s >= 70 else "#eab308" if s >= 45 else "#ef4444"
                    col.markdown(
                        f"**{label}**: {s:.0f}/100<br>"
                        f"<div style='background:#e3e9f0;border-radius:4px;height:8px;width:100%'>"
                        f"<div style='background:{color};height:8px;width:{s}%;border-radius:4px'></div></div>",
                        unsafe_allow_html=True,
                    )
                    for h in hints:
                        col.caption(h)

                # Valuation
                _pb    = row.get("price_to_book")
                _ev    = row.get("ev_ebitda")
                _pcf   = row.get("p_cf")
                _pnav  = row.get("spg_p_nav")
                _val_hints = []
                if pd.notna(_pb):   _val_hints.append(f"P/B = {_pb:.2f}x {'✅' if _pb < 1.5 else '⚠️' if _pb < 3 else '❌'}")
                if pd.notna(_ev):   _val_hints.append(f"EV/EBITDA = {_ev:.1f}x {'✅' if _ev < 8 else '⚠️' if _ev < 15 else '❌'}")
                if pd.notna(_pcf):  _val_hints.append(f"P/CF = {_pcf:.1f}x {'✅' if _pcf < 10 else '⚠️'}")
                if pd.notna(_pnav): _val_hints.append(f"P/NAV = {_pnav:.2f}x {'✅' if _pnav < 1 else '⚠️' if _pnav < 1.5 else '❌'}")
                if not _val_hints:  _val_hints = ["No valuation data available"]
                _score_bar("Valuation", row.get("score_valuation"), _expl_cols[0], _val_hints)

                # Health
                _cr   = row.get("current_ratio")
                _de   = row.get("debt_to_equity")
                _cpct = row.get("cash_pct_mcap")
                _fcf  = row.get("freeCashflow") if "freeCashflow" in row.index else None
                _hlt_hints = []
                if pd.notna(_cr):   _hlt_hints.append(f"Current ratio = {_cr:.1f} {'✅' if _cr > 2 else '⚠️' if _cr > 1 else '❌'}")
                if pd.notna(_de):   _hlt_hints.append(f"D/E = {_de:.0f}% {'✅' if _de < 50 else '⚠️' if _de < 100 else '❌'}")
                if pd.notna(_cpct): _hlt_hints.append(f"Cash = {_cpct:.0f}% of mkt cap {'✅' if _cpct > 15 else '⚠️'}")
                if not _hlt_hints:  _hlt_hints = ["No health data available"]
                _score_bar("Health", row.get("score_health"), _expl_cols[1], _hlt_hints)

                # Momentum
                _rsi   = row.get("rsi")
                _wk52  = row.get("wk52_position")
                _vsma  = row.get("price_vs_ma200")
                _mom_hints = []
                if pd.notna(_rsi):  _mom_hints.append(f"RSI = {_rsi:.0f} {'✅ oversold' if _rsi < 35 else '⚠️' if _rsi < 50 else '—'}")
                if pd.notna(_wk52): _mom_hints.append(f"52wk pos = {_wk52:.0f}% {'✅ near low' if _wk52 < 25 else '—'}")
                if pd.notna(_vsma): _mom_hints.append(f"vs 200MA = {_vsma:+.0f}% {'✅' if _vsma < -15 else '—'}")
                if not _mom_hints:  _mom_hints = ["No momentum data"]
                _score_bar("Momentum", row.get("score_momentum"), _expl_cols[0], _mom_hints)

                # Mining (S&P) — 6 sub-components
                _min_hints = []
                # ① AISC Margin (30%)
                if pd.notna(row.get("spg_aisc_margin")):
                    _m = row["spg_aisc_margin"]
                    _icon = "✅" if _m > 40 else "⚠️" if _m > 20 else "❌"
                    _min_hints.append(f"① AISC margin = {_m:.0f}%  {_icon}  (wt 30%)")
                # ② NAV Discount (20%)
                if pd.notna(row.get("spg_p_nav")):
                    _pn = row["spg_p_nav"]
                    _icon = "✅" if _pn < 0.9 else "⚠️" if _pn < 1.25 else "❌"
                    _min_hints.append(f"② P/NAV = {_pn:.2f}x  {_icon}  (wt 20%)")
                # ③ Reserves Backing (15%)
                if pd.notna(row.get("spg_reserves_m")) and pd.notna(row.get("market_cap")) and row.get("market_cap", 0) > 0:
                    _bk = row["spg_reserves_m"] / (row["market_cap"] / 1e6)
                    _icon = "✅" if _bk > 2 else "⚠️" if _bk > 1 else "❌"
                    _min_hints.append(f"③ Reserves backing = {_bk:.1f}x  {_icon}  (wt 15%)")
                # ④ EV/oz Production (15%) — peer-percentile
                _ev_oz_p = row.get("ev_per_oz_prod")
                if pd.notna(_ev_oz_p):
                    _icon = "✅" if _ev_oz_p < 8000 else "⚠️" if _ev_oz_p < 15000 else "❌"
                    _min_hints.append(f"④ EV/oz produced = ${_ev_oz_p:,.0f}  {_icon}  (wt 15%)")
                # ⑤ Reserve Life (10%)
                _rli = row.get("spg_reserve_life")
                if pd.notna(_rli):
                    _icon = "✅" if _rli > 10 else "⚠️" if _rli > 6 else "❌"
                    _min_hints.append(f"⑤ Reserve life = {_rli:.1f} yr  {_icon}  (wt 10%)")
                # ⑥ Grade Quality (10%)
                _gg = row.get("spg_grade_gpt")
                _gp = row.get("spg_grade_pct")
                if pd.notna(_gg):
                    _icon = "✅" if _gg > 2 else "⚠️" if _gg > 0.8 else "❌"
                    _min_hints.append(f"⑥ Grade = {_gg:.3f} g/t  {_icon}  (wt 10%)")
                elif pd.notna(_gp):
                    _icon = "✅" if _gp > 0.8 else "⚠️" if _gp > 0.3 else "❌"
                    _min_hints.append(f"⑥ Grade = {_gp:.3f}%  {_icon}  (wt 10%)")
                if not _min_hints:
                    _min_hints = ["No S&P/SNL data — score is neutral 50"]
                _score_bar("⛏️ Mining (S&P)", row.get("score_mining"), _expl_cols[1], _min_hints)

                # Commodity + Stage
                _comm_hints = [
                    f"Commodity: {row.get('commodity', '—')}",
                    f"Outlook score: {row.get('score_commodity', '—')}/100",
                ]
                _stage_hints = [
                    f"Stage: {row.get('stage', '—')}",
                    f"Stage score: {row.get('score_stage', '—')}/100",
                ]
                _score_bar("Commodity", row.get("score_commodity"), _expl_cols[2], _comm_hints)
                _score_bar("Stage", row.get("score_stage"), _expl_cols[2], _stage_hints)

                # Peer rank context
                _pr   = row.get("peer_rank")
                _pn   = row.get("peer_n")
                _pgrp = row.get("peer_group", "")
                _ppct = row.get("peer_pct")
                if pd.notna(_pr) and pd.notna(_pn) and _pn > 0:
                    _pr_color = "#b45309" if _pr == 1 else "#16a34a" if _pr <= max(2, _pn * 0.25) else "#5b6b7f"
                    st.markdown(
                        f"<div style='margin-top:12px;padding:10px 14px;background:#ffffff;"
                        f"border:1px solid #e3e9f0;border-radius:8px;border-left:3px solid {_pr_color}'>"
                        f"<span style='color:{_pr_color};font-weight:700;font-size:15px'>"
                        f"#{int(_pr)} of {int(_pn)}</span>"
                        f"<span style='color:#5b6b7f;font-size:13px'> in peer group: "
                        f"<b style='color:#172033'>{_pgrp}</b></span>"
                        + (f"<span style='color:#5b6b7f;font-size:12px'> · top {100-int(_ppct)}%</span>"
                           if pd.notna(_ppct) else "")
                        + "</div>",
                        unsafe_allow_html=True,
                    )

            # Historical charts
            hist = load_history(sel_ticker, days=90)
            if not hist.empty:
                hist = hist.sort_values("snap_date")
                h_col1, h_col2 = st.columns(2)

                with h_col1:
                    st.markdown("#### Price History (90d)")
                    if hist["price"].notna().any():
                        fig_price = px.line(
                            hist, x="snap_date", y="price",
                            labels={"price": f"Price ({config.CURRENCY})", "snap_date": "Date"},
                        )
                        fig_price.update_traces(line_color="#3b82f6")
                        fig_price.update_layout(height=260, margin=dict(t=10, b=10))
                        st.plotly_chart(fig_price, width="stretch")
                    else:
                        st.info("No price history.")

                with h_col2:
                    st.markdown("#### Score History (90d)")
                    score_cols = [c for c in
                        ["score_composite","score_valuation","score_health","score_momentum"]
                        if c in hist.columns and hist[c].notna().any()]
                    fig_scores = px.line(
                        hist, x="snap_date", y=score_cols,
                        labels={"value": "Score", "snap_date": "Date", "variable": ""},
                        color_discrete_map={
                            "score_composite":  "#22c55e",
                            "score_valuation":  "#3b82f6",
                            "score_health":     "#a855f7",
                            "score_momentum":   "#f59e0b",
                        },
                    )
                    fig_scores.update_layout(height=260, margin=dict(t=10, b=10))
                    st.plotly_chart(fig_scores, width="stretch")
            else:
                st.info("No history yet — needs multiple daily refreshes to build history.")

            # ── Peer Comparison ───────────────────────────────────────────────
            st.divider()
            st.markdown("#### 📊 Peer Comparison")

            # Determine primary commodity (before "/")
            _primary_comm = str(row.get("commodity", "")).split("/")[0].strip()
            _sel_stage    = str(row.get("stage", ""))

            # Strict peers: same primary commodity + same stage
            _peer_mask_strict = (
                df["commodity"].str.startswith(_primary_comm) &
                (df["stage"] == _sel_stage) &
                (df["ticker"] != sel_ticker)
            )
            _peers_strict = df[_peer_mask_strict]

            # Broad peers: same primary commodity, any stage
            _peer_mask_broad = (
                df["commodity"].str.startswith(_primary_comm) &
                (df["ticker"] != sel_ticker)
            )
            _peers = _peers_strict if len(_peers_strict) >= 3 else df[_peer_mask_broad]
            _peer_label = (
                f"{_primary_comm} · {_sel_stage}"
                if len(_peers_strict) >= 3
                else f"{_primary_comm} (all stages)"
            )

            if _peers.empty:
                st.info("Not enough peers to compare.")
            else:
                st.caption(
                    f"Comparing **{row.get('name', sel_ticker)}** against "
                    f"**{len(_peers)}** peers in {_peer_label}"
                )

                # ── Comparison table ──────────────────────────────────────────
                _peer_display_cols = {
                    "grade":               "Grade",
                    "name":                "Company",
                    "stage":               "Stage",
                    "score_composite":     "Score",
                    "price_to_book":       "P/B",
                    "ev_ebitda":           "EV/EBITDA",
                    "spg_p_nav":           "P/NAV",
                    "spg_aisc_margin":     "AISC Margin%",
                    "spg_cash_cost_oz":    "Cash Cost($/oz)",
                    "ev_per_oz_prod":      "EV/oz Prod",
                    "ev_per_oz_reserve":   "EV/oz Rsv",
                    "spg_production_oz":   "Prod(koz/yr)",
                    "spg_reserve_life":    "Rsv Life(yr)",
                    "spg_reserves_m":      "Reserves($M)",
                    "upside_to_nav":       "↑ NAV Upside%",
                    "rsi":                 "RSI",
                }
                _avail = {k: v for k, v in _peer_display_cols.items() if k in df.columns}

                # Build table: selected company on top, then peers sorted by score
                _sel_row_df = df[df["ticker"] == sel_ticker][list(_avail.keys())].copy()
                _sel_row_df.insert(0, "_highlight", "► Selected")

                _peers_tbl = _peers[list(_avail.keys())].copy()
                _peers_tbl.insert(0, "_highlight", "Peer")
                _peers_tbl = _peers_tbl.sort_values("score_composite", ascending=False).head(15)

                _combined = pd.concat([_sel_row_df, _peers_tbl], ignore_index=True)
                _combined_display = _combined.rename(columns=_avail)

                # Track which row index is "selected" before dropping the flag column
                _sel_indices = _combined_display.index[_combined_display["_highlight"] == "► Selected"].tolist()

                def _highlight_selected(row_s):
                    if row_s.name in _sel_indices:
                        return ["background-color: #dbeafe; color: #1e3a5f; font-weight: bold"] * len(row_s)
                    return [""] * len(row_s)

                # Convert production oz → koz for readability
                if "Prod(koz/yr)" in _combined_display.columns:
                    _combined_display["Prod(koz/yr)"] = (
                        pd.to_numeric(_combined_display["Prod(koz/yr)"], errors="coerce") / 1000
                    )

                _combined_styled = (
                    _combined_display
                    .drop(columns=["_highlight"])
                    .style
                    .apply(_highlight_selected, axis=1)
                    .format({
                        "Score":           lambda x: f"{x:.1f}" if pd.notna(x) else "—",
                        "P/B":             lambda x: f"{x:.2f}" if pd.notna(x) else "—",
                        "EV/EBITDA":       lambda x: f"{x:.1f}x" if pd.notna(x) else "—",
                        "P/NAV":           lambda x: f"{x:.2f}x" if pd.notna(x) else "—",
                        "AISC Margin%":    lambda x: f"{x:.1f}%" if pd.notna(x) else "—",
                        "Cash Cost($/oz)": lambda x: f"${x:,.0f}" if pd.notna(x) else "—",
                        "EV/oz Prod":      lambda x: f"${x:,.0f}" if pd.notna(x) else "—",
                        "EV/oz Rsv":       lambda x: f"${x:,.0f}" if pd.notna(x) else "—",
                        "Prod(koz/yr)":    lambda x: f"{x:,.1f}" if pd.notna(x) else "—",
                        "Rsv Life(yr)":    lambda x: f"{x:.1f} yr" if pd.notna(x) else "—",
                        "Reserves($M)":    lambda x: f"${x:,.0f}M" if pd.notna(x) else "—",
                        "↑ NAV Upside%":   lambda x: f"{x:+.0f}%" if pd.notna(x) else "—",
                        "RSI":             lambda x: f"{x:.0f}" if pd.notna(x) else "—",
                    }, na_rep="—")
                )
                st.dataframe(_combined_styled, width="stretch", height=400)

                # ── Visual: selected vs peer median bar chart ─────────────────
                _bar_metrics = [
                    ("score_composite",   "Score",          1.0,  "Score"),
                    ("price_to_book",     "P/B",            -1.0, "Ratio"),   # lower = better
                    ("ev_ebitda",         "EV/EBITDA",      -1.0, "Ratio"),   # lower = better
                    ("spg_p_nav",         "P/NAV",          -1.0, "Ratio"),   # lower = better
                    ("spg_aisc_margin",   "AISC Margin%",   1.0,  "%"),       # higher = better
                    ("ev_per_oz_prod",    "EV/oz Prod",     -1.0, "$"),       # lower = better
                    ("ev_per_oz_reserve", "EV/oz Rsv",      -1.0, "$"),       # lower = better
                    ("spg_reserve_life",  "Reserve Life",   1.0,  "yr"),      # higher = better
                    ("upside_to_nav",     "↑NAV Upside%",   1.0,  "%"),       # higher = better
                    ("rsi",               "RSI",            0.0,  ""),        # neutral
                ]
                _bar_data = []
                for _col, _label, _direction, _unit in _bar_metrics:
                    if _col not in df.columns:
                        continue
                    _sel_val  = row.get(_col)
                    _peer_med = _peers[_col].median()
                    if pd.isna(_sel_val) or pd.isna(_peer_med) or _peer_med == 0:
                        continue
                    _pct_vs_peer = (_sel_val - _peer_med) / abs(_peer_med) * 100
                    # For "lower is better" metrics, flip the sign for display
                    _display_pct = _pct_vs_peer * (_direction if _direction != 0 else 1.0)
                    _bar_data.append({
                        "Metric":        _label,
                        "Selected":      round(_sel_val, 2),
                        "Peer Median":   round(_peer_med, 2),
                        "vs Peer (%)":   round(_display_pct, 1),
                        "color":         "#22c55e" if _display_pct >= 0 else "#ef4444",
                    })

                if _bar_data:
                    _bar_df = pd.DataFrame(_bar_data)
                    _fig_peer = go.Figure()
                    _fig_peer.add_bar(
                        x=_bar_df["Metric"],
                        y=_bar_df["vs Peer (%)"],
                        marker_color=_bar_df["color"].tolist(),
                        text=[f"{v:+.1f}%" for v in _bar_df["vs Peer (%)"]],
                        textposition="outside",
                        customdata=_bar_df[["Selected", "Peer Median"]].values,
                        hovertemplate=(
                            "<b>%{x}</b><br>"
                            "Selected: %{customdata[0]}<br>"
                            "Peer Median: %{customdata[1]}<br>"
                            "vs Peer: %{y:+.1f}%<extra></extra>"
                        ),
                    )
                    _fig_peer.add_hline(y=0, line_dash="dot", line_color="gray")
                    _fig_peer.update_layout(
                        title=f"{row.get('name', sel_ticker)} vs Peer Median (green = better than peers)",
                        yaxis_title="% vs Peer Median (sign-adjusted: up = better)",
                        height=320,
                        margin=dict(t=40, b=40),
                        showlegend=False,
                    )
                    st.plotly_chart(_fig_peer, width="stretch")

            # ── Score & Price History ─────────────────────────────────────────
            _hist_df = load_history(sel_ticker, days=90)
            if not _hist_df.empty and len(_hist_df) > 1:
                with st.expander("📈 Score & Price History (snapshots)", expanded=False):
                    _hist_df["snap_date"] = pd.to_datetime(_hist_df["snap_date"])
                    _hist_df = _hist_df.sort_values("snap_date")

                    _hcol1, _hcol2 = st.columns(2)

                    with _hcol1:
                        _fig_sh = go.Figure()
                        _fig_sh.add_trace(go.Scatter(
                            x=_hist_df["snap_date"], y=_hist_df["score_composite"],
                            name="Composite",
                            line=dict(color="#3b82f6", width=2.5),
                            mode="lines+markers", marker=dict(size=5),
                        ))
                        if _hist_df["score_valuation"].notna().any():
                            _fig_sh.add_trace(go.Scatter(
                                x=_hist_df["snap_date"], y=_hist_df["score_valuation"],
                                name="Valuation",
                                line=dict(color="#22c55e", width=1.5, dash="dot"),
                                mode="lines",
                            ))
                        if _hist_df["score_momentum"].notna().any():
                            _fig_sh.add_trace(go.Scatter(
                                x=_hist_df["snap_date"], y=_hist_df["score_momentum"],
                                name="Momentum",
                                line=dict(color="#eab308", width=1.5, dash="dot"),
                                mode="lines",
                            ))
                        _fig_sh.add_hline(y=55, line_dash="dash", line_color="gray",
                                          annotation_text="Buy threshold",
                                          annotation_position="bottom right")
                        _fig_sh.update_layout(
                            title="Score History",
                            yaxis=dict(range=[0, 100], title="Score"),
                            height=280,
                            margin=dict(l=0, r=0, t=40, b=0),
                            legend=dict(orientation="h", y=-0.25),
                        )
                        st.plotly_chart(_fig_sh, width="stretch")

                    with _hcol2:
                        if _hist_df["price"].notna().any():
                            _fig_ph = go.Figure()
                            _fig_ph.add_trace(go.Scatter(
                                x=_hist_df["snap_date"], y=_hist_df["price"],
                                name="Price",
                                line=dict(color="#f59e0b", width=2),
                                mode="lines+markers", marker=dict(size=5),
                                fill="tozeroy",
                                fillcolor="rgba(245,158,11,0.08)",
                            ))
                            # Price target reference line if set
                            _hist_pt = get_price_target(sel_ticker)
                            if _hist_pt:
                                _fig_ph.add_hline(
                                    y=_hist_pt, line_dash="dash", line_color="#22c55e",
                                    annotation_text=f"Target ${_hist_pt:.3f}",
                                    annotation_position="top right",
                                )
                            _fig_ph.update_layout(
                                title="Price History (from snapshots)",
                                yaxis=dict(title=f"Price ({config.CURRENCY})"),
                                height=280,
                                margin=dict(l=0, r=0, t=40, b=0),
                                showlegend=False,
                            )
                            st.plotly_chart(_fig_ph, width="stretch")
                        else:
                            st.info("No price data in snapshots yet.")
            elif len(_hist_df) <= 1:
                st.caption("💡 Score history will appear once data has been refreshed on multiple days.")

            # ── Score Change Waterfall ────────────────────────────────────────
            _wf_hist = load_history(sel_ticker, days=15)
            if not _wf_hist.empty:
                _wf_hist["snap_date"] = pd.to_datetime(_wf_hist["snap_date"])
                _wf_hist = _wf_hist.sort_values("snap_date")

            _wf_sub_cols = [
                ("score_valuation",  "Valuation"),
                ("score_health",     "Health"),
                ("score_momentum",   "Momentum"),
                ("score_mining",     "Mining"),
                ("score_commodity",  "Commodity"),
                ("score_stage",      "Stage"),
                ("score_composite",  "Composite"),
            ]

            if not _wf_hist.empty and len(_wf_hist) >= 2:
                with st.expander("📊 Score Change vs Last Snapshot", expanded=False):
                    _wf_curr = _wf_hist.iloc[-1]
                    _wf_prev = _wf_hist.iloc[-2]
                    _wf_prev_date = _wf_prev["snap_date"].strftime("%Y-%m-%d")
                    _wf_curr_date = _wf_curr["snap_date"].strftime("%Y-%m-%d")

                    _wf_labels, _wf_deltas, _wf_colors = [], [], []
                    for _col, _label in _wf_sub_cols:
                        _curr_val = _wf_curr.get(_col)
                        _prev_val = _wf_prev.get(_col)
                        if _curr_val is not None and _prev_val is not None:
                            try:
                                _delta = float(_curr_val) - float(_prev_val)
                            except (TypeError, ValueError):
                                _delta = 0.0
                            _wf_labels.append(_label)
                            _wf_deltas.append(round(_delta, 1))
                            if _col == "score_composite":
                                _wf_colors.append("#3b82f6" if _delta >= 0 else "#ef4444")
                            else:
                                _wf_colors.append("#22c55e" if _delta >= 0 else "#f87171")

                    if _wf_labels:
                        # Build waterfall: sub-scores as absolute bars, composite separately
                        _sub_labels = _wf_labels[:-1]
                        _sub_deltas = _wf_deltas[:-1]
                        _sub_colors = _wf_colors[:-1]
                        _comp_delta = _wf_deltas[-1] if _wf_deltas else 0.0
                        _comp_color = _wf_colors[-1] if _wf_colors else "#3b82f6"

                        _fig_wf = go.Figure()

                        # Sub-score bars
                        _fig_wf.add_trace(go.Bar(
                            x=_sub_labels,
                            y=_sub_deltas,
                            marker_color=_sub_colors,
                            text=[f"{d:+.1f}" for d in _sub_deltas],
                            textposition="outside",
                            name="Sub-score Δ",
                            width=0.55,
                        ))

                        _fig_wf.update_layout(
                            title=dict(
                                text=f"Score Δ: {_wf_prev_date} → {_wf_curr_date}",
                                font=dict(size=14),
                            ),
                            yaxis=dict(title="Δ Score (pts)", zeroline=True,
                                       zerolinecolor="#6b7280", zerolinewidth=1.5),
                            xaxis=dict(title="Sub-score"),
                            height=320,
                            margin=dict(l=0, r=0, t=50, b=0),
                            showlegend=False,
                            plot_bgcolor="rgba(0,0,0,0)",
                        )
                        _fig_wf.add_hline(y=0, line_width=1, line_color="#6b7280")

                        _wf_c1, _wf_c2 = st.columns([3, 1])
                        with _wf_c1:
                            st.plotly_chart(_fig_wf, width="stretch")
                        with _wf_c2:
                            st.markdown("**Composite Score**")
                            _comp_arrow = "▲" if _comp_delta >= 0 else "▼"
                            _comp_css   = "color:#22c55e" if _comp_delta >= 0 else "color:#ef4444"
                            st.markdown(
                                f"<span style='font-size:2rem;{_comp_css}'>"
                                f"{_comp_arrow} {_comp_delta:+.1f}</span>",
                                unsafe_allow_html=True,
                            )
                            _curr_comp = _wf_curr.get("score_composite")
                            _prev_comp = _wf_prev.get("score_composite")
                            if _curr_comp is not None and _prev_comp is not None:
                                st.caption(f"{float(_prev_comp):.1f} → {float(_curr_comp):.1f}")
                            st.markdown(f"*vs {_wf_prev_date}*")

                        # Summary table
                        _wf_tbl = []
                        for _label, _delta in zip(_wf_labels, _wf_deltas):
                            _icon = "🟢" if _delta > 0 else ("🔴" if _delta < 0 else "⚪")
                            _wf_tbl.append({"Metric": _label, "Δ": f"{_delta:+.1f}", "": _icon})
                        st.dataframe(
                            pd.DataFrame(_wf_tbl),
                            hide_index=True,
                            column_config={
                                "Metric": st.column_config.TextColumn("Metric", width=120),
                                "Δ":      st.column_config.TextColumn("Change (pts)", width=110),
                                "":       st.column_config.TextColumn("", width=40),
                            },
                        )
                    else:
                        st.info("Sub-score data not available for waterfall comparison.")
            elif not _wf_hist.empty and len(_wf_hist) < 2:
                st.caption("💡 Score change waterfall will appear once data has been refreshed on at least 2 days.")

            # ── Research Links (ASX Announcements+, IR, etc.) ────────────────────────────
            with st.expander("🔗 Research & Filing Links", expanded=False):
                _company_name  = row.get("name", sel_ticker)
                _ticker_clean  = sel_ticker.replace(".TO", "").replace(".V", "").replace(".CN", "")

                # ASX Announcements search URL
                import urllib.parse as _urlparse
                _sedar_q   = _urlparse.quote(_company_name)
                _sedar_url = f"https://www.asx.com.au/asx/research/companiesInfo.do?by=asxCodes&allCompanies=false&name={_sedar_q}"

                # ASX — Australian mining news aggregator
                _sw_url = f"https://www.stockwatch.com/Quote/Detail/?U={_ticker_clean}"

                # Mining.com company search
                _mining_q   = _urlparse.quote(_company_name.split(" ")[0])
                _mining_url = f"https://www.mining.com/?s={_mining_q}"

                # Junior Stock Review — ASX focused
                _jsr_url = f"https://www.juniorstockreview.com/?s={_mining_q}"

                # ASX company page
                _exchange = str(row.get("exchange", ""))
                _tsx_url = f"https://www.asx.com.au/asx/research/companiesInfo.do?by=asxCodes&allCompanies=false&name={_ticker_clean}"

                _rl_cols = st.columns(3)
                _rl_cols[0].markdown(
                    f"[📂 **ASX Announcements**]({_sedar_url})  \n"
                    f"<span style='font-size:11px;color:#64748b'>Annual Reports · Quarterly Activities Reports · Press releases</span>",
                    unsafe_allow_html=True,
                )
                _rl_cols[1].markdown(
                    f"[📈 **Stockwatch**]({_sw_url})  \n"
                    f"<span style='font-size:11px;color:#64748b'>ASX news & quotes</span>",
                    unsafe_allow_html=True,
                )
                _rl_cols[2].markdown(
                    f"[🏛️ **ASX Listing**]({_tsx_url})  \n"
                    f"<span style='font-size:11px;color:#64748b'>Exchange profile</span>",
                    unsafe_allow_html=True,
                )
                _rl_cols2 = st.columns(3)
                _rl_cols2[0].markdown(
                    f"[⛏️ **Mining.com**]({_mining_url})  \n"
                    f"<span style='font-size:11px;color:#64748b'>Industry news</span>",
                    unsafe_allow_html=True,
                )
                _rl_cols2[1].markdown(
                    f"[🔬 **Junior Stock Review**]({_jsr_url})  \n"
                    f"<span style='font-size:11px;color:#64748b'>ASX focused analysis</span>",
                    unsafe_allow_html=True,
                )
                # Yahoo Finance link (always available)
                _yf_url = f"https://finance.yahoo.com/quote/{sel_ticker}"
                _rl_cols2[2].markdown(
                    f"[📊 **Yahoo Finance**]({_yf_url})  \n"
                    f"<span style='font-size:11px;color:#64748b'>Financials · Options · Holders</span>",
                    unsafe_allow_html=True,
                )

                st.caption(
                    "Links open in browser. ASX Announcements is the official Australian securities filing database "
                    "(replaced legacy ASX Announcements in 2023). JORC Code technical reports are filed there."
                )

            # ── News Feed ─────────────────────────────────────────────────────
            with st.expander("📰 Latest News", expanded=False):
                _news_items = _fetch_news(sel_ticker)
                if _news_items:
                    for _ni in _news_items:
                        _title     = _ni.get("title", "")
                        _link      = _ni.get("link", "") or _ni.get("url", "")
                        _publisher = _ni.get("publisher", "")
                        _pub_time  = _ni.get("providerPublishTime")
                        _time_str  = ""
                        if _pub_time:
                            try:
                                from datetime import datetime
                                _dt = datetime.fromtimestamp(int(_pub_time))
                                _time_str = _dt.strftime("%b %d, %Y")
                            except Exception:
                                pass
                        if _title and _link:
                            st.markdown(
                                f"**[{_title}]({_link})**  \n"
                                f"<span style='font-size:11px;color:#64748b'>"
                                f"{_publisher}" + (f" · {_time_str}" if _time_str else "") +
                                f"</span>",
                                unsafe_allow_html=True,
                            )
                            st.markdown("---")
                else:
                    st.info("No recent news found for this ticker.")

            # ── Side-by-side Comparison ───────────────────────────────────────
            with st.expander("⚖️ Compare with other companies", expanded=False):
                st.caption("Select up to 4 companies to compare side-by-side.")
                _cmp_all_opts = (
                    df.sort_values("score_composite", ascending=False)
                    .apply(lambda r: f"{r['name']} ({r['ticker']})", axis=1)
                    .tolist()
                )
                _cmp_default = [f"{row.get('name', sel_ticker)} ({sel_ticker})"]
                _cmp_sel = st.multiselect(
                    "Companies to compare",
                    options=_cmp_all_opts,
                    default=_cmp_default,
                    max_selections=4,
                    key="cmp_multisel",
                )
                if len(_cmp_sel) >= 2:
                    _cmp_tickers = [s.split("(")[-1].rstrip(")") for s in _cmp_sel]
                    _cmp_df = df[df["ticker"].isin(_cmp_tickers)].set_index("ticker")

                    # Metric rows to compare
                    _cmp_metrics = [
                        ("name",             "Company",        lambda v: str(v)),
                        ("grade",            "Grade",          lambda v: str(v)),
                        ("score_composite",  "Score /100",     lambda v: f"{v:.1f}" if pd.notna(v) else "—"),
                        ("score_valuation",  "Valuation",      lambda v: f"{v:.1f}" if pd.notna(v) else "—"),
                        ("score_health",     "Health",         lambda v: f"{v:.1f}" if pd.notna(v) else "—"),
                        ("score_momentum",   "Momentum",       lambda v: f"{v:.1f}" if pd.notna(v) else "—"),
                        ("score_mining",     "⛏️ Mining",      lambda v: f"{v:.1f}" if pd.notna(v) else "—"),
                        ("price",            "Price",          lambda v: f"${v:.3f}" if pd.notna(v) else "—"),
                        ("market_cap",       "Mkt Cap",        lambda v: (f"${v/1e9:.2f}B" if v >= 1e9 else f"${v/1e6:.0f}M") if pd.notna(v) else "—"),
                        ("price_to_book",    "P/B",            lambda v: f"{v:.2f}x" if pd.notna(v) else "—"),
                        ("ev_ebitda",        "EV/EBITDA",      lambda v: f"{v:.1f}x" if pd.notna(v) else "—"),
                        ("spg_p_nav",        "P/NAV (S&P)",    lambda v: f"{v:.2f}x" if pd.notna(v) else "—"),
                        ("upside_to_nav",    "↑ Upside %",     lambda v: f"{v:+.0f}%" if pd.notna(v) else "—"),
                        ("spg_aisc_margin",  "AISC Margin%",   lambda v: f"{v:.1f}%" if pd.notna(v) else "—"),
                        ("debt_to_equity",   "D/E %",          lambda v: f"{v:.0f}%" if pd.notna(v) else "—"),
                        ("current_ratio",    "Current Ratio",  lambda v: f"{v:.2f}" if pd.notna(v) else "—"),
                        ("rsi",              "RSI",            lambda v: f"{v:.0f}" if pd.notna(v) else "—"),
                        ("return_1m",        "1M Return",      lambda v: f"{v:+.1f}%" if pd.notna(v) else "—"),
                        ("return_3m",        "3M Return",      lambda v: f"{v:+.1f}%" if pd.notna(v) else "—"),
                        ("wk52_position",    "52wk Pos%",      lambda v: f"{v:.0f}%" if pd.notna(v) else "—"),
                        ("spg_reserves_m",   "Reserves $M",    lambda v: f"${v:,.0f}M" if pd.notna(v) else "—"),
                        ("dividend_yield",   "Div Yield",      lambda v: f"{v:.2f}%" if pd.notna(v) else "—"),
                    ]

                    # Build comparison table
                    _cmp_rows = {}
                    for _col, _label, _fmt in _cmp_metrics:
                        _cmp_rows[_label] = {
                            tk: _fmt(_cmp_df.loc[tk, _col])
                            if tk in _cmp_df.index and _col in _cmp_df.columns
                            else "—"
                            for tk in _cmp_tickers
                        }
                    _cmp_tbl = pd.DataFrame(_cmp_rows).T
                    _cmp_tbl.columns = [
                        f"{_cmp_df.loc[tk, 'name']} ({tk})" if tk in _cmp_df.index else tk
                        for tk in _cmp_tickers
                    ]
                    st.dataframe(_cmp_tbl, width="stretch")

                    # Radar overlay chart
                    _radar_dims = ["score_valuation", "score_health", "score_momentum",
                                   "score_mining", "score_commodity", "score_stage"]
                    _radar_labels = ["Valuation", "Health", "Momentum",
                                     "⛏️ Mining", "Commodity", "Stage"]
                    _radar_colors = ["#3b82f6", "#22c55e", "#f59e0b", "#a855f7"]
                    fig_cmp_radar = go.Figure()
                    for _ri, _tk in enumerate(_cmp_tickers):
                        if _tk not in _cmp_df.index:
                            continue
                        _r_vals = [
                            float(_cmp_df.loc[_tk, d]) if d in _cmp_df.columns
                            and pd.notna(_cmp_df.loc[_tk, d]) else 0
                            for d in _radar_dims
                        ]
                        _tk_name = _cmp_df.loc[_tk, "name"] if "name" in _cmp_df.columns else _tk
                        fig_cmp_radar.add_trace(go.Scatterpolar(
                            r=_r_vals + [_r_vals[0]],
                            theta=_radar_labels + [_radar_labels[0]],
                            fill="toself",
                            fillcolor=f"rgba{tuple(int(_radar_colors[_ri % 4].lstrip('#')[i:i+2], 16) for i in (0, 2, 4)) + (0.12,)}",
                            line=dict(color=_radar_colors[_ri % 4], width=2),
                            name=f"{_tk_name} ({_tk})",
                        ))
                    fig_cmp_radar.update_layout(
                        polar=dict(radialaxis=dict(range=[0, 100])),
                        height=380,
                        margin=dict(l=40, r=40, t=20, b=20),
                        legend=dict(orientation="h", y=-0.15),
                    )
                    st.plotly_chart(fig_cmp_radar, width="stretch")
                else:
                    st.info("Select at least 2 companies to compare.")

# ── TAB 4: Report ──────────────────────────────────────────────────────────────
with tab_report:
    from datetime import date as _date

    COMMODITY_OUTLOOK_EN = {
        "Gold":       ("strongly bullish", "gold is trading near all-time highs (~$4,600/oz) driven by central bank buying, de-dollarization and macro uncertainty"),
        "Silver":     ("bullish",          "silver is being pulled higher alongside gold while industrial demand from solar and EVs builds a structural floor"),
        "Copper":     ("bullish",          "copper trades above $13,000/t on tight mine supply and strong electrification demand; near-term macro risks cap upside"),
        "Uranium":    ("strongly bullish",  "uranium demand is surging with global nuclear revival — utilities remain under-contracted, supporting a sustained price floor"),
        "Lithium":    ("cautious",         "lithium faces persistent oversupply from Chinese capacity and sluggish EV demand growth outside China"),
        "Nickel":     ("cautious",         "nickel is under significant pressure from Indonesian Class 1 supply growth and weak stainless steel demand"),
        "Zinc":       ("neutral",          "zinc fundamentals are broadly balanced; LME inventories remain manageable but demand recovery is slow"),
        "Iron Ore":   ("cautious",         "iron ore faces structural headwinds from China's property sector deleveraging and decarbonization of steel"),
        "Diversified":("neutral",          "diversified miners offer commodity-mix resilience and may benefit from higher gold/copper prices partially offset by base metal softness"),
        "PGM":        ("cautious",         "PGM prices remain depressed by EV transition uncertainty and abundant Russian-origin supply"),
        "Potash":     ("neutral",          "potash supply/demand is in gradual rebalancing with Brazilian imports supporting floor prices"),
        "Coal":       ("cautious",         "thermal coal faces long-term structural decline; near-term cashflows remain strong for low-cost producers"),
    }
    COMMODITY_OUTLOOK_ZH = {
        "Gold":       ("强烈看涨", "黄金价格接近历史高位（约4600美元/盎司），受央行购金、去美元化与宏观不确定性三重驱动"),
        "Silver":     ("看涨",    "白银随金价上行，同时受到太阳能与新能源汽车工业需求的结构性支撑"),
        "Copper":     ("看涨",    "铜价突破13000美元/吨，矿山供应偏紧叠加电气化需求旺盛；短期宏观风险限制上行空间"),
        "Uranium":    ("强烈看涨", "全球核电复兴推动铀需求大幅攀升，公用事业采购仍严重不足，价格底部坚实"),
        "Lithium":    ("谨慎",    "中国产能持续扩张和电动车需求增速放缓导致锂市场供过于求"),
        "Nickel":     ("谨慎",    "印尼一级镍大幅增产叠加不锈钢需求疲软，镍价承压显著"),
        "Zinc":       ("中性",    "锌基本面大体平衡，LME库存尚可，但需求复苏缓慢"),
        "Iron Ore":   ("谨慎",    "中国房地产去杠杆与钢铁绿色转型对铁矿石形成结构性压力"),
        "Diversified":("中性",    "多元化矿企提供大宗商品组合韧性，黄金铜价上涨部分对冲基本金属疲软"),
        "PGM":        ("谨慎",    "俄罗斯供给充足加之电动车转型不确定性，PGM价格持续承压"),
        "Potash":     ("中性",    "钾肥市场缓慢再平衡，巴西进口需求提供价格底部支撑"),
        "Coal":       ("谨慎",    "动力煤面临长期结构性下行，但低成本生产商短期现金流依然强劲"),
    }

    def _commodity_view(commodity: str, lang: str = "en") -> tuple[str, str]:
        table = COMMODITY_OUTLOOK_ZH if lang == "zh" else COMMODITY_OUTLOOK_EN
        for key, val in table.items():
            if key.lower() in commodity.lower():
                return val
        return ("中性", "大宗商品基本面整体平衡") if lang == "zh" else ("neutral", "commodity fundamentals are broadly balanced")

    def _bullet(text: str) -> str:
        return f"  • {text}"

    def _generate_thesis(row: pd.Series, peer_pb: float, peer_evebitda: float) -> list[str]:
        """Return a list of bullet-point strings explaining undervaluation."""
        bullets: list[str] = []

        # ── Valuation ──────────────────────────────────────────────────────────
        pb = row.get("price_to_book")
        evebitda = row.get("ev_ebitda")
        pcf = row.get("p_cf")
        sv = row.get("score_valuation", 0)

        if pd.notna(pb) and pb > 0:
            if pb < 1.0:
                bullets.append(_bullet(
                    f"Trades below book value (P/B {pb:.2f}x) — market is pricing the assets "
                    f"at a discount to their stated net worth."
                ))
            elif pb < peer_pb * 0.75:
                bullets.append(_bullet(
                    f"P/B of {pb:.2f}x is significantly below the peer median ({peer_pb:.2f}x), "
                    f"suggesting the stock is cheap relative to sector peers."
                ))
            elif pb < peer_pb:
                bullets.append(_bullet(
                    f"P/B of {pb:.2f}x is below the peer median ({peer_pb:.2f}x)."
                ))

        if pd.notna(evebitda) and evebitda > 0:
            if evebitda < 5:
                bullets.append(_bullet(
                    f"EV/EBITDA of {evebitda:.1f}x is very low — the enterprise is priced at "
                    f"less than 5 years of operating earnings."
                ))
            elif pd.notna(peer_evebitda) and evebitda < peer_evebitda * 0.75:
                bullets.append(_bullet(
                    f"EV/EBITDA of {evebitda:.1f}x is well below the peer median ({peer_evebitda:.1f}x), "
                    f"indicating relative cheapness on an earnings basis."
                ))

        if pd.notna(pcf) and 0 < pcf < 8:
            bullets.append(_bullet(
                f"P/Cash Flow of {pcf:.1f}x indicates the stock is cheap relative to cash "
                f"generation — a strong signal for value investors."
            ))

        if sv >= 70 and not bullets:
            bullets.append(_bullet(
                "Strong valuation score driven by multiple metrics trading at a "
                "discount to sector peers."
            ))

        # ── Health ─────────────────────────────────────────────────────────────
        cr = row.get("current_ratio")
        de = row.get("debt_to_equity")
        cash_pct = row.get("cash_pct_mcap")
        sh = row.get("score_health", 0)

        if pd.notna(cr) and cr >= 2.0:
            bullets.append(_bullet(
                f"Current ratio of {cr:.1f}x signals strong short-term liquidity — "
                f"the company can comfortably cover near-term obligations."
            ))
        elif pd.notna(cr) and cr >= 1.5:
            bullets.append(_bullet(f"Solid current ratio of {cr:.1f}x indicates adequate liquidity."))

        if pd.notna(de) and de < 30:
            bullets.append(_bullet(
                f"Debt/equity of {de:.0f}% is low — a clean balance sheet reduces downside risk "
                f"and preserves financial flexibility."
            ))
        elif pd.notna(de) and de < 60:
            bullets.append(_bullet(f"Manageable debt/equity of {de:.0f}%."))

        if pd.notna(cash_pct) and cash_pct >= 15:
            bullets.append(_bullet(
                f"Cash represents {cash_pct:.0f}% of market cap, providing a meaningful "
                f"floor and reducing effective downside from the current price."
            ))

        fcf_en = row.get("free_cf")
        mcap_en = row.get("market_cap")
        if pd.notna(fcf_en) and pd.notna(mcap_en) and mcap_en > 0 and fcf_en > 0:
            _fcf_yld = fcf_en / mcap_en * 100
            if _fcf_yld >= 8:
                bullets.append(_bullet(
                    f"FCF yield of {_fcf_yld:.1f}% — exceptional free cash flow generation "
                    f"relative to market cap; the company is self-funding at current prices."
                ))
            elif _fcf_yld >= 4:
                bullets.append(_bullet(
                    f"FCF yield of {_fcf_yld:.1f}% — solid cash generation supporting "
                    f"balance sheet strength and potential shareholder returns."
                ))

        div_yld_en = row.get("dividend_yield")
        if pd.notna(div_yld_en) and div_yld_en >= 2:
            bullets.append(_bullet(
                f"Dividend yield of {div_yld_en:.2f}% provides income while you wait for "
                f"capital appreciation — uncommon for a mining stock and signals confidence "
                f"from management in cash flow sustainability."
            ))

        # Profitability metrics (producers only — meaningful when data exists)
        _roe_en = row.get("return_on_equity")
        _op_en  = row.get("operating_margins")
        if pd.notna(_roe_en) and _roe_en >= 0.10:
            bullets.append(_bullet(
                f"Return on equity of {_roe_en*100:.1f}% demonstrates that management is "
                f"generating meaningful returns from shareholder capital — a hallmark of "
                f"operationally excellent producers."
            ))
        if pd.notna(_op_en) and _op_en >= 0.20:
            bullets.append(_bullet(
                f"Operating margin of {_op_en*100:.1f}% indicates the company is earning "
                f"significant cash from operations, providing cushion against commodity price "
                f"volatility and funding growth internally."
            ))

        # ── Momentum / Contrarian ──────────────────────────────────────────────
        rsi = row.get("rsi")
        wk52 = row.get("wk52_position")
        vs_hi = row.get("pct_from_52hi")
        vs_ma200 = row.get("price_vs_ma200")

        if pd.notna(rsi) and rsi < 35:
            bullets.append(_bullet(
                f"RSI of {rsi:.0f} is in oversold territory — the stock has been "
                f"sold down aggressively and may be due for a mean-reversion bounce."
            ))
        elif pd.notna(rsi) and rsi < 45:
            bullets.append(_bullet(f"RSI of {rsi:.0f} leans oversold, suggesting bearish sentiment may be overdone."))

        if pd.notna(wk52) and wk52 < 25:
            bullets.append(_bullet(
                f"Trading near its 52-week low (position: {wk52:.0f}% of annual range) — "
                f"contrarian opportunity with significant upside to the annual high."
            ))
        elif pd.notna(vs_hi) and vs_hi < -30:
            bullets.append(_bullet(
                f"Down {abs(vs_hi):.0f}% from its 52-week high, creating a "
                f"substantial discount relative to where the market recently valued it."
            ))

        if pd.notna(vs_ma200) and vs_ma200 < -15:
            bullets.append(_bullet(
                f"Trading {abs(vs_ma200):.0f}% below its 200-day moving average — "
                f"deeply oversold on a long-term trend basis."
            ))

        # ── S&P/SNL Mining Data ────────────────────────────────────────────────
        p_nav   = row.get("spg_p_nav")
        aisc    = row.get("spg_aisc_per_oz")
        aisc_m  = row.get("spg_aisc_margin")
        resv_m  = row.get("spg_reserves_m")
        mcap    = row.get("market_cap")

        if pd.notna(p_nav) and p_nav > 0:
            if p_nav < 0.75:
                bullets.append(_bullet(
                    f"Trading at a deep P/NAV discount of {p_nav:.2f}x — the stock is priced at "
                    f"{(1-p_nav)*100:.0f}% below its S&P-estimated Net Asset Value, "
                    f"implying significant upside if assets are fairly valued."
                ))
            elif p_nav < 1.0:
                bullets.append(_bullet(
                    f"P/NAV of {p_nav:.2f}x — trading below net asset value per S&P Capital IQ "
                    f"estimates, a classic undervaluation signal for mining stocks."
                ))
            elif p_nav < 1.3:
                bullets.append(_bullet(
                    f"P/NAV of {p_nav:.2f}x — trading close to fair value by NAV; "
                    f"any gold price appreciation would expand the implied discount."
                ))

        aisc_t_en  = row.get("spg_aisc_per_t")
        aisc_lb_en = row.get("spg_aisc_per_lb")
        if pd.notna(aisc) and pd.notna(aisc_m):
            _comm_en = str(row.get("commodity", "Gold")).split("/")[0]
            _spot_en = config.COMMODITY_SPOT.get(_comm_en, config.COMMODITY_SPOT.get("Gold", 4821.0))
            _unit_en = "oz"
            _aisc_val_en = aisc
            if aisc_m > 50:
                bullets.append(_bullet(
                    f"Exceptional AISC margin of {aisc_m:.0f}% (AISC ${_aisc_val_en:,.0f}/{_unit_en} vs. ~${_spot_en:,.0f} spot) — "
                    f"among the most cost-efficient producers globally, "
                    f"generating strong free cash flow at current prices."
                ))
            elif aisc_m > 35:
                bullets.append(_bullet(
                    f"Solid AISC margin of {aisc_m:.0f}% (AISC ${_aisc_val_en:,.0f}/{_unit_en} vs. ~${_spot_en:,.0f} spot) — "
                    f"healthy operating leverage to any price upside."
                ))
        elif pd.notna(aisc_t_en) and pd.notna(aisc_m):
            _comm_en = str(row.get("commodity", "Copper")).split("/")[0]
            _spot_en = config.COMMODITY_SPOT.get(_comm_en, 9200.0)
            if aisc_m > 40:
                bullets.append(_bullet(
                    f"Exceptional AISC margin of {aisc_m:.0f}% (AISC ${aisc_t_en:,.0f}/t vs. ~${_spot_en:,.0f}/t spot) — "
                    f"highly cost-competitive in the {_comm_en} sector."
                ))
            elif aisc_m > 20:
                bullets.append(_bullet(
                    f"Solid AISC margin of {aisc_m:.0f}% (AISC ${aisc_t_en:,.0f}/t vs. ~${_spot_en:,.0f}/t spot)."
                ))
        elif pd.notna(aisc_lb_en) and pd.notna(aisc_m):
            _spot_u_en = config.COMMODITY_SPOT.get("Uranium", 86.65)
            if aisc_m > 40:
                bullets.append(_bullet(
                    f"Exceptional AISC margin of {aisc_m:.0f}% (AISC ${aisc_lb_en:.2f}/lb vs. ~${_spot_u_en:.2f}/lb uranium spot) — "
                    f"highly cost-competitive uranium producer."
                ))
            elif aisc_m > 20:
                bullets.append(_bullet(
                    f"Solid AISC margin of {aisc_m:.0f}% (AISC ${aisc_lb_en:.2f}/lb vs. ~${_spot_u_en:.2f}/lb uranium spot)."
                ))

        if pd.notna(resv_m) and pd.notna(mcap) and mcap > 0:
            backing = resv_m / (mcap / 1e6)
            if backing > 3:
                bullets.append(_bullet(
                    f"Reserves in-situ value of ${resv_m:,.0f}M is {backing:.1f}x market cap — "
                    f"the ground value of its resource base substantially exceeds its equity price."
                ))
            elif backing > 1.5:
                bullets.append(_bullet(
                    f"Reserves in-situ value of ${resv_m:,.0f}M represents {backing:.1f}x market cap, "
                    f"indicating strong asset backing relative to current valuation."
                ))

        # ── Cash cost vs AISC (operating leverage floor) ─────────────────────
        _cc_oz_en = row.get("spg_cash_cost_oz")
        _cc_t_en  = row.get("spg_cash_cost_t")
        _cc_lb_en = row.get("spg_cash_cost_lb")
        _aisc_en  = row.get("spg_aisc_per_oz") or row.get("spg_aisc_per_t") or row.get("spg_aisc_per_lb")
        if pd.notna(_cc_oz_en) and pd.notna(_aisc_en) and _aisc_en > 0:
            _sust_en = _aisc_en - _cc_oz_en
            _sust_pct = _sust_en / _aisc_en * 100
            if _cc_oz_en < 800:
                bullets.append(_bullet(
                    f"Cash cost of ${_cc_oz_en:,.0f}/oz is notably low — even in a severe "
                    f"downturn the operation remains profitable on a cash basis. "
                    f"Sustaining capex adds ${_sust_en:,.0f}/oz ({_sust_pct:.0f}% of AISC)."
                ))
            elif _cc_oz_en < 1200:
                bullets.append(_bullet(
                    f"Cash cost of ${_cc_oz_en:,.0f}/oz indicates a resilient cost floor "
                    f"with ${_sust_en:,.0f}/oz in sustaining capex to reach full AISC."
                ))
        elif pd.notna(_cc_lb_en) and pd.notna(_aisc_en):
            _sust_lb_en = max(_aisc_en - _cc_lb_en, 0)
            bullets.append(_bullet(
                f"Cash cost of ${_cc_lb_en:.3f}/lb provides cost visibility; "
                f"${_sust_lb_en:.3f}/lb sustaining capex brings total AISC to ${_aisc_en:.3f}/lb."
            ))

        # ── EV / production multiples (commodity-aware) ───────────────────────
        _ev_comm_en = str(row.get("commodity", "")).split("/")[0].strip()
        _ev_oz_p_en  = row.get("ev_per_oz_prod")
        _ev_lb_p_en  = row.get("ev_per_lb_prod")
        _ev_t_p_en   = row.get("ev_per_t_prod")
        _ev_oz_r_en  = row.get("ev_per_oz_reserve")
        _ev_lb_r_en  = row.get("ev_per_lb_reserve")

        if _ev_comm_en in ("Gold", "Silver") and pd.notna(_ev_oz_p_en):
            # Gold/Silver: benchmarks ~$1,500–$3,000/oz for seniors; <$1,500 cheap
            if _ev_oz_p_en < 1500:
                bullets.append(_bullet(
                    f"EV/oz produced of ${_ev_oz_p_en:,.0f} is deep-value territory for a "
                    f"gold/silver producer — well below the $1,500–$3,000 range typical of "
                    f"mid-tier peers."
                ))
            elif _ev_oz_p_en < 3000:
                bullets.append(_bullet(
                    f"EV/oz produced of ${_ev_oz_p_en:,.0f} is in line with mid-tier gold/silver "
                    f"trading ranges, suggesting the market is fairly pricing production capacity."
                ))
            if pd.notna(_ev_oz_r_en) and _ev_oz_r_en < 200:
                bullets.append(_bullet(
                    f"EV/oz reserve of ${_ev_oz_r_en:,.0f} implies the market is attributing "
                    f"very low in-ground value — well below typical M&A acquisition premiums."
                ))
        elif _ev_comm_en in ("Copper",) and pd.notna(_ev_lb_p_en):
            # Copper: EV/lb produced; typical range $0.05–$0.20/lb for mid-tier
            if _ev_lb_p_en < 0.05:
                bullets.append(_bullet(
                    f"EV/lb produced of ${_ev_lb_p_en:.4f}/lb is exceptionally cheap for a "
                    f"copper producer — typical mid-tier range is $0.05–$0.20/lb."
                ))
            elif _ev_lb_p_en < 0.15:
                bullets.append(_bullet(
                    f"EV/lb produced of ${_ev_lb_p_en:.4f}/lb is in the lower half of the "
                    f"copper producer trading range ($0.05–$0.20/lb), suggesting modest valuation."
                ))
            if pd.notna(_ev_lb_r_en) and _ev_lb_r_en < 0.02:
                bullets.append(_bullet(
                    f"EV/lb reserve of ${_ev_lb_r_en:.4f}/lb implies the market is attributing "
                    f"minimal value to copper in-ground — well below typical acquisition premiums."
                ))
        elif _ev_comm_en in ("Uranium",) and pd.notna(_ev_lb_p_en):
            # Uranium: EV/lb produced; typical range $5–$25/lb for mid-tier
            if _ev_lb_p_en < 5:
                bullets.append(_bullet(
                    f"EV/lb produced of ${_ev_lb_p_en:.2f}/lb is very low for a uranium producer "
                    f"— typical mid-tier range is $5–$25/lb, implying meaningful undervaluation."
                ))
            elif _ev_lb_p_en < 15:
                bullets.append(_bullet(
                    f"EV/lb produced of ${_ev_lb_p_en:.2f}/lb sits in the lower tier of uranium "
                    f"producer multiples ($5–$25/lb), suggesting reasonable value."
                ))
        elif pd.notna(_ev_t_p_en):
            # Iron ore / base metals in $/t
            if _ev_t_p_en < 50:
                bullets.append(_bullet(
                    f"EV/t produced of ${_ev_t_p_en:,.1f}/t is low relative to typical bulk-commodity "
                    f"producer ranges, suggesting the production base is underpriced."
                ))

        # ── Reserve life ─────────────────────────────────────────────────────
        _rli_en = row.get("spg_reserve_life")
        if pd.notna(_rli_en) and _rli_en >= 15:
            bullets.append(_bullet(
                f"Reserve life of {_rli_en:.1f} years provides long-term production "
                f"visibility and reduces development/replacement capital risk."
            ))
        elif pd.notna(_rli_en) and _rli_en >= 10:
            bullets.append(_bullet(
                f"Reserve life of {_rli_en:.1f} years is healthy, exceeding the "
                f"10-year threshold considered sustainable for a mid-tier producer."
            ))
        elif pd.notna(_rli_en) and 0 < _rli_en < 5:
            bullets.append(_bullet(
                f"⚠️ Reserve life of only {_rli_en:.1f} years signals near-term "
                f"depletion risk — new discovery or resource conversion will be critical."
            ))

        # ── Analyst Consensus ─────────────────────────────────────────────────
        _an_target = row.get("analyst_target_mean")
        _an_count  = row.get("analyst_count")
        _an_rec    = row.get("analyst_rec_key")
        _an_up     = row.get("analyst_upside")
        _price_en  = row.get("price")
        if pd.notna(_an_target) and pd.notna(_an_count) and _an_count >= 1:
            _n_int = int(_an_count)
            _an_hi = row.get("analyst_target_high")
            _an_lo = row.get("analyst_target_low")
            _range_str = ""
            if pd.notna(_an_hi) and pd.notna(_an_lo):
                _range_str = f" (range: ${_an_lo:.2f}–${_an_hi:.2f})"
            _rec_str = ""
            if pd.notna(_an_rec):
                _rec_str = f", consensus recommendation: **{_an_rec}**"
            if pd.notna(_an_up) and _an_up >= 50:
                bullets.append(_bullet(
                    f"Analyst consensus target of ${_an_target:.2f}{_range_str} "
                    f"implies {_an_up:.0f}% upside from current price "
                    f"({_n_int} analyst{'s' if _n_int != 1 else ''}){_rec_str}. "
                    f"Strong institutional conviction aligns with the quantitative case."
                ))
            elif pd.notna(_an_up) and _an_up >= 20:
                bullets.append(_bullet(
                    f"Analyst consensus target of ${_an_target:.2f}{_range_str} "
                    f"implies {_an_up:.0f}% upside "
                    f"({_n_int} analyst{'s' if _n_int != 1 else ''}){_rec_str}."
                ))
            elif pd.notna(_an_up) and _an_up > 0:
                bullets.append(_bullet(
                    f"Analyst consensus target of ${_an_target:.2f} implies {_an_up:.0f}% upside "
                    f"({_n_int} analyst{'s' if _n_int != 1 else ''}){_rec_str}."
                ))

        # ── Commodity ─────────────────────────────────────────────────────────
        commodity = str(row.get("commodity", ""))
        outlook_word, outlook_reason = _commodity_view(commodity)
        bullets.append(_bullet(
            f"Commodity outlook is {outlook_word}: {outlook_reason}."
        ))

        # ── Stage ─────────────────────────────────────────────────────────────
        stage = str(row.get("stage", ""))
        if stage == "Explorer":
            bullets.append(_bullet(
                "Early-stage explorer — highest risk/reward profile; "
                "a discovery or resource upgrade could re-rate the stock substantially."
            ))
        elif stage == "Developer":
            bullets.append(_bullet(
                "Developer stage — project construction or feasibility in progress; "
                "de-risking milestones ahead can drive significant re-rating."
            ))
        elif stage == "Royalty/Streaming":
            bullets.append(_bullet(
                "Royalty/streaming model provides commodity upside with low operating risk "
                "and no direct mine-cost exposure."
            ))

        return bullets

    def _generate_thesis_zh(row: pd.Series, peer_pb: float, peer_evebitda: float) -> list[str]:
        """中文版：返回解释低估值原因的子弹点列表。"""
        bullets: list[str] = []

        # ── 估值 ───────────────────────────────────────────────────────────────
        pb       = row.get("price_to_book")
        evebitda = row.get("ev_ebitda")
        pcf      = row.get("p_cf")
        sv       = row.get("score_valuation", 0)

        if pd.notna(pb) and pb > 0:
            if pb < 1.0:
                bullets.append(_bullet(
                    f"市净率（P/B）为 {pb:.2f}x，低于账面价值——市场对其资产的定价低于账面净值，存在明显折价。"
                ))
            elif pb < peer_pb * 0.75:
                bullets.append(_bullet(
                    f"市净率 {pb:.2f}x 显著低于同业中位数（{peer_pb:.2f}x），"
                    f"相对行业同行估值偏低。"
                ))
            elif pb < peer_pb:
                bullets.append(_bullet(f"市净率 {pb:.2f}x 低于同业中位数（{peer_pb:.2f}x）。"))

        if pd.notna(evebitda) and evebitda > 0:
            if evebitda < 5:
                bullets.append(_bullet(
                    f"EV/EBITDA 仅 {evebitda:.1f}x——企业价值不足 5 年经营利润，估值极为低廉。"
                ))
            elif pd.notna(peer_evebitda) and evebitda < peer_evebitda * 0.75:
                bullets.append(_bullet(
                    f"EV/EBITDA {evebitda:.1f}x 大幅低于同业中位数（{peer_evebitda:.1f}x），"
                    f"以盈利能力衡量相对便宜。"
                ))

        if pd.notna(pcf) and 0 < pcf < 8:
            bullets.append(_bullet(
                f"市现率（P/CF）{pcf:.1f}x，相对现金流生成能力估值低廉，"
                f"对价值投资者具有较强吸引力。"
            ))

        if sv >= 70 and not bullets:
            bullets.append(_bullet("多项估值指标均低于同业，综合估值评分突出。"))

        # ── 财务健康 ───────────────────────────────────────────────────────────
        cr       = row.get("current_ratio")
        de       = row.get("debt_to_equity")
        cash_pct = row.get("cash_pct_mcap")

        if pd.notna(cr) and cr >= 2.0:
            bullets.append(_bullet(
                f"流动比率 {cr:.1f}x，流动性充裕，能够轻松覆盖短期债务。"
            ))
        elif pd.notna(cr) and cr >= 1.5:
            bullets.append(_bullet(f"流动比率 {cr:.1f}x，流动性状况良好。"))

        if pd.notna(de) and de < 30:
            bullets.append(_bullet(
                f"资产负债率（D/E）仅 {de:.0f}%，资产负债表干净，"
                f"财务灵活性强，下行风险有限。"
            ))
        elif pd.notna(de) and de < 60:
            bullets.append(_bullet(f"资产负债率 {de:.0f}%，负债水平可控。"))

        if pd.notna(cash_pct) and cash_pct >= 15:
            bullets.append(_bullet(
                f"账面现金占市值的 {cash_pct:.0f}%，为股价提供实质性支撑，"
                f"有效降低当前价位的实际下行风险。"
            ))

        fcf_zh = row.get("free_cf")
        mcap_zh = row.get("market_cap")
        if pd.notna(fcf_zh) and pd.notna(mcap_zh) and mcap_zh > 0 and fcf_zh > 0:
            _fcf_yld_zh = fcf_zh / mcap_zh * 100
            if _fcf_yld_zh >= 8:
                bullets.append(_bullet(
                    f"自由现金流收益率高达 {_fcf_yld_zh:.1f}%——相对市值的现金生成能力极强，"
                    f"公司在当前价格下完全自给自足。"
                ))
            elif _fcf_yld_zh >= 4:
                bullets.append(_bullet(
                    f"自由现金流收益率 {_fcf_yld_zh:.1f}%，现金生成能力稳健，"
                    f"支撑资产负债表健康并具备股东回报潜力。"
                ))

        div_yld_zh = row.get("dividend_yield")
        if pd.notna(div_yld_zh) and div_yld_zh >= 2:
            bullets.append(_bullet(
                f"股息收益率 {div_yld_zh:.2f}%，在等待资本增值的同时提供现金回报——"
                f"矿业股中较为罕见，传递出管理层对持续现金流的信心。"
            ))

        # 盈利能力指标（仅适用于有实际经营数据的生产商）
        _roe_zh = row.get("return_on_equity")
        _op_zh  = row.get("operating_margins")
        if pd.notna(_roe_zh) and _roe_zh >= 0.10:
            bullets.append(_bullet(
                f"净资产收益率（ROE）达 {_roe_zh*100:.1f}%，表明管理层高效运用股东资本创造回报——"
                f"是优质矿业生产商的核心标志之一。"
            ))
        if pd.notna(_op_zh) and _op_zh >= 0.20:
            bullets.append(_bullet(
                f"经营利润率 {_op_zh*100:.1f}%，公司从运营中获取可观利润，"
                f"有效缓冲大宗商品价格波动并支持内生性成长。"
            ))

        # ── 动能 / 逆向信号 ───────────────────────────────────────────────────
        rsi    = row.get("rsi")
        wk52   = row.get("wk52_position")
        vs_hi  = row.get("pct_from_52hi")
        vs_ma200 = row.get("price_vs_ma200")

        if pd.notna(rsi) and rsi < 35:
            bullets.append(_bullet(
                f"RSI 为 {rsi:.0f}，处于超卖区域——股价已被大幅抛售，"
                f"存在均值回归反弹机会。"
            ))
        elif pd.notna(rsi) and rsi < 45:
            bullets.append(_bullet(f"RSI 为 {rsi:.0f}，偏向超卖，市场悲观情绪可能已过度。"))

        if pd.notna(wk52) and wk52 < 25:
            bullets.append(_bullet(
                f"股价处于 52 周区间的低位（位置：{wk52:.0f}%），"
                f"逆向布局机会突出，上行空间可观。"
            ))
        elif pd.notna(vs_hi) and vs_hi < -30:
            bullets.append(_bullet(
                f"较 52 周高点回落 {abs(vs_hi):.0f}%，相对市场近期认可的估值"
                f"形成大幅折价。"
            ))

        if pd.notna(vs_ma200) and vs_ma200 < -15:
            bullets.append(_bullet(
                f"股价较 200 日均线低 {abs(vs_ma200):.0f}%，"
                f"长期趋势维度深度超卖。"
            ))

        # ── S&P/SNL 矿业数据 ──────────────────────────────────────────────────
        p_nav  = row.get("spg_p_nav")
        aisc   = row.get("spg_aisc_per_oz")
        aisc_m = row.get("spg_aisc_margin")
        resv_m = row.get("spg_reserves_m")
        mcap   = row.get("market_cap")

        if pd.notna(p_nav) and p_nav > 0:
            if p_nav < 0.75:
                bullets.append(_bullet(
                    f"P/NAV 仅 {p_nav:.2f}x——股价相对 S&P 预估净资产价值折价 "
                    f"{(1-p_nav)*100:.0f}%，资产低估信号显著。"
                ))
            elif p_nav < 1.0:
                bullets.append(_bullet(
                    f"P/NAV {p_nav:.2f}x——股价低于 S&P Capital IQ 估算的每股净资产，"
                    f"符合矿业股低估的经典特征。"
                ))
            elif p_nav < 1.3:
                bullets.append(_bullet(
                    f"P/NAV {p_nav:.2f}x——接近公允价值；金价若进一步上涨，"
                    f"隐含折价将进一步扩大。"
                ))

        aisc_t_zh  = row.get("spg_aisc_per_t")
        aisc_lb_zh = row.get("spg_aisc_per_lb")
        if pd.notna(aisc) and pd.notna(aisc_m):
            _comm_zh = str(row.get("commodity", "Gold")).split("/")[0]
            _spot_zh = config.COMMODITY_SPOT.get(_comm_zh, config.COMMODITY_SPOT.get("Gold", 4821.0))
            if aisc_m > 50:
                bullets.append(_bullet(
                    f"AISC 利润率高达 {aisc_m:.0f}%（AISC ${aisc:,.0f}/盎司，现货约 ${_spot_zh:,.0f}）——"
                    f"全球成本效率领先，当前价格下具备强劲的自由现金流生成能力。"
                ))
            elif aisc_m > 35:
                bullets.append(_bullet(
                    f"AISC 利润率 {aisc_m:.0f}%（AISC ${aisc:,.0f}/盎司，现货约 ${_spot_zh:,.0f}），"
                    f"对商品价格上涨具备健康的经营杠杆。"
                ))
        elif pd.notna(aisc_t_zh) and pd.notna(aisc_m):
            _comm_zh = str(row.get("commodity", "Copper")).split("/")[0]
            _spot_zh = config.COMMODITY_SPOT.get(_comm_zh, 9200.0)
            if aisc_m > 40:
                bullets.append(_bullet(
                    f"AISC 利润率高达 {aisc_m:.0f}%（AISC ${aisc_t_zh:,.0f}/吨，现货约 ${_spot_zh:,.0f}/吨）——"
                    f"在{_comm_zh}行业中成本竞争力突出。"
                ))
            elif aisc_m > 20:
                bullets.append(_bullet(
                    f"AISC 利润率 {aisc_m:.0f}%（AISC ${aisc_t_zh:,.0f}/吨，现货约 ${_spot_zh:,.0f}/吨）。"
                ))
        elif pd.notna(aisc_lb_zh) and pd.notna(aisc_m):
            _spot_u_zh = config.COMMODITY_SPOT.get("Uranium", 86.65)
            if aisc_m > 40:
                bullets.append(_bullet(
                    f"AISC 利润率高达 {aisc_m:.0f}%（AISC ${aisc_lb_zh:.2f}/磅，铀现货约 ${_spot_u_zh:.2f}/磅）——"
                    f"成本效率领先的铀矿生产商。"
                ))
            elif aisc_m > 20:
                bullets.append(_bullet(
                    f"AISC 利润率 {aisc_m:.0f}%（AISC ${aisc_lb_zh:.2f}/磅，铀现货约 ${_spot_u_zh:.2f}/磅）。"
                ))

        if pd.notna(resv_m) and pd.notna(mcap) and mcap > 0:
            backing = resv_m / (mcap / 1e6)
            if backing > 3:
                bullets.append(_bullet(
                    f"储量在地价值 ${resv_m:,.0f}M，相当于市值的 {backing:.1f} 倍——"
                    f"资源本身的地下价值已大幅超过当前股权市值，安全边际突出。"
                ))
            elif backing > 1.5:
                bullets.append(_bullet(
                    f"储量在地价值 ${resv_m:,.0f}M，为市值的 {backing:.1f} 倍，"
                    f"相对当前估值资产支撑充足。"
                ))

        # ── 现金成本 vs AISC ────────────────────────────────────────────────
        _cc_oz_zh = row.get("spg_cash_cost_oz")
        _cc_lb_zh = row.get("spg_cash_cost_lb")
        _aisc_zh  = row.get("spg_aisc_per_oz") or row.get("spg_aisc_per_t") or row.get("spg_aisc_per_lb")
        if pd.notna(_cc_oz_zh) and pd.notna(_aisc_zh) and _aisc_zh > 0:
            _sust_zh = _aisc_zh - _cc_oz_zh
            _sust_pct_zh = _sust_zh / _aisc_zh * 100
            if _cc_oz_zh < 800:
                bullets.append(_bullet(
                    f"现金成本仅 ${_cc_oz_zh:,.0f}/盎司，在价格大幅下行时仍可保持盈利，"
                    f"持续性资本开支为 ${_sust_zh:,.0f}/盎司（占 AISC 的 {_sust_pct_zh:.0f}%）。"
                ))
            elif _cc_oz_zh < 1200:
                bullets.append(_bullet(
                    f"现金成本 ${_cc_oz_zh:,.0f}/盎司，成本底线具备韧性；"
                    f"加入 ${_sust_zh:,.0f}/盎司持续性资本开支后 AISC 为 ${_aisc_zh:,.0f}/盎司。"
                ))
        elif pd.notna(_cc_lb_zh) and pd.notna(_aisc_zh):
            _sust_lb_zh = max(_aisc_zh - _cc_lb_zh, 0)
            bullets.append(_bullet(
                f"现金成本 ${_cc_lb_zh:.3f}/磅，成本结构透明；"
                f"含 ${_sust_lb_zh:.3f}/磅持续性资本开支，AISC 合计 ${_aisc_zh:.3f}/磅。"
            ))

        # ── EV/产量乘数（按大宗商品分类）────────────────────────────────────
        _ev_comm_zh = str(row.get("commodity", "")).split("/")[0].strip()
        _ev_oz_p_zh  = row.get("ev_per_oz_prod")
        _ev_lb_p_zh  = row.get("ev_per_lb_prod")
        _ev_t_p_zh   = row.get("ev_per_t_prod")
        _ev_oz_r_zh  = row.get("ev_per_oz_reserve")
        _ev_lb_r_zh  = row.get("ev_per_lb_reserve")

        if _ev_comm_zh in ("Gold", "Silver") and pd.notna(_ev_oz_p_zh):
            if _ev_oz_p_zh < 1500:
                bullets.append(_bullet(
                    f"EV/年产量仅 ${_ev_oz_p_zh:,.0f}/盎司，处于黄金/白银矿企深度价值区间——"
                    f"中型矿企正常区间为 1,500–3,000 美元/盎司，当前估值明显偏低。"
                ))
            elif _ev_oz_p_zh < 3000:
                bullets.append(_bullet(
                    f"EV/年产量 ${_ev_oz_p_zh:,.0f}/盎司，与中型黄金/白银矿企同行区间吻合，"
                    f"市场对产能的定价较为合理。"
                ))
            if pd.notna(_ev_oz_r_zh) and _ev_oz_r_zh < 200:
                bullets.append(_bullet(
                    f"EV/储量仅 ${_ev_oz_r_zh:,.0f}/盎司，市场对地下储量的隐含估值极低，"
                    f"远低于典型并购收购溢价，具备重大重估潜力。"
                ))
        elif _ev_comm_zh in ("Copper",) and pd.notna(_ev_lb_p_zh):
            if _ev_lb_p_zh < 0.05:
                bullets.append(_bullet(
                    f"EV/年产量仅 ${_ev_lb_p_zh:.4f}/磅，铜矿企业中属极低估值——"
                    f"中型铜矿通常区间为 0.05–0.20 美元/磅。"
                ))
            elif _ev_lb_p_zh < 0.15:
                bullets.append(_bullet(
                    f"EV/年产量 ${_ev_lb_p_zh:.4f}/磅，处于铜矿估值区间下半段（0.05–0.20 美元/磅），估值合理偏低。"
                ))
            if pd.notna(_ev_lb_r_zh) and _ev_lb_r_zh < 0.02:
                bullets.append(_bullet(
                    f"EV/铜储量仅 ${_ev_lb_r_zh:.4f}/磅，市场对地下铜储量的定价极低，远低于典型并购溢价。"
                ))
        elif _ev_comm_zh in ("Uranium",) and pd.notna(_ev_lb_p_zh):
            if _ev_lb_p_zh < 5:
                bullets.append(_bullet(
                    f"EV/年产量仅 ${_ev_lb_p_zh:.2f}/磅，铀矿企业中属低估——"
                    f"中型铀矿通常区间为 5–25 美元/磅，当前隐含明显低估。"
                ))
            elif _ev_lb_p_zh < 15:
                bullets.append(_bullet(
                    f"EV/年产量 ${_ev_lb_p_zh:.2f}/磅，处于铀矿估值区间偏低位（5–25 美元/磅）。"
                ))
        elif pd.notna(_ev_t_p_zh):
            if _ev_t_p_zh < 50:
                bullets.append(_bullet(
                    f"EV/年产量 ${_ev_t_p_zh:,.1f}/吨，低于大宗矿商通常区间，产能基础估值偏低。"
                ))

        # ── 储量年限 ─────────────────────────────────────────────────────────
        _rli_zh = row.get("spg_reserve_life")
        if pd.notna(_rli_zh) and _rli_zh >= 15:
            bullets.append(_bullet(
                f"储量年限达 {_rli_zh:.1f} 年，长期产量可见性强，"
                f"大幅降低后续勘探/资源补充的资本风险。"
            ))
        elif pd.notna(_rli_zh) and _rli_zh >= 10:
            bullets.append(_bullet(
                f"储量年限 {_rli_zh:.1f} 年，超过中等规模矿企可持续经营的 10 年基准，"
                f"资源保障充足。"
            ))
        elif pd.notna(_rli_zh) and 0 < _rli_zh < 5:
            bullets.append(_bullet(
                f"⚠️ 储量年限仅 {_rli_zh:.1f} 年，近期耗尽风险值得关注——"
                f"新勘探发现或资源量转化将至关重要。"
            ))

        # ── 分析师共识 ───────────────────────────────────────────────────────────
        _an_target_zh = row.get("analyst_target_mean")
        _an_count_zh  = row.get("analyst_count")
        _an_rec_zh    = row.get("analyst_rec_key")
        _an_up_zh     = row.get("analyst_upside")
        if pd.notna(_an_target_zh) and pd.notna(_an_count_zh) and _an_count_zh >= 1:
            _n_int_zh = int(_an_count_zh)
            _an_hi_zh = row.get("analyst_target_high")
            _an_lo_zh = row.get("analyst_target_low")
            _range_str_zh = ""
            if pd.notna(_an_hi_zh) and pd.notna(_an_lo_zh):
                _range_str_zh = f"（区间：${_an_lo_zh:.2f}–${_an_hi_zh:.2f}）"
            _rec_str_zh = f"，共识评级：**{_an_rec_zh}**" if pd.notna(_an_rec_zh) else ""
            if pd.notna(_an_up_zh) and _an_up_zh >= 50:
                bullets.append(_bullet(
                    f"分析师共识目标价 ${_an_target_zh:.2f}{_range_str_zh}，"
                    f"较当前价格隐含 {_an_up_zh:.0f}% 上行空间"
                    f"（{_n_int_zh} 位分析师）{_rec_str_zh}。"
                    f"机构高度认可，与量化评分形成共鸣。"
                ))
            elif pd.notna(_an_up_zh) and _an_up_zh >= 20:
                bullets.append(_bullet(
                    f"分析师共识目标价 ${_an_target_zh:.2f}{_range_str_zh}，"
                    f"隐含 {_an_up_zh:.0f}% 上行空间"
                    f"（{_n_int_zh} 位分析师）{_rec_str_zh}。"
                ))
            elif pd.notna(_an_up_zh) and _an_up_zh > 0:
                bullets.append(_bullet(
                    f"分析师共识目标价 ${_an_target_zh:.2f}，"
                    f"隐含 {_an_up_zh:.0f}% 上行空间"
                    f"（{_n_int_zh} 位分析师）{_rec_str_zh}。"
                ))

        # ── 大宗商品 ──────────────────────────────────────────────────────────
        commodity = str(row.get("commodity", ""))
        outlook_word, outlook_reason = _commodity_view(commodity, lang="zh")
        bullets.append(_bullet(f"大宗商品前景{outlook_word}：{outlook_reason}。"))

        # ── 开发阶段 ──────────────────────────────────────────────────────────
        stage = str(row.get("stage", ""))
        if stage == "Explorer":
            bullets.append(_bullet(
                "早期勘探阶段——风险与回报并存；一旦有新发现或资源量升级，"
                "股价有望大幅重估。"
            ))
        elif stage == "Developer":
            bullets.append(_bullet(
                "开发阶段——项目建设或可行性研究推进中；"
                "后续去风险里程碑将驱动显著重估。"
            ))
        elif stage == "Royalty/Streaming":
            bullets.append(_bullet(
                "特许权/流媒体模式在享受大宗商品价格上行的同时，"
                "规避了直接运营成本风险，商业模式优越。"
            ))

        return bullets

    def _generate_report(data: pd.DataFrame, lang: str = "en") -> str:
        """Generate the full report in English or Chinese (lang='en'|'zh')."""
        zh = (lang == "zh")
        today = _date.today().strftime("%Y年%m月%d日") if zh else _date.today().strftime("%B %d, %Y")
        lines: list[str] = []

        if zh:
            lines += [
                "# 澳大利亚矿业股票低估值分析报告",
                f"*生成日期：{today} | 数据截至最近一次刷新*",
                "", "---", "",
            ]
        else:
            lines += [
                f"# {config.MARKET_NAME} Mining Stock Undervaluation Report",
                f"*Generated: {today} | Data as of last refresh*",
                "", "---", "",
            ]

        # ── Summary ───────────────────────────────────────────────────────────
        strong_buys = data[data["grade"] == "🟢 Strong Buy"]
        buys        = data[data["grade"] == "🔵 Buy"]
        top_score   = data["score_composite"].max()
        avg_score   = data["score_composite"].mean()

        # Dynamic weights from session state (user may have adjusted sliders)
        _ss = st.session_state
        _tw = max(
            _ss.get("w_val", 30) + _ss.get("w_hlt", 20) + _ss.get("w_mom", 15) +
            _ss.get("w_min", 25) + _ss.get("w_com", 5)  + _ss.get("w_stg", 5), 1
        )
        def _wpct(key: str, default: int) -> int:
            return round(_ss.get(key, default) / _tw * 100)

        _w_val = _wpct("w_val", 30)
        _w_hlt = _wpct("w_hlt", 20)
        _w_mom = _wpct("w_mom", 15)
        _w_min = _wpct("w_min", 25)
        _w_com = _wpct("w_com", 5)
        _w_stg = _wpct("w_stg", 5)

        if zh:
            lines += [
                "## 执行摘要", "",
                f"本次筛选覆盖 ASX 合计 **{len(data)} 家** 矿业公司，模型识别出：",
                f"- **{len(strong_buys)} 只强烈买入**候选标的（综合评分 ≥ 75/100）",
                f"- **{len(buys)} 只买入**候选标的（综合评分 ≥ 60/100）",
                f"- 所有筛选标的平均综合评分：**{avg_score:.1f}/100**",
                f"- 单只个股最高评分：**{top_score:.1f}/100**",
                "",
                f"评分基于六大加权维度（当前权重）："
                f"**估值**（{_w_val}%，含 P/NAV）、**财务健康**（{_w_hlt}%）、"
                f"**价格动能/逆向信号**（{_w_mom}%）、"
                f"**矿业质量**（{_w_min}%，S&P/SNL：AISC利润率、储量支撑、NAV折价、EV/盎司、储量年限、品位）、"
                f"**大宗商品前景**（{_w_com}%）、**开发阶段**（{_w_stg}%）。",
                "", "---", "",
            ]
        else:
            lines += [
                "## Executive Summary", "",
                f"Out of **{len(data)} companies** screened across ASX, the model identifies:",
                f"- **{len(strong_buys)} Strong Buy** candidates (composite score ≥ 75/100)",
                f"- **{len(buys)} Buy** candidates (composite score ≥ 60/100)",
                f"- Average composite score across all screened names: **{avg_score:.1f}/100**",
                f"- Highest individual score: **{top_score:.1f}/100**",
                "",
                f"Scores are based on six weighted factors (current weights): "
                f"**Valuation** ({_w_val}%, incl. P/NAV), "
                f"**Financial Health** ({_w_hlt}%), **Price Momentum** ({_w_mom}%), "
                f"**Mining Quality** ({_w_min}%, S&P/SNL: AISC margin, reserves backing, "
                f"NAV discount, EV/oz, reserve life, grade), "
                f"**Commodity Outlook** ({_w_com}%), and **Development Stage** ({_w_stg}%).",
                "", "---", "",
            ]

        # ── Universe Mining Stats (SPG) ───────────────────────────────────────
        _has_spg = any(c.startswith("spg_") for c in data.columns)
        if _has_spg:
            _med_aisc_mgn = data["spg_aisc_margin"].median() if "spg_aisc_margin" in data.columns else None
            _med_pnav_r   = data["spg_p_nav"].median()       if "spg_p_nav"       in data.columns else None
            _n_rli10      = int((data["spg_reserve_life"] >= 10).sum()) if "spg_reserve_life" in data.columns else None
            _n_ev8k       = int((data["ev_per_oz_prod"] < 8_000).sum()) if "ev_per_oz_prod"   in data.columns else None
            _n_spg_total  = int(data[[c for c in data.columns if c.startswith("spg_")]].notna().any(axis=1).sum())
            _med_cc_oz    = data["spg_cash_cost_oz"].median() if "spg_cash_cost_oz" in data.columns else None

            if zh:
                lines += [
                    "## 矿业行业统计概览", "",
                    f"当前筛选范围内共有 **{_n_spg_total}** 家公司拥有 S&P/SNL 矿业基本面数据。",
                    "",
                ]
                _stat_rows_zh = []
                if _med_aisc_mgn is not None and not np.isnan(_med_aisc_mgn):
                    _stat_rows_zh.append(f"- **中位数 AISC 利润率：** {_med_aisc_mgn:.1f}%")
                if _med_pnav_r is not None and not np.isnan(_med_pnav_r):
                    _stat_rows_zh.append(f"- **中位数 P/NAV：** {_med_pnav_r:.2f}x")
                if _med_cc_oz is not None and not np.isnan(_med_cc_oz):
                    _stat_rows_zh.append(f"- **中位数现金成本：** ${_med_cc_oz:,.0f}/盎司")
                if _n_rli10 is not None:
                    _stat_rows_zh.append(f"- **储量寿命 ≥ 10 年的公司数：** {_n_rli10} 家")
                if _n_ev8k is not None:
                    _stat_rows_zh.append(f"- **EV/盎司产量 < $8,000 的公司数（低估值信号）：** {_n_ev8k} 家")
                lines += _stat_rows_zh
                lines += ["", "---", ""]
            else:
                lines += [
                    "## Mining Universe Statistics", "",
                    f"**{_n_spg_total}** companies in this screen have S&P/SNL fundamental mining data.",
                    "",
                ]
                _stat_rows_en = []
                if _med_aisc_mgn is not None and not np.isnan(_med_aisc_mgn):
                    _stat_rows_en.append(f"- **Median AISC Margin:** {_med_aisc_mgn:.1f}%")
                if _med_pnav_r is not None and not np.isnan(_med_pnav_r):
                    _stat_rows_en.append(f"- **Median P/NAV:** {_med_pnav_r:.2f}x")
                if _med_cc_oz is not None and not np.isnan(_med_cc_oz):
                    _stat_rows_en.append(f"- **Median Cash Cost:** ${_med_cc_oz:,.0f}/oz")
                if _n_rli10 is not None:
                    _stat_rows_en.append(f"- **Companies with Reserve Life ≥ 10 yr:** {_n_rli10}")
                if _n_ev8k is not None:
                    _stat_rows_en.append(f"- **Companies with EV/oz Produced < $8,000 (value signal):** {_n_ev8k}")
                lines += _stat_rows_en
                lines += ["", "---", ""]

        # ── Analyst Consensus Summary ─────────────────────────────────────────
        _has_analyst_data = (
            "analyst_upside" in data.columns and
            "analyst_count" in data.columns and
            data["analyst_upside"].notna().any()
        )
        if _has_analyst_data:
            _an_covered = data[data["analyst_count"].fillna(0) >= 2]
            _n_an_cov   = len(_an_covered)
            _med_an_up  = _an_covered["analyst_upside"].median() if not _an_covered.empty else None
            _n_an_bull  = int((_an_covered["analyst_upside"].fillna(0) >= 30).sum()) if not _an_covered.empty else 0
            _n_strong_conv = int((
                (_an_covered["score_composite"] >= 60) &
                (_an_covered["analyst_upside"].fillna(0) >= 30)
            ).sum()) if not _an_covered.empty else 0

            if zh:
                lines += [
                    "## 分析师共识概览", "",
                    f"共有 **{_n_an_cov}** 家公司拥有 ≥2 位分析师追踪覆盖：",
                    "",
                ]
                _an_rows_zh = []
                if _med_an_up is not None and not np.isnan(_med_an_up):
                    _an_rows_zh.append(f"- **分析师中位上行空间：** {_med_an_up:+.0f}%")
                _an_rows_zh.append(f"- **分析师目标上行 ≥ 30% 的公司数：** {_n_an_bull}")
                _an_rows_zh.append(f"- **量化评分 ≥ 60 且分析师上行 ≥ 30% 的高置信度标的：** {_n_strong_conv}")
                lines += _an_rows_zh
                lines += ["", "---", ""]
            else:
                lines += [
                    "## Analyst Consensus Summary", "",
                    f"**{_n_an_cov}** companies with ≥ 2 analyst opinions:",
                    "",
                ]
                _an_rows_en = []
                if _med_an_up is not None and not np.isnan(_med_an_up):
                    _an_rows_en.append(f"- **Median analyst upside:** {_med_an_up:+.0f}%")
                _an_rows_en.append(f"- **Companies with analyst upside ≥ 30%:** {_n_an_bull}")
                _an_rows_en.append(f"- **High-conviction names (score ≥ 60 AND analyst upside ≥ 30%):** {_n_strong_conv}")
                lines += _an_rows_en
                lines += ["", "---", ""]

        # ── Peer medians ──────────────────────────────────────────────────────
        peer_pb       = data["price_to_book"].median()
        peer_evebitda = data["ev_ebitda"].median()

        # ── Top picks ─────────────────────────────────────────────────────────
        candidates = data[data["grade"].isin(["🟢 Strong Buy", "🔵 Buy"])].sort_values(
            "score_composite", ascending=False
        )

        if candidates.empty:
            lines.append("*当前筛选条件下无买入或强烈买入候选标的。*" if zh else
                         "*No Buy or Strong Buy candidates match the current filters.*")
        else:
            if zh:
                lines += [
                    "## 低估值精选标的", "",
                    "> 以下列出所有评级为**强烈买入**或**买入**的公司，"
                    "并逐条说明模型认为其低估的具体依据。",
                    "",
                ]
            else:
                lines += [
                    "## Top Undervalued Picks", "",
                    "> Companies graded **Strong Buy** or **Buy** are listed below with "
                    "specific reasons why the model considers them undervalued.",
                    "",
                ]

            for _, row in candidates.iterrows():
                name      = row.get("name", row.get("ticker", "Unknown"))
                ticker    = row.get("ticker", "")
                grade     = row.get("grade", "")
                score     = row.get("score_composite", 0)
                commodity = row.get("commodity", "")
                stage     = row.get("stage", "")
                price     = row.get("price")
                mcap      = row.get("market_cap")

                price_str = f"${price:.3f}" if pd.notna(price) else "N/A"
                mcap_str  = (f"${mcap/1e9:.2f}B" if pd.notna(mcap) and mcap >= 1e9
                             else f"${mcap/1e6:.0f}M" if pd.notna(mcap) else "N/A")

                if zh:
                    lines += [
                        f"### {grade} {name} ({ticker})",
                        f"**评分：{score:.1f}/100** | {commodity} · {stage} | "
                        f"股价：{price_str} | 市值：{mcap_str}",
                        "",
                        "**该股票被认为低估的具体原因：**",
                        "",
                    ]
                    bullets = _generate_thesis_zh(row, peer_pb, peer_evebitda)
                else:
                    lines += [
                        f"### {grade} {name} ({ticker})",
                        f"**Score: {score:.1f}/100** | {commodity} · {stage} | "
                        f"Price: {price_str} | Mkt Cap: {mcap_str}",
                        "",
                        "**Why this stock appears undervalued:**",
                        "",
                    ]
                    bullets = _generate_thesis(row, peer_pb, peer_evebitda)

                lines += bullets
                lines += ["", "---", ""]

        # ── Commodity breakdown ────────────────────────────────────────────────
        lines += (["## 大宗商品分布", ""] if zh else ["## Commodity Breakdown", ""])
        comm_stats = (
            data.groupby("commodity")["score_composite"]
            .agg(["count", "mean", "max"])
            .sort_values("mean", ascending=False)
        )
        for comm, row in comm_stats.iterrows():
            outlook_word, _ = _commodity_view(str(comm), lang=lang)
            if zh:
                lines.append(
                    f"- **{comm}**（{int(row['count'])} 只）— "
                    f"平均分 {row['mean']:.0f}，最高 {row['max']:.0f} | "
                    f"前景：*{outlook_word}*"
                )
            else:
                lines.append(
                    f"- **{comm}** ({int(row['count'])} stocks) — "
                    f"Avg score {row['mean']:.0f}, Best {row['max']:.0f} | "
                    f"Outlook: *{outlook_word}*"
                )
        lines += ["", "---", ""]

        # ── Disclaimer ────────────────────────────────────────────────────────
        if zh:
            lines += [
                "## 免责声明", "",
                "*本报告由系统根据公开市场数据自动生成，仅供参考，不构成任何投资建议。"
                "矿业股票属于高风险投资品种，投资前请务必进行独立尽职调查。*",
            ]
        else:
            lines += [
                "## Disclaimer", "",
                "*This report is generated automatically from publicly available market data. "
                "It is for informational purposes only and does not constitute investment advice. "
                "Mining stocks are high-risk investments. Always conduct your own due diligence "
                "before making any investment decisions.*",
            ]

        return "\n".join(lines)

    # ── Render the report tab ──────────────────────────────────────────────────
    st.markdown("### Undervaluation Report  |  低估值分析报告")

    col_rl, col_rr = st.columns([3, 1])
    with col_rr:
        lang = st.radio(
            "Language / 语言",
            ["Both / 双语", "English", "中文"],
            index=0,
            horizontal=True,
            key="report_lang",
        )
        top_n = st.selectbox(
            "Top N / 显示前N只",
            [10, 20, 50, 999],
            index=1,
            format_func=lambda x: "All / 全部" if x == 999 else str(x),
            key="report_topn",
        )
        grade_filter = st.multiselect(
            "Grades / 评级筛选",
            ["🟢 Strong Buy", "🔵 Buy", "🟡 Watch"],
            default=["🟢 Strong Buy", "🔵 Buy"],
            key="report_grade",
        )

    report_data = filtered[filtered["grade"].isin(grade_filter)].nlargest(
        top_n if top_n < 999 else len(filtered), "score_composite"
    )
    src = report_data if not report_data.empty else filtered

    with col_rl:
        st.caption(
            f"Showing {len(src)} companies | 显示 {len(src)} 家公司 | "
            f"Grades: {', '.join(grade_filter) or 'none'}"
        )

    # ── Universe Score Trend ──────────────────────────────────────────────────
    _trend_data = load_sector_trends()
    if not _trend_data.empty and _trend_data["snap_date"].nunique() >= 2:
        with st.expander("📈 Universe Score Trend  |  行业评分走势", expanded=True):
            # Universe aggregate: weighted avg score across all groups per date
            _td = _trend_data.copy()
            _td["_weighted"] = _td["avg_score"] * _td["n_companies"]
            _univ_agg = _td.groupby("snap_date").agg(
                _weighted_sum=("_weighted", "sum"),
                n_companies=("n_companies", "sum"),
            ).reset_index()
            _univ_agg["avg_score"] = (_univ_agg["_weighted_sum"] / _univ_agg["n_companies"]).round(1)
            _univ_agg = _univ_agg.drop(columns=["_weighted_sum"])
            _univ_agg["commodity_group"] = "Universe (All)"
            _univ_trend = _univ_agg

            _trend_plot_df = pd.concat([_trend_data, _univ_trend], ignore_index=True)
            _trend_plot_df["snap_date"] = pd.to_datetime(_trend_plot_df["snap_date"])
            _trend_plot_df = _trend_plot_df.sort_values("snap_date")

            _top_comms = (
                _trend_data.groupby("commodity_group")["n_companies"].sum()
                .nlargest(5).index.tolist()
            )
            _trend_filtered = _trend_plot_df[
                _trend_plot_df["commodity_group"].isin(_top_comms + ["Universe (All)"])
            ]

            _fig_trend = go.Figure()
            _comm_palette = {
                "Universe (All)": "#f59e0b",
                "Gold":     "#eab308", "Silver":   "#94a3b8",
                "Copper":   "#f97316", "Uranium":  "#22c55e",
                "Nickel":   "#06b6d4", "Lithium":  "#a855f7",
                "Iron Ore": "#ef4444", "Zinc":     "#3b82f6",
            }
            for _grp in ["Universe (All)"] + _top_comms:
                _sub = _trend_filtered[_trend_filtered["commodity_group"] == _grp]
                if _sub.empty:
                    continue
                _is_univ = _grp == "Universe (All)"
                _fig_trend.add_trace(go.Scatter(
                    x=_sub["snap_date"], y=_sub["avg_score"],
                    mode="lines+markers",
                    name=_grp,
                    line=dict(
                        color=_comm_palette.get(_grp, "#6b7280"),
                        width=3 if _is_univ else 1.5,
                        dash="solid" if _is_univ else "dot",
                    ),
                    marker=dict(size=6 if _is_univ else 4),
                    customdata=_sub[["n_companies"]].values,
                    hovertemplate=(
                        f"<b>{_grp}</b><br>"
                        "Date: %{x|%Y-%m-%d}<br>"
                        "Avg Score: %{y:.1f}<br>"
                        "Companies: %{customdata[0]:.0f}<extra></extra>"
                    ),
                ))

            _fig_trend.add_hline(
                y=60, line_dash="dash", line_color="rgba(34,197,94,0.4)",
                annotation_text="Buy threshold (60)", annotation_position="bottom right",
                annotation_font_size=10,
            )
            _fig_trend.add_hline(
                y=75, line_dash="dash", line_color="rgba(59,130,246,0.4)",
                annotation_text="Strong Buy threshold (75)", annotation_position="top right",
                annotation_font_size=10,
            )
            _fig_trend.update_layout(
                title="Average Composite Score by Commodity Group Over Time  |  各大宗商品组平均综合评分历史走势",
                xaxis_title="Snapshot Date / 快照日期",
                yaxis_title="Average Score / 平均评分",
                yaxis=dict(range=[0, 100]),
                height=380,
                legend=dict(orientation="h", y=-0.22, font=dict(size=11)),
                margin=dict(t=50, b=80, l=50, r=20),
                hovermode="x unified",
            )
            st.plotly_chart(_fig_trend, use_container_width=True)
            st.caption(
                f"Based on {_trend_data['snap_date'].nunique()} snapshots across "
                f"{int(_trend_data['n_companies'].max())} companies. "
                "Refresh daily to accumulate trend history. / "
                f"基于 {_trend_data['snap_date'].nunique()} 个快照，"
                "每日刷新可积累更多趋势数据。"
            )
    elif not _trend_data.empty:
        st.info(
            "Universe trend chart requires ≥ 2 daily snapshots — "
            "refresh again tomorrow to start building the trend. / "
            "趋势图需要至少 2 个快照，明日再次刷新即可开始积累数据。"
        )

    if lang == "Both / 双语":
        # ── Side-by-side bilingual layout ─────────────────────────────────────
        col_en, col_zh = st.columns(2)
        report_en = _generate_report(src, lang="en")
        report_zh = _generate_report(src, lang="zh")
        with col_en:
            st.markdown(report_en, unsafe_allow_html=False)
        with col_zh:
            st.markdown(report_zh, unsafe_allow_html=False)

        # Combined bilingual download
        combined_md = (
            "<!-- English -->\n" + report_en +
            "\n\n---\n\n<!-- 中文 -->\n" + report_zh
        )
        _dl_col1, _dl_col2 = st.columns(2)
        _dl_col1.download_button(
            "⬇️ Download Bilingual Report / 下载双语报告 (.md)",
            data=combined_md.encode("utf-8"),
            file_name=f"mining_report_bilingual_{_date.today()}.md",
            mime="text/markdown",
        )
        try:
            _pdf_bytes = _report_to_pdf(combined_md, title="Mining Screener — Bilingual Report")
            _dl_col2.download_button(
                "⬇️ Download PDF / 下载 PDF",
                data=_pdf_bytes,
                file_name=f"mining_report_bilingual_{_date.today()}.pdf",
                mime="application/pdf",
            )
        except Exception as _pdf_err:
            _dl_col2.error(f"PDF error: {_pdf_err}")
    else:
        lang_code = "zh" if lang == "中文" else "en"
        report_md = _generate_report(src, lang=lang_code)
        st.markdown(report_md, unsafe_allow_html=False)
        fname_md = f"矿业报告_{_date.today()}.md" if lang_code == "zh" else f"mining_report_{_date.today()}.md"
        fname_pdf = f"矿业报告_{_date.today()}.pdf" if lang_code == "zh" else f"mining_report_{_date.today()}.pdf"
        btn_label = "⬇️ 下载报告（.md）" if lang_code == "zh" else "⬇️ Download Report (.md)"
        btn_pdf   = "⬇️ 下载 PDF" if lang_code == "zh" else "⬇️ Download PDF"
        _dl_col1, _dl_col2 = st.columns(2)
        _dl_col1.download_button(
            btn_label,
            data=report_md.encode("utf-8"),
            file_name=fname_md,
            mime="text/markdown",
        )
        try:
            _pdf_bytes = _report_to_pdf(report_md, title="Mining Screener Report")
            _dl_col2.download_button(
                btn_pdf,
                data=_pdf_bytes,
                file_name=fname_pdf,
                mime="application/pdf",
            )
        except Exception as _pdf_err:
            _dl_col2.error(f"PDF error: {_pdf_err}")

# ── TAB 5: Watchlist ───────────────────────────────────────────────────────────
with tab_watchlist:
    _wl_tickers  = get_watchlist()
    _positions_df = get_positions()   # ticker, shares, avg_cost, currency

    if not _wl_tickers:
        st.info("No companies in your watchlist yet.  \n"
                "Go to **🔍 Company Detail** → select a company → click **⭐ Add to Watchlist**.")
    else:
        # Merge market data
        _wl_df = df[df["ticker"].isin(_wl_tickers)].copy()
        if "score_composite" in _wl_df.columns:
            _wl_df = _wl_df.sort_values("score_composite", ascending=False)

        # ── Grade-change alert banner ──────────────────────────────────────────
        _GRADE_ORD = {"🟢 Strong Buy": 5, "🔵 Buy": 4, "🟡 Watch": 3, "🟠 Neutral": 2, "🔴 Avoid": 1}
        if "grade_prev" in _wl_df.columns and _wl_df["grade_prev"].notna().any():
            _wl_ups = _wl_df[
                _wl_df["grade_prev"].notna() &
                (_wl_df["grade"].map(_GRADE_ORD).fillna(0) >
                 _wl_df["grade_prev"].map(_GRADE_ORD).fillna(0))
            ]
            _wl_dns = _wl_df[
                _wl_df["grade_prev"].notna() &
                (_wl_df["grade"].map(_GRADE_ORD).fillna(0) <
                 _wl_df["grade_prev"].map(_GRADE_ORD).fillna(0))
            ]
            if not _wl_ups.empty:
                _up_parts = [
                    f"**{r['name']}** ({r['ticker']}): {r['grade_prev']} → {r['grade']}"
                    for _, r in _wl_ups.iterrows()
                ]
                st.success("⬆️ **Grade upgrades in your watchlist:**  " + "  ·  ".join(_up_parts))
            if not _wl_dns.empty:
                _dn_parts = [
                    f"**{r['name']}** ({r['ticker']}): {r['grade_prev']} → {r['grade']}"
                    for _, r in _wl_dns.iterrows()
                ]
                st.warning("⬇️ **Grade downgrades in your watchlist:**  " + "  ·  ".join(_dn_parts))

        # ── Price-target proximity alerts ──────────────────────────────────────
        _pt_alerts = []
        for _t in _wl_tickers:
            _pt = get_price_target(_t)
            if not _pt:
                continue
            _px_row = _wl_df[_wl_df["ticker"] == _t]
            if _px_row.empty:
                continue
            _px = _px_row.iloc[0].get("price")
            if pd.isna(_px) or _px <= 0:
                continue
            _dist = (_pt / _px - 1) * 100
            if abs(_dist) <= 5:
                _nm = _px_row.iloc[0].get("name", _t)
                _arrow = "↑" if _dist > 0 else "↓"
                _pt_alerts.append(
                    f"**{_nm}** ({_t}): ${_px:.3f} — {_arrow}{abs(_dist):.1f}% to target ${_pt:.3f}"
                )
        if _pt_alerts:
            st.info("🎯 **Near price target (±5%):**  " + "  ·  ".join(_pt_alerts))

        # Merge position data
        if not _positions_df.empty:
            _wl_df = _wl_df.merge(
                _positions_df[["ticker", "shares", "avg_cost", "currency"]],
                on="ticker", how="left",
            )
        else:
            _wl_df["shares"]   = float("nan")
            _wl_df["avg_cost"] = float("nan")
            _wl_df["currency"] = None

        # Compute P&L per row (where position data exists)
        _has_pos = _wl_df["shares"].notna() & (_wl_df["shares"] > 0)
        _wl_df["mkt_value"]  = np.where(
            _has_pos & _wl_df["price"].notna(),
            _wl_df["shares"] * _wl_df["price"], np.nan,
        )
        _wl_df["cost_basis"] = np.where(
            _has_pos & _wl_df["avg_cost"].notna(),
            _wl_df["shares"] * _wl_df["avg_cost"], np.nan,
        )
        _wl_df["pnl_abs"]  = _wl_df["mkt_value"] - _wl_df["cost_basis"]
        _wl_df["pnl_pct"]  = np.where(
            _wl_df["cost_basis"].notna() & (_wl_df["cost_basis"] > 0),
            (_wl_df["pnl_abs"] / _wl_df["cost_basis"] * 100).round(1),
            np.nan,
        )
        _wl_df["port_weight"] = np.where(
            _wl_df["mkt_value"].notna() & (_wl_df["mkt_value"].sum() > 0),
            (_wl_df["mkt_value"] / _wl_df["mkt_value"].sum() * 100).round(1),
            np.nan,
        )

        # ── Portfolio summary strip ────────────────────────────────────────────
        _total_val  = _wl_df["mkt_value"].sum()
        _total_cost = _wl_df["cost_basis"].sum()
        _total_pnl  = (_total_val - _total_cost) if (_total_val > 0 and _total_cost > 0) else float("nan")
        _total_pnl_pct = (_total_pnl / _total_cost * 100) if (pd.notna(_total_pnl) and _total_cost > 0) else float("nan")
        _n_positions = int(_has_pos.sum())

        st.markdown(f"### ⭐ Watchlist · 💼 Portfolio")

        _ps = st.columns(6)
        _ps[0].metric("Watchlisted", len(_wl_df))
        _ps[1].metric("Positions", _n_positions)
        _val_fmt  = f"${_total_val:,.0f}" if _total_val > 0 else "—"
        _pnl_fmt  = (f"{_total_pnl_pct:+.1f}%" if pd.notna(_total_pnl_pct) else None)
        _ps[2].metric("Market Value", _val_fmt, delta=_pnl_fmt)
        _pnl_abs_fmt = f"${_total_pnl:+,.0f}" if pd.notna(_total_pnl) else "—"
        _ps[3].metric("Total P&L", _pnl_abs_fmt)
        _avg_score = _wl_df["score_composite"].mean()
        _ps[4].metric("Avg Score", f"{_avg_score:.1f}" if pd.notna(_avg_score) else "—")
        _avg_delta = _wl_df["score_delta"].mean() if "score_delta" in _wl_df.columns else float("nan")
        _ps[5].metric("Avg Δ Score", f"{_avg_delta:+.1f}" if pd.notna(_avg_delta) else "—")

        # ── Portfolio exposure charts ──────────────────────────────────────────
        _has_any_positions = _n_positions > 0 and _total_val > 0
        if _has_any_positions:
            st.divider()
            st.markdown("#### Portfolio Exposure")
            _exp_cols = st.columns(3)

            # Commodity breakdown (by market value, fallback to equal weight)
            with _exp_cols[0]:
                _comm_val = (
                    _wl_df[_wl_df["mkt_value"].notna()]
                    .groupby("commodity")["mkt_value"].sum()
                    .reset_index()
                    .rename(columns={"mkt_value": "Value", "commodity": "Commodity"})
                    .sort_values("Value", ascending=False)
                )
                if not _comm_val.empty:
                    _fig_comm = px.pie(
                        _comm_val, values="Value", names="Commodity",
                        title="By Commodity",
                        hole=0.45,
                        color_discrete_sequence=px.colors.qualitative.Set2,
                    )
                    _fig_comm.update_traces(textposition="outside", textinfo="percent+label")
                    _fig_comm.update_layout(showlegend=False, height=300,
                                            margin=dict(l=10, r=10, t=40, b=10))
                    st.plotly_chart(_fig_comm, width="stretch")

            # Stage breakdown
            with _exp_cols[1]:
                _stage_val = (
                    _wl_df[_wl_df["mkt_value"].notna()]
                    .groupby("stage")["mkt_value"].sum()
                    .reset_index()
                    .rename(columns={"mkt_value": "Value", "stage": "Stage"})
                    .sort_values("Value", ascending=False)
                )
                if not _stage_val.empty:
                    _fig_stage = px.bar(
                        _stage_val, x="Value", y="Stage",
                        orientation="h",
                        title="By Stage",
                        color="Stage",
                        color_discrete_sequence=px.colors.qualitative.Pastel,
                    )
                    _fig_stage.update_layout(showlegend=False, height=300,
                                             margin=dict(l=10, r=10, t=40, b=10),
                                             xaxis_title="Value ($)")
                    _fig_stage.update_yaxes(autorange="reversed")
                    st.plotly_chart(_fig_stage, width="stretch")

            # P&L waterfall per position
            with _exp_cols[2]:
                _pnl_df = (
                    _wl_df[_wl_df["pnl_abs"].notna()]
                    .sort_values("pnl_abs", ascending=False)
                    [["name", "pnl_abs", "pnl_pct"]]
                    .copy()
                )
                if not _pnl_df.empty:
                    _fig_pnl = go.Figure()
                    _fig_pnl.add_bar(
                        x=_pnl_df["pnl_abs"],
                        y=_pnl_df["name"],
                        orientation="h",
                        marker_color=_pnl_df["pnl_abs"].apply(
                            lambda v: "#22c55e" if v >= 0 else "#ef4444"
                        ).tolist(),
                        text=_pnl_df["pnl_pct"].apply(
                            lambda v: f"{v:+.1f}%" if pd.notna(v) else ""
                        ).tolist(),
                        textposition="outside",
                    )
                    _fig_pnl.update_layout(
                        title="P&L by Position",
                        height=300,
                        margin=dict(l=10, r=60, t=40, b=10),
                        xaxis_title="P&L ($)",
                        showlegend=False,
                    )
                    _fig_pnl.update_yaxes(autorange="reversed")
                    st.plotly_chart(_fig_pnl, width="stretch")

        st.divider()

        # ── Company cards ──────────────────────────────────────────────────────
        st.markdown(f"#### Positions & Notes")
        for _, _wrow in _wl_df.iterrows():
            _wtk    = _wrow.get("ticker", "")
            _wname  = _wrow.get("name", _wtk)
            _wgrade = _wrow.get("grade", "—")
            _wscore = _wrow.get("score_composite")
            _wdelta = _wrow.get("score_delta")
            _wprice = _wrow.get("price")
            _wmcap  = _wrow.get("market_cap")
            _wcomm  = _wrow.get("commodity", "—")
            _wstage = _wrow.get("stage", "—")

            # Position data
            _w_shares   = _wrow.get("shares")
            _w_avgcost  = _wrow.get("avg_cost")
            _w_mktval   = _wrow.get("mkt_value")
            _w_pnl_abs  = _wrow.get("pnl_abs")
            _w_pnl_pct  = _wrow.get("pnl_pct")
            _w_weight   = _wrow.get("port_weight")

            _score_str = f"{_wscore:.1f}/100" if pd.notna(_wscore) else "—"
            _delta_str = f"{_wdelta:+.1f}" if pd.notna(_wdelta) else None
            _price_str = f"${_wprice:.3f}" if pd.notna(_wprice) else "—"
            _has_pos_row = pd.notna(_w_shares) and float(_w_shares) > 0

            _header_suffix = ""
            if _has_pos_row and pd.notna(_w_mktval):
                _pnl_tag = f"  [{_w_pnl_pct:+.1f}%]" if pd.notna(_w_pnl_pct) else ""
                _header_suffix = f"  ·  💼 ${_w_mktval:,.0f}{_pnl_tag}"

            with st.expander(
                f"{_wgrade}  {_wname} ({_wtk}) — Score {_score_str}{_header_suffix}",
                expanded=True,
            ):
                _wc1, _wc2, _wc3, _wc4 = st.columns([2, 1, 1, 2])

                with _wc1:
                    st.markdown(f"**{_wcomm}** · {_wstage}")
                    _pb     = _wrow.get("price_to_book")
                    _pnv    = _wrow.get("spg_p_nav")
                    _aim    = _wrow.get("spg_aisc_margin")
                    _up     = _wrow.get("upside_to_nav") or _wrow.get("pb_peer_upside")
                    _w_an_up = _wrow.get("analyst_upside")
                    _w_an_ct = _wrow.get("analyst_count")
                    _kpi_parts = []
                    if pd.notna(_up):  _kpi_parts.append(f"↑{_up:+.0f}% upside")
                    if pd.notna(_pb):  _kpi_parts.append(f"P/B {_pb:.2f}x")
                    if pd.notna(_pnv): _kpi_parts.append(f"P/NAV {_pnv:.2f}x")
                    if pd.notna(_aim): _kpi_parts.append(f"AISC margin {_aim:.0f}%")
                    if _kpi_parts:
                        st.caption("  ·  ".join(_kpi_parts))
                    # Analyst consensus line
                    if pd.notna(_w_an_up) and pd.notna(_w_an_ct) and float(_w_an_ct) >= 1:
                        _an_col = "#22c55e" if _w_an_up >= 20 else "#eab308" if _w_an_up >= 0 else "#ef4444"
                        _an_rec = _wrow.get("analyst_rec_key", "")
                        _an_rec_str = f" · {_an_rec.replace('-',' ').title()}" if _an_rec else ""
                        st.markdown(
                            f"<span style='font-size:11px;color:{_an_col}'>"
                            f"Analyst: <b>{_w_an_up:+.0f}%</b> ({int(_w_an_ct)} analysts){_an_rec_str}</span>",
                            unsafe_allow_html=True,
                        )

                _wc2.metric("Score", _score_str, delta=_delta_str)
                _wc3.metric("Price", _price_str)
                # Price target proximity badge
                _wpt = get_price_target(_wtk)
                if _wpt and pd.notna(_wprice) and _wprice > 0:
                    _pt_dist = (_wpt / _wprice - 1) * 100
                    _pt_icon = "🎯"
                    _pt_color = ("#22c55e" if _pt_dist > 5
                                 else "#ef4444" if _pt_dist < -5
                                 else "#eab308")
                    _pt_dir = "↑" if _pt_dist > 0 else "↓"
                    _wc3.markdown(
                        f"<span style='color:{_pt_color};font-size:12px'>"
                        f"{_pt_icon} Target: ${_wpt:.3f} ({_pt_dir}{abs(_pt_dist):.1f}%)</span>",
                        unsafe_allow_html=True,
                    )
                elif _wpt:
                    _wc3.caption(f"🎯 Target: ${_wpt:.3f}")
                if _has_pos_row and pd.notna(_w_weight):
                    _wc3.caption(f"Weight: {_w_weight:.1f}%")

                # ── Position editor ───────────────────────────────────────────
                with st.expander("💼 Position", expanded=_has_pos_row):
                    _pos_cols = st.columns([2, 2, 1])
                    _inp_shares = _pos_cols[0].number_input(
                        "Shares held",
                        min_value=0.0,
                        value=float(_w_shares) if pd.notna(_w_shares) else 0.0,
                        step=100.0,
                        format="%.0f",
                        key=f"pos_shares_{_wtk}",
                    )
                    _inp_cost = _pos_cols[1].number_input(
                        "Avg cost / share",
                        min_value=0.0,
                        value=float(_w_avgcost) if pd.notna(_w_avgcost) else 0.0,
                        step=0.01,
                        format="%.3f",
                        key=f"pos_cost_{_wtk}",
                    )
                    _pos_cols[2].markdown("<br>", unsafe_allow_html=True)
                    if _pos_cols[2].button("💾", key=f"pos_save_{_wtk}",
                                           help="Save position"):
                        upsert_position(
                            _wtk,
                            shares=_inp_shares,
                            avg_cost=_inp_cost if _inp_cost > 0 else None,
                        )
                        st.toast(f"Position saved for {_wtk} ✓", icon="💾")
                        st.rerun()

                    # Live P&L preview (based on current input, not saved values)
                    if _inp_shares > 0 and pd.notna(_wprice) and _wprice > 0:
                        _preview_val  = _inp_shares * _wprice
                        _preview_cost = _inp_shares * _inp_cost if _inp_cost > 0 else None
                        _preview_pnl  = ((_preview_val - _preview_cost) / _preview_cost * 100
                                         if _preview_cost else None)
                        _pnl_color = ("green" if _preview_pnl and _preview_pnl >= 0
                                      else "red" if _preview_pnl else "gray")
                        _pnl_str = (f" · P&L: <span style='color:{_pnl_color}'>"
                                    f"{_preview_pnl:+.1f}%</span>"
                                    if _preview_pnl is not None else "")
                        st.markdown(
                            f"<small>Market value: **${_preview_val:,.0f}**{_pnl_str}</small>",
                            unsafe_allow_html=True,
                        )

                # ── Note ──────────────────────────────────────────────────────
                with _wc4:
                    _cur_note = get_watchlist_note(_wtk)
                    _new_note = st.text_area(
                        "📝 Note", value=_cur_note, height=80,
                        key=f"wl_note_{_wtk}",
                        placeholder="Your thesis / reminder…",
                        label_visibility="collapsed",
                    )
                    if _new_note != _cur_note:
                        update_watchlist_note(_wtk, _new_note)

                # Remove buttons
                _rm_cols = st.columns([1, 1, 4])
                if _rm_cols[0].button(f"✖️ Remove", key=f"wl_rm_{_wtk}",
                                      help="Remove from watchlist"):
                    remove_from_watchlist(_wtk)
                    st.rerun()
                if _has_pos_row and _rm_cols[1].button(
                    "🗑️ Clear pos", key=f"pos_del_{_wtk}", help="Zero out position"
                ):
                    delete_position(_wtk)
                    st.toast(f"Position cleared for {_wtk}", icon="🗑️")
                    st.rerun()

        # ── Radar comparison ───────────────────────────────────────────────────
        if len(_wl_df) >= 2:
            st.divider()
            st.markdown("#### Score Comparison (Watchlist)")
            _radar_metrics = ["score_valuation", "score_health", "score_momentum",
                              "score_mining", "score_commodity", "score_stage"]
            _radar_labels  = ["Valuation", "Health", "Momentum", "⛏️ Mining", "Commodity", "Stage"]
            _available_m   = [m for m in _radar_metrics if m in _wl_df.columns]

            _fig_wl = go.Figure()
            for _, _wr in _wl_df.iterrows():
                _vals = [float(_wr.get(m, 50) or 50) for m in _available_m]
                _labels_used = [_radar_labels[_radar_metrics.index(m)] for m in _available_m]
                _fig_wl.add_trace(go.Scatterpolar(
                    r=_vals + [_vals[0]],
                    theta=_labels_used + [_labels_used[0]],
                    fill="toself",
                    opacity=0.4,
                    name=f"{_wr.get('name', _wr.get('ticker', ''))} ({_wr.get('ticker','')})",
                ))
            _fig_wl.update_layout(
                polar=dict(radialaxis=dict(range=[0, 100])),
                height=420,
                margin=dict(l=60, r=60, t=40, b=40),
            )
            st.plotly_chart(_fig_wl, width="stretch")

        # ── Mining Metrics Snapshot ───────────────────────────────────────────
        st.divider()
        st.markdown("#### ⛏️ Mining Metrics Snapshot")
        st.caption("Key S&P/SNL metrics for all watchlisted companies — sorted by composite score.")
        _snap_cols = {
            "name":                  "Company",
            "ticker":                "Ticker",
            "commodity":             "Commodity",
            "grade":                 "Grade",
            "score_composite":       "Score",
            "spg_p_nav":             "P/NAV",
            "spg_aisc_margin":       "AISC Mgn%",
            "spg_cash_cost_oz":      "Cash Cost($/oz)",
            "ev_per_oz_prod":        "EV/oz Prod",
            "ev_per_oz_reserve":     "EV/oz Rsv",
            "spg_production_oz":     "Prod(koz/yr)",
            "spg_reserve_life":      "Rsv Life(yr)",
            "upside_to_nav":         "↑NAV%",
            "analyst_upside":        "↑Analyst%",
            "analyst_count":         "#Analysts",
            "return_3m":             "3M Ret%",
        }
        _snap_avail = {k: v for k, v in _snap_cols.items() if k in _wl_df.columns}
        if len(_snap_avail) >= 4:
            _snap_tbl = _wl_df[list(_snap_avail.keys())].copy()
            if "spg_production_oz" in _snap_tbl.columns:
                _snap_tbl["spg_production_oz"] = pd.to_numeric(
                    _snap_tbl["spg_production_oz"], errors="coerce") / 1000
            _snap_tbl = _snap_tbl.rename(columns=_snap_avail)
            _snap_tbl.index = range(1, len(_snap_tbl) + 1)

            def _snap_color_score(val):
                if pd.isna(val): return ""
                if val >= 70: return "background-color:#dcfce7;color:#14532d;font-weight:700"
                if val >= 55: return "background-color:#dbeafe;color:#1e3a5f"
                if val >= 40: return "background-color:#fef9c3;color:#713f12"
                return "background-color:#fee2e2;color:#7f1d1d"

            def _snap_color_ret(val):
                if pd.isna(val) or val == 0: return ""
                return "color:#22c55e;font-weight:600" if val > 0 else "color:#ef4444;font-weight:600"

            _snap_fmt = {
                "Score":          lambda x: f"{x:.1f}" if pd.notna(x) else "—",
                "P/NAV":          lambda x: f"{x:.2f}x" if pd.notna(x) else "—",
                "AISC Mgn%":      lambda x: f"{x:.1f}%" if pd.notna(x) else "—",
                "Cash Cost($/oz)":lambda x: f"${x:,.0f}" if pd.notna(x) else "—",
                "EV/oz Prod":     lambda x: f"${x:,.0f}" if pd.notna(x) else "—",
                "EV/oz Rsv":      lambda x: f"${x:,.0f}" if pd.notna(x) else "—",
                "Prod(koz/yr)":   lambda x: f"{x:,.1f}" if pd.notna(x) else "—",
                "Rsv Life(yr)":   lambda x: f"{x:.1f}" if pd.notna(x) else "—",
                "↑NAV%":          lambda x: f"{x:+.0f}%" if pd.notna(x) else "—",
                "↑Analyst%":      lambda x: f"{x:+.0f}%" if pd.notna(x) else "—",
                "#Analysts":      lambda x: f"{int(x)}" if pd.notna(x) else "—",
                "3M Ret%":        lambda x: f"{x:+.1f}%" if pd.notna(x) else "—",
            }
            _snap_score_col = ["Score"] if "Score" in _snap_tbl.columns else []
            _snap_ret_col   = ["3M Ret%", "↑NAV%", "↑Analyst%"] if "3M Ret%" in _snap_tbl.columns else []
            _snap_styled = (
                _snap_tbl.style
                .map(_snap_color_score, subset=_snap_score_col)
                .map(_snap_color_ret,   subset=_snap_ret_col)
                .format({k: v for k, v in _snap_fmt.items() if k in _snap_tbl.columns}, na_rep="—")
            )
            st.dataframe(_snap_styled, width="stretch", height=min(80 + len(_snap_tbl) * 35, 500))
        else:
            st.info("Run a data refresh to populate S&P/SNL mining metrics.")

        # ── Analyst Consensus Chart ────────────────────────────────────────────
        _wl_an_df = _wl_df[
            _wl_df["analyst_upside"].notna() &
            (_wl_df["analyst_count"].fillna(0) >= 1)
        ].copy() if "analyst_upside" in _wl_df.columns else pd.DataFrame()

        if not _wl_an_df.empty:
            st.divider()
            st.markdown("#### 🎙️ Analyst Consensus Upside (Watchlist)")
            st.caption("Consensus price target vs current price — bars show upside/downside implied by analyst coverage.")
            _wl_an_df = _wl_an_df.sort_values("analyst_upside", ascending=True)
            _wl_an_colors = _wl_an_df["analyst_upside"].apply(
                lambda v: "#22c55e" if v >= 20 else "#3b82f6" if v >= 0 else "#ef4444"
            ).tolist()
            _wl_an_labels = _wl_an_df.apply(
                lambda r: (
                    f"{r.get('name', r['ticker'])} ({r['ticker']})"
                    + (f" — {int(r['analyst_count'])}A" if pd.notna(r.get('analyst_count')) else "")
                ),
                axis=1,
            ).tolist()
            _fig_wl_an = go.Figure()
            _fig_wl_an.add_bar(
                x=_wl_an_df["analyst_upside"].tolist(),
                y=_wl_an_labels,
                orientation="h",
                marker_color=_wl_an_colors,
                text=_wl_an_df["analyst_upside"].apply(
                    lambda v: f"{v:+.0f}%"
                ).tolist(),
                textposition="outside",
                customdata=_wl_an_df[["analyst_target_mean", "price"]].values if "analyst_target_mean" in _wl_an_df.columns else None,
                hovertemplate=(
                    "<b>%{y}</b><br>Analyst upside: %{x:+.1f}%<br>"
                    "Target: $%{customdata[0]:.3f}  |  Price: $%{customdata[1]:.3f}<extra></extra>"
                ) if "analyst_target_mean" in _wl_an_df.columns else None,
            )
            _fig_wl_an.add_vline(x=0, line_color="rgba(100,100,100,0.5)", line_width=1)
            _fig_wl_an.add_vline(x=20, line_color="rgba(34,197,94,0.3)", line_dash="dot", line_width=1)
            _fig_wl_an.update_layout(
                height=max(250, 50 + len(_wl_an_df) * 35),
                margin=dict(l=10, r=80, t=10, b=10),
                xaxis_title="Analyst Upside (%)",
                showlegend=False,
                xaxis=dict(zeroline=False),
            )
            st.plotly_chart(_fig_wl_an, width="stretch")

        # ── Portfolio table ────────────────────────────────────────────────────
        if _n_positions > 0:
            st.divider()
            st.markdown("#### Portfolio Summary Table")
            _ptbl = (
                _wl_df[_has_pos]
                [[c for c in ["name", "ticker", "commodity", "stage",
                               "price", "shares", "avg_cost",
                               "mkt_value", "cost_basis", "pnl_abs", "pnl_pct",
                               "port_weight", "score_composite", "grade"]
                  if c in _wl_df.columns]]
                .copy()
                .sort_values("mkt_value", ascending=False)
                .rename(columns={
                    "name": "Company", "ticker": "Ticker",
                    "commodity": "Commodity", "stage": "Stage",
                    "price": "Price", "shares": "Shares",
                    "avg_cost": "Avg Cost", "mkt_value": "Mkt Value",
                    "cost_basis": "Cost Basis", "pnl_abs": "P&L ($)",
                    "pnl_pct": "P&L %", "port_weight": "Weight %",
                    "score_composite": "Score", "grade": "Grade",
                })
            )

            def _color_pnl(val):
                if pd.isna(val): return ""
                return "color: #22c55e; font-weight:600" if val >= 0 else "color: #ef4444; font-weight:600"

            _ptbl_styled = (
                _ptbl.style
                .map(_color_pnl, subset=["P&L ($)", "P&L %"] if "P&L ($)" in _ptbl.columns else [])
                .format({
                    "Price":      lambda x: f"${x:.3f}" if pd.notna(x) else "—",
                    "Shares":     lambda x: f"{x:,.0f}" if pd.notna(x) else "—",
                    "Avg Cost":   lambda x: f"${x:.3f}" if pd.notna(x) else "—",
                    "Mkt Value":  lambda x: f"${x:,.0f}" if pd.notna(x) else "—",
                    "Cost Basis": lambda x: f"${x:,.0f}" if pd.notna(x) else "—",
                    "P&L ($)":    lambda x: f"${x:+,.0f}" if pd.notna(x) else "—",
                    "P&L %":      lambda x: f"{x:+.1f}%" if pd.notna(x) else "—",
                    "Weight %":   lambda x: f"{x:.1f}%" if pd.notna(x) else "—",
                    "Score":      lambda x: f"{x:.1f}" if pd.notna(x) else "—",
                }, na_rep="—")
            )
            st.dataframe(_ptbl_styled, width="stretch",
                         height=min(100 + len(_ptbl) * 35, 500))

        # ── Portfolio Rebalancing ──────────────────────────────────────────────
        if _n_positions > 0:
            st.divider()
            st.markdown("#### ⚖️ Rebalancing Suggestions")
            st.caption(
                "Target weight = each position's score / total score of all held positions. "
                "Positive gap = underweight (consider adding), negative gap = overweight (consider trimming)."
            )

            _rb_rows = _wl_df[_has_pos].copy()
            _rb_rows = _rb_rows[_rb_rows["score_composite"].notna() &
                                 _rb_rows["mkt_value"].notna()].copy()

            if not _rb_rows.empty:
                _total_score = _rb_rows["score_composite"].sum()
                _total_mktval = _rb_rows["mkt_value"].sum()

                _rb_rows["target_weight"] = (
                    _rb_rows["score_composite"] / _total_score * 100
                ).round(1)
                _rb_rows["current_weight"] = (
                    _rb_rows["mkt_value"] / _total_mktval * 100
                ).round(1)
                _rb_rows["weight_gap"] = (
                    _rb_rows["target_weight"] - _rb_rows["current_weight"]
                ).round(1)
                _rb_rows["action"] = _rb_rows["weight_gap"].apply(
                    lambda g: "➕ Add" if g > 3 else "➖ Trim" if g < -3 else "✅ Hold"
                )

                # Compute suggested $ to add/trim to reach target (based on total portfolio value)
                _rb_rows["suggested_$"] = (
                    _rb_rows["weight_gap"] / 100 * _total_mktval
                ).round(0)

                _rb_display = (
                    _rb_rows[["name", "ticker", "grade", "score_composite",
                               "current_weight", "target_weight", "weight_gap",
                               "action", "suggested_$"]]
                    .sort_values("weight_gap", ascending=False)
                    .rename(columns={
                        "name": "Company", "ticker": "Ticker",
                        "grade": "Grade", "score_composite": "Score",
                        "current_weight": "Current %",
                        "target_weight": "Target %",
                        "weight_gap": "Gap %",
                        "action": "Action",
                        "suggested_$": "Suggested $",
                    })
                )

                def _color_action(val):
                    if val == "➕ Add":   return "color: #22c55e; font-weight: 700"
                    if val == "➖ Trim":  return "color: #ef4444; font-weight: 700"
                    return "color: #94a3b8"

                def _color_gap(val):
                    if pd.isna(val) or val == 0: return ""
                    return "color: #22c55e; font-weight:600" if val > 0 else "color: #ef4444; font-weight:600"

                _rb_styled = (
                    _rb_display.style
                    .map(_color_action, subset=["Action"])
                    .map(_color_gap,    subset=["Gap %", "Suggested $"])
                    .format({
                        "Score":       "{:.1f}",
                        "Current %":   "{:.1f}%",
                        "Target %":    "{:.1f}%",
                        "Gap %":       "{:+.1f}%",
                        "Suggested $": lambda x: f"${x:+,.0f}" if pd.notna(x) else "—",
                    }, na_rep="—")
                )
                st.dataframe(_rb_styled, width="stretch",
                             height=min(100 + len(_rb_display) * 35, 450))

                # Quick visual: waterfall-style bar chart of gaps
                _rb_chart = _rb_display.sort_values("Gap %")
                _rb_colors = [
                    "#22c55e" if g > 3 else "#ef4444" if g < -3 else "#94a3b8"
                    for g in _rb_chart["Gap %"]
                ]
                fig_rb = go.Figure()
                fig_rb.add_bar(
                    x=_rb_chart["Gap %"],
                    y=_rb_chart["Company"] + " (" + _rb_chart["Ticker"] + ")",
                    orientation="h",
                    marker_color=_rb_colors,
                    text=[f"{g:+.1f}%" for g in _rb_chart["Gap %"]],
                    textposition="outside",
                    customdata=np.column_stack([
                        _rb_chart["Score"],
                        _rb_chart["Current %"],
                        _rb_chart["Target %"],
                    ]),
                    hovertemplate=(
                        "<b>%{y}</b><br>"
                        "Score: %{customdata[0]:.1f}<br>"
                        "Current: %{customdata[1]:.1f}%<br>"
                        "Target: %{customdata[2]:.1f}%<br>"
                        "Gap: %{x:+.1f}%<extra></extra>"
                    ),
                )
                fig_rb.add_vline(x=0, line_color="gray", line_width=1)
                fig_rb.update_layout(
                    title="Portfolio Weight Gap (Target − Current)",
                    xaxis_title="Weight Gap % (positive = underweight)",
                    height=max(280, len(_rb_chart) * 36 + 80),
                    margin=dict(l=0, r=60, t=40, b=30),
                    showlegend=False,
                )
                st.plotly_chart(fig_rb, width="stretch")
                st.caption(
                    f"Total portfolio value: **${_total_mktval:,.0f}** · "
                    f"Threshold: ±3% gap triggers Add/Trim recommendation."
                )

        # ── Benchmark Comparison ───────────────────────────────────────────────
        if _n_positions > 0:
            st.divider()
            st.markdown("#### 📊 Portfolio vs Benchmark")
            st.caption("Compare estimated portfolio cost-basis return against GDX / GDXJ since your first purchase.")

            _bench_opts = {
                "GDX — VanEck Gold Miners":        "GDX",
                "GDXJ — VanEck Junior Gold Miners": "GDXJ",
                "GDX — VanEck Gold Miners ETF": "GDX",
            }
            _bench_sel = st.selectbox(
                "Benchmark", list(_bench_opts.keys()), key="bench_selector"
            )
            _bench_sym = _bench_opts[_bench_sel]

            # Compute portfolio average cost-basis date (use earliest position update as proxy)
            _pos_with_cost = _wl_df[_has_pos & _wl_df["avg_cost"].notna() & (_wl_df["avg_cost"] > 0)]
            if not _pos_with_cost.empty:
                _bench_data = _fetch_benchmark(_bench_sym, period="2y")
                if not _bench_data.empty:
                    _bench_data["date"] = pd.to_datetime(_bench_data["date"])
                    _bench_data = _bench_data.sort_values("date")

                    # Normalise benchmark to 100 at start
                    _bench_start_px = _bench_data.iloc[0]["price"]
                    _bench_data["bench_idx"] = _bench_data["price"] / _bench_start_px * 100

                    # Portfolio return: current mkt value vs total cost basis
                    _port_cost  = _wl_df[_has_pos]["cost_basis"].sum()
                    _port_val   = _wl_df[_has_pos]["mkt_value"].sum()
                    _port_ret   = (_port_val / _port_cost - 1) * 100 if _port_cost > 0 else None

                    fig_bench = go.Figure()
                    fig_bench.add_trace(go.Scatter(
                        x=_bench_data["date"],
                        y=_bench_data["bench_idx"],
                        name=_bench_sel.split("—")[0].strip(),
                        line=dict(color="#94a3b8", width=2, dash="dot"),
                        mode="lines",
                    ))

                    if _port_ret is not None:
                        # Show portfolio as a single endpoint vs benchmark
                        _bench_latest = _bench_data.iloc[-1]["bench_idx"]
                        _port_idx = 100 + _port_ret
                        _port_color = "#22c55e" if _port_idx >= _bench_latest else "#ef4444"
                        # Horizontal reference line for portfolio return
                        fig_bench.add_hline(
                            y=_port_idx,
                            line_color=_port_color, line_dash="solid", line_width=2,
                            annotation_text=f"Portfolio: {_port_ret:+.1f}%",
                            annotation_position="right",
                            annotation_font_color=_port_color,
                        )
                        fig_bench.add_hline(
                            y=100, line_color="gray", line_dash="dash", line_width=1,
                            annotation_text="Cost basis",
                            annotation_position="left",
                        )

                        _outperform = _port_idx - _bench_latest
                        _vs_str = f"{'outperforming' if _outperform >= 0 else 'underperforming'} {_bench_sym} by {abs(_outperform):.1f}pts"
                        st.caption(
                            f"Portfolio return: **{_port_ret:+.1f}%** vs cost basis — {_vs_str}"
                        )

                    fig_bench.update_layout(
                        yaxis_title="Indexed to 100 at start",
                        xaxis_title="Date",
                        height=350,
                        margin=dict(t=10, b=40, l=0, r=100),
                        showlegend=True,
                        legend=dict(orientation="h", y=-0.2),
                    )
                    st.plotly_chart(fig_bench, width="stretch")
                else:
                    st.warning(f"Could not fetch benchmark data for {_bench_sym}.")
            else:
                st.info("Enter average cost for your positions to enable benchmark comparison.")

        # ── Transaction Log ────────────────────────────────────────────────────
        st.divider()
        st.markdown("#### 📋 Transaction Log")

        _all_tx = get_transactions()
        # Filter to watchlist tickers only
        _wl_tx = _all_tx[_all_tx["ticker"].isin(_wl_tickers)] if not _all_tx.empty else _all_tx

        # Add transaction form
        with st.expander("➕ Add Transaction", expanded=False):
            _tx_cols = st.columns([2, 1, 1, 1, 1, 2])
            _tx_tk   = _tx_cols[0].selectbox(
                "Ticker", sorted(_wl_tickers) if _wl_tickers else ["—"],
                key="tx_ticker",
            )
            _tx_type = _tx_cols[1].selectbox("Type", ["buy", "sell"], key="tx_type")
            _tx_date = _tx_cols[2].date_input("Date", key="tx_date")
            _tx_sh   = _tx_cols[3].number_input("Shares", min_value=0.0, step=100.0,
                                                  format="%.0f", key="tx_shares")
            _tx_px   = _tx_cols[4].number_input("Price", min_value=0.0, step=0.01,
                                                  format="%.3f", key="tx_price")
            _tx_note = _tx_cols[5].text_input("Note (optional)", key="tx_note")
            if st.button("💾 Save Transaction", key="tx_save"):
                if _tx_sh > 0 and _tx_px > 0 and _tx_tk not in ("—", ""):
                    add_transaction(_tx_tk, _tx_date, _tx_sh, _tx_px, _tx_type, _tx_note)
                    st.toast(f"Transaction saved: {_tx_type.upper()} {_tx_sh:.0f} {_tx_tk} @ ${_tx_px:.3f}", icon="✅")
                    st.rerun()
                else:
                    st.warning("Fill in ticker, shares and price.")

        if not _wl_tx.empty:
            _tx_display = _wl_tx.rename(columns={
                "id": "ID", "ticker": "Ticker", "trans_date": "Date",
                "shares": "Shares", "price": "Price",
                "trans_type": "Type", "note": "Note",
            })[["ID", "Ticker", "Date", "Type", "Shares", "Price", "Note"]].copy()
            _tx_display["Value"] = (_tx_display["Shares"] * _tx_display["Price"]).round(0)

            _tx_styled = (
                _tx_display.style
                .map(lambda v: "color:#22c55e;font-weight:700" if v == "buy"
                     else "color:#ef4444;font-weight:700" if v == "sell" else "",
                     subset=["Type"])
                .format({
                    "Shares": "{:,.0f}",
                    "Price":  "${:.3f}",
                    "Value":  "${:,.0f}",
                }, na_rep="—")
            )
            st.dataframe(_tx_styled, width="stretch",
                         height=min(100 + len(_tx_display) * 35, 400))

            # Delete a transaction
            _del_id = st.number_input("Delete transaction by ID", min_value=0,
                                       step=1, value=0, key="tx_del_id")
            if st.button("🗑️ Delete", key="tx_del_btn") and _del_id > 0:
                delete_transaction(int(_del_id))
                st.toast(f"Transaction #{_del_id} deleted", icon="🗑️")
                st.rerun()
        else:
            st.info("No transactions recorded yet. Use the form above to log your trades.")

        # ── Import Positions from CSV ──────────────────────────────────────────
        st.divider()
        st.markdown("#### 📥 Import Positions from CSV")
        st.caption(
            "Upload a CSV with columns: `ticker`, `shares`, `avg_cost` (optional: `currency`). "
            "Tickers must match the universe (e.g. ABX.TO, WPM.TO)."
        )
        _upload_col1, _upload_col2 = st.columns([3, 1])
        _uploaded = _upload_col1.file_uploader(
            "Choose CSV", type=["csv"], key="pos_csv_upload",
            label_visibility="collapsed",
        )
        if _uploaded is not None:
            try:
                _csv_df = pd.read_csv(_uploaded)
                _csv_df.columns = [c.strip().lower() for c in _csv_df.columns]
                _required = {"ticker", "shares"}
                if not _required.issubset(set(_csv_df.columns)):
                    st.error(f"CSV must have columns: {_required}. Found: {list(_csv_df.columns)}")
                else:
                    # Validate tickers
                    _valid_tks = set(df["ticker"].tolist())
                    _csv_df["ticker"] = _csv_df["ticker"].str.strip().str.upper()
                    _known   = _csv_df[_csv_df["ticker"].isin(_valid_tks)]
                    _unknown = _csv_df[~_csv_df["ticker"].isin(_valid_tks)]

                    st.markdown(f"**Preview** — {len(_known)} valid / {len(_unknown)} unrecognised rows")
                    st.dataframe(_known[["ticker", "shares",
                                        *[c for c in ["avg_cost", "currency"] if c in _known.columns]]],
                                 height=min(100 + len(_known) * 35, 300))
                    if not _unknown.empty:
                        st.warning(f"Unrecognised tickers (skipped): {', '.join(_unknown['ticker'].tolist())}")

                    if _upload_col2.button("📥 Import", key="pos_import_btn",
                                           use_container_width=True):
                        _imported = 0
                        for _, _ir in _known.iterrows():
                            _i_cost = float(_ir["avg_cost"]) if "avg_cost" in _ir and pd.notna(_ir["avg_cost"]) else None
                            _i_cur  = str(_ir["currency"]) if "currency" in _ir and pd.notna(_ir["currency"]) else config.CURRENCY
                            upsert_position(_ir["ticker"], float(_ir["shares"]),
                                            _i_cost, _i_cur)
                            _imported += 1
                        st.toast(f"Imported {_imported} positions ✓", icon="📥")
                        st.rerun()
            except Exception as _csv_err:
                st.error(f"CSV parse error: {_csv_err}")


# ── TAB 6: SNL Data Browser ────────────────────────────────────────────────────
with tab_snl:
    import sqlite3 as _sqlite3
    import pandas as _pd2

    st.markdown("### SNL Metals & Mining — 本地数据库")

    _snl_db_path = str(config.DB_PATH)

    def _snl_conn():
        c = _sqlite3.connect(_snl_db_path)
        c.row_factory = _sqlite3.Row
        return c

    def _snl_df(sql, params=()):
        with _snl_conn() as _c:
            return _pd2.read_sql_query(sql, _c, params=params)

    # ── Sync log summary ──────────────────────────────────────────────────────
    with _snl_conn() as _sc:
        try:
            _sync_rows = _sc.execute(
                "SELECT table_name, last_sync_at, row_count, status FROM snl_sync_log ORDER BY table_name"
            ).fetchall()
        except Exception:
            _sync_rows = []

    if _sync_rows:
        _total_rows = sum(r["row_count"] or 0 for r in _sync_rows)
        _ok = sum(1 for r in _sync_rows if r["status"] == "ok")
        _last = max((r["last_sync_at"] or "") for r in _sync_rows)[:16]
        st.caption(f"✅ {_ok}/{len(_sync_rows)} tables synced · {_total_rows:,} total rows · last sync {_last} UTC")
    else:
        st.warning("SNL 数据尚未同步。请运行：python snl_sync.py")
        st.stop()

    # ── Mapping ───────────────────────────────────────────────────────────────
    _snl_mapping = {}
    _snl_key2ticker = {}
    _snl_key2name = {}
    try:
        import json as _json2
        with open(os.path.join(os.path.dirname(__file__), "_asx_snl_ticker_mapping.json")) as _mf:
            _snl_mapping = _json2.load(_mf)
        _snl_key2ticker = {str(v["snl_key"]): k for k, v in _snl_mapping.items()}
        _snl_key2name   = {str(v["snl_key"]): v["name"] for k, v in _snl_mapping.items()}
    except Exception:
        pass

    def _add_ticker(df, key_col="snl_key"):
        df.insert(0, "Ticker", df[key_col].map(_snl_key2ticker).fillna("?"))
        df.insert(1, "Company", df[key_col].map(_snl_key2name).fillna(""))
        return df

    # ── Ticker filter ─────────────────────────────────────────────────────────
    _all_tickers = sorted(_snl_mapping.keys())
    _sel_tickers = st.multiselect(
        "筛选公司（留空 = 全部）", _all_tickers,
        default=[], key="snl_ticker_filter", placeholder="选择公司…"
    )
    _sel_keys = [str(_snl_mapping[t]["snl_key"]) for t in _sel_tickers] if _sel_tickers else None

    def _kf():
        if _sel_keys:
            ph = ",".join("?" * len(_sel_keys))
            return f"AND snl_key IN ({ph})", tuple(_sel_keys)
        return "", ()

    def _kfn(col):
        if _sel_keys:
            ph = ",".join("?" * len(_sel_keys))
            return f"AND {col} IN ({ph})", tuple(_sel_keys)
        return "", ()

    # ── Section tabs ──────────────────────────────────────────────────────────
    _s1, _s2, _s3, _s4, _s5, _s6, _s7, _s8, _s9 = st.tabs([
        "生产与成本", "储量资源 R&R", "原位价值",
        "全球排名", "产量预测", "矿权 & 项目",
        "可研/FS 研究", "矿山成本拆解", "勘探预算"
    ])

    # ── 1. 生产与成本 ──────────────────────────────────────────────────────────
    with _s1:
        _c1, _c2, _c3 = st.columns(3)
        _per_opts = [r[0] for r in _snl_conn().execute(
            "SELECT DISTINCT period FROM snl_company_production ORDER BY period DESC").fetchall()]
        _sel_period = _c1.selectbox("期间", _per_opts, key="prod_per")
        _comm_opts  = [r[0] for r in _snl_conn().execute(
            "SELECT DISTINCT commodity FROM snl_company_production ORDER BY commodity").fetchall()]
        _sel_comm   = _c2.selectbox("矿种", ["全部"] + _comm_opts, key="prod_comm")
        _unit       = _c3.selectbox("产量单位", ["oz (贵金属)", "t (公吨)", "lb (磅)"], key="prod_unit")

        _kf_s, _kp_s = _kf()
        _cf  = "AND commodity=?" if _sel_comm != "全部" else ""
        _cps = (_sel_comm,) if _sel_comm != "全部" else ()

        _df = _snl_df(f"""
            SELECT snl_key, period, commodity,
                   prod_oz, prod_t, prod_lb,
                   cash_cost_oz, cash_cost_t, cash_cost_lb,
                   aic_oz, aisc_oz, aisc_t, aisc_lb,
                   realized_price_oz, realized_price_t, realized_price_lb,
                   revenue_m
            FROM snl_company_production
            WHERE period=? {_cf} {_kf_s}
            ORDER BY prod_oz DESC NULLS LAST
        """, (_sel_period,) + _cps + _kp_s)

        if _df.empty:
            st.info("无数据")
        else:
            _add_ticker(_df)
            if _unit == "oz (贵金属)":
                _show_cols = ["Ticker","Company","commodity","prod_oz",
                              "cash_cost_oz","aisc_oz","aic_oz","realized_price_oz","revenue_m"]
                _labels    = ["Ticker","公司","矿种","产量(oz)",
                              "C1($/oz)","AISC($/oz)","AIC($/oz)","实现价格($/oz)","收入($M)"]
            elif _unit == "t (公吨)":
                _show_cols = ["Ticker","Company","commodity","prod_t",
                              "cash_cost_t","aisc_t","realized_price_t","revenue_m"]
                _labels    = ["Ticker","公司","矿种","产量(t)",
                              "C1($/t)","AISC($/t)","实现价格($/t)","收入($M)"]
            else:
                _show_cols = ["Ticker","Company","commodity","prod_lb",
                              "cash_cost_lb","aisc_lb","realized_price_lb","revenue_m"]
                _labels    = ["Ticker","公司","矿种","产量(lb)",
                              "C1($/lb)","AISC($/lb)","实现价格($/lb)","收入($M)"]

            _show = _df[[c for c in _show_cols if c in _df.columns]].copy()
            _show.columns = _labels[:len(_show.columns)]
            st.dataframe(_show, use_container_width=True, hide_index=True,
                         column_config={
                             "产量(oz)": st.column_config.NumberColumn(format="%,.0f"),
                             "产量(t)":  st.column_config.NumberColumn(format="%,.0f"),
                             "产量(lb)": st.column_config.NumberColumn(format="%,.0f"),
                             "收入($M)": st.column_config.NumberColumn(format="$%,.0f"),
                         })

    # ── 2. 储量资源 R&R ────────────────────────────────────────────────────────
    with _s2:
        _c1, _c2 = st.columns(2)
        _comm_opts2 = [r[0] for r in _snl_conn().execute(
            "SELECT DISTINCT commodity FROM snl_company_rr ORDER BY commodity").fetchall()]
        _sel_comm2 = _c1.selectbox("矿种", _comm_opts2,
                                   index=_comm_opts2.index("Gold") if "Gold" in _comm_opts2 else 0,
                                   key="rr_comm")
        _per_opts2 = [r[0] for r in _snl_conn().execute(
            "SELECT DISTINCT period FROM snl_company_rr ORDER BY period DESC LIMIT 12").fetchall()]
        _sel_per2  = _c2.selectbox("期间", _per_opts2, key="rr_per")

        _kf_s, _kp_s = _kf()
        _df = _snl_df(f"""
            SELECT snl_key, period, commodity,
                   grade_gpt, grade_pct,
                   contained_reserves_oz/1000       AS reserves_koz,
                   contained_mi_oz/1000             AS mi_koz,
                   contained_inferred_oz/1000       AS inferred_koz,
                   contained_rr_oz/1000             AS rr_koz,
                   contained_reserves_lb/1e6        AS reserves_mlb,
                   contained_rr_lb/1e6              AS rr_mlb,
                   ore_tonnes_reserves/1e6          AS ore_mt_res,
                   ore_tonnes_rr/1e6                AS ore_mt_rr
            FROM snl_company_rr
            WHERE commodity=? AND period=? {_kf_s}
            ORDER BY contained_rr_oz DESC NULLS LAST
        """, (_sel_comm2, _sel_per2) + _kp_s)

        if _df.empty:
            st.info("无数据")
        else:
            _add_ticker(_df)
            _is_precious = _sel_comm2 in ("Gold","Silver","PGM","Platinum","Palladium")
            if _is_precious:
                _show = _df[["Ticker","Company","grade_gpt",
                             "reserves_koz","mi_koz","inferred_koz","rr_koz","ore_mt_rr"]].copy()
                _show.columns = ["Ticker","公司","品位(g/t)","储量(koz)","M+I(koz)","Inferred(koz)","R&R(koz)","矿石量(Mt)"]
                _cfg = {"储量(koz)": st.column_config.NumberColumn(format="%,.0f"),
                        "M+I(koz)":  st.column_config.NumberColumn(format="%,.0f"),
                        "Inferred(koz)": st.column_config.NumberColumn(format="%,.0f"),
                        "R&R(koz)":  st.column_config.NumberColumn(format="%,.0f")}
            else:
                _show = _df[["Ticker","Company","grade_pct",
                             "reserves_mlb","rr_mlb","ore_mt_res","ore_mt_rr"]].copy()
                _show.columns = ["Ticker","公司","品位(%)","储量(Mlb)","R&R(Mlb)","储量矿石(Mt)","R&R矿石(Mt)"]
                _cfg = {}
            st.dataframe(_show, use_container_width=True, hide_index=True, column_config=_cfg)

    # ── 3. 原位价值 ────────────────────────────────────────────────────────────
    with _s3:
        _kf_s, _kp_s = _kf()
        _df = _snl_df(f"""
            SELECT i.snl_key, i.period,
                   round(i.insitu_reserves_m)  reserves_m,
                   round(i.insitu_mi_m)         mi_m,
                   round(i.insitu_inferred_m)   inferred_m,
                   round(i.insitu_resources_m)  resources_m,
                   round(i.insitu_rr_m)         rr_m
            FROM snl_company_insitu i
            WHERE i.period = (
                SELECT MAX(period) FROM snl_company_insitu i2 WHERE i2.snl_key=i.snl_key
            ) {_kf_s}
            ORDER BY i.insitu_rr_m DESC NULLS LAST
        """, _kp_s)

        if _df.empty:
            st.info("无数据")
        else:
            _add_ticker(_df)
            _show = _df[["Ticker","Company","period","reserves_m","mi_m","inferred_m","resources_m","rr_m"]].copy()
            _show.columns = ["Ticker","公司","期间","储量($M)","M+I($M)","Inferred($M)","总资源($M)","R&R合计($M)"]
            st.dataframe(_show, use_container_width=True, hide_index=True,
                         column_config={c: st.column_config.NumberColumn(format="$%,.0f")
                                        for c in ["储量($M)","M+I($M)","Inferred($M)","总资源($M)","R&R合计($M)"]})

    # ── 4. 全球排名 ────────────────────────────────────────────────────────────
    with _s4:
        _c1, _c2, _c3 = st.columns(3)
        _per_opts4 = [r[0] for r in _snl_conn().execute(
            "SELECT DISTINCT period FROM snl_company_ranking ORDER BY period DESC LIMIT 8").fetchall()]
        _sel_per4  = _c1.selectbox("期间", _per_opts4, key="rank_per")
        _comm_opts4 = [r[0] for r in _snl_conn().execute(
            "SELECT DISTINCT commodity FROM snl_company_ranking ORDER BY commodity").fetchall()]
        _sel_comm4 = _c2.selectbox("矿种", _comm_opts4,
                                   index=_comm_opts4.index("Gold") if "Gold" in _comm_opts4 else 0, key="rank_comm")
        _meth_opts4 = [r[0] for r in _snl_conn().execute(
            "SELECT DISTINCT ownership_method FROM snl_company_ranking").fetchall()]
        _sel_meth4 = _c3.selectbox("持股方式", _meth_opts4,
                                   index=_meth_opts4.index("Controlled") if "Controlled" in _meth_opts4 else 0, key="rank_meth")

        _kf_s, _kp_s = _kf()
        _df = _snl_df(f"""
            SELECT snl_key, global_rank,
                   prod_oz/1000 koz, prod_t/1000 kt, prod_lb/1e6 mlb
            FROM snl_company_ranking
            WHERE period=? AND commodity=? AND ownership_method=? {_kf_s}
            ORDER BY global_rank NULLS LAST
        """, (_sel_per4, _sel_comm4, _sel_meth4) + _kp_s)

        if _df.empty:
            st.info("无数据")
        else:
            _add_ticker(_df)
            _show = _df[["global_rank","Ticker","Company","koz","kt","mlb"]].copy()
            _show.columns = ["全球排名","Ticker","公司","产量(koz)","产量(kt)","产量(Mlb)"]
            st.dataframe(_show, use_container_width=True, hide_index=True,
                         column_config={"全球排名": st.column_config.NumberColumn(format="%d")})

    # ── 5. 产量预测 ────────────────────────────────────────────────────────────
    with _s5:
        _kf_s, _kp_s = _kf()
        _df = _snl_df(f"""
            SELECT snl_key, estimate_period, description,
                   prod_high_oz/1000 hi_koz, prod_low_oz/1000 lo_koz,
                   prod_high_t/1000  hi_kt,  prod_low_t/1000  lo_kt,
                   aisc_high_oz, aisc_low_oz,
                   cash_cost_high_oz, cash_cost_low_oz
            FROM snl_company_projections
            WHERE estimate_period >= '2024' {_kf_s}
            ORDER BY snl_key, estimate_period
        """, _kp_s)
        if _df.empty:
            st.info("无数据")
        else:
            _add_ticker(_df)
            _show = _df[["Ticker","Company","estimate_period","description",
                          "hi_koz","lo_koz","hi_kt","lo_kt",
                          "aisc_high_oz","aisc_low_oz","cash_cost_high_oz","cash_cost_low_oz"]].copy()
            _show.columns = ["Ticker","公司","预测期","描述",
                              "产量上限(koz)","产量下限(koz)","产量上限(kt)","产量下限(kt)",
                              "AISC上限($/oz)","AISC下限($/oz)","C1上限($/oz)","C1下限($/oz)"]
            st.dataframe(_show, use_container_width=True, hide_index=True)

    # ── 6. 矿权 & 项目 ─────────────────────────────────────────────────────────
    with _s6:
        _c1, _c2 = st.columns(2)
        _stage_opts6 = ["全部"] + [r[0] for r in _snl_conn().execute(
            "SELECT DISTINCT stage FROM snl_property_general WHERE stage IS NOT NULL ORDER BY stage").fetchall()]
        _sel_stage6 = _c1.selectbox("开发阶段", _stage_opts6, key="prop_stage")
        _comm_opts6 = ["全部"] + [r[0] for r in _snl_conn().execute(
            "SELECT DISTINCT primary_commodity FROM snl_property_general WHERE primary_commodity IS NOT NULL ORDER BY primary_commodity").fetchall()]
        _sel_comm6 = _c2.selectbox("主矿种", _comm_opts6, key="prop_comm")

        _sf6 = "AND g.stage=?" if _sel_stage6 != "全部" else ""
        _sp6 = (_sel_stage6,) if _sel_stage6 != "全部" else ()
        _cf6 = "AND g.primary_commodity=?" if _sel_comm6 != "全部" else ""
        _cp6 = (_sel_comm6,) if _sel_comm6 != "全部" else ()
        _kf_s6, _kp_s6 = _kfn("o.snl_key")

        _df = _snl_df(f"""
            SELECT o.snl_key, g.property_name, g.stage, g.status,
                   g.primary_commodity, g.country, g.state_province,
                   round(o.pct_own,1) pct_own,
                   c.mill_capacity_tpd, c.actual_startup_year
            FROM snl_property_general g
            JOIN snl_property_owner o ON o.property_id=g.property_id
            LEFT JOIN snl_property_capacity c ON c.property_id=g.property_id
            WHERE 1=1 {_sf6} {_cf6} {_kf_s6}
            ORDER BY o.snl_key, g.stage, g.property_name
        """, _sp6 + _cp6 + _kp_s6)

        if _df.empty:
            st.info("无数据")
        else:
            _add_ticker(_df)
            _show = _df[["Ticker","Company","property_name","stage","status",
                          "primary_commodity","country","state_province",
                          "pct_own","mill_capacity_tpd","actual_startup_year"]].copy()
            _show.columns = ["Ticker","公司","矿权名称","开发阶段","状态",
                              "主矿种","国家","省/州","持股%","产能(t/d)","投产年"]
            st.dataframe(_show, use_container_width=True, hide_index=True,
                         column_config={"产能(t/d)": st.column_config.NumberColumn(format="%,.0f")})
            st.caption(f"共 {len(_show):,} 条矿权记录")

    # ── 7. 可研/FS 研究 ────────────────────────────────────────────────────────
    with _s7:
        _c1, _c2 = st.columns(2)
        _min_year7 = _c1.slider("研究年份 ≥", 2015, 2025, 2020, key="study_yr")
        _type_opts7 = ["全部"] + [r[0] for r in _snl_conn().execute(
            "SELECT DISTINCT study_type FROM snl_property_studies WHERE study_type IS NOT NULL ORDER BY study_type").fetchall()]
        _sel_type7 = _c2.selectbox("研究类型", _type_opts7, key="study_type")
        _tf7 = "AND s.study_type=?" if _sel_type7 != "全部" else ""
        _tp7 = (_sel_type7,) if _sel_type7 != "全部" else ()
        _kf_s7, _kp_s7 = _kfn("o.snl_key")

        _df = _snl_df(f"""
            SELECT o.snl_key, s.property_name, s.study_type, s.study_year,
                   round(s.posttax_npv_m,0)    npv_m,
                   round(s.pretax_npv_m,0)     pretax_m,
                   round(s.posttax_irr_pct,1)  irr,
                   round(s.mine_life_yrs,0)    life_yr,
                   round(s.initial_capex_m,0)  capex_m,
                   round(s.lom_sustaining_m,0) sustaining_m,
                   s.npv_discount_pct, s.currency
            FROM snl_property_studies s
            JOIN snl_property_owner o ON o.property_id=s.property_id
            WHERE s.study_year >= ? {_tf7} {_kf_s7}
            ORDER BY s.posttax_npv_m DESC NULLS LAST
        """, (_min_year7,) + _tp7 + _kp_s7)

        if _df.empty:
            st.info("无数据")
        else:
            _add_ticker(_df)
            _show = _df[["Ticker","Company","property_name","study_type","study_year",
                          "npv_m","irr","life_yr","capex_m","sustaining_m","npv_discount_pct","currency"]].copy()
            _show.columns = ["Ticker","公司","矿权","研究类型","年份",
                              "税后NPV($M)","IRR(%)","寿命(年)","初始资本($M)","LOM持续资本($M)","折现率(%)","币种"]
            st.dataframe(_show, use_container_width=True, hide_index=True,
                         column_config={
                             "税后NPV($M)":    st.column_config.NumberColumn(format="$%,.0f"),
                             "初始资本($M)":   st.column_config.NumberColumn(format="$%,.0f"),
                             "LOM持续资本($M)": st.column_config.NumberColumn(format="$%,.0f"),
                         })

    # ── 8. 矿山成本拆解 ────────────────────────────────────────────────────────
    with _s8:
        _c1, _c2, _c3 = st.columns(3)
        _tbl8 = _c1.radio("数据表", ["贵金属", "贱金属"], horizontal=True, key="econ_tbl")
        _per_opts8 = [r[0] for r in _snl_conn().execute(
            "SELECT DISTINCT period FROM snl_mine_econ_precious ORDER BY period DESC LIMIT 10").fetchall()]
        _sel_per8  = _c2.selectbox("期间", _per_opts8, key="econ_per8")
        _comm_opts8 = ["全部"] + [r[0] for r in _snl_conn().execute(
            f"SELECT DISTINCT commodity FROM snl_mine_econ_precious ORDER BY commodity").fetchall()]
        _sel_comm8 = _c3.selectbox("矿种", _comm_opts8, key="econ_comm8")
        _cf8 = "AND e.commodity=?" if _sel_comm8 != "全部" else ""
        _cp8 = (_sel_comm8,) if _sel_comm8 != "全部" else ()
        _tbl8_name = "snl_mine_econ_precious" if _tbl8 == "贵金属" else "snl_mine_econ_base"
        _kf_s8, _kp_s8 = _kfn("o.snl_key")

        _df = _snl_df(f"""
            SELECT o.snl_key, e.property_name, e.commodity, e.basis,
                   round(e.mine_total_cost,0)   mine_c,
                   round(e.mill_total_cost,0)   mill_c,
                   round(e.byproduct_credits,0) bp_cr,
                   round(e.cash_op_cost,0)       cash_op,
                   round(e.total_cash_cost,0)    c2,
                   round(e.aisc_oz,0)            aisc,
                   round(e.sustaining_capex_oz,0) sust_capex,
                   round(e.commodity_price,2)    spot_price
            FROM {_tbl8_name} e
            JOIN snl_property_owner o ON o.property_id=e.property_id
            WHERE e.period=? {_cf8} {_kf_s8}
            ORDER BY e.aisc_oz NULLS LAST
        """, (_sel_per8,) + _cp8 + _kp_s8)

        if _df.empty:
            st.info("无数据")
        else:
            _add_ticker(_df)
            _show = _df[["Ticker","Company","property_name","commodity","basis",
                          "mine_c","mill_c","bp_cr","cash_op","c2","aisc","sust_capex","spot_price"]].copy()
            _show.columns = ["Ticker","公司","矿权","矿种","计价基准",
                              "矿山成本","选厂成本","副产品抵扣","现金运营成本","C2总成本",
                              "AISC($/oz)","持续资本($/oz)","矿产品价格"]
            st.dataframe(_show, use_container_width=True, hide_index=True)

    # ── 9. 勘探预算 ────────────────────────────────────────────────────────────
    with _s9:
        _kf_s, _kp_s = _kf()
        _yr_opts9 = [r[0] for r in _snl_conn().execute(
            "SELECT DISTINCT fiscal_year FROM snl_exploration_budget ORDER BY fiscal_year DESC").fetchall()]
        _sel_yr9 = st.selectbox("财政年度", ["全部"] + [str(y) for y in _yr_opts9], key="bud_yr")
        _yf9 = "AND fiscal_year=?" if _sel_yr9 != "全部" else ""
        _yp9 = (int(_sel_yr9),) if _sel_yr9 != "全部" else ()

        _df = _snl_df(f"""
            SELECT snl_key, fiscal_year,
                   round(total_budget_m,1) budget_m,
                   round(actual_spent_m,1) spent_m,
                   commodity, company_class
            FROM snl_exploration_budget
            WHERE 1=1 {_yf9} {_kf_s}
            ORDER BY actual_spent_m DESC NULLS LAST
        """, _yp9 + _kp_s)

        if _df.empty:
            st.info("无数据")
        else:
            _add_ticker(_df)
            _show = _df[["Ticker","Company","fiscal_year","budget_m","spent_m","commodity","company_class"]].copy()
            _show.columns = ["Ticker","公司","财政年度","预算($M)","实际支出($M)","勘探矿种","公司规模"]
            st.dataframe(_show, use_container_width=True, hide_index=True,
                         column_config={
                             "预算($M)":    st.column_config.NumberColumn(format="$%.1f"),
                             "实际支出($M)": st.column_config.NumberColumn(format="$%.1f"),
                         })

            try:
                import plotly.express as _px2
                _plot = _df[_df["spent_m"].notna()].copy()
                _plot["Ticker"] = _plot["snl_key"].map(_snl_key2ticker).fillna("?")
                if not _plot.empty:
                    _fig9 = _px2.bar(
                        _plot.sort_values("spent_m", ascending=True).tail(20),
                        x="spent_m", y="Ticker", orientation="h",
                        color="company_class", title="勘探实际支出 Top 20 ($M)",
                        labels={"spent_m": "支出($M)", "Ticker": ""},
                    )
                    _fig9.update_layout(height=480)
                    st.plotly_chart(_fig9, use_container_width=True)
            except Exception:
                pass
